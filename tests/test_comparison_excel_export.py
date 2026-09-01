from __future__ import annotations

from io import BytesIO
import numpy as np
import openpyxl
import pandas as pd
import pytest

from mcdm.comparison_excel import (
    COMPARISON_EXCEL_EXPORT_FILENAME,
    COMPARISON_EXCEL_EXPORT_REVISION,
    build_comparison_excel_workbook,
)


@pytest.fixture
def sample_data():
    matrix = pd.DataFrame(
        [
            [250.0, 16.0, 12.0, 5.0],
            [200.0, 20.0, 8.0, 3.0],
            [300.0, 12.0, 16.0, 4.0],
            [275.0, 14.0, 10.0, 4.5],
        ],
        index=["Alt 1", "Alt 2", "Alt 3", "Alt 4"],
        columns=["Cost", "Quality", "Support", "Reliability"],
    )
    weights = {"Cost": 0.35, "Quality": 0.25, "Support": 0.20, "Reliability": 0.20}
    directions = {
        "Cost": "minimize",
        "Quality": "maximize",
        "Support": "maximize",
        "Reliability": "maximize",
    }
    return matrix, weights, directions


def test_comparison_workbook_structure_and_metadata(sample_data):
    matrix, weights, directions = sample_data
    methods = ["AURA", "MOORA", "TOPSIS", "SAW", "VIKOR", "WASPAS"]

    excel_bytes = build_comparison_excel_workbook(
        matrix,
        weights,
        directions,
        methods=methods,
        benchmark_method="AURA",
    )
    assert excel_bytes.startswith(b"PK")

    wb = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=False)
    assert wb.properties.version == COMPARISON_EXCEL_EXPORT_REVISION
    assert COMPARISON_EXCEL_EXPORT_FILENAME == "mcdm_method_comparison_v1.xlsx"

    # Verify sheet names: primary summary tab first, then every method tab
    assert wb.sheetnames[0] == "Comparison & Agreement"
    for m in methods:
        assert m in wb.sheetnames


def test_comparison_sheet_contains_live_formula_links_and_agreement(sample_data):
    matrix, weights, directions = sample_data
    methods = ["AURA", "MOORA", "TOPSIS", "SAW"]

    excel_bytes = build_comparison_excel_workbook(
        matrix,
        weights,
        directions,
        methods=methods,
        benchmark_method="AURA",
    )
    wb = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=False)
    sheet = wb["Comparison & Agreement"]

    # Check title
    assert sheet["A1"].value == "MCDM Cross-Method Comparison & Agreement Analysis"
    assert sheet["B8"].value == "AURA"  # Benchmark method parameter

    # Check method ranking comparison header and formula links
    assert sheet["A11"].value == "Alternative"
    assert sheet["B11"].value == "AURA"
    assert sheet["C11"].value == "MOORA"
    assert sheet["D11"].value == "TOPSIS"
    assert sheet["E11"].value == "SAW"

    # Check first alternative ranking formulas
    assert sheet["A12"].value == "Alt 1"
    assert sheet["B12"].value.startswith("='AURA'!")
    assert sheet["C12"].value.startswith("='MOORA'!")
    assert sheet["D12"].value.startswith("='TOPSIS'!")
    assert sheet["E12"].value.startswith("='SAW'!")

    # Check Table 3 Agreement section
    # Search for "Agreement of MCDM Methods"
    found_agree = False
    for row in range(1, 40):
        cell_val = str(sheet.cell(row, 1).value or "")
        if "Agreement of MCDM Methods" in cell_val:
            found_agree = True
            header_row = row + 1
            assert sheet.cell(header_row, 1).value == "Method"
            assert sheet.cell(header_row, 2).value == "Spearman ρ"
            assert sheet.cell(header_row, 3).value == "Kendall τ-b"
            assert sheet.cell(header_row, 4).value == "MARD"
            break
    assert found_agree


def test_comparison_workbook_with_all_methods(sample_data):
    matrix, weights, directions = sample_data
    all_methods = ["AURA", "ARAS", "SYAI", "ARIE", "MOORA", "TOPSIS", "SAW", "VIKOR", "WASPAS"]

    excel_bytes = build_comparison_excel_workbook(
        matrix,
        weights,
        directions,
        methods=all_methods,
        benchmark_method="AURA",
        parameters={"alpha": 0.6, "p": 1, "beta": 0.4, "gamma": 0.5, "kappa": 0.5, "lambda": 0.5, "v": 0.5},
    )
    wb = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=False)

    # Every method should have its full calculation sheet
    for m in all_methods:
        assert m in wb.sheetnames

    # Check that there are no RANK.EQ formulas in any sheet
    for name in wb.sheetnames:
        s = wb[name]
        for row in s.iter_rows(values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    assert "RANK.EQ" not in cell.value, f"Found RANK.EQ in sheet {name}, cell {cell.coordinate}"
