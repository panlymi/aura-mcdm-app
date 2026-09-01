from __future__ import annotations

from io import BytesIO

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from mcdm.moora_excel import (
    MOORA_EXCEL_EXPORT_FILENAME,
    MOORA_EXCEL_EXPORT_REVISION,
    build_moora_excel_workbook,
)
from moora_calculator import calculate_moora


MATRIX = pd.DataFrame(
    {
        "Benefit": [9.0, 7.0, 5.0, 7.0],
        "Cost": [2.0, 4.0, 7.0, 4.0],
        "Benefit2": [6.0, 8.0, 5.0, 8.0],
    },
    index=["A1", "A2", "A3", "A4"],
)
WEIGHTS = {"Benefit": 0.4, "Cost": 0.35, "Benefit2": 0.25}
DIRECTIONS = {"Benefit": "maximize", "Cost": "minimize", "Benefit2": "maximize"}


def _find_cell(sheet, value: str):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == value:
                return cell
    raise AssertionError(f"Could not find {value!r} in {sheet.title}")


def _build_workbook():
    content = build_moora_excel_workbook(
        MATRIX,
        WEIGHTS,
        DIRECTIONS,
    )
    return content, load_workbook(BytesIO(content), data_only=False)


def test_export_contains_live_model_summary_verified_values_and_guide():
    content, workbook = _build_workbook()

    assert content.startswith(b"PK")
    assert workbook.sheetnames == [
        "MOORA",
        "Decision Summary",
        "Verified Values",
        "Formula Guide",
    ]
    assert workbook.calculation.calcMode == "auto"
    assert workbook.calculation.fullCalcOnLoad is True
    assert workbook.calculation.forceFullCalc is True
    assert workbook.properties.version == MOORA_EXCEL_EXPORT_REVISION
    assert MOORA_EXCEL_EXPORT_FILENAME == "moora_complete_formula_calculation_v3.xlsx"


def test_every_moora_stage_is_driven_by_live_excel_formulas():
    _, workbook = _build_workbook()
    sheet = workbook["MOORA"]

    orig_title = _find_cell(
        sheet, "Step 1 — Original Decision Matrix and Vector Denominators"
    )
    den_row = orig_title.row + 3 + len(MATRIX)
    assert sheet.cell(den_row, 1).value == "Denominator √(∑x²)"
    assert sheet.cell(den_row, 2).value.startswith("=SQRT(SUMSQ(")

    normalized_title = _find_cell(
        sheet, "Step 2 — Vector (Ratio) Normalized Decision Matrix (x*_ij)"
    )
    normalized_formula = sheet.cell(normalized_title.row + 3, 2).value
    assert normalized_formula.startswith("=IF(ABS(")
    assert str(den_row) in normalized_formula

    assert sheet.cell(orig_title.row + 1, 2).value.startswith("=IF($B$9<=0,0,")
    assert sheet["E14"].value.startswith('=IF($B$9<=0,"Invalid:')
    assert any(
        "B14:B16" in str(validation.sqref)
        for validation in sheet.data_validations.dataValidation
    )

    weighted_title = _find_cell(
        sheet, "Step 3 — Weighted Normalized Decision Matrix (v_ij = w_j × x*_ij)"
    )
    weighted_formula = sheet.cell(weighted_title.row + 3, 2).value
    assert weighted_formula.startswith("=")
    assert "*" in weighted_formula

    result_title = _find_cell(
        sheet,
        "Step 4 — Benefit/Cost Sums, Assessment Value, Rank, and Tie-Safe Sort Order",
    )
    first_result_row = result_title.row + 2
    assert "SUM(" in sheet.cell(first_result_row, 2).value or "=" in sheet.cell(first_result_row, 2).value
    assert sheet.cell(first_result_row, 4).value == f"=B{first_result_row}-C{first_result_row}"
    assert sheet.cell(first_result_row, 5).value.startswith("=RANK(")
    assert "RANK.EQ" not in sheet.cell(first_result_row, 5).value
    assert "COUNTIF" in sheet.cell(first_result_row, 6).value

    formulas = [
        cell.value
        for row in sheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    ]
    assert len(formulas) >= (len(MATRIX) * (len(MATRIX.columns) * 2 + 4))


def test_moora_sorted_ranking_table_matches_canonical_order():
    content, workbook = _build_workbook()
    sheet = workbook["MOORA"]
    results = calculate_moora(MATRIX, WEIGHTS, DIRECTIONS)

    sorted_title = _find_cell(sheet, "Step 5 — Final Ranking (Sorted by Score)")
    first_sorted_row = sorted_title.row + 2
    for index, alternative in enumerate(results.index):
        row = first_sorted_row + index
        assert sheet.cell(row, 1).value.startswith("=INDEX(")
        assert f"MATCH({index + 1}," in sheet.cell(row, 1).value
        assert "$F$" in sheet.cell(row, 1).value
        assert sheet.cell(row, 5).number_format == "0"

    data_only = load_workbook(BytesIO(content), data_only=True)["MOORA"]
    assert data_only.sheet_properties.pageSetUpPr.fitToPage is True


def test_decision_summary_sheet_contains_winner_cards_table_and_chart():
    _, workbook = _build_workbook()
    sheet = workbook["Decision Summary"]

    assert sheet["A4"].value == "First Rank-1 Alternative"
    assert "='MOORA'!A" in str(sheet["A5"].value)
    assert sheet["C4"].value == "Winning Assessment Value"
    assert "='MOORA'!D" in str(sheet["C5"].value)
    assert sheet["E4"].value == "Alternatives"
    assert sheet["G4"].value == "Total Criteria"
    assert sheet["I4"].value == "Benefit Criteria"

    assert len(sheet._charts) == 1
    chart = sheet._charts[0]
    assert chart.type == "col"
    assert len(chart.series) == 1
    assert chart.series[0].val.numRef.f.startswith("'Decision Summary'!$D$")


def test_verified_values_reconciles_canonical_results():
    _, workbook = _build_workbook()
    sheet = workbook["Verified Values"]
    canonical_results, canonical_steps = calculate_moora(
        MATRIX, WEIGHTS, DIRECTIONS, return_steps=True
    )

    raw_title = _find_cell(sheet, "Step 1 — Original Decision Matrix")
    first_raw_row = raw_title.row + 2
    for row_offset, (alt, values) in enumerate(MATRIX.iterrows()):
        row = first_raw_row + row_offset
        assert sheet.cell(row, 1).value == alt
        for col_idx, val in enumerate(values, start=2):
            assert np.isclose(sheet.cell(row, col_idx).value, val)

    norm_title = _find_cell(
        sheet, "Step 2 — Vector (Ratio) Normalized Decision Matrix"
    )
    first_norm_row = norm_title.row + 2
    canonical_norm = canonical_steps["Step 2: Ratio Normalized Matrix ($x^*_{ij}$)"]
    for row_offset, (alt, values) in enumerate(canonical_norm.iterrows()):
        row = first_norm_row + row_offset
        assert sheet.cell(row, 1).value == alt
        for col_idx, val in enumerate(values, start=2):
            assert np.isclose(sheet.cell(row, col_idx).value, val)

    step4_title = _find_cell(
        sheet, "Step 4 — Benefit Sum, Cost Sum, Assessment Value (y_i), and Rank"
    )
    first_step4_row = step4_title.row + 2
    step4_df = canonical_steps["Step 4: Normalized Assessment Value ($y_i$)"]
    for row_offset, (alt, values) in enumerate(step4_df.iterrows()):
        row = first_step4_row + row_offset
        assert sheet.cell(row, 1).value == alt
        assert np.isclose(sheet.cell(row, 2).value, values["Sum (Maximize)"])
        assert np.isclose(sheet.cell(row, 3).value, values["Sum (Minimize)"])
        assert np.isclose(sheet.cell(row, 4).value, values["Assessment Value ($y_i$)"])
        assert sheet.cell(row, 5).value == canonical_results.loc[alt, "Rank"]


def test_user_controlled_labels_are_stored_as_text_not_excel_formulas():
    matrix = pd.DataFrame(
        {"=1+1": [1.0, 2.0], "Normal": [2.0, 1.0]},
        index=["=HYPERLINK(\"https://example.com\")", "Safe"],
    )
    content = build_moora_excel_workbook(
        matrix,
        {"=1+1": 0.5, "Normal": 0.5},
        {"=1+1": "maximize", "Normal": "minimize"},
    )
    workbook = load_workbook(BytesIO(content), data_only=False)
    sheet = workbook["MOORA"]
    raw_title = _find_cell(
        sheet, "Step 1 — Original Decision Matrix and Vector Denominators"
    )
    header_row = raw_title.row + 2
    data_row = raw_title.row + 3

    assert sheet.cell(header_row, 2).value == "=1+1"
    assert sheet.cell(header_row, 2).data_type == "s"
    assert sheet.cell(data_row, 1).value.startswith("=HYPERLINK")
    assert sheet.cell(data_row, 1).data_type == "s"


def test_formula_guide_contains_complete_equations_and_primary_reference():
    _, workbook = _build_workbook()
    sheet = workbook["Formula Guide"]
    values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value]

    assert any("x*_ij = x_ij / √(∑" in str(value) for value in values)
    assert any("y_i = ∑_{j ∈ B} v_ij - ∑_{j ∈ C} v_ij" in str(value) for value in values)
    assert any("Brauers" in str(value) for value in values)
    assert any(str(value).startswith("https://") for value in values)
    assert any("preserves the sign" in str(value) for value in values)


def test_decimal_format_suppresses_trailing_zeros():
    _, workbook = _build_workbook()
    sheet = workbook["MOORA"]
    raw_title = _find_cell(
        sheet, "Step 1 — Original Decision Matrix and Vector Denominators"
    )
    data_cell = sheet.cell(raw_title.row + 3, 2)
    assert data_cell.number_format == "General"
