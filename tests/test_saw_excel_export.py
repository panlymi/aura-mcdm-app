from __future__ import annotations

from io import BytesIO

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from mcdm.saw_excel import (
    SAW_EXCEL_EXPORT_FILENAME,
    SAW_EXCEL_EXPORT_REVISION,
    build_saw_excel_workbook,
)
from saw_calculator import calculate_saw


MATRIX = pd.DataFrame(
    {
        "Benefit": [9.0, 7.0, 5.0, 7.0],
        "Cost": [2.0, 4.0, 7.0, 4.0],
        "Benefit2": [6.0, 8.0, 5.0, 8.0],
    },
    index=["A1", "A2", "A3", "A4"],
)
WEIGHTS = {"Benefit": 0.4, "Cost": 0.35, "Benefit2": 0.25}
DIRECTIONS = {"Benefit": "maximize", "Cost": "minimize", "Benefit2": "maximize"}


def _find_cell(sheet, value: str):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == value:
                return cell
    raise AssertionError(f"Could not find {value!r} in {sheet.title}")


def _build_workbook():
    content = build_saw_excel_workbook(MATRIX, WEIGHTS, DIRECTIONS)
    return content, load_workbook(BytesIO(content), data_only=False)


def test_export_contains_complete_formula_model_and_metadata():
    content, workbook = _build_workbook()

    assert content.startswith(b"PK")
    assert workbook.sheetnames == [
        "SAW",
        "Decision Summary",
        "Verified Values",
        "Formula Guide",
    ]
    assert workbook.calculation.calcMode == "auto"
    assert workbook.calculation.fullCalcOnLoad is True
    assert workbook.calculation.forceFullCalc is True
    assert workbook.properties.version == SAW_EXCEL_EXPORT_REVISION
    assert SAW_EXCEL_EXPORT_FILENAME == "saw_complete_formula_calculation_v1.xlsx"


def test_live_saw_formulas_cover_normalization_weighting_scoring_and_ties():
    _, workbook = _build_workbook()
    sheet = workbook["SAW"]

    raw_title = _find_cell(sheet, "Step 1 — Original Decision Matrix")
    raw_weights_row = raw_title.row + 1
    assert sheet.cell(raw_weights_row, 2).value == "=$C$14"
    assert sheet["C14"].value == "=IF($B$9<=0,0,B14/$B$9)"

    norm_title = _find_cell(sheet, "Step 2 — Ratio-Normalized Decision Matrix (r_ij)")
    norm_formula = sheet.cell(norm_title.row + 3, 2).value
    assert norm_formula.startswith("=IF(ABS(")
    assert "/$E$14" in norm_formula

    weighted_title = _find_cell(
        sheet, "Step 3 — Weighted Normalized Decision Matrix (w_j × r_ij)"
    )
    assert "*B$" in sheet.cell(weighted_title.row + 3, 2).value

    result_title = _find_cell(
        sheet, "Step 4 — SAW Score, Competition Rank, and Tie-Safe Sort Order"
    )
    first_result_row = result_title.row + 2
    assert sheet.cell(first_result_row, 2).value.startswith("=SUM(")
    assert sheet.cell(first_result_row, 3).value.startswith("=RANK.EQ(")
    assert "COUNTIF" in sheet.cell(first_result_row, 4).value

    ranking_title = _find_cell(sheet, "Final Ranking — Sorted by SAW Score")
    first_ranking_row = ranking_title.row + 2
    assert f"MATCH(1,$D${first_result_row}:" in sheet.cell(first_ranking_row, 1).value
    for offset in range(len(MATRIX)):
        assert f"MATCH({offset + 1}," in sheet.cell(first_ranking_row + offset, 1).value


def test_verified_values_match_canonical_saw_calculator():
    _, workbook = _build_workbook()
    sheet = workbook["Verified Values"]
    canonical, steps = calculate_saw(MATRIX, WEIGHTS, DIRECTIONS, return_steps=True)

    normalized_title = _find_cell(sheet, "Step 2 — Ratio-Normalized Decision Matrix")
    first_normalized_row = normalized_title.row + 2
    expected_normalized = steps["Step 2: Normalized Decision Matrix"]
    for row_offset, (_alternative, values) in enumerate(expected_normalized.iterrows()):
        row = first_normalized_row + row_offset
        for column_index, value in enumerate(values, start=2):
            assert np.isclose(sheet.cell(row, column_index).value, value)

    final_title = _find_cell(sheet, "Step 4 — Final SAW Score and Ranking")
    first_final_row = final_title.row + 2
    for row_offset, (alternative, values) in enumerate(canonical.iterrows()):
        row = first_final_row + row_offset
        assert sheet.cell(row, 1).value == alternative
        assert np.isclose(sheet.cell(row, 2).value, values["V_i (SAW Score)"])
        assert sheet.cell(row, 3).value == values["Rank"]


def test_summary_chart_guide_validation_and_formula_injection_protection():
    _, workbook = _build_workbook()
    summary = workbook["Decision Summary"]
    guide = workbook["Formula Guide"]
    model = workbook["SAW"]

    assert summary["A4"].value == "First Rank-1 Alternative"
    assert len(summary._charts) == 1
    assert "10.1287/opre.15.3.537" in " ".join(
        str(cell.value) for row in guide.iter_rows() for cell in row if cell.value
    )
    assert any("B14:B16" in str(validation.sqref) for validation in model.data_validations.dataValidation)

    injected = pd.DataFrame(
        {"=1+1": [2.0, 1.0], "Cost": [1.0, 2.0]},
        index=['=HYPERLINK("https://example.com")', "Safe"],
    )
    content = build_saw_excel_workbook(
        injected,
        {"=1+1": 0.5, "Cost": 0.5},
        {"=1+1": "maximize", "Cost": "minimize"},
    )
    injected_sheet = load_workbook(BytesIO(content), data_only=False)["SAW"]
    raw_title = _find_cell(injected_sheet, "Step 1 — Original Decision Matrix")
    assert injected_sheet.cell(raw_title.row + 2, 2).data_type == "s"
    assert injected_sheet.cell(raw_title.row + 3, 1).data_type == "s"
