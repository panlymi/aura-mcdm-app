from __future__ import annotations

import numpy as np
import pandas as pd
import pytest

from mcdm.weighting import (
    calculate_critic_weights,
    calculate_max_displacement,
    calculate_pca_weights,
    calculate_sd_weights,
    evaluate_weight_robustness,
    generate_deterministic_weight_scenarios,
)


@pytest.fixture
def sample_matrix():
    return pd.DataFrame(
        {
            "C1_Benefit": [9.0, 7.0, 5.0, 8.0],
            "C2_Cost": [2.0, 4.0, 6.0, 3.0],
            "C3_Benefit": [6.0, 8.0, 5.0, 7.0],
        },
        index=["A1", "A2", "A3", "A4"],
    )


@pytest.fixture
def sample_directions():
    return {"C1_Benefit": "maximize", "C2_Cost": "minimize", "C3_Benefit": "maximize"}


def test_critic_weights_sum_to_one(sample_matrix, sample_directions):
    weights, steps = calculate_critic_weights(sample_matrix, sample_directions)
    assert len(weights) == 3
    assert np.isclose(sum(weights.values()), 1.0)
    assert all(w >= 0 for w in weights.values())
    assert "Information Quantity (C_j)" in steps


def test_sd_weights_sum_to_one(sample_matrix, sample_directions):
    weights, steps = calculate_sd_weights(sample_matrix, sample_directions)
    assert len(weights) == 3
    assert np.isclose(sum(weights.values()), 1.0)
    assert all(w >= 0 for w in weights.values())
    assert "Standard Deviations (σ_j)" in steps


def test_pca_weights_sum_to_one(sample_matrix, sample_directions):
    weights, steps = calculate_pca_weights(sample_matrix, sample_directions)
    assert len(weights) == 3
    assert np.isclose(sum(weights.values()), 1.0)
    assert all(w >= 0 for w in weights.values())
    assert "Eigenvalues (Variance)" in steps


def test_max_displacement():
    r1 = pd.Series([1, 2, 3, 4])
    r2 = pd.Series([1, 3, 2, 4])
    assert calculate_max_displacement(r1, r2) == 1
    r3 = pd.Series([4, 3, 2, 1])
    assert calculate_max_displacement(r1, r3) == 3


def test_generate_deterministic_weight_scenarios(sample_matrix, sample_directions):
    baseline = {"C1_Benefit": 0.4, "C2_Cost": 0.35, "C3_Benefit": 0.25}
    scenarios = generate_deterministic_weight_scenarios(
        sample_matrix, sample_directions, baseline_weights=baseline
    )
    assert "Official" in scenarios
    assert "Equal" in scenarios
    assert "CRITIC" in scenarios
    assert "Entropy" in scenarios
    assert "MEREC" in scenarios
    assert "Standard Deviation" in scenarios
    assert "PCA" in scenarios

    for name, w_map in scenarios.items():
        assert np.isclose(sum(w_map.values()), 1.0), f"Scenario {name} weights do not sum to 1.0"


@pytest.mark.parametrize("method", ["AURA", "TOPSIS", "VIKOR", "WASPAS", "SAW", "ARAS", "MOORA"])
def test_evaluate_weight_robustness(method, sample_matrix, sample_directions):
    baseline = {"C1_Benefit": 0.4, "C2_Cost": 0.35, "C3_Benefit": 0.25}
    scenarios = generate_deterministic_weight_scenarios(
        sample_matrix, sample_directions, baseline_weights=baseline
    )
    table_4, rankings_df, weights_df = evaluate_weight_robustness(
        method,
        sample_matrix,
        sample_directions,
        scenarios,
        baseline_scenario="Official",
        top_k=2,
    )
    assert not table_4.empty
    assert "Official" in table_4.index
    assert "Equal" in table_4.index
    # Spearman rho with official for Official scenario is 1.0
    assert np.isclose(table_4.loc["Official", "ρ with official"], 1.0)
    assert np.isclose(table_4.loc["Official", "Top-2 overlap with official"], 1.0)
    assert len(rankings_df) == len(sample_matrix)
    assert len(weights_df) == len(sample_matrix.columns)


def test_build_weight_robustness_excel_workbook(sample_matrix, sample_directions):
    from io import BytesIO
    from openpyxl import load_workbook
    from mcdm.weighting import build_weight_robustness_excel_workbook

    baseline = {"C1_Benefit": 0.4, "C2_Cost": 0.35, "C3_Benefit": 0.25}
    scenarios = generate_deterministic_weight_scenarios(
        sample_matrix, sample_directions, baseline_weights=baseline
    )
    table_4, rankings_df, weights_df = evaluate_weight_robustness(
        "AURA",
        sample_matrix,
        sample_directions,
        scenarios,
        baseline_scenario="Official",
        top_k=2,
    )
    xlsx_bytes = build_weight_robustness_excel_workbook(
        table_4, rankings_df, weights_df, method="AURA", baseline_name="Official"
    )
    assert xlsx_bytes.startswith(b"PK")
    wb = load_workbook(BytesIO(xlsx_bytes))
    assert wb.sheetnames == [
        "Table 4 - Robustness Summary",
        "Rankings by Scenario",
        "Criteria Weights by Scenario",
    ]


def test_build_weight_calculation_excel_workbook(sample_matrix, sample_directions):
    from io import BytesIO
    from openpyxl import load_workbook
    from mcdm.weight_excel import build_weight_calculation_excel_workbook

    baseline = {"C1_Benefit": 0.4, "C2_Cost": 0.35, "C3_Benefit": 0.25}
    xlsx_bytes = build_weight_calculation_excel_workbook(
        sample_matrix, baseline, sample_directions, baseline_name="Official"
    )
    assert xlsx_bytes.startswith(b"PK")
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=False)
    expected_sheets = [
        "Weight Summary",
        "Entropy (EWM)",
        "MEREC",
        "CRITIC",
        "Standard Deviation",
        "PCA Loadings",
    ]
    assert wb.sheetnames == expected_sheets

    # Verify Summary Sheet formulas
    ws_summary = wb["Weight Summary"]
    # Row 6 has C1 formulas
    assert "=" in str(ws_summary.cell(6, 4).value) # Equal weight formula
    assert "=" in str(ws_summary.cell(6, 5).value) # Entropy link formula
    assert "=" in str(ws_summary.cell(6, 6).value) # MEREC link formula
    assert "=" in str(ws_summary.cell(6, 7).value) # CRITIC link formula
    assert "=" in str(ws_summary.cell(6, 8).value) # SD link formula
    assert "=" in str(ws_summary.cell(6, 9).value) # PCA link formula

    # Verify Entropy sheet live formulas
    ws_ewm = wb["Entropy (EWM)"]
    assert any("SUM" in str(cell.value) for row in ws_ewm.iter_rows() for cell in row if str(cell.value).startswith("="))

    # Verify CRITIC sheet CORREL formula
    ws_critic = wb["CRITIC"]
    assert any("CORREL" in str(cell.value) for row in ws_critic.iter_rows() for cell in row if str(cell.value).startswith("="))


