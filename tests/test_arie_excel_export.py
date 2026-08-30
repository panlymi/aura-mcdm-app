from __future__ import annotations

from io import BytesIO

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from arie_calculator import calculate_arie
from mcdm.arie_excel import (
    ARIE_EXCEL_EXPORT_FILENAME,
    ARIE_EXCEL_EXPORT_REVISION,
    build_arie_excel_workbook,
)


MATRIX = pd.DataFrame(
    {
        "Benefit": [9.0, 7.0, 5.0, 7.0],
        "Cost": [2.0, 4.0, 7.0, 4.0],
        "Target": [6.0, 8.0, 5.0, 8.0],
    },
    index=["A1", "A2", "A3", "A4"],
)
WEIGHTS = {"Benefit": 0.4, "Cost": 0.35, "Target": 0.25}
DIRECTIONS = {
    "Benefit": "maximize",
    "Cost": "minimize",
    "Target": {"type": "target", "value": 6.5},
}


def _find_cell(sheet, value: str):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == value:
                return cell
    raise AssertionError(f"Could not find {value!r} in {sheet.title}")


def _build_workbook():
    content = build_arie_excel_workbook(
        MATRIX,
        WEIGHTS,
        DIRECTIONS,
        gamma=1.3,
        kappa=0.35,
    )
    return content, load_workbook(BytesIO(content), data_only=False)


def test_export_contains_live_model_summary_verified_values_and_guide():
    content, workbook = _build_workbook()

    assert content.startswith(b"PK")
    assert workbook.sheetnames == [
        "ARIE",
        "Decision Summary",
        "Verified Values",
        "Formula Guide",
    ]
    assert workbook.calculation.calcMode == "auto"
    assert workbook.calculation.fullCalcOnLoad is True
    assert workbook.calculation.forceFullCalc is True
    assert workbook.properties.version == ARIE_EXCEL_EXPORT_REVISION
    assert ARIE_EXCEL_EXPORT_FILENAME == "arie_complete_formula_calculation_v1.xlsx"


def test_every_arie_stage_is_driven_by_live_excel_formulas():
    _, workbook = _build_workbook()
    sheet = workbook["ARIE"]

    normalized_title = _find_cell(sheet, "Step 2 — Normalized Decision Matrix (r_ij)")
    first_normalized_row = normalized_title.row + 3
    benefit_formula = sheet.cell(first_normalized_row, 2).value
    cost_formula = sheet.cell(first_normalized_row, 3).value
    target_formula = sheet.cell(first_normalized_row, 4).value
    assert benefit_formula.startswith("=IF(")
    assert "$B$7" in benefit_formula
    assert cost_formula.startswith("=")
    assert "/(C" in cost_formula
    assert "+$B$7" in cost_formula
    assert target_formula.startswith("=IF(")
    assert "ABS(" in target_formula

    weighted_title = _find_cell(
        sheet, "Step 3 & 4a — Weighted Matrix and Ideal Solutions"
    )
    weighted_formula = sheet.cell(weighted_title.row + 3, 2).value
    assert weighted_formula.startswith("=")
    assert "*B$" in weighted_formula

    ideal = _find_cell(sheet, "Ideal solution (v_max)")
    anti_ideal = _find_cell(sheet, "Anti-ideal solution (v_min)")
    assert sheet.cell(ideal.row, 2).value.startswith("=MAX(")
    assert sheet.cell(anti_ideal.row, 2).value.startswith("=MIN(")

    best_title = _find_cell(sheet, "Step 4b — Similarity-to-Ideal Contributions")
    worst_title = _find_cell(sheet, "Step 4b — Similarity-to-Anti-Ideal Contributions")
    best_formula = sheet.cell(best_title.row + 2, 2).value
    worst_formula = sheet.cell(worst_title.row + 2, 2).value
    assert best_formula.startswith("=POWER(")
    assert "$B$5" in best_formula
    assert "$B$7" in best_formula
    assert worst_formula.startswith("=POWER(")
    assert "$B$5" in worst_formula
    assert "$B$7" in worst_formula

    results_title = _find_cell(
        sheet, "Step 5 — Relative Closeness, Competition Rank, and Sort Order"
    )
    first_result_row = results_title.row + 2
    assert "$B$6" in sheet.cell(first_result_row, 4).value
    assert "$B$7" in sheet.cell(first_result_row, 4).value
    assert sheet.cell(first_result_row, 5).value.startswith("=RANK(")
    assert "COUNTIF" not in sheet.cell(first_result_row, 5).value
    assert "COUNTIF" in sheet.cell(first_result_row, 6).value

    formulas = [
        cell.value
        for row in sheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    ]
    expected_minimum = len(MATRIX) * len(MATRIX.columns) * 4 + len(MATRIX) * 6
    assert len(formulas) >= expected_minimum


def test_live_ranking_preserves_ties_and_uses_a_separate_sort_helper():
    _, workbook = _build_workbook()
    sheet = workbook["ARIE"]

    sorted_title = _find_cell(sheet, "Final Ranking — Sorted by Score")
    first_rank_row = sorted_title.row + 2
    alternative_formula = sheet.cell(first_rank_row, 1).value
    rank_formula = sheet.cell(first_rank_row, 5).value

    assert alternative_formula.startswith("=INDEX(")
    assert "MATCH(1,$F$" in alternative_formula
    assert rank_formula.startswith("=INDEX($E$")


def test_verified_snapshot_matches_the_canonical_arie_calculator():
    _, workbook = _build_workbook()
    sheet = workbook["Verified Values"]
    expected = calculate_arie(
        MATRIX,
        WEIGHTS,
        DIRECTIONS,
        gamma=1.3,
        kappa=0.35,
    )

    final_title = _find_cell(sheet, "Final Ranking — Sorted by Rank")
    first_data_row = final_title.row + 2
    exported_alternatives = []
    exported_scores = []
    exported_ranks = []
    for row in range(first_data_row, first_data_row + len(expected)):
        exported_alternatives.append(sheet.cell(row, 1).value)
        exported_scores.append(sheet.cell(row, 4).value)
        exported_ranks.append(sheet.cell(row, 5).value)

    assert exported_alternatives == [str(value) for value in expected.index]
    np.testing.assert_allclose(
        exported_scores,
        expected["Relative Closeness (RC_i)"].to_numpy(dtype=float),
        rtol=0,
        atol=1e-15,
    )
    np.testing.assert_array_equal(exported_ranks, expected["Rank"].to_numpy(dtype=int))


def test_verified_similarity_contributions_reconcile_to_exported_totals():
    _, workbook = _build_workbook()
    sheet = workbook["Verified Values"]

    best_title = _find_cell(sheet, "Step 4b — Similarity-to-Ideal Contributions")
    worst_title = _find_cell(sheet, "Step 4b — Similarity-to-Anti-Ideal Contributions")
    totals_title = _find_cell(
        sheet, "Step 5 — Similarity Totals, Relative Closeness, and Rank"
    )
    for row_offset in range(len(MATRIX)):
        best_parts = [
            sheet.cell(best_title.row + 2 + row_offset, column).value
            for column in range(2, 2 + len(MATRIX.columns))
        ]
        worst_parts = [
            sheet.cell(worst_title.row + 2 + row_offset, column).value
            for column in range(2, 2 + len(MATRIX.columns))
        ]
        totals_row = totals_title.row + 2 + row_offset
        np.testing.assert_allclose(sum(best_parts), sheet.cell(totals_row, 2).value)
        np.testing.assert_allclose(sum(worst_parts), sheet.cell(totals_row, 3).value)


def test_decision_summary_is_live_and_contains_a_native_ranking_chart():
    _, workbook = _build_workbook()
    sheet = workbook["Decision Summary"]

    assert sheet["A5"].value.startswith("='ARIE'!")
    assert sheet["C5"].value.startswith("='ARIE'!")
    assert sheet["E5"].value == "='ARIE'!B5"
    assert sheet["G5"].value == "='ARIE'!B6"
    assert "Ideal emphasis" in sheet["I5"].value
    assert len(sheet._charts) == 1
    chart = sheet._charts[0]
    assert len(chart.series) == 1
    assert chart.series[0].val.numRef.f.startswith("'Decision Summary'!$D$")


def test_reference_style_inputs_and_parameter_validation_are_applied():
    _, workbook = _build_workbook()
    sheet = workbook["ARIE"]
    raw_title = _find_cell(sheet, "Step 1 — Original Decision Matrix")
    weight_row = raw_title.row + 1
    header_row = raw_title.row + 2

    assert sheet.cell(weight_row, 1).fill.fgColor.rgb.endswith("F8CBAD")
    assert sheet.cell(header_row, 1).fill.fgColor.rgb.endswith("FFF200")
    assert sheet.cell(header_row, 2).fill.fgColor.rgb.endswith("00B0F0")
    assert sheet.cell(header_row, 3).fill.fgColor.rgb.endswith("FF3B30")
    assert sheet.cell(header_row, 4).fill.fgColor.rgb.endswith("F4B183")
    assert sheet["B5"].fill.fgColor.rgb.endswith("FFF2CC")
    assert sheet["B6"].fill.fgColor.rgb.endswith("FFF2CC")
    assert sheet.sheet_view.showGridLines is False
    assert sheet.freeze_panes == "B1"

    validations = list(sheet.data_validations.dataValidation)
    gamma_validation = next(
        validation for validation in validations if "B5" in str(validation.sqref)
    )
    kappa_validation = next(
        validation for validation in validations if "B6" in str(validation.sqref)
    )
    assert gamma_validation.formula1 == "AND(ISNUMBER(B5),B5>0)"
    assert kappa_validation.formula1 == "0"
    assert kappa_validation.formula2 == "1"


def test_user_controlled_labels_are_stored_as_text_not_excel_formulas():
    matrix = pd.DataFrame(
        {"=1+1": [1.0, 2.0], "Normal": [2.0, 1.0]},
        index=["=HYPERLINK(\"https://example.com\")", "Safe"],
    )
    content = build_arie_excel_workbook(
        matrix,
        {"=1+1": 0.5, "Normal": 0.5},
        {"=1+1": "maximize", "Normal": "minimize"},
    )
    workbook = load_workbook(BytesIO(content), data_only=False)
    sheet = workbook["ARIE"]
    raw_title = _find_cell(sheet, "Step 1 — Original Decision Matrix")
    header_row = raw_title.row + 2
    data_row = raw_title.row + 3

    assert sheet.cell(header_row, 2).value == "=1+1"
    assert sheet.cell(header_row, 2).data_type == "s"
    assert sheet.cell(data_row, 1).value.startswith("=HYPERLINK")
    assert sheet.cell(data_row, 1).data_type == "s"


def test_formula_guide_contains_complete_equations_and_primary_reference():
    _, workbook = _build_workbook()
    sheet = workbook["Formula Guide"]
    values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value]

    assert any("x_j^min / (x_ij + ε)" in str(value) for value in values)
    assert any("κ Sim_i^best" in str(value) for value in values)
    assert "https://www.ejpam.com/index.php/ejpam/article/view/6578" in values
    assert "https://doi.org/10.29020/nybg.ejpam.v18i4.6578" in values
