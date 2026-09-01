from __future__ import annotations

from io import BytesIO

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from mcdm.vikor_excel import (
    VIKOR_EXCEL_EXPORT_FILENAME,
    VIKOR_EXCEL_EXPORT_REVISION,
    build_vikor_excel_workbook,
)
from vikor_calculator import calculate_vikor


MATRIX = pd.DataFrame(
    {
        "Benefit": [9.0, 7.0, 5.0, 7.0],
        "Cost": [2.0, 4.0, 7.0, 4.0],
        "Benefit2": [6.0, 8.0, 5.0, 8.0],
    },
    index=["A1", "A2", "A3", "A4"],
)
WEIGHTS = {"Benefit": 0.4, "Cost": 0.35, "Benefit2": 0.25}
DIRECTIONS = {"Benefit": "maximize", "Cost": "minimize", "Benefit2": "maximize"}
V_PARAM = 0.35


def _find_cell(sheet, value: str):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == value:
                return cell
    raise AssertionError(f"Could not find {value!r} in {sheet.title}")


def _build_workbook():
    content = build_vikor_excel_workbook(
        MATRIX,
        WEIGHTS,
        DIRECTIONS,
        v_param=V_PARAM,
    )
    return content, load_workbook(BytesIO(content), data_only=False)


def test_export_contains_complete_formula_model_and_metadata():
    content, workbook = _build_workbook()

    assert content.startswith(b"PK")
    assert workbook.sheetnames == [
        "VIKOR",
        "Decision Summary",
        "Verified Values",
        "Formula Guide",
    ]
    assert workbook.calculation.calcMode == "auto"
    assert workbook.calculation.fullCalcOnLoad is True
    assert workbook.calculation.forceFullCalc is True
    assert workbook.properties.version == VIKOR_EXCEL_EXPORT_REVISION
    assert VIKOR_EXCEL_EXPORT_FILENAME == "vikor_complete_formula_calculation_v1.xlsx"


def test_live_vikor_formulas_cover_losses_utility_regret_q_and_ties():
    _, workbook = _build_workbook()
    sheet = workbook["VIKOR"]

    assert sheet["E5"].value == V_PARAM
    assert any("E5" in str(validation.sqref) for validation in sheet.data_validations.dataValidation)
    assert sheet["E14"].value.startswith("=MAX(")
    assert sheet["F14"].value.startswith("=MIN(")
    assert sheet["E15"].value.startswith("=MIN(")
    assert sheet["F15"].value.startswith("=MAX(")

    distance_title = _find_cell(
        sheet, "Step 2 — Normalized Distance from the Criterion Best (d_ij)"
    )
    first_distance_row = distance_title.row + 3
    assert sheet.cell(first_distance_row, 2).value.startswith("=IF(ABS(")

    weighted_title = _find_cell(
        sheet, "Step 3 — Weighted Normalized Distance Matrix (w_j × d_ij)"
    )
    assert "*B$" in sheet.cell(weighted_title.row + 3, 2).value

    sr_title = _find_cell(
        sheet, "Step 4 — Group Utility (S_i) and Individual Regret (R_i)"
    )
    first_sr_row = sr_title.row + 2
    assert sheet.cell(first_sr_row, 2).value.startswith("=SUM(")
    assert sheet.cell(first_sr_row, 3).value.startswith("=MAX(")

    result_title = _find_cell(
        sheet,
        "Step 6 — Normalized Utility/Regret Terms, Q Index, Rank, and Sort Order",
    )
    first_result_row = result_title.row + 2
    assert sheet.cell(first_result_row, 4).value.startswith("=IF(ABS(")
    assert sheet.cell(first_result_row, 5).value.startswith("=IF(ABS(")
    assert "*(1-" not in sheet.cell(first_result_row, 6).value
    assert sheet.cell(first_result_row, 6).value.startswith("=$B$")
    assert sheet.cell(first_result_row, 7).value.endswith(",1)")
    assert "COUNTIF" in sheet.cell(first_result_row, 8).value

    ranking_title = _find_cell(sheet, "Step 7 — Final Ranking Sorted by VIKOR Q Index")
    first_ranking_row = ranking_title.row + 2
    for offset in range(len(MATRIX)):
        assert f"MATCH({offset + 1},$H$" in sheet.cell(first_ranking_row + offset, 1).value


def test_verified_values_match_canonical_vikor_calculator():
    _, workbook = _build_workbook()
    sheet = workbook["Verified Values"]
    canonical, steps = calculate_vikor(
        MATRIX,
        WEIGHTS,
        DIRECTIONS,
        v_param=V_PARAM,
        return_steps=True,
    )

    weighted_title = _find_cell(sheet, "Step 4 — Weighted Normalized Distance Matrix")
    first_weighted_row = weighted_title.row + 2
    expected_weighted = steps["Step 3: Weighted Normalized Distance Matrix"]
    for row_offset, (_alternative, values) in enumerate(expected_weighted.iterrows()):
        row = first_weighted_row + row_offset
        for column_index, value in enumerate(values, start=2):
            assert np.isclose(sheet.cell(row, column_index).value, value)

    final_title = _find_cell(sheet, "Step 7 — Final VIKOR Index and Ranking")
    first_final_row = final_title.row + 2
    for row_offset, (alternative, values) in enumerate(canonical.iterrows()):
        row = first_final_row + row_offset
        assert sheet.cell(row, 1).value == alternative
        assert np.isclose(sheet.cell(row, 2).value, values["S_i (Utility)"])
        assert np.isclose(sheet.cell(row, 3).value, values["R_i (Regret)"])
        assert np.isclose(sheet.cell(row, 4).value, values["Q_i (VIKOR Index)"])
        assert sheet.cell(row, 5).value == values["Rank"]


def test_summary_guide_weight_validation_and_formula_injection_protection():
    _, workbook = _build_workbook()
    summary = workbook["Decision Summary"]
    guide = workbook["Formula Guide"]
    model = workbook["VIKOR"]

    assert summary["A4"].value == "First Rank-1 Alternative"
    assert summary["E4"].value == "v (Majority Weight)"
    assert len(summary._charts) == 1
    guide_text = " ".join(
        str(cell.value) for row in guide.iter_rows() for cell in row if cell.value
    )
    assert "10.1016/S0377-2217(03)00020-1" in guide_text
    assert "are not computed" in guide_text
    assert any("B14:B16" in str(validation.sqref) for validation in model.data_validations.dataValidation)

    injected = pd.DataFrame(
        {"=1+1": [1.0, 2.0], "Cost": [2.0, 1.0]},
        index=['=HYPERLINK("https://example.com")', "Safe"],
    )
    content = build_vikor_excel_workbook(
        injected,
        {"=1+1": 0.5, "Cost": 0.5},
        {"=1+1": "maximize", "Cost": "minimize"},
        v_param=0.5,
    )
    injected_sheet = load_workbook(BytesIO(content), data_only=False)["VIKOR"]
    raw_title = _find_cell(injected_sheet, "Step 1 — Original Decision Matrix")
    assert injected_sheet.cell(raw_title.row + 2, 2).data_type == "s"
    assert injected_sheet.cell(raw_title.row + 3, 1).data_type == "s"


def test_v_parameter_rejects_values_outside_closed_unit_interval():
    for invalid in (-0.01, 1.01):
        try:
            build_vikor_excel_workbook(MATRIX, WEIGHTS, DIRECTIONS, v_param=invalid)
        except ValueError as exc:
            assert "between 0 and 1" in str(exc)
        else:
            raise AssertionError(f"Expected v_param={invalid} to be rejected")
