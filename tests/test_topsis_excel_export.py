from __future__ import annotations

from io import BytesIO

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from mcdm.topsis_excel import (
    TOPSIS_EXCEL_EXPORT_FILENAME,
    TOPSIS_EXCEL_EXPORT_REVISION,
    build_topsis_excel_workbook,
)
from topsis_calculator import calculate_topsis


MATRIX = pd.DataFrame(
    {
        "Benefit": [9.0, 7.0, 5.0, 7.0],
        "Cost": [2.0, 4.0, 7.0, 4.0],
        "Benefit2": [6.0, 8.0, 5.0, 8.0],
    },
    index=["A1", "A2", "A3", "A4"],
)
WEIGHTS = {"Benefit": 0.4, "Cost": 0.35, "Benefit2": 0.25}
DIRECTIONS = {"Benefit": "maximize", "Cost": "minimize", "Benefit2": "maximize"}


def _find_cell(sheet, value: str):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == value:
                return cell
    raise AssertionError(f"Could not find {value!r} in {sheet.title}")


def _build_workbook():
    content = build_topsis_excel_workbook(MATRIX, WEIGHTS, DIRECTIONS)
    return content, load_workbook(BytesIO(content), data_only=False)


def test_export_contains_complete_formula_model_and_metadata():
    content, workbook = _build_workbook()

    assert content.startswith(b"PK")
    assert workbook.sheetnames == [
        "TOPSIS",
        "Decision Summary",
        "Verified Values",
        "Formula Guide",
    ]
    assert workbook.calculation.calcMode == "auto"
    assert workbook.calculation.fullCalcOnLoad is True
    assert workbook.calculation.forceFullCalc is True
    assert workbook.properties.version == TOPSIS_EXCEL_EXPORT_REVISION
    assert TOPSIS_EXCEL_EXPORT_FILENAME == "topsis_complete_formula_calculation_v1.xlsx"


def test_live_topsis_formulas_cover_every_stage_and_tie_safe_sorting():
    _, workbook = _build_workbook()
    sheet = workbook["TOPSIS"]

    assert sheet["E14"].value.startswith("=SQRT(SUMSQ(")
    norm_title = _find_cell(sheet, "Step 2 — Vector-Normalized Decision Matrix (r_ij)")
    assert sheet.cell(norm_title.row + 3, 2).value.startswith("=IF(ABS(")

    weighted_title = _find_cell(
        sheet, "Step 3 — Weighted Normalized Decision Matrix (v_ij = w_j × r_ij)"
    )
    assert "*B$" in sheet.cell(weighted_title.row + 3, 2).value

    ideal_title = _find_cell(
        sheet, "Step 4 — Positive-Ideal (A+) and Negative-Ideal (A−) Solutions"
    )
    pis_row = ideal_title.row + 2
    nis_row = ideal_title.row + 3
    assert sheet.cell(pis_row, 2).value.startswith("=MAX(")
    assert sheet.cell(nis_row, 2).value.startswith("=MIN(")
    assert sheet.cell(pis_row, 3).value.startswith("=MIN(")
    assert sheet.cell(nis_row, 3).value.startswith("=MAX(")

    result_title = _find_cell(
        sheet,
        "Step 5 — Separation, Relative Closeness, Competition Rank, and Sort Order",
    )
    first_result_row = result_title.row + 2
    assert sheet.cell(first_result_row, 2).value.startswith("=SQRT(SUM(")
    assert sheet.cell(first_result_row, 3).value.startswith("=SQRT(SUM(")
    assert sheet.cell(first_result_row, 4).value.startswith("=IF(")
    assert sheet.cell(first_result_row, 5).value.startswith("=RANK.EQ(")
    assert "COUNTIF" in sheet.cell(first_result_row, 6).value

    ranking_title = _find_cell(sheet, "Final Ranking — Sorted by Relative Closeness")
    first_ranking_row = ranking_title.row + 2
    for offset in range(len(MATRIX)):
        assert f"MATCH({offset + 1},$F$" in sheet.cell(first_ranking_row + offset, 1).value


def test_verified_values_match_canonical_topsis_calculator():
    _, workbook = _build_workbook()
    sheet = workbook["Verified Values"]
    canonical, steps = calculate_topsis(MATRIX, WEIGHTS, DIRECTIONS, return_steps=True)

    ideal_title = _find_cell(sheet, "Step 4 — Positive-Ideal and Negative-Ideal Solutions")
    first_ideal_row = ideal_title.row + 2
    expected_ideals = steps["Step 4: Ideal and Anti-Ideal Solutions"]
    for row_offset, (_solution, values) in enumerate(expected_ideals.iterrows()):
        row = first_ideal_row + row_offset
        for column_index, value in enumerate(values, start=2):
            assert np.isclose(sheet.cell(row, column_index).value, value)

    final_title = _find_cell(sheet, "Step 6 — Relative Closeness and Final Ranking")
    first_final_row = final_title.row + 2
    for row_offset, (alternative, values) in enumerate(canonical.iterrows()):
        row = first_final_row + row_offset
        assert sheet.cell(row, 1).value == alternative
        assert np.isclose(sheet.cell(row, 2).value, values["D+ (Ideal)"])
        assert np.isclose(sheet.cell(row, 3).value, values["D- (Anti-Ideal)"])
        assert np.isclose(sheet.cell(row, 4).value, values["Relative Closeness (C_i)"])
        assert sheet.cell(row, 5).value == values["Rank"]


def test_summary_guide_validation_and_formula_injection_protection():
    _, workbook = _build_workbook()
    summary = workbook["Decision Summary"]
    guide = workbook["Formula Guide"]
    model = workbook["TOPSIS"]

    assert summary["A4"].value == "First Rank-1 Alternative"
    assert len(summary._charts) == 1
    guide_text = " ".join(
        str(cell.value) for row in guide.iter_rows() for cell in row if cell.value
    )
    assert "10.1007/978-3-642-48318-9" in guide_text
    assert "10.1016/S0377-2217(03)00020-1" in guide_text
    assert any("B14:B16" in str(validation.sqref) for validation in model.data_validations.dataValidation)

    injected = pd.DataFrame(
        {"=1+1": [1.0, 2.0], "Cost": [2.0, 1.0]},
        index=['=HYPERLINK("https://example.com")', "Safe"],
    )
    content = build_topsis_excel_workbook(
        injected,
        {"=1+1": 0.5, "Cost": 0.5},
        {"=1+1": "maximize", "Cost": "minimize"},
    )
    injected_sheet = load_workbook(BytesIO(content), data_only=False)["TOPSIS"]
    raw_title = _find_cell(injected_sheet, "Step 1 — Original Decision Matrix")
    assert injected_sheet.cell(raw_title.row + 2, 2).data_type == "s"
    assert injected_sheet.cell(raw_title.row + 3, 1).data_type == "s"
