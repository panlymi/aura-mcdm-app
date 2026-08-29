from __future__ import annotations

from io import BytesIO

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from aura_calculator import calculate_aura
from mcdm.aura_excel import (
    AURA_EXCEL_EXPORT_FILENAME,
    AURA_EXCEL_EXPORT_REVISION,
    build_aura_excel_workbook,
)


MATRIX = pd.DataFrame(
    {
        "Benefit": [9.0, 7.0, 5.0, 7.0],
        "Cost": [2.0, 4.0, 7.0, 4.0],
        "Target": [6.0, 8.0, 5.0, 8.0],
    },
    index=["A1", "A2", "A3", "A4"],
)
WEIGHTS = {"Benefit": 0.4, "Cost": 0.35, "Target": 0.25}
DIRECTIONS = {
    "Benefit": "maximize",
    "Cost": "minimize",
    "Target": {"type": "target", "value": 6.5},
}


def _find_cell(sheet, value: str):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == value:
                return cell
    raise AssertionError(f"Could not find {value!r} in {sheet.title}")


def _build_workbook():
    content = build_aura_excel_workbook(
        MATRIX,
        WEIGHTS,
        DIRECTIONS,
        alpha=0.4,
        p=2,
    )
    return content, load_workbook(BytesIO(content), data_only=False)


def test_export_contains_formula_model_verified_values_and_formula_guide():
    content, workbook = _build_workbook()

    assert content.startswith(b"PK")
    assert workbook.sheetnames == ["AURA", "Verified Values", "Formula Guide"]
    assert workbook.calculation.calcMode == "auto"
    assert workbook.calculation.fullCalcOnLoad is True
    assert workbook.calculation.forceFullCalc is True
    assert workbook.properties.version == AURA_EXCEL_EXPORT_REVISION
    assert AURA_EXCEL_EXPORT_FILENAME == "aura_complete_formula_calculation_v5.xlsx"


def test_every_aura_stage_is_driven_by_live_excel_formulas():
    _, workbook = _build_workbook()
    sheet = workbook["AURA"]

    normalized_title = _find_cell(sheet, "Step 1 — Normalized Decision Matrix (r_ij)")
    normalized_formula = sheet.cell(normalized_title.row + 3, 2).value
    assert normalized_formula.startswith("=IF(")
    assert "ABS(" in normalized_formula
    assert "MAX(" in normalized_formula
    assert "MIN(" in normalized_formula
    assert "1E-9" in normalized_formula

    weighted_title = _find_cell(
        sheet, "Step 2 & 3 — Weighted Matrix and Ideal Solutions"
    )
    weighted_formula = sheet.cell(weighted_title.row + 3, 2).value
    assert weighted_formula.startswith("=")
    assert "*B$" in weighted_formula

    pis = _find_cell(sheet, "PIS")
    nis = _find_cell(sheet, "NIS")
    average = _find_cell(sheet, "AS (Average)")
    assert sheet.cell(pis.row, 2).value.startswith("=MAX(")
    assert sheet.cell(nis.row, 2).value.startswith("=MIN(")
    assert sheet.cell(average.row, 2).value.startswith("=AVERAGE(")

    distances_title = _find_cell(
        sheet, "Step 4 & 5 — Distances, Correction, Utility Score, and Rank"
    )
    first_result_row = distances_title.row + 4
    assert "SUMPRODUCT" in sheet.cell(first_result_row, 2).value
    assert "ABS" in sheet.cell(first_result_row, 2).value
    assert sheet.cell(first_result_row, 5).value.startswith("=B")
    assert "$C$" in sheet.cell(first_result_row, 5).value
    assert "$B$5" in sheet.cell(first_result_row, 8).value
    rank_formula = sheet.cell(first_result_row, 9).value
    assert rank_formula.startswith("=RANK(")
    assert "RANK.EQ" not in rank_formula
    assert "@" not in rank_formula

    formulas = [
        cell.value
        for row in sheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    ]
    expected_minimum = len(MATRIX) * len(MATRIX.columns) * 2 + len(MATRIX) * 9
    assert len(formulas) >= expected_minimum


def test_verified_value_snapshot_matches_the_canonical_aura_calculator():
    _, workbook = _build_workbook()
    sheet = workbook["Verified Values"]
    expected = calculate_aura(
        MATRIX,
        WEIGHTS,
        DIRECTIONS,
        alpha=0.4,
        p=2,
    )

    final_title = _find_cell(sheet, "Final Ranking — Sorted by Rank")
    first_data_row = final_title.row + 2
    exported_alternatives = []
    exported_utility = []
    exported_ranks = []
    for row in range(first_data_row, first_data_row + len(expected)):
        exported_alternatives.append(sheet.cell(row, 1).value)
        exported_utility.append(sheet.cell(row, 5).value)
        exported_ranks.append(sheet.cell(row, 6).value)

    assert exported_alternatives == [str(value) for value in expected.index]
    np.testing.assert_allclose(
        exported_utility,
        expected["Utility Score"].to_numpy(dtype=float),
        rtol=0,
        atol=1e-15,
    )
    np.testing.assert_array_equal(
        exported_ranks,
        expected["Rank"].to_numpy(dtype=int),
    )


def test_reference_style_is_applied_to_weights_and_criterion_types():
    _, workbook = _build_workbook()
    sheet = workbook["AURA"]
    raw_title = _find_cell(sheet, "Step 0 — Original Decision Matrix")
    weight_row = raw_title.row + 1
    header_row = raw_title.row + 2

    assert sheet.cell(weight_row, 1).fill.fgColor.rgb.endswith("C6EAF8")
    assert sheet.cell(header_row, 1).fill.fgColor.rgb.endswith("FFF200")
    assert sheet.cell(header_row, 2).fill.fgColor.rgb.endswith("FFF200")
    assert sheet.cell(header_row, 3).fill.fgColor.rgb.endswith("FF3B30")
    assert sheet.cell(header_row, 4).fill.fgColor.rgb.endswith("F4B183")
    assert sheet.sheet_view.showGridLines is False
    assert sheet.freeze_panes == "B1"
    assert workbook["Verified Values"].freeze_panes is None


def test_crisp_template_values_fit_without_hash_placeholders():
    matrix = pd.DataFrame(
        {
            "Cost": [20_000.0, 25_000.0, 18_000.0],
            "Quality": [8.0, 9.0, 6.0],
            "Durability": [5.0, 7.0, 4.0],
        },
        index=["Car A", "Car B", "Car C"],
    )
    content = build_aura_excel_workbook(
        matrix,
        {"Cost": 0.3, "Quality": 0.3, "Durability": 0.4},
        {"Cost": "minimize", "Quality": "maximize", "Durability": "maximize"},
    )
    workbook = load_workbook(BytesIO(content), data_only=False)
    sheet = workbook["AURA"]

    assert sheet.column_dimensions["B"].width >= len("25000.000000000") + 2


def test_user_controlled_labels_are_stored_as_text_not_excel_formulas():
    matrix = pd.DataFrame(
        {"=1+1": [1.0, 2.0], "Normal": [2.0, 1.0]},
        index=["=HYPERLINK(\"https://example.com\")", "Safe"],
    )
    content = build_aura_excel_workbook(
        matrix,
        {"=1+1": 0.5, "Normal": 0.5},
        {"=1+1": "maximize", "Normal": "minimize"},
    )
    workbook = load_workbook(BytesIO(content), data_only=False)
    sheet = workbook["AURA"]
    raw_title = _find_cell(sheet, "Step 0 — Original Decision Matrix")
    header_row = raw_title.row + 2
    data_row = raw_title.row + 3

    assert sheet.cell(header_row, 2).value == "=1+1"
    assert sheet.cell(header_row, 2).data_type == "s"
    assert sheet.cell(data_row, 1).value.startswith("=HYPERLINK")
    assert sheet.cell(data_row, 1).data_type == "s"


def test_formula_guide_contains_the_method_reference_and_complete_equations():
    _, workbook = _build_workbook()
    sheet = workbook["Formula Guide"]
    values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value]

    assert any("1 - |x_ij - r_j|" in str(value) for value in values)
    assert any("α(D+_i-D-_i)" in str(value) for value in values)
    assert "https://doi.org/10.1016/j.softx.2025.102395" in values


def test_decimal_format_suppresses_trailing_zeros():
    _, workbook = _build_workbook()
    sheet = workbook["AURA"]
    raw_title = _find_cell(sheet, "Step 0 — Original Decision Matrix")
    data_cell = sheet.cell(raw_title.row + 3, 2)
    assert data_cell.number_format == "0.#########"

