"""Formula-rich Excel export for complete, auditable ARAS calculations."""

from __future__ import annotations

from io import BytesIO
from typing import Any, Mapping

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from aras_calculator import calculate_aras

from .criteria import CriterionPreference, CriterionType, validate_method_capabilities
from .syai_excel import (
    _ALTERNATIVE_FILL,
    _BODY_FONT,
    _CALC_NUMBER_FORMAT,
    _FORMULA_FILL,
    _GRID_BORDER,
    _HEADER_FONT,
    _INPUT_FILL,
    _RAW_NUMBER_FORMAT,
    _SECTION_FILL,
    _SUMMARY_FILL,
    _WEIGHT_FORMAT,
    _WINNER_FILL,
    _WHITE_FILL,
    _autofit_column_widths,
    _preference_fill,
    _preference_label,
    _set_section_title,
    _set_text,
    _set_title,
    _style_grid,
    _style_matrix_header,
    _write_indexed_dataframe,
    _write_live_weights_row,
)
from .validation import validate_crisp_matrix, validate_method_matrix, validate_weights


_NUMERICAL_GUARD = 1e-9

ARAS_EXCEL_EXPORT_REVISION = "v1"
ARAS_EXCEL_EXPORT_FILENAME = (
    f"aras_complete_formula_calculation_{ARAS_EXCEL_EXPORT_REVISION}.xlsx"
)


def _build_formula_sheet(
    workbook: Workbook,
    frame: pd.DataFrame,
    weights: Mapping[str, float],
    preferences: Mapping[str, CriterionPreference],
) -> dict[str, int]:
    sheet = workbook.create_sheet("ARAS")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 82

    columns = [str(column) for column in frame.columns]
    alternatives = [str(index) for index in frame.index]
    criteria_count = len(columns)
    alternatives_count = len(alternatives)
    last_matrix_column = criteria_count + 1
    model_last_column = max(7, last_matrix_column)

    _set_title(
        sheet,
        1,
        model_last_column,
        "ARAS — Complete Formula-Driven Calculation Workbook",
    )
    sheet.merge_cells(
        start_row=2, start_column=1, end_row=2, end_column=model_last_column
    )
    sheet.cell(
        2,
        1,
        "Edit the criterion weights and matrix inputs to explore the decision model. "
        "Every optimal reference value, ratio normalization, weighted component, overall score (S_i), "
        "utility degree (K_i), rank, and sort order is dynamically linked to the live model.",
    )
    sheet.cell(2, 1).font = Font(name="Aptos", size=9, italic=True, color="666666")
    sheet.cell(2, 1).alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[2].height = 30

    settings_title_row = 12
    settings_header_row = 13
    settings_data_start = 14
    settings_data_end = settings_data_start + criteria_count - 1

    _set_section_title(sheet, 4, 5, "Model Parameters and Audit Checks")
    parameter_rows = [
        (5, "Alternatives", alternatives_count, "Number of decision alternatives"),
        (6, "Criteria", criteria_count, "Number of decision criteria"),
        (
            7,
            "Benefit criteria",
            sum(1 for p in preferences.values() if p.kind is CriterionType.BENEFIT),
            "Number of criteria to maximize",
        ),
        (
            8,
            "Cost criteria",
            sum(1 for p in preferences.values() if p.kind is CriterionType.COST),
            "Number of criteria to minimize",
        ),
    ]
    for row, label, value, description in parameter_rows:
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
        sheet.cell(row, 3, description)
    sheet.cell(9, 1, "Weight sum check")
    sheet.cell(9, 2, f"=SUM(B{settings_data_start}:B{settings_data_end})")
    sheet.cell(9, 3, '="Expected 1.0000; current "&TEXT(B9,"0.0000")')
    _style_grid(sheet, 5, 9, 1, 5)
    sheet.cell(5, 2).fill = _FORMULA_FILL
    sheet.cell(6, 2).fill = _FORMULA_FILL
    sheet.cell(7, 2).fill = _FORMULA_FILL
    sheet.cell(8, 2).fill = _FORMULA_FILL
    sheet.cell(9, 2).fill = _SUMMARY_FILL
    sheet.cell(9, 2).number_format = _WEIGHT_FORMAT

    _set_section_title(sheet, settings_title_row, 5, "Criterion Settings")
    headers = [
        "Criterion",
        "Weight (w_j)",
        "Preference",
        "Audit Note",
    ]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(settings_header_row, column_index, header)
        sheet.cell(settings_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(settings_header_row, column_index).font = _HEADER_FONT
    sheet.row_dimensions[settings_header_row].height = 24

    for row_offset, criterion in enumerate(columns):
        row = settings_data_start + row_offset
        preference = preferences[criterion]
        _set_text(sheet, row, 1, criterion)
        sheet.cell(row, 2, float(weights[criterion]))
        sheet.cell(row, 2).number_format = _WEIGHT_FORMAT
        sheet.cell(row, 2).fill = _INPUT_FILL
        _set_text(sheet, row, 3, _preference_label(preference))
        sheet.cell(row, 3).fill = _preference_fill(preference)
        sheet.cell(
            row,
            4,
            f'="Normalized weight: "&TEXT(B{row}/$B$9,"0.0000")',
        )
        sheet.cell(row, 4).font = Font(
            name="Aptos", size=9, italic=True, color="555555"
        )
    _style_grid(
        sheet,
        settings_data_start,
        settings_data_end,
        1,
        4,
        number_format=_CALC_NUMBER_FORMAT,
    )

    # Step 1: Decision Matrix with Optimal Alternative (x_0)
    raw_title_row = settings_data_end + 3
    raw_weights_row = raw_title_row + 1
    raw_header_row = raw_title_row + 2
    raw_opt_row = raw_title_row + 3
    raw_data_start = raw_opt_row + 1
    raw_data_end = raw_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        raw_title_row,
        last_matrix_column,
        "Step 1 — Decision Matrix Extended with Optimal Alternative (x_0)",
    )
    _write_live_weights_row(sheet, raw_weights_row, columns, settings_data_start)
    sheet.cell(raw_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, raw_header_row, column_index, criterion)
    _style_grid(
        sheet,
        raw_header_row,
        raw_data_end,
        1,
        last_matrix_column,
        number_format=_RAW_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, raw_header_row, columns, preferences)

    # Optimal row x_0
    sheet.cell(raw_opt_row, 1, "Optimal (x_0)")
    sheet.cell(raw_opt_row, 1).fill = _SUMMARY_FILL
    sheet.cell(raw_opt_row, 1).font = _HEADER_FONT
    for column_index, criterion in enumerate(columns, start=2):
        letter = get_column_letter(column_index)
        if preferences[criterion].kind is CriterionType.BENEFIT:
            formula = f"=MAX({letter}{raw_data_start}:{letter}{raw_data_end})"
        else:
            formula = f"=MIN({letter}{raw_data_start}:{letter}{raw_data_end})"
        sheet.cell(raw_opt_row, column_index, formula)
        sheet.cell(raw_opt_row, column_index).fill = _SUMMARY_FILL
        sheet.cell(raw_opt_row, column_index).number_format = _RAW_NUMBER_FORMAT

    # Data rows
    for row_offset, (alternative, row_values) in enumerate(frame.iterrows()):
        row = raw_data_start + row_offset
        _set_text(sheet, row, 1, str(alternative))
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index, value in enumerate(row_values, start=2):
            sheet.cell(row, column_index, float(value))
            sheet.cell(row, column_index).number_format = _RAW_NUMBER_FORMAT
            sheet.cell(row, column_index).fill = _WHITE_FILL

    # Step 2: Ratio-Normalized Decision Matrix
    norm_title_row = raw_data_end + 3
    norm_weights_row = norm_title_row + 1
    norm_header_row = norm_title_row + 2
    norm_opt_row = norm_title_row + 3
    norm_data_start = norm_opt_row + 1
    norm_data_end = norm_data_start + alternatives_count - 1
    norm_sum_row = norm_data_end + 1

    _set_section_title(
        sheet,
        norm_title_row,
        last_matrix_column,
        "Step 2 — Ratio-Normalized Decision Matrix (r_ij)",
    )
    _write_live_weights_row(sheet, norm_weights_row, columns, settings_data_start)
    sheet.cell(norm_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, norm_header_row, column_index, criterion)
    _style_grid(
        sheet,
        norm_header_row,
        norm_sum_row,
        1,
        last_matrix_column,
        number_format=_CALC_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, norm_header_row, columns, preferences)

    # Normalization formulas for x_0 and all alternatives
    for row_offset in range(alternatives_count + 1):
        target_row = norm_opt_row + row_offset
        source_row = raw_opt_row + row_offset
        if row_offset == 0:
            sheet.cell(target_row, 1, "Optimal (x_0)")
            sheet.cell(target_row, 1).fill = _SUMMARY_FILL
            sheet.cell(target_row, 1).font = _HEADER_FONT
        else:
            _set_text(sheet, target_row, 1, alternatives[row_offset - 1])
            sheet.cell(target_row, 1).fill = _ALTERNATIVE_FILL

        for column_index, criterion in enumerate(columns, start=2):
            letter = get_column_letter(column_index)
            if preferences[criterion].kind is CriterionType.BENEFIT:
                formula = f"=IF(SUM({letter}${raw_opt_row}:{letter}${raw_data_end})=0,0,{letter}{source_row}/SUM({letter}${raw_opt_row}:{letter}${raw_data_end}))"
            else:
                # Cost: (1/x_ij) / SUM(1/x_kj)
                formula = (
                    f"=IF({letter}{source_row}<=0,0,"
                    f"(1/{letter}{source_row})/SUMPRODUCT(1/{letter}${raw_opt_row}:{letter}${raw_data_end}))"
                )
            sheet.cell(target_row, column_index, formula)
            sheet.cell(target_row, column_index).fill = _FORMULA_FILL
            sheet.cell(target_row, column_index).number_format = _CALC_NUMBER_FORMAT

    # Sum check row
    sheet.cell(norm_sum_row, 1, "Column Sum")
    sheet.cell(norm_sum_row, 1).font = _HEADER_FONT
    sheet.cell(norm_sum_row, 1).fill = _SECTION_FILL
    for column_index in range(2, last_matrix_column + 1):
        letter = get_column_letter(column_index)
        sheet.cell(norm_sum_row, column_index, f"=SUM({letter}{norm_opt_row}:{letter}{norm_data_end})")
        sheet.cell(norm_sum_row, column_index).fill = _SUMMARY_FILL
        sheet.cell(norm_sum_row, column_index).number_format = _CALC_NUMBER_FORMAT

    # Step 3: Weighted Normalized Matrix
    weighted_title_row = norm_sum_row + 3
    weighted_weights_row = weighted_title_row + 1
    weighted_header_row = weighted_title_row + 2
    weighted_opt_row = weighted_title_row + 3
    weighted_data_start = weighted_opt_row + 1
    weighted_data_end = weighted_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        weighted_title_row,
        last_matrix_column,
        "Step 3 — Weighted Normalized Matrix (v_ij = w_j × r_ij)",
    )
    _write_live_weights_row(sheet, weighted_weights_row, columns, settings_data_start)
    sheet.cell(weighted_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, weighted_header_row, column_index, criterion)
    _style_grid(
        sheet,
        weighted_header_row,
        weighted_data_end,
        1,
        last_matrix_column,
        number_format=_CALC_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, weighted_header_row, columns, preferences)

    for row_offset in range(alternatives_count + 1):
        target_row = weighted_opt_row + row_offset
        source_norm_row = norm_opt_row + row_offset
        if row_offset == 0:
            sheet.cell(target_row, 1, "Optimal (x_0)")
            sheet.cell(target_row, 1).fill = _SUMMARY_FILL
            sheet.cell(target_row, 1).font = _HEADER_FONT
        else:
            _set_text(sheet, target_row, 1, alternatives[row_offset - 1])
            sheet.cell(target_row, 1).fill = _ALTERNATIVE_FILL

        for column_index in range(2, last_matrix_column + 1):
            letter = get_column_letter(column_index)
            sheet.cell(
                target_row,
                column_index,
                f"={letter}{source_norm_row}*{letter}${weighted_weights_row}",
            )
            sheet.cell(target_row, column_index).fill = _FORMULA_FILL
            sheet.cell(target_row, column_index).number_format = _CALC_NUMBER_FORMAT

    # Step 4: Overall Performance Value (S_i), Utility Degree (K_i), Rank
    results_title_row = weighted_data_end + 3
    results_header_row = results_title_row + 1
    results_data_start = results_header_row + 1
    results_data_end = results_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        results_title_row,
        5,
        "Step 4 — Overall Performance (S_i), Utility Degree (K_i), and Rank",
    )
    res_headers = [
        "Alternative",
        "Overall Score (S_i)",
        "Utility Degree (K_i)",
        "Rank",
        "Sort Order",
    ]
    for column_index, header in enumerate(res_headers, start=1):
        sheet.cell(results_header_row, column_index, header)
        sheet.cell(results_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(results_header_row, column_index).font = _HEADER_FONT
    sheet.row_dimensions[results_header_row].height = 24

    for row_offset in range(alternatives_count):
        row = results_data_start + row_offset
        w_row = weighted_data_start + row_offset
        _set_text(sheet, row, 1, alternatives[row_offset])
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL

        # S_i = SUM(B{w_row}:last_col{w_row})
        last_letter = get_column_letter(last_matrix_column)
        sheet.cell(row, 2, f"=SUM(B{w_row}:{last_letter}{w_row})")
        # K_i = S_i / S_0
        sheet.cell(row, 3, f"=IF(SUM(B{weighted_opt_row}:{last_letter}{weighted_opt_row})=0,0,B{row}/SUM(B{weighted_opt_row}:{last_letter}{weighted_opt_row}))")
        sheet.cell(
            row,
            4,
            f"=RANK(C{row},$C${results_data_start}:$C${results_data_end},0)",
        )
        sheet.cell(
            row,
            5,
            f"=D{row}+COUNTIF($C${results_data_start}:C{row},C{row})-1",
        )
        sheet.cell(row, 2).fill = _FORMULA_FILL
        sheet.cell(row, 3).fill = _SUMMARY_FILL
        sheet.cell(row, 4).fill = _WHITE_FILL
        sheet.cell(row, 5).fill = _WHITE_FILL
        sheet.cell(row, 4).number_format = "0"
        sheet.cell(row, 5).number_format = "0"

    _style_grid(
        sheet,
        results_data_start,
        results_data_end,
        1,
        5,
        number_format=_CALC_NUMBER_FORMAT,
    )
    for row in range(results_data_start, results_data_end + 1):
        sheet.cell(row, 4).number_format = "0"
        sheet.cell(row, 5).number_format = "0"

    # Step 5: Final Ranking
    ranking_title_row = results_data_end + 3
    ranking_header_row = ranking_title_row + 1
    ranking_data_start = ranking_header_row + 1
    ranking_data_end = ranking_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        ranking_title_row,
        4,
        "Step 5 — Final Ranking (Sorted by Utility Degree K_i)",
    )
    rank_headers = ["Alternative", "Overall Score (S_i)", "Utility Degree (K_i)", "Rank"]
    for column_index, header in enumerate(rank_headers, start=1):
        sheet.cell(ranking_header_row, column_index, header)
        sheet.cell(ranking_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(ranking_header_row, column_index).font = _HEADER_FONT
    _style_grid(
        sheet,
        ranking_data_start,
        ranking_data_end,
        1,
        4,
        number_format=_CALC_NUMBER_FORMAT,
    )
    sheet.row_dimensions[ranking_header_row].height = 31

    for sort_position in range(1, alternatives_count + 1):
        row = ranking_data_start + sort_position - 1
        match_formula = (
            f"MATCH({sort_position},$E${results_data_start}:$E${results_data_end},0)"
        )
        for column_index, source_letter in enumerate(["A", "B", "C", "D"], start=1):
            sheet.cell(
                row,
                column_index,
                f"=INDEX(${source_letter}${results_data_start}:"
                f"${source_letter}${results_data_end},{match_formula})",
            )
            if column_index > 1:
                sheet.cell(row, column_index).fill = _WHITE_FILL
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        sheet.cell(row, 4).number_format = "0"

    sheet.conditional_formatting.add(
        f"C{ranking_data_start}:C{ranking_data_end}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )
    sheet.conditional_formatting.add(
        f"A{ranking_data_start}:D{ranking_data_end}",
        FormulaRule(formula=[f"$D{ranking_data_start}=1"], fill=_WINNER_FILL),
    )

    sheet.freeze_panes = "B1"
    sheet.print_area = f"A1:{get_column_letter(model_last_column)}{ranking_data_end}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    _autofit_column_widths(sheet)

    return {
        "results_data_start": results_data_start,
        "results_data_end": results_data_end,
        "ranking_data_start": ranking_data_start,
        "ranking_data_end": ranking_data_end,
        "rank_column_letter": "D",
        "score_column_letter": "C",
    }


def build_aras_excel_workbook(
    data: pd.DataFrame,
    weights: Mapping[str, Any],
    directions: Mapping[str, Any],
) -> bytes:
    """Return a complete formula-driven ARAS workbook as XLSX bytes."""
    frame = validate_crisp_matrix(data)
    columns = [str(column) for column in frame.columns]
    preferences = validate_method_capabilities("ARAS", columns, directions)
    validate_method_matrix("ARAS", frame, directions)
    normalized_weights = validate_weights(weights, columns, normalize=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.properties.version = ARAS_EXCEL_EXPORT_REVISION
    workbook.properties.creator = "AURA MCDM App"
    workbook.properties.title = "Complete ARAS calculation workbook"

    _build_formula_sheet(workbook, frame, normalized_weights, preferences)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


__all__ = [
    "ARAS_EXCEL_EXPORT_FILENAME",
    "ARAS_EXCEL_EXPORT_REVISION",
    "build_aras_excel_workbook",
    "_build_formula_sheet",
]
