"""Pure, bounded parsing helpers for user-uploaded decision matrices."""

from __future__ import annotations

import csv
import hashlib
from io import BytesIO, TextIOWrapper
from pathlib import Path
from zipfile import BadZipFile, ZipFile

import pandas as pd
from openpyxl import load_workbook

from .validation import MCDMValidationError


MAX_UPLOAD_MB = 10
MAX_UPLOAD_MEBIBYTES = MAX_UPLOAD_MB
MAX_UPLOAD_BYTES = MAX_UPLOAD_MEBIBYTES * 1024 * 1024
MAX_ALTERNATIVES = 500
MAX_CRITERIA = 50
MAX_CELLS = 25_000
# A normal 500 x 50 numeric or short-text workbook expands to only a few MiB.
# This ceiling leaves ample headroom while preventing highly compressed XLSX bombs.
MAX_XLSX_UNCOMPRESSED_MEBIBYTES = 32
MAX_XLSX_UNCOMPRESSED_BYTES = MAX_XLSX_UNCOMPRESSED_MEBIBYTES * 1024 * 1024
SUPPORTED_SUFFIXES = frozenset({".csv", ".xlsx"})


def _coerce_content(content: bytes) -> bytes:
    if not isinstance(content, (bytes, bytearray, memoryview)):
        raise MCDMValidationError("The uploaded file content must be binary data.")
    return bytes(content)


def content_fingerprint(content: bytes) -> str:
    """Return a stable SHA-256 fingerprint of the exact uploaded bytes."""

    return hashlib.sha256(_coerce_content(content)).hexdigest()


def validate_upload_size(content: bytes) -> None:
    """Reject content larger than the application upload ceiling."""

    size = len(_coerce_content(content))
    if size > MAX_UPLOAD_BYTES:
        raise MCDMValidationError(
            f"The uploaded file is too large ({size / (1024 * 1024):.2f} MiB). "
            f"The maximum is {MAX_UPLOAD_MB} MB."
        )


def _validate_dimensions(alternatives: int, criteria: int) -> None:
    if alternatives < 1:
        raise MCDMValidationError(
            "The decision matrix must contain at least one alternative data row."
        )
    if criteria < 1:
        raise MCDMValidationError(
            "The decision matrix must contain at least one criterion column."
        )
    cells = alternatives * criteria
    if cells > MAX_CELLS:
        raise MCDMValidationError(
            f"The decision matrix contains {cells:,} cells; the maximum is {MAX_CELLS:,}."
        )
    if alternatives > MAX_ALTERNATIVES:
        raise MCDMValidationError(
            f"The decision matrix contains {alternatives:,} alternatives; "
            f"the maximum is {MAX_ALTERNATIVES:,}."
        )
    if criteria > MAX_CRITERIA:
        raise MCDMValidationError(
            f"The decision matrix contains {criteria:,} criteria; "
            f"the maximum is {MAX_CRITERIA:,}."
        )


def validate_matrix_limits(data: pd.DataFrame) -> None:
    """Enforce the bounded matrix workload accepted by the public app."""

    if not isinstance(data, pd.DataFrame):
        raise MCDMValidationError("The uploaded decision matrix must be a pandas DataFrame.")
    _validate_dimensions(*data.shape)


def _preflight_csv(content: bytes) -> None:
    with TextIOWrapper(BytesIO(content), encoding="utf-8-sig", newline="") as text_stream:
        reader = csv.reader(text_stream, strict=True)
        try:
            header = next(reader)
        except StopIteration as exc:
            raise csv.Error("The CSV is empty or has no header row.") from exc

        criteria = len(header) - 1
        if criteria < 1:
            _validate_dimensions(1, criteria)
        if criteria > MAX_CRITERIA:
            _validate_dimensions(1, criteria)

        alternatives = 0
        expected_fields = len(header)
        for row in reader:
            if not row:
                continue
            if len(row) != expected_fields:
                raise csv.Error(
                    f"Row {alternatives + 2} has {len(row)} fields; expected {expected_fields}."
                )
            alternatives += 1
            if alternatives > MAX_ALTERNATIVES or alternatives * criteria > MAX_CELLS:
                _validate_dimensions(alternatives, criteria)

        _validate_dimensions(alternatives, criteria)


def _preflight_xlsx(content: bytes) -> None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            expanded_size = sum(member.file_size for member in archive.infolist())
    except BadZipFile:
        raise

    if expanded_size > MAX_XLSX_UNCOMPRESSED_BYTES:
        raise MCDMValidationError(
            f"The XLSX archive expands to {expanded_size / (1024 * 1024):.2f} MiB; "
            f"the safe maximum is {MAX_XLSX_UNCOMPRESSED_MEBIBYTES} MiB."
        )

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        if not workbook.sheetnames:
            raise ValueError("The workbook has no worksheets.")
        first_sheet = workbook[workbook.sheetnames[0]]
        alternatives = max(int(first_sheet.max_row or 0) - 1, 0)
        criteria = max(int(first_sheet.max_column or 0) - 1, 0)
        _validate_dimensions(alternatives, criteria)
    finally:
        workbook.close()


def _friendly_read_error(display_name: str, suffix: str) -> MCDMValidationError:
    if suffix == ".csv":
        guidance = (
            "Use a valid UTF-8 CSV with one header row and alternative names "
            "in the first column."
        )
        format_name = "CSV"
    else:
        guidance = (
            "Use a valid, unencrypted XLSX workbook with one header row and "
            "alternative names in the first column."
        )
        format_name = "XLSX"
    return MCDMValidationError(
        f"Could not read {display_name!r} as {format_name}. {guidance}"
    )


def load_decision_matrix(filename: str, content: bytes) -> pd.DataFrame:
    """Parse a case-insensitive CSV/XLSX upload and enforce public-app limits."""

    raw_content = _coerce_content(content)
    validate_upload_size(raw_content)

    display_name = Path(str(filename)).name
    suffix = Path(display_name).suffix.casefold()
    if suffix not in SUPPORTED_SUFFIXES:
        raise MCDMValidationError(
            "Unsupported file type. Upload a decision matrix with a .csv or .xlsx extension."
        )

    try:
        if suffix == ".csv":
            _preflight_csv(raw_content)
        else:
            _preflight_xlsx(raw_content)
    except MCDMValidationError:
        raise
    except Exception as exc:
        raise _friendly_read_error(display_name, suffix) from exc

    try:
        source = BytesIO(raw_content)
        if suffix == ".csv":
            data = pd.read_csv(
                source,
                index_col=0,
                encoding="utf-8-sig",
                nrows=MAX_ALTERNATIVES + 1,
            )
        else:
            data = pd.read_excel(
                source,
                index_col=0,
                engine="openpyxl",
                nrows=MAX_ALTERNATIVES + 1,
            )
    except Exception as exc:
        raise _friendly_read_error(display_name, suffix) from exc

    validate_matrix_limits(data)
    return data
