"""Formula-rich Excel export for complete, auditable WASPAS calculations."""

from __future__ import annotations

from io import BytesIO
from typing import Any, Mapping

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from waspas_calculator import calculate_waspas

from .criteria import CriterionPreference, CriterionType, validate_method_capabilities
from .syai_excel import (
    _ALTERNATIVE_FILL,
    _BODY_FONT,
    _CALC_NUMBER_FORMAT,
    _FORMULA_FILL,
    _GRID_BORDER,
    _HEADER_FONT,
    _INPUT_FILL,
    _RAW_NUMBER_FORMAT,
    _SECTION_FILL,
    _SUMMARY_FILL,
    _WEIGHT_FORMAT,
    _WINNER_FILL,
    _WHITE_FILL,
    _autofit_column_widths,
    _preference_fill,
    _preference_label,
    _set_section_title,
    _set_text,
    _set_title,
    _style_grid,
    _style_matrix_header,
    _write_indexed_dataframe,
    _write_live_weights_row,
)
from .validation import validate_crisp_matrix, validate_weights


_NUMERICAL_GUARD = 1e-12

# Included in Streamlit's cache key so workbook changes invalidate old bytes.
WASPAS_EXCEL_EXPORT_REVISION = "v2"
WASPAS_EXCEL_EXPORT_FILENAME = (
    f"waspas_complete_formula_calculation_{WASPAS_EXCEL_EXPORT_REVISION}.xlsx"
)


def _build_formula_sheet(
    workbook: Workbook,
    frame: pd.DataFrame,
    weights: Mapping[str, float],
    preferences: Mapping[str, CriterionPreference],
    *,
    lambda_value: float,
) -> dict[str, int]:
    sheet = workbook.create_sheet("WASPAS")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 82

    columns = [str(column) for column in frame.columns]
    alternatives = [str(index) for index in frame.index]
    criteria_count = len(columns)
    alternatives_count = len(alternatives)
    last_matrix_column = criteria_count + 1
    last_matrix_letter = get_column_letter(last_matrix_column)
    model_last_column = max(7, last_matrix_column)

    _set_title(
        sheet,
        1,
        model_last_column,
        "WASPAS — Complete Formula-Driven Calculation Workbook",
    )
    sheet.merge_cells(
        start_row=2, start_column=1, end_row=2, end_column=model_last_column
    )
    sheet.cell(
        2,
        1,
        "Edit the yellow lambda and weight cells to explore the WSM–WPM balance. "
        "Every normalized value, component score, aggregate score, rank, KPI, and "
        "chart is linked to the live model.",
    )
    sheet.cell(2, 1).font = Font(name="Aptos", size=9, italic=True, color="666666")
    sheet.cell(2, 1).alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[2].height = 30

    settings_title_row = 12
    settings_header_row = 13
    settings_data_start = 14
    settings_data_end = settings_data_start + criteria_count - 1

    _set_section_title(sheet, 4, 5, "Model Parameters and Audit Checks")
    parameter_rows = [
        (5, "Lambda (λ)", float(lambda_value), "WSM share; 1 = WSM and 0 = WPM"),
        (6, "Numerical guard", _NUMERICAL_GUARD, "Protects benefit normalization"),
        (7, "Alternatives", alternatives_count, "Number of decision alternatives"),
        (8, "Criteria", criteria_count, "Number of decision criteria"),
    ]
    for row, label, value, description in parameter_rows:
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
        sheet.cell(row, 3, description)
    sheet.cell(9, 1, "Weight sum check")
    sheet.cell(9, 2, f"=SUM(B{settings_data_start}:B{settings_data_end})")
    sheet.cell(9, 3, '="Expected 1.0000; current "&TEXT(B9,"0.0000")')
    _style_grid(sheet, 5, 9, 1, 5)
    sheet.cell(5, 2).fill = _INPUT_FILL
    sheet.cell(5, 2).number_format = "0.00"
    sheet.cell(6, 2).fill = _FORMULA_FILL
    sheet.cell(6, 2).number_format = "0.0E+00"
    sheet.cell(9, 2).fill = _SUMMARY_FILL
    sheet.cell(9, 2).number_format = _WEIGHT_FORMAT
    sheet.cell(5, 2).comment = Comment(
        "Lambda blends the two WASPAS components. Use 1 for pure WSM, 0 for pure "
        "WPM, or 0.5 for the conventional equal blend.",
        "User",
    )
    sheet.cell(6, 2).comment = Comment(
        "This small guard is only used if a benefit-column maximum becomes zero. "
        "The downloaded decision matrix has already passed the app's validation.",
        "User",
    )

    lambda_validation = DataValidation(
        type="decimal",
        operator="between",
        formula1="0",
        formula2="1",
        allow_blank=False,
    )
    lambda_validation.error = "Lambda must be between 0 and 1."
    lambda_validation.errorTitle = "Invalid lambda"
    lambda_validation.prompt = "Enter the WSM share from 0 through 1."
    lambda_validation.promptTitle = "WASPAS lambda"
    lambda_validation.showErrorMessage = True
    lambda_validation.showInputMessage = True
    sheet.add_data_validation(lambda_validation)
    lambda_validation.add(sheet["B5"])

    raw_title_row = settings_data_end + 3
    raw_weights_row = raw_title_row + 1
    raw_header_row = raw_title_row + 2
    raw_data_start = raw_title_row + 3
    raw_data_end = raw_data_start + alternatives_count - 1
    raw_min_row = raw_data_end + 1
    raw_max_row = raw_data_end + 2

    _set_section_title(sheet, settings_title_row, 4, "Criterion Settings")
    settings_headers = ["Criterion", "Normalized Weight", "Preference", "Reference x*"]
    for column_index, header in enumerate(settings_headers, start=1):
        sheet.cell(settings_header_row, column_index, header)
        sheet.cell(settings_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(settings_header_row, column_index).font = _HEADER_FONT
    _style_grid(sheet, settings_header_row, settings_data_end, 1, 4)

    for criterion_index, criterion in enumerate(columns):
        row = settings_data_start + criterion_index
        matrix_letter = get_column_letter(criterion_index + 2)
        preference = preferences[criterion]
        _set_text(sheet, row, 1, criterion)
        sheet.cell(row, 2, float(weights[criterion]))
        sheet.cell(row, 2).number_format = _WEIGHT_FORMAT
        sheet.cell(row, 2).fill = _INPUT_FILL
        sheet.cell(row, 3, _preference_label(preference))
        sheet.cell(row, 3).fill = _preference_fill(preference)
        reference_row = (
            raw_max_row
            if preference.kind is CriterionType.BENEFIT
            else raw_min_row
        )
        sheet.cell(row, 4, f"={matrix_letter}${reference_row}")
        sheet.cell(row, 4).fill = _FORMULA_FILL
        sheet.cell(row, 4).number_format = _RAW_NUMBER_FORMAT

    weight_validation = DataValidation(
        type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=False
    )
    weight_validation.error = "Weights must be non-negative numbers."
    weight_validation.errorTitle = "Invalid weight"
    weight_validation.showErrorMessage = True
    sheet.add_data_validation(weight_validation)
    weight_validation.add(f"B{settings_data_start}:B{settings_data_end}")

    _set_section_title(
        sheet, raw_title_row, last_matrix_column, "Step 1 — Original Decision Matrix"
    )
    _write_live_weights_row(sheet, raw_weights_row, columns, settings_data_start)
    sheet.cell(raw_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, raw_header_row, column_index, criterion)
    _style_grid(
        sheet,
        raw_header_row,
        raw_max_row,
        1,
        last_matrix_column,
        number_format=_RAW_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, raw_header_row, columns, preferences)
    for row_offset, (alternative, values) in enumerate(frame.iterrows()):
        row = raw_data_start + row_offset
        _set_text(sheet, row, 1, alternative)
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index, value in enumerate(values, start=2):
            sheet.cell(row, column_index, float(value))
            sheet.cell(row, column_index).fill = _WHITE_FILL
            sheet.cell(row, column_index).number_format = _RAW_NUMBER_FORMAT
    for summary_row, label, function_name in [
        (raw_min_row, "Minimum", "MIN"),
        (raw_max_row, "Maximum", "MAX"),
    ]:
        sheet.cell(summary_row, 1, label)
        for column_index in range(1, last_matrix_column + 1):
            sheet.cell(summary_row, column_index).fill = _SUMMARY_FILL
            sheet.cell(summary_row, column_index).font = _HEADER_FONT
        for column_index in range(2, last_matrix_column + 1):
            letter = get_column_letter(column_index)
            sheet.cell(
                summary_row,
                column_index,
                f"={function_name}({letter}${raw_data_start}:{letter}${raw_data_end})",
            )

    normalized_title_row = raw_max_row + 3
    normalized_weights_row = normalized_title_row + 1
    normalized_header_row = normalized_title_row + 2
    normalized_data_start = normalized_title_row + 3
    normalized_data_end = normalized_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        normalized_title_row,
        last_matrix_column,
        "Step 2 — Ratio-Normalized Decision Matrix (r_ij)",
    )
    _write_live_weights_row(sheet, normalized_weights_row, columns, settings_data_start)
    sheet.cell(normalized_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, normalized_header_row, column_index, criterion)
    _style_grid(
        sheet,
        normalized_header_row,
        normalized_data_end,
        1,
        last_matrix_column,
        number_format=_CALC_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, normalized_header_row, columns, preferences)
    for row_offset, _alternative in enumerate(alternatives):
        row = normalized_data_start + row_offset
        raw_row = raw_data_start + row_offset
        sheet.cell(row, 1, f"=A{raw_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index in range(2, last_matrix_column + 1):
            letter = get_column_letter(column_index)
            criterion = columns[column_index - 2]
            if preferences[criterion].kind is CriterionType.BENEFIT:
                formula = (
                    f"=IF(ABS({letter}${raw_max_row})<$B$6,0,"
                    f"{letter}{raw_row}/{letter}${raw_max_row})"
                )
            else:
                formula = f"={letter}${raw_min_row}/{letter}{raw_row}"
            sheet.cell(row, column_index, formula)
    sheet.cell(normalized_data_start, 2).comment = Comment(
        "Benefit criteria use x_ij / max(x_j); cost criteria use min(x_j) / x_ij.",
        "User",
    )

    wsm_title_row = normalized_data_end + 3
    wsm_weights_row = wsm_title_row + 1
    wsm_header_row = wsm_title_row + 2
    wsm_data_start = wsm_title_row + 3
    wsm_data_end = wsm_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        wsm_title_row,
        last_matrix_column,
        "Step 3 — Weighted Sum Components (w_j × r_ij)",
    )
    _write_live_weights_row(sheet, wsm_weights_row, columns, settings_data_start)
    sheet.cell(wsm_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, wsm_header_row, column_index, criterion)
    _style_grid(
        sheet,
        wsm_header_row,
        wsm_data_end,
        1,
        last_matrix_column,
        number_format=_CALC_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, wsm_header_row, columns, preferences)
    for row_offset, _alternative in enumerate(alternatives):
        row = wsm_data_start + row_offset
        normalized_row = normalized_data_start + row_offset
        sheet.cell(row, 1, f"=A{normalized_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index in range(2, last_matrix_column + 1):
            letter = get_column_letter(column_index)
            sheet.cell(
                row,
                column_index,
                f"={letter}{normalized_row}*{letter}${wsm_weights_row}",
            )

    wpm_title_row = wsm_data_end + 3
    wpm_weights_row = wpm_title_row + 1
    wpm_header_row = wpm_title_row + 2
    wpm_data_start = wpm_title_row + 3
    wpm_data_end = wpm_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        wpm_title_row,
        last_matrix_column,
        "Step 4 — Weighted Product Components (r_ij ^ w_j)",
    )
    _write_live_weights_row(sheet, wpm_weights_row, columns, settings_data_start)
    sheet.cell(wpm_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, wpm_header_row, column_index, criterion)
    _style_grid(
        sheet,
        wpm_header_row,
        wpm_data_end,
        1,
        last_matrix_column,
        number_format=_CALC_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, wpm_header_row, columns, preferences)
    for row_offset, _alternative in enumerate(alternatives):
        row = wpm_data_start + row_offset
        normalized_row = normalized_data_start + row_offset
        sheet.cell(row, 1, f"=A{normalized_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index in range(2, last_matrix_column + 1):
            letter = get_column_letter(column_index)
            sheet.cell(
                row,
                column_index,
                f"=IF({letter}${wpm_weights_row}=0,1,"
                f"POWER({letter}{normalized_row},{letter}${wpm_weights_row}))",
            )

    results_title_row = wpm_data_end + 3
    results_header_row = results_title_row + 1
    results_data_start = results_title_row + 2
    results_data_end = results_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        results_title_row,
        6,
        "Step 5 — WASPAS Aggregation, Competition Rank, and Sort Order",
    )
    result_headers = [
        "Alternative",
        "Q_i (WSM)",
        "Q_i (WPM)",
        "Q_i (WASPAS Score)",
        "Rank",
        "Sort Order",
    ]
    for column_index, header in enumerate(result_headers, start=1):
        sheet.cell(results_header_row, column_index, header)
        sheet.cell(results_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(results_header_row, column_index).font = _HEADER_FONT
    _style_grid(
        sheet,
        results_header_row,
        results_data_end,
        1,
        6,
        number_format=_CALC_NUMBER_FORMAT,
    )
    sheet.row_dimensions[results_header_row].height = 31
    for row_offset, _alternative in enumerate(alternatives):
        row = results_data_start + row_offset
        wsm_row = wsm_data_start + row_offset
        wpm_row = wpm_data_start + row_offset
        sheet.cell(row, 1, f"=A{wsm_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        sheet.cell(row, 2, f"=SUM(B{wsm_row}:{last_matrix_letter}{wsm_row})")
        sheet.cell(row, 3, f"=PRODUCT(B{wpm_row}:{last_matrix_letter}{wpm_row})")
        sheet.cell(row, 4, f"=$B$5*B{row}+(1-$B$5)*C{row}")
        sheet.cell(
            row,
            5,
            f"=RANK(D{row},$D${results_data_start}:$D${results_data_end},0)",
        )
        sheet.cell(
            row,
            6,
            f"=E{row}+COUNTIF($D${results_data_start}:D{row},D{row})-1",
        )
        sheet.cell(row, 5).number_format = "0"
        sheet.cell(row, 6).number_format = "0"
    sheet.cell(results_data_start, 4).comment = Comment(
        "Higher WASPAS scores are preferred. Lambda is read from B5; tied scores "
        "share a competition rank.",
        "User",
    )
    sheet.conditional_formatting.add(
        f"D{results_data_start}:D{results_data_end}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )
    sheet.conditional_formatting.add(
        f"A{results_data_start}:F{results_data_end}",
        FormulaRule(formula=[f"$E{results_data_start}=1"], fill=_WINNER_FILL),
    )

    ranking_title_row = results_data_end + 3
    ranking_header_row = ranking_title_row + 1
    ranking_data_start = ranking_title_row + 2
    ranking_data_end = ranking_data_start + alternatives_count - 1
    _set_section_title(sheet, ranking_title_row, 5, "Final Ranking — Sorted by Score")
    ranking_headers = result_headers[:5]
    for column_index, header in enumerate(ranking_headers, start=1):
        sheet.cell(ranking_header_row, column_index, header)
        sheet.cell(ranking_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(ranking_header_row, column_index).font = _HEADER_FONT
    _style_grid(
        sheet,
        ranking_header_row,
        ranking_data_end,
        1,
        5,
        number_format=_CALC_NUMBER_FORMAT,
    )
    sheet.row_dimensions[ranking_header_row].height = 31
    for sort_position in range(1, alternatives_count + 1):
        row = ranking_data_start + sort_position - 1
        match_formula = (
            f"MATCH({sort_position},$F${results_data_start}:$F${results_data_end},0)"
        )
        for column_index, source_letter in enumerate(["A", "B", "C", "D", "E"], start=1):
            sheet.cell(
                row,
                column_index,
                f"=INDEX(${source_letter}${results_data_start}:"
                f"${source_letter}${results_data_end},{match_formula})",
            )
            if column_index > 1:
                sheet.cell(row, column_index).fill = _WHITE_FILL
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        sheet.cell(row, 5).number_format = "0"
    sheet.conditional_formatting.add(
        f"D{ranking_data_start}:D{ranking_data_end}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )
    sheet.conditional_formatting.add(
        f"A{ranking_data_start}:E{ranking_data_end}",
        FormulaRule(formula=[f"$E{ranking_data_start}=1"], fill=_WINNER_FILL),
    )

    sheet.freeze_panes = "B1"
    sheet.print_area = f"A1:{get_column_letter(model_last_column)}{ranking_data_end}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.oddFooter.center.text = "WASPAS complete formula model"
    sheet.oddFooter.right.text = "Page &P of &N"
    sheet.column_dimensions["A"].width = max(
        25, min(38, max(len(value) for value in alternatives) + 3)
    )
    for column_index, criterion in enumerate(columns, start=2):
        letter = get_column_letter(column_index)
        raw_value_width = max(len(f"{float(value):.9f}") for value in frame[criterion])
        sheet.column_dimensions[letter].width = max(
            16, min(30, max(len(criterion) + 3, raw_value_width + 2))
        )
    for row in range(1, ranking_data_end + 1):
        if sheet.row_dimensions[row].height is None:
            sheet.row_dimensions[row].height = 19
    _autofit_column_widths(sheet)

    return {
        "ranking_data_start": ranking_data_start,
        "ranking_data_end": ranking_data_end,
    }


def _build_summary_sheet(
    workbook: Workbook,
    formula_positions: Mapping[str, int],
    alternatives_count: int,
) -> None:
    sheet = workbook.create_sheet("Decision Summary")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 90
    _set_title(sheet, 1, 14, "WASPAS Decision Summary")
    sheet.merge_cells("A2:N2")
    sheet["A2"] = (
        "A live view of the blended WSM–WPM ranking. Change lambda or weights on "
        "the WASPAS sheet and Excel will refresh the KPIs, table, and chart."
    )
    sheet["A2"].font = Font(name="Aptos", size=9, italic=True, color="666666")
    sheet["A2"].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[2].height = 29

    ranking_start = formula_positions["ranking_data_start"]
    cards = [
        ("A", "B", "Winner", f"='WASPAS'!A{ranking_start}", "@"),
        ("C", "D", "Winning score", f"='WASPAS'!D{ranking_start}", "0.0000"),
        ("E", "F", "Lambda (λ)", "='WASPAS'!B5", "0.00"),
        (
            "G",
            "H",
            "Blend stance",
            '=IF(\'WASPAS\'!B5>0.5,"WSM emphasis",IF(\'WASPAS\'!B5<0.5,'
            '"WPM emphasis","Balanced"))',
            "@",
        ),
        ("I", "J", "Alternatives", "='WASPAS'!B7", "0"),
    ]
    for start_letter, end_letter, label, formula, number_format in cards:
        sheet.merge_cells(f"{start_letter}4:{end_letter}4")
        sheet.merge_cells(f"{start_letter}5:{end_letter}6")
        label_cell = sheet[f"{start_letter}4"]
        value_cell = sheet[f"{start_letter}5"]
        label_cell.value = label
        value_cell.value = formula
        label_cell.fill = _SECTION_FILL
        value_cell.fill = _SUMMARY_FILL
        label_cell.font = _HEADER_FONT
        value_cell.font = Font(name="Aptos Display", size=15, bold=True, color="0B4F6C")
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        value_cell.number_format = number_format
        for row in range(4, 7):
            for column in range(
                sheet[start_letter + "1"].column,
                sheet[end_letter + "1"].column + 1,
            ):
                sheet.cell(row, column).border = _GRID_BORDER

    summary_header_row = 9
    summary_data_start = 10
    summary_data_end = summary_data_start + alternatives_count - 1
    headers = ["Alternative", "Q_i (WSM)", "Q_i (WPM)", "WASPAS Score", "Rank"]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(summary_header_row, column_index, header)
        sheet.cell(summary_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(summary_header_row, column_index).font = _HEADER_FONT
    _style_grid(
        sheet,
        summary_header_row,
        summary_data_end,
        1,
        5,
        number_format=_CALC_NUMBER_FORMAT,
    )
    for row_offset in range(alternatives_count):
        row = summary_data_start + row_offset
        source_row = ranking_start + row_offset
        for column_index, source_letter in enumerate(["A", "B", "C", "D", "E"], start=1):
            sheet.cell(row, column_index, f"='WASPAS'!{source_letter}{source_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        sheet.cell(row, 5).number_format = "0"
    sheet.conditional_formatting.add(
        f"D{summary_data_start}:D{summary_data_end}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )
    sheet.conditional_formatting.add(
        f"A{summary_data_start}:E{summary_data_end}",
        FormulaRule(formula=[f"$E{summary_data_start}=1"], fill=_WINNER_FILL),
    )

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "WASPAS score by alternative"
    chart.x_axis.title = "Alternative"
    chart.y_axis.title = "WASPAS score"
    chart.height = 7.2
    chart.width = 14.2
    chart.legend = None
    data = Reference(
        sheet, min_col=4, min_row=summary_header_row, max_row=summary_data_end
    )
    categories = Reference(
        sheet, min_col=1, min_row=summary_data_start, max_row=summary_data_end
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.numFmt = "0.000"
    chart.y_axis.numFmt = "0.00"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    sheet.add_chart(chart, "G9")

    notes_row = max(summary_data_end + 3, 27)
    _set_section_title(sheet, notes_row, 10, "How to read this summary")
    notes = [
        "Higher WASPAS scores indicate stronger alternatives.",
        "Lambda = 1 gives pure WSM; lambda = 0 gives pure WPM.",
        "Lambda = 0.5 gives the conventional equal blend of the two components.",
        "Tied scores share a competition rank; sort order only controls display order.",
        "Verified Values preserves the canonical calculation at download time.",
    ]
    for row_offset, note in enumerate(notes, start=1):
        sheet.merge_cells(
            start_row=notes_row + row_offset,
            start_column=1,
            end_row=notes_row + row_offset,
            end_column=10,
        )
        sheet.cell(notes_row + row_offset, 1, f"• {note}")
        sheet.cell(notes_row + row_offset, 1).font = _BODY_FONT
        sheet.cell(notes_row + row_offset, 1).alignment = Alignment(vertical="center")

    widths = {
        "A": 28,
        "B": 17,
        "C": 17,
        "D": 19,
        "E": 12,
        "F": 3,
        "G": 14,
        "H": 14,
        "I": 18,
        "J": 18,
        "K": 14,
        "L": 14,
        "M": 14,
        "N": 14,
    }
    for letter, width in widths.items():
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A10"
    sheet.print_area = f"A1:N{notes_row + len(notes)}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def _build_verified_values_sheet(
    workbook: Workbook,
    frame: pd.DataFrame,
    weights: Mapping[str, float],
    preferences: Mapping[str, CriterionPreference],
    results: pd.DataFrame,
    steps: Mapping[str, Any],
    *,
    lambda_value: float,
) -> None:
    sheet = workbook.create_sheet("Verified Values")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 88
    max_columns = max(len(frame.columns) + 1, 6)
    _set_title(sheet, 1, max_columns, "WASPAS — Verified Numerical Values")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_columns)
    sheet.cell(
        2,
        1,
        "Static values from the canonical Python implementation at download time. "
        "Use them to reconcile every live Excel calculation stage.",
    )
    sheet.cell(2, 1).font = Font(name="Aptos", size=9, italic=True, color="666666")
    sheet.cell(2, 1).alignment = Alignment(wrap_text=True)

    _set_section_title(sheet, 4, 6, "Parameters")
    parameter_values = [
        ["Lambda (λ)", float(lambda_value), "Alternatives", len(frame), "Criteria", len(frame.columns)],
        ["WSM share", float(lambda_value), "WPM share", 1.0 - float(lambda_value), None, None],
    ]
    for row_offset, values in enumerate(parameter_values, start=5):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row_offset, column_index, value)
    _style_grid(sheet, 5, 6, 1, 6)
    for coordinate in ("B5", "B6", "D6"):
        sheet[coordinate].number_format = "0.00"

    settings_title_row = 8
    settings_header_row = 9
    settings_data_start = 10
    _set_section_title(sheet, settings_title_row, 4, "Criterion Settings")
    settings_headers = ["Criterion", "Weight", "Preference", "Reference x*"]
    for column_index, header in enumerate(settings_headers, start=1):
        sheet.cell(settings_header_row, column_index, header)
        sheet.cell(settings_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(settings_header_row, column_index).font = _HEADER_FONT
    for row_offset, criterion in enumerate(frame.columns, start=settings_data_start):
        name = str(criterion)
        preference = preferences[name]
        _set_text(sheet, row_offset, 1, name)
        sheet.cell(row_offset, 2, float(weights[name]))
        sheet.cell(row_offset, 2).number_format = _WEIGHT_FORMAT
        _set_text(sheet, row_offset, 3, _preference_label(preference))
        sheet.cell(row_offset, 3).fill = _preference_fill(preference)
        reference = (
            float(frame[criterion].max())
            if preference.kind is CriterionType.BENEFIT
            else float(frame[criterion].min())
        )
        sheet.cell(row_offset, 4, reference)
        sheet.cell(row_offset, 4).number_format = _RAW_NUMBER_FORMAT
    settings_data_end = settings_data_start + len(frame.columns) - 1
    _style_grid(
        sheet,
        settings_header_row,
        settings_data_end,
        1,
        4,
        number_format=_CALC_NUMBER_FORMAT,
    )

    next_row = settings_data_end + 3
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 1 — Original Decision Matrix",
        frame,
        matrix_preferences=preferences,
    )
    normalized = steps["Step 2: Normalized Decision Matrix"].reindex(frame.index)
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 2 — Ratio-Normalized Decision Matrix",
        normalized,
        matrix_preferences=preferences,
    )
    wsm_components = steps["Step 3: Weighted Sum Components"].reindex(frame.index)
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 3 — Weighted Sum Components",
        wsm_components,
        matrix_preferences=preferences,
    )
    wpm_components = steps["Step 4: Weighted Product Components"].reindex(frame.index)
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 4 — Weighted Product Components",
        wpm_components,
        matrix_preferences=preferences,
    )
    aggregate = steps["Step 5: Aggregated Scores"].reindex(frame.index)
    aggregate_with_rank = pd.concat(
        [aggregate, results.reindex(frame.index)[["Rank"]]], axis=1
    )
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 5 — WSM, WPM, WASPAS Score, and Rank",
        aggregate_with_rank,
    )
    aggregate_data_start = next_row - len(aggregate_with_rank) - 1
    for row in range(aggregate_data_start, aggregate_data_start + len(aggregate_with_rank)):
        sheet.cell(row, 5).number_format = "0"

    final_values = results[
        ["Q_i (WSM)", "Q_i (WPM)", "Q_i (WASPAS Score)", "Rank"]
    ]
    final_start = next_row
    _write_indexed_dataframe(
        sheet,
        final_start,
        "Final Ranking — Sorted by Rank",
        final_values,
    )
    final_data_start = final_start + 2
    for row in range(final_data_start, final_data_start + len(results)):
        sheet.cell(row, 5).number_format = "0"

    sheet.column_dimensions["A"].width = max(
        26,
        min(
            40,
            max(
                max(len(str(value)) for value in frame.index),
                max(len(str(value)) for value in frame.columns),
            )
            + 3,
        ),
    )
    for column_index in range(2, max_columns + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 22
    _autofit_column_widths(sheet)


def _build_formula_guide_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Formula Guide")
    sheet.sheet_view.showGridLines = False
    _set_title(sheet, 1, 5, "WASPAS Formula Guide and Audit Trail")
    sheet.merge_cells("A2:E2")
    sheet["A2"] = (
        "The WASPAS sheet contains live Excel formulas. Verified Values preserves "
        "the canonical numerical snapshot generated at download time."
    )
    sheet["A2"].alignment = Alignment(wrap_text=True)
    sheet["A2"].font = Font(name="Aptos", size=9, italic=True, color="666666")

    headers = ["Step", "Symbol", "Mathematical formula", "Meaning", "Excel implementation"]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(4, column_index, header)
        sheet.cell(4, column_index).fill = _SECTION_FILL
        sheet.cell(4, column_index).font = _HEADER_FONT

    rows = [
        (
            "2",
            "r_ij benefit",
            "x_ij / max_i(x_ij)",
            "Ratio-normalize a criterion where larger values are preferred.",
            "Input divided by visible column maximum",
        ),
        (
            "2",
            "r_ij cost",
            "min_i(x_ij) / x_ij",
            "Ratio-normalize a strictly positive criterion where smaller is preferred.",
            "Visible column minimum divided by input",
        ),
        (
            "3",
            "Q_i^(1)",
            "Σ_j w_j r_ij",
            "Weighted Sum Model component.",
            "SUM of the weighted-sum contribution row",
        ),
        (
            "4",
            "Q_i^(2)",
            "Π_j (r_ij)^(w_j)",
            "Weighted Product Model component.",
            "PRODUCT of POWER(normalized value, live weight)",
        ),
        (
            "5",
            "Q_i",
            "λ Q_i^(1) + (1-λ) Q_i^(2)",
            "WASPAS aggregate score; higher values are preferred.",
            "Live lambda reference in cell B5",
        ),
        (
            "5",
            "λ",
            "0 ≤ λ ≤ 1",
            "WSM share; 0 gives WPM, 1 gives WSM, and 0.5 is balanced.",
            "Validated editable input",
        ),
        (
            "5",
            "Rank_i",
            "descending competition rank of Q_i",
            "Equal scores share a rank and skipped positions remain skipped.",
            "RANK(score, score range, 0)",
        ),
    ]
    for row_index, values in enumerate(rows, start=5):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row_index, column_index, value)
    _style_grid(sheet, 4, 4 + len(rows), 1, 5)
    for row in range(5, 5 + len(rows)):
        sheet.row_dimensions[row].height = 43

    audit_row = 6 + len(rows)
    _set_section_title(sheet, audit_row, 5, "Workbook Audit Notes")
    audit_notes = [
        "Yellow cells are intended inputs; gray cells are formulas or fixed method values.",
        "Changing lambda or weights updates all live outputs, the summary, and the chart.",
        "Weights are normalized by the app before export and should sum to 1.",
        "Verified Values remains a static reference to the calculation at download time.",
        "WASPAS natively supports benefit and cost criteria, not target criteria.",
    ]
    for row_offset, note in enumerate(audit_notes, start=1):
        sheet.cell(audit_row + row_offset, 1, f"• {note}")
        sheet.merge_cells(
            start_row=audit_row + row_offset,
            start_column=1,
            end_row=audit_row + row_offset,
            end_column=5,
        )
        sheet.cell(audit_row + row_offset, 1).font = _BODY_FONT

    source_row = audit_row + len(audit_notes) + 2
    _set_section_title(sheet, source_row, 5, "Method Reference")
    article_url = "https://eejournal.ktu.lt/index.php/elt/article/view/1810"
    doi_url = "https://doi.org/10.5755/j01.eee.122.6.1810"
    sheet.cell(source_row + 1, 1, "Original WASPAS article")
    sheet.cell(source_row + 1, 2, article_url)
    sheet.merge_cells(
        start_row=source_row + 1,
        start_column=2,
        end_row=source_row + 1,
        end_column=5,
    )
    sheet.cell(source_row + 1, 2).hyperlink = article_url
    sheet.cell(source_row + 1, 2).style = "Hyperlink"
    sheet.cell(source_row + 2, 1, "DOI")
    sheet.cell(source_row + 2, 2, doi_url)
    sheet.merge_cells(
        start_row=source_row + 2,
        start_column=2,
        end_row=source_row + 2,
        end_column=5,
    )
    sheet.cell(source_row + 2, 2).hyperlink = doi_url
    sheet.cell(source_row + 2, 2).style = "Hyperlink"

    widths = {"A": 13, "B": 20, "C": 48, "D": 55, "E": 46}
    for letter, width in widths.items():
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A5"


def build_waspas_excel_workbook(
    matrix: pd.DataFrame,
    weights: Mapping[str, Any],
    directions: Mapping[str, Any],
    *,
    lambda_value: float = 0.5,
) -> bytes:
    """Return a complete WASPAS workbook with live formulas and verified values."""

    frame = validate_crisp_matrix(matrix).copy()
    frame.columns = [str(column) for column in frame.columns]
    normalized_weights = validate_weights(weights, frame.columns, normalize=True)
    preferences = validate_method_capabilities("WASPAS", frame.columns, directions)
    results, steps = calculate_waspas(
        frame,
        normalized_weights,
        directions,
        lambda_value=float(lambda_value),
        return_steps=True,
    )

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "MCDM Calculator"
    workbook.properties.title = "Complete WASPAS Formula-Driven Calculation"
    workbook.properties.subject = "Auditable WASPAS decision model"
    workbook.properties.version = WASPAS_EXCEL_EXPORT_REVISION
    workbook.properties.description = (
        "Live Excel formulas, a decision summary, and canonical numerical values for "
        f"every WASPAS calculation stage. Export revision {WASPAS_EXCEL_EXPORT_REVISION}."
    )
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    formula_positions = _build_formula_sheet(
        workbook,
        frame,
        normalized_weights,
        preferences,
        lambda_value=float(lambda_value),
    )
    _build_summary_sheet(workbook, formula_positions, len(frame))
    _build_verified_values_sheet(
        workbook,
        frame,
        normalized_weights,
        preferences,
        results,
        steps,
        lambda_value=float(lambda_value),
    )
    _build_formula_guide_sheet(workbook)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


__all__ = [
    "WASPAS_EXCEL_EXPORT_FILENAME",
    "WASPAS_EXCEL_EXPORT_REVISION",
    "build_waspas_excel_workbook",
]
