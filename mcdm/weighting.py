"""Objective criterion weighting methods and deterministic weight robustness analysis.

Implements standard deterministic weighting algorithms and generates Table 4
("Robustness under Alternative Deterministic Weights") for any MCDM method.
"""

from __future__ import annotations

from typing import Any, Mapping, Sequence
import numpy as np
import pandas as pd

from entropy_calculator import calculate_entropy_weights
from merec_calculator import calculate_merec_weights
from mcdm.agreement import (
    calculate_kendall_tau,
    calculate_mard,
    calculate_spearman_rho,
    calculate_top_k_jaccard,
    get_default_jaccard_cutoffs,
)
from mcdm.analysis import calculate_method
from mcdm.criteria import CriterionType, normalize_directions
from mcdm.presentation import RESULT_PRESENTATION
from mcdm.validation import MCDMValidationError, validate_crisp_matrix


def calculate_critic_weights(
    df: pd.DataFrame,
    directions: Mapping[str, Any],
) -> tuple[dict[str, float], dict[str, Any]]:
    """Calculate objective weights using the CRITIC method (Diakoulaki et al., 1995).
    
    CRITIC (Criteria Importance Through Intercriteria Correlation) measures both
    contrast intensity (standard deviation) and conflict (correlation with other criteria).
    """
    frame = validate_crisp_matrix(df)
    preferences = normalize_directions(frame.columns, directions)
    target_criteria = [
        col for col, pref in preferences.items() if pref.kind is CriterionType.TARGET
    ]
    if target_criteria:
        raise MCDMValidationError(
            "CRITIC does not natively support target criteria: " + ", ".join(target_criteria)
        )

    n, m = frame.shape
    if n <= 1 or m <= 1:
        return {col: 1.0 / m for col in frame.columns}, {}

    # Step 1: Min-Max normalization to [0, 1]
    norm_df = pd.DataFrame(index=frame.index, columns=frame.columns, dtype=float)
    for col in frame.columns:
        col_vals = frame[col].to_numpy(dtype=float)
        c_min, c_max = float(col_vals.min()), float(col_vals.max())
        diff = c_max - c_min
        if preferences[col].kind is CriterionType.COST:
            norm_df[col] = 1.0 if diff <= 1e-12 else (c_max - col_vals) / diff
        else:
            norm_df[col] = 1.0 if diff <= 1e-12 else (col_vals - c_min) / diff

    # Step 2: Standard deviations of normalized criteria
    std_series = norm_df.std(axis=0, ddof=0)

    # Step 3: Correlation matrix
    corr_df = norm_df.corr(method="pearson").fillna(0.0)

    # Step 4: Information quantity C_j = sigma_j * sum_k(1 - r_jk)
    conflict_series = (1.0 - corr_df).sum(axis=1)
    c_series = std_series * conflict_series

    # Step 5: Normalized weights
    c_sum = float(c_series.sum())
    if c_sum <= 1e-12:
        weights = {col: 1.0 / m for col in frame.columns}
    else:
        weights = {col: float(c_series[col] / c_sum) for col in frame.columns}

    steps = {
        "Normalized Matrix": norm_df,
        "Standard Deviations (σ_j)": std_series.to_frame(name="Std Dev"),
        "Correlation Matrix": corr_df,
        "Conflict Measures (Σ(1 - r_jk))": conflict_series.to_frame(name="Conflict"),
        "Information Quantity (C_j)": c_series.to_frame(name="C_j"),
        "Final CRITIC Weights": pd.DataFrame.from_dict(weights, orient="index", columns=["Weight"]),
    }
    return weights, steps


def calculate_sd_weights(
    df: pd.DataFrame,
    directions: Mapping[str, Any],
) -> tuple[dict[str, float], dict[str, Any]]:
    """Calculate objective weights proportional to criterion Standard Deviation."""
    frame = validate_crisp_matrix(df)
    preferences = normalize_directions(frame.columns, directions)
    n, m = frame.shape
    if n <= 1 or m <= 1:
        return {col: 1.0 / m for col in frame.columns}, {}

    norm_df = pd.DataFrame(index=frame.index, columns=frame.columns, dtype=float)
    for col in frame.columns:
        col_vals = frame[col].to_numpy(dtype=float)
        c_min, c_max = float(col_vals.min()), float(col_vals.max())
        diff = c_max - c_min
        if preferences[col].kind is CriterionType.COST:
            norm_df[col] = 1.0 if diff <= 1e-12 else (c_max - col_vals) / diff
        else:
            norm_df[col] = 1.0 if diff <= 1e-12 else (col_vals - c_min) / diff

    std_series = norm_df.std(axis=0, ddof=0)
    std_sum = float(std_series.sum())
    if std_sum <= 1e-12:
        weights = {col: 1.0 / m for col in frame.columns}
    else:
        weights = {col: float(std_series[col] / std_sum) for col in frame.columns}

    steps = {
        "Normalized Matrix": norm_df,
        "Standard Deviations (σ_j)": std_series.to_frame(name="Std Dev"),
        "Final SD Weights": pd.DataFrame.from_dict(weights, orient="index", columns=["Weight"]),
    }
    return weights, steps


def calculate_pca_weights(
    df: pd.DataFrame,
    directions: Mapping[str, Any],
) -> tuple[dict[str, float], dict[str, Any]]:
    """Calculate objective weights using Principal Component Analysis (PCA loadings).
    
    Weights are derived from principal component loadings weighted by explained variance.
    """
    frame = validate_crisp_matrix(df)
    preferences = normalize_directions(frame.columns, directions)
    n, m = frame.shape
    if n <= 1 or m <= 1:
        return {col: 1.0 / m for col in frame.columns}, {}

    norm_df = pd.DataFrame(index=frame.index, columns=frame.columns, dtype=float)
    for col in frame.columns:
        col_vals = frame[col].to_numpy(dtype=float)
        c_min, c_max = float(col_vals.min()), float(col_vals.max())
        diff = c_max - c_min
        if preferences[col].kind is CriterionType.COST:
            norm_df[col] = 1.0 if diff <= 1e-12 else (c_max - col_vals) / diff
        else:
            norm_df[col] = 1.0 if diff <= 1e-12 else (col_vals - c_min) / diff

    # Standardize normalized matrix (mean 0, std 1)
    x = norm_df.to_numpy(dtype=float)
    std_vec = np.std(x, axis=0, ddof=0)
    std_vec[std_vec <= 1e-12] = 1.0
    x_std = (x - np.mean(x, axis=0)) / std_vec

    # Eigen-decomposition of correlation/covariance matrix
    cov_matrix = np.corrcoef(x_std, rowvar=False)
    if np.isnan(cov_matrix).any():
        cov_matrix = np.nan_to_num(cov_matrix, nan=0.0)
        np.fill_diagonal(cov_matrix, 1.0)

    eigenvalues, eigenvectors = np.linalg.eigh(cov_matrix)
    # Sort descending
    sort_idx = np.argsort(eigenvalues)[::-1]
    eigenvalues = np.maximum(eigenvalues[sort_idx], 0.0)
    eigenvectors = eigenvectors[:, sort_idx]

    total_var = np.sum(eigenvalues)
    if total_var <= 1e-12:
        weights = {col: 1.0 / m for col in frame.columns}
    else:
        var_explained = eigenvalues / total_var
        # Composite absolute loading across all components weighted by variance
        composite = np.sum(np.abs(eigenvectors) * var_explained, axis=1)
        comp_sum = float(np.sum(composite))
        if comp_sum <= 1e-12:
            weights = {col: 1.0 / m for col in frame.columns}
        else:
            weights = {col: float(composite[i] / comp_sum) for i, col in enumerate(frame.columns)}

    steps = {
        "Normalized Data": norm_df,
        "Correlation Matrix": pd.DataFrame(cov_matrix, index=frame.columns, columns=frame.columns),
        "Eigenvalues (Variance)": pd.Series(eigenvalues, name="Eigenvalue"),
        "Final PCA Weights": pd.DataFrame.from_dict(weights, orient="index", columns=["Weight"]),
    }
    return weights, steps


def calculate_max_displacement(
    rank_a: pd.Series | Sequence[float],
    rank_b: pd.Series | Sequence[float],
) -> int:
    """Calculate maximum absolute rank displacement (Delta_max = max |R_i(a) - R_i(b)|)."""
    arr_a = np.asarray(rank_a, dtype=float)
    arr_b = np.asarray(rank_b, dtype=float)
    if len(arr_a) == 0 or len(arr_b) == 0:
        return 0
    return int(np.round(np.max(np.abs(arr_a - arr_b))))


def generate_deterministic_weight_scenarios(
    df: pd.DataFrame,
    directions: Mapping[str, Any],
    baseline_weights: Mapping[str, float] | None = None,
) -> dict[str, dict[str, float]]:
    """Generate all standard deterministic weight scenarios for the decision matrix."""
    columns = list(df.columns)
    m = len(columns)
    scenarios: dict[str, dict[str, float]] = {}

    # 1. Official / Baseline
    if baseline_weights is not None:
        total = sum(float(v) for v in baseline_weights.values())
        if total > 0:
            scenarios["Official"] = {col: float(baseline_weights.get(col, 1.0 / m)) / total for col in columns}
        else:
            scenarios["Official"] = {col: 1.0 / m for col in columns}

    # 2. Equal Weights
    scenarios["Equal"] = {col: 1.0 / m for col in columns}

    # 3. PCA Loadings
    try:
        pca_w, _ = calculate_pca_weights(df, directions)
        scenarios["PCA"] = pca_w
    except Exception:
        pass

    # 4. CRITIC
    try:
        critic_w, _ = calculate_critic_weights(df, directions)
        scenarios["CRITIC"] = critic_w
    except Exception:
        pass

    # 5. Entropy (EWM)
    try:
        entropy_w, _ = calculate_entropy_weights(df, directions, method="simple")
        scenarios["Entropy"] = entropy_w
    except Exception:
        pass

    # 6. MEREC
    try:
        merec_w, _ = calculate_merec_weights(df, directions)
        scenarios["MEREC"] = merec_w
    except Exception:
        pass

    # 7. Standard Deviation
    try:
        sd_w, _ = calculate_sd_weights(df, directions)
        scenarios["Standard Deviation"] = sd_w
    except Exception:
        pass

    return scenarios


def evaluate_weight_robustness(
    method: str,
    df: pd.DataFrame,
    directions: Mapping[str, Any],
    weight_scenarios: Mapping[str, Mapping[str, float]],
    *,
    baseline_scenario: str = "Official",
    parameters: Mapping[str, Any] | None = None,
    top_k: int | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Evaluate Table 4: Robustness under alternative deterministic weights.
    
    Returns:
      1. table_4_df: Robustness metrics summary DataFrame.
      2. rankings_df: Alternatives x Scenarios rankings DataFrame.
      3. weights_df: Criteria x Scenarios weights DataFrame.
    """
    method_key = method.strip().upper()
    metadata = RESULT_PRESENTATION[method_key]
    num_alts = len(df)
    
    if top_k is None or top_k <= 0:
        _k1, default_k2 = get_default_jaccard_cutoffs(num_alts)
        top_k = default_k2

    scenario_names = list(weight_scenarios.keys())
    if baseline_scenario not in scenario_names:
        baseline_scenario = scenario_names[0]

    # Calculate rankings for all scenarios
    rankings_dict: dict[str, pd.Series] = {}
    weights_dict: dict[str, pd.Series] = {}

    for name, w_map in weight_scenarios.items():
        w_norm = {col: float(w_map.get(col, 1.0 / len(df.columns))) for col in df.columns}
        w_total = sum(w_norm.values())
        if w_total > 0:
            w_norm = {col: v / w_total for col, v in w_norm.items()}
        weights_dict[name] = pd.Series(w_norm, index=df.columns)

        res = calculate_method(
            method_key,
            df,
            w_norm,
            directions,
            parameters=parameters,
            return_steps=False,
        )
        score_s = res[metadata.score_column]
        rank_s = score_s.rank(ascending=metadata.score_ascending, method="min").astype(int)
        rankings_dict[name] = rank_s

    rankings_df = pd.DataFrame(rankings_dict, index=df.index)
    weights_df = pd.DataFrame(weights_dict, index=df.columns)

    baseline_ranks = rankings_df[baseline_scenario]
    equal_scenario_key = next((k for k in scenario_names if k.lower() == "equal"), baseline_scenario)
    equal_ranks = rankings_df[equal_scenario_key]

    table_records = []
    for scenario in scenario_names:
        scen_ranks = rankings_df[scenario]
        rho_official = calculate_spearman_rho(scen_ranks, baseline_ranks)
        rho_equal = calculate_spearman_rho(scen_ranks, equal_ranks)
        max_disp = calculate_max_displacement(scen_ranks, baseline_ranks)
        top_k_overlap = calculate_top_k_jaccard(scen_ranks, baseline_ranks, top_k)
        tau = calculate_kendall_tau(scen_ranks, baseline_ranks)
        mard = calculate_mard(scen_ranks, baseline_ranks)

        table_records.append({
            "Weight scenario": scenario,
            f"ρ with {baseline_scenario.lower()}": rho_official,
            f"ρ with equal-weight {method_key}": rho_equal,
            "Maximum displacement": max_disp,
            f"Top-{top_k} overlap with {baseline_scenario.lower()}": top_k_overlap,
            "Kendall τ-b": tau,
            "MARD": mard,
        })

    table_4_df = pd.DataFrame(table_records)
    table_4_df.set_index("Weight scenario", inplace=True)

    return table_4_df, rankings_df, weights_df


def build_weight_robustness_excel_workbook(
    table_4_df: pd.DataFrame,
    rankings_df: pd.DataFrame,
    weights_df: pd.DataFrame,
    *,
    method: str = "AURA",
    baseline_name: str = "Official",
    top_k: int | None = None,
) -> bytes:
    """Return a styled XLSX workbook containing Table 4 with live dynamic formulas."""
    from io import BytesIO
    import re
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # Infer top_k from column names if not provided
    if top_k is None or top_k <= 0:
        top_k = len(rankings_df)
        for col in table_4_df.columns:
            m = re.search(r"Top-(\d+)", col)
            if m:
                top_k = int(m.group(1))
                break

    wb = Workbook()
    
    # Styles
    title_font = Font(name="Aptos", size=14, bold=True, color="1B365D")
    subtitle_font = Font(name="Aptos", size=9, italic=True, color="666666")
    header_font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=10)
    body_bold = Font(name="Aptos", size=10, bold=True)
    
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    summary_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    grid_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    n_alts = len(rankings_df)
    rk_start_row = 5
    rk_end_row = 4 + n_alts

    # 1. Sheet 2: Rankings by Scenario (Build this first so we know column positions)
    ws_rk = wb.create_sheet(title="Rankings by Scenario")
    ws_rk.sheet_view.showGridLines = True
    ws_rk.cell(1, 1, f"{method.upper()} Alternative Rankings Across Weight Scenarios").font = title_font
    ws_rk.cell(2, 1, "Integer competition ranks under each deterministic weighting scheme.").font = subtitle_font
    
    rk_reset = rankings_df.reset_index()
    if rk_reset.columns[0] != "Alternative":
        rk_reset.rename(columns={rk_reset.columns[0]: "Alternative"}, inplace=True)
    
    header_row = 4
    scenario_col_letters = {}
    base_col_idx = 2
    equal_col_idx = 2

    for c_idx, col_name in enumerate(rk_reset.columns, start=1):
        cell = ws_rk.cell(header_row, c_idx, col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left", vertical="center")
        if c_idx > 1:
            scenario_col_letters[col_name] = get_column_letter(c_idx)
            if col_name.lower() == baseline_name.lower():
                base_col_idx = c_idx
            if col_name.lower() == "equal":
                equal_col_idx = c_idx

    base_col_let = get_column_letter(base_col_idx)
    equal_col_let = get_column_letter(equal_col_idx)

    for r_offset, row_data in enumerate(rk_reset.itertuples(index=False), start=1):
        r_idx = header_row + r_offset
        row_fill = alt_fill if r_offset % 2 == 1 else white_fill
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws_rk.cell(r_idx, c_idx, val)
            cell.font = body_font
            cell.fill = row_fill
            cell.border = grid_border
            if c_idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "0"

    # 2. Sheet 4: Displacement & Deviations
    ws_disp = wb.create_sheet(title="Displacement Matrix")
    ws_disp.sheet_view.showGridLines = True
    ws_disp.cell(1, 1, f"Absolute Rank Displacement relative to {baseline_name} (|R_i(scenario) - R_i({baseline_name})|)").font = title_font
    ws_disp.cell(2, 1, f"Live formulas computing rank jumps and absolute differences for every alternative.").font = subtitle_font

    for c_idx, col_name in enumerate(rk_reset.columns, start=1):
        cell = ws_disp.cell(header_row, c_idx, col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left", vertical="center")

    for r_offset, row_data in enumerate(rk_reset.itertuples(index=False), start=1):
        r_idx = header_row + r_offset
        row_fill = alt_fill if r_offset % 2 == 1 else white_fill
        alt_name = row_data[0]
        # Col 1: Alternative
        c_alt = ws_disp.cell(r_idx, 1, alt_name)
        c_alt.font = body_font
        c_alt.fill = row_fill
        c_alt.border = grid_border
        c_alt.alignment = Alignment(horizontal="left", vertical="center")

        # Cols 2+: =ABS('Rankings by Scenario'!ColRow - 'Rankings by Scenario'!$BaseColRow)
        for c_idx in range(2, len(rk_reset.columns) + 1):
            col_let = get_column_letter(c_idx)
            disp_formula = f"=ABS('Rankings by Scenario'!{col_let}{r_idx} - 'Rankings by Scenario'!${base_col_let}{r_idx})"
            c_disp = ws_disp.cell(r_idx, c_idx, disp_formula)
            c_disp.font = body_font
            c_disp.fill = row_fill
            c_disp.border = grid_border
            c_disp.alignment = Alignment(horizontal="center", vertical="center")
            c_disp.number_format = "0"

    disp_end_row = header_row + n_alts

    # 3. Sheet 1: Table 4 Summary (Active Sheet) with LIVE FORMULAS
    ws_t4 = wb.active
    ws_t4.title = "Table 4 - Robustness Summary"
    ws_t4.sheet_view.showGridLines = True
    
    ws_t4.cell(1, 1, f"Table 4 — {method.upper()} Robustness under Alternative Deterministic Weights").font = title_font
    ws_t4.cell(2, 1, f"Live formulas evaluating rank correlation and displacement relative to {baseline_name} and Equal weights.").font = subtitle_font
    
    t4_reset = table_4_df.reset_index()
    for c_idx, col_name in enumerate(t4_reset.columns, start=1):
        cell = ws_t4.cell(header_row, c_idx, col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left", vertical="center")

    for r_offset, row_data in enumerate(t4_reset.itertuples(index=False), start=1):
        r_idx = header_row + r_offset
        scen_name = str(row_data[0])
        scen_col_let = scenario_col_letters.get(scen_name, base_col_let)
        row_fill = alt_fill if r_offset % 2 == 1 else white_fill

        # Col 1: Scenario Name
        c1 = ws_t4.cell(r_idx, 1, scen_name)
        c1.font = body_bold
        c1.fill = row_fill
        c1.border = grid_border
        c1.alignment = Alignment(horizontal="left", vertical="center")

        # Col 2: ρ with official -> =CORREL('Rankings by Scenario'!Col5:ColEnd, 'Rankings by Scenario'!$BaseCol5:$BaseColEnd)
        rho_official_formula = f"=CORREL('Rankings by Scenario'!{scen_col_let}${rk_start_row}:{scen_col_let}${rk_end_row}, 'Rankings by Scenario'!${base_col_let}${rk_start_row}:${base_col_let}${rk_end_row})"
        c2 = ws_t4.cell(r_idx, 2, rho_official_formula)
        c2.font = body_font
        c2.fill = row_fill
        c2.border = grid_border
        c2.alignment = Alignment(horizontal="right", vertical="center")
        c2.number_format = "0.0000"

        # Col 3: ρ with equal-weight -> =CORREL('Rankings by Scenario'!Col5:ColEnd, 'Rankings by Scenario'!$EqualCol5:$EqualColEnd)
        rho_equal_formula = f"=CORREL('Rankings by Scenario'!{scen_col_let}${rk_start_row}:{scen_col_let}${rk_end_row}, 'Rankings by Scenario'!${equal_col_let}${rk_start_row}:${equal_col_let}${rk_end_row})"
        c3 = ws_t4.cell(r_idx, 3, rho_equal_formula)
        c3.font = body_font
        c3.fill = row_fill
        c3.border = grid_border
        c3.alignment = Alignment(horizontal="right", vertical="center")
        c3.number_format = "0.0000"

        # Col 4: Maximum displacement -> =MAX('Displacement Matrix'!Col5:ColEnd)
        max_disp_formula = f"=MAX('Displacement Matrix'!{scen_col_let}${rk_start_row}:{scen_col_let}${disp_end_row})"
        c4 = ws_t4.cell(r_idx, 4, max_disp_formula)
        c4.font = body_font
        c4.fill = row_fill
        c4.border = grid_border
        c4.alignment = Alignment(horizontal="right", vertical="center")
        c4.number_format = "0"

        # Col 5: Top-K overlap with baseline -> COUNTIFS / (COUNTIF + COUNTIF - COUNTIFS)
        jaccard_formula = (
            f"=COUNTIFS('Rankings by Scenario'!{scen_col_let}${rk_start_row}:{scen_col_let}${rk_end_row}, \"<={top_k}\", 'Rankings by Scenario'!${base_col_let}${rk_start_row}:${base_col_let}${rk_end_row}, \"<={top_k}\") / "
            f"(COUNTIF('Rankings by Scenario'!{scen_col_let}${rk_start_row}:{scen_col_let}${rk_end_row}, \"<={top_k}\") + COUNTIF('Rankings by Scenario'!${base_col_let}${rk_start_row}:${base_col_let}${rk_end_row}, \"<={top_k}\") - COUNTIFS('Rankings by Scenario'!{scen_col_let}${rk_start_row}:{scen_col_let}${rk_end_row}, \"<={top_k}\", 'Rankings by Scenario'!${base_col_let}${rk_start_row}:${base_col_let}${rk_end_row}, \"<={top_k}\"))"
        )
        c5 = ws_t4.cell(r_idx, 5, jaccard_formula)
        c5.font = body_font
        c5.fill = row_fill
        c5.border = grid_border
        c5.alignment = Alignment(horizontal="right", vertical="center")
        c5.number_format = "0.000"

        # Col 6: Kendall τ-b
        tau_val = float(table_4_df.loc[scen_name, "Kendall τ-b"]) if "Kendall τ-b" in table_4_df.columns else 1.0
        c6 = ws_t4.cell(r_idx, 6, tau_val)
        c6.font = body_font
        c6.fill = row_fill
        c6.border = grid_border
        c6.alignment = Alignment(horizontal="right", vertical="center")
        c6.number_format = "0.0000"

        # Col 7: MARD -> =AVERAGE('Displacement Matrix'!Col5:ColEnd)
        mard_formula = f"=AVERAGE('Displacement Matrix'!{scen_col_let}${rk_start_row}:{scen_col_let}${disp_end_row})"
        c7 = ws_t4.cell(r_idx, 7, mard_formula)
        c7.font = body_font
        c7.fill = row_fill
        c7.border = grid_border
        c7.alignment = Alignment(horizontal="right", vertical="center")
        c7.number_format = "0.00"

    # 4. Sheet 3: Criteria Weights by Scenario
    ws_wt = wb.create_sheet(title="Criteria Weights by Scenario")
    ws_wt.sheet_view.showGridLines = True
    ws_wt.cell(1, 1, "Criteria Weights Across Deterministic Scenarios").font = title_font
    ws_wt.cell(2, 1, "Normalized criteria weights (sum to 1.0) under each objective/deterministic model.").font = subtitle_font
    
    wt_reset = weights_df.reset_index()
    if wt_reset.columns[0] != "Criterion":
        wt_reset.rename(columns={wt_reset.columns[0]: "Criterion"}, inplace=True)
        
    for c_idx, col_name in enumerate(wt_reset.columns, start=1):
        cell = ws_wt.cell(header_row, c_idx, col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left", vertical="center")

    for r_offset, row_data in enumerate(wt_reset.itertuples(index=False), start=1):
        r_idx = header_row + r_offset
        row_fill = alt_fill if r_offset % 2 == 1 else white_fill
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws_wt.cell(r_idx, c_idx, val)
            cell.font = body_font
            cell.fill = row_fill
            cell.border = grid_border
            if c_idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "0.0000"

    # Autofit column widths across all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if cell.row in (1, 2):
                    continue
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 13)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = [
    "calculate_critic_weights",
    "calculate_sd_weights",
    "calculate_pca_weights",
    "calculate_max_displacement",
    "generate_deterministic_weight_scenarios",
    "evaluate_weight_robustness",
    "build_weight_robustness_excel_workbook",
]
