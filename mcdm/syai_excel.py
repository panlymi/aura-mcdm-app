"""Formula-rich Excel export for complete, auditable SYAI calculations."""

from __future__ import annotations

from io import BytesIO
import re
from typing import Any, Mapping

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from syai_calculator import calculate_syai

from .criteria import CriterionPreference, CriterionType, validate_method_capabilities
from .validation import validate_crisp_matrix, validate_weights


_TITLE_FILL = PatternFill("solid", fgColor="0B4F6C")
_SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
_WEIGHT_FILL = PatternFill("solid", fgColor="F8CBAD")
_BENEFIT_FILL = PatternFill("solid", fgColor="00B0F0")
_ALTERNATIVE_FILL = PatternFill("solid", fgColor="FFF200")
_COST_FILL = PatternFill("solid", fgColor="FF3B30")
_TARGET_FILL = PatternFill("solid", fgColor="F4B183")
_SUMMARY_FILL = PatternFill("solid", fgColor="E2F0D9")
_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
_FORMULA_FILL = PatternFill("solid", fgColor="F3F6F8")
_WINNER_FILL = PatternFill("solid", fgColor="C6E0B4")
_WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
_THIN_GRAY = Side(style="thin", color="B7C9D6")
_MEDIUM_TEAL = Side(style="medium", color="0B4F6C")
_GRID_BORDER = Border(
    left=_THIN_GRAY,
    right=_THIN_GRAY,
    top=_THIN_GRAY,
    bottom=_THIN_GRAY,
)
_SECTION_BORDER = Border(bottom=_MEDIUM_TEAL)
_BODY_FONT = Font(name="Aptos", size=10, color="1F1F1F")
_HEADER_FONT = Font(name="Aptos", size=10, bold=True, color="1F1F1F")
_TITLE_FONT = Font(name="Aptos Display", size=16, bold=True, color="FFFFFF")
_ILLEGAL_EXCEL_TEXT = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

_RAW_NUMBER_FORMAT = "General"
_CALC_NUMBER_FORMAT = "0.000000"
_WEIGHT_FORMAT = "0.0000"

# Included in Streamlit's cache key so workbook changes invalidate old bytes.
SYAI_EXCEL_EXPORT_REVISION = "v2"
SYAI_EXCEL_EXPORT_FILENAME = (
    f"syai_complete_formula_calculation_{SYAI_EXCEL_EXPORT_REVISION}.xlsx"
)


def _safe_excel_text(value: Any) -> str:
    """Return user text that Excel stores as text rather than a formula."""

    return _ILLEGAL_EXCEL_TEXT.sub(" ", str(value))[:32767]


def _set_text(sheet, row: int, column: int, value: Any) -> None:
    cell = sheet.cell(row, column)
    cell.value = _safe_excel_text(value)
    cell.data_type = "s"


def _preference_label(preference: CriterionPreference) -> str:
    if preference.kind is CriterionType.BENEFIT:
        return "Benefit (maximize)"
    if preference.kind is CriterionType.COST:
        return "Cost (minimize)"
    return "Target"


def _preference_fill(preference: CriterionPreference) -> PatternFill:
    if preference.kind is CriterionType.BENEFIT:
        return _BENEFIT_FILL
    if preference.kind is CriterionType.COST:
        return _COST_FILL
    return _TARGET_FILL


def _set_title(sheet, row: int, last_column: int, text: str) -> None:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_column)
    cell = sheet.cell(row=row, column=1, value=text)
    cell.fill = _TITLE_FILL
    cell.font = _TITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _SECTION_BORDER
    sheet.row_dimensions[row].height = 29


def _set_section_title(sheet, row: int, last_column: int, text: str) -> None:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_column)
    cell = sheet.cell(row=row, column=1, value=text)
    cell.fill = _SECTION_FILL
    cell.font = Font(name="Aptos Display", size=11, bold=True, color="1F1F1F")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _SECTION_BORDER
    sheet.row_dimensions[row].height = 23


def _style_grid(
    sheet,
    min_row: int,
    max_row: int,
    min_column: int,
    max_column: int,
    *,
    number_format: str | None = None,
) -> None:
    for row in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_column,
        max_col=max_column,
    ):
        for cell in row:
            cell.border = _GRID_BORDER
            cell.font = _BODY_FONT
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            if number_format and cell.column > min_column:
                cell.number_format = number_format


def _style_matrix_header(
    sheet,
    header_row: int,
    columns: list[str],
    preferences: Mapping[str, CriterionPreference],
) -> None:
    alternative_cell = sheet.cell(header_row, 1)
    alternative_cell.fill = _ALTERNATIVE_FILL
    alternative_cell.font = _HEADER_FONT
    for offset, criterion in enumerate(columns, start=2):
        cell = sheet.cell(header_row, offset)
        cell.fill = _preference_fill(preferences[criterion])
        cell.font = _HEADER_FONT
    sheet.row_dimensions[header_row].height = 31


def _write_live_weights_row(
    sheet,
    row: int,
    columns: list[str],
    settings_data_start: int,
) -> None:
    sheet.cell(row, 1, "weightage")
    for offset, _criterion in enumerate(columns, start=2):
        settings_row = settings_data_start + offset - 2
        sheet.cell(row, offset, f"=$B${settings_row}")
    for cell in sheet[row][: len(columns) + 1]:
        cell.fill = _WEIGHT_FILL
        cell.font = _HEADER_FONT
        cell.border = _GRID_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.column > 1:
            cell.number_format = _WEIGHT_FORMAT


def _write_indexed_dataframe(
    sheet,
    start_row: int,
    title: str,
    frame: pd.DataFrame,
    *,
    header_label: str = "Alternative",
    matrix_preferences: Mapping[str, CriterionPreference] | None = None,
) -> int:
    columns = [str(column) for column in frame.columns]
    last_column = len(columns) + 1
    _set_section_title(sheet, start_row, last_column, title)
    header_row = start_row + 1
    sheet.cell(header_row, 1, header_label)
    for column_index, column in enumerate(columns, start=2):
        _set_text(sheet, header_row, column_index, column)
    _style_grid(
        sheet,
        header_row,
        header_row + len(frame),
        1,
        last_column,
        number_format=_CALC_NUMBER_FORMAT,
    )
    if matrix_preferences is None:
        for cell in sheet[header_row][:last_column]:
            cell.fill = _SECTION_FILL
            cell.font = _HEADER_FONT
    else:
        _style_matrix_header(sheet, header_row, columns, matrix_preferences)

    for row_offset, (index, values) in enumerate(frame.iterrows(), start=1):
        excel_row = header_row + row_offset
        _set_text(sheet, excel_row, 1, index)
        sheet.cell(excel_row, 1).fill = _ALTERNATIVE_FILL
        for column_index, value in enumerate(values, start=2):
            sheet.cell(excel_row, column_index, float(value))
            sheet.cell(excel_row, column_index).number_format = _CALC_NUMBER_FORMAT
    return header_row + len(frame) + 2


def _autofit_column_widths(
    sheet,
    *,
    min_width: int = 14,
    padding: int = 3,
    max_width: int = 62,
) -> None:
    merged_ranges = list(sheet.merged_cells.ranges)

    def is_wide_merged_cell(row_idx: int, col_idx: int) -> bool:
        for cell_range in merged_ranges:
            if (
                cell_range.min_row <= row_idx <= cell_range.max_row
                and cell_range.min_col <= col_idx <= cell_range.max_col
                and cell_range.max_col > cell_range.min_col
            ):
                return True
        return False

    for column in sheet.columns:
        column_index = column[0].column
        letter = get_column_letter(column_index)
        max_length = 0
        for cell in column:
            if cell.value is None or is_wide_merged_cell(cell.row, column_index):
                continue
            value = str(cell.value)
            visible_length = 15 if value.startswith("=") else max(
                len(line) for line in value.split("\n")
            )
            max_length = max(max_length, visible_length)
        width = max(min_width, min(max_width, max_length + padding))
        existing_width = sheet.column_dimensions[letter].width or 0
        sheet.column_dimensions[letter].width = max(existing_width, width)


def _build_formula_sheet(
    workbook: Workbook,
    frame: pd.DataFrame,
    weights: Mapping[str, float],
    preferences: Mapping[str, CriterionPreference],
    *,
    beta: float,
) -> dict[str, int]:
    sheet = workbook.create_sheet("SYAI")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 85

    columns = [str(column) for column in frame.columns]
    alternative_names = [str(index) for index in frame.index]
    last_matrix_column = len(columns) + 1
    last_matrix_letter = get_column_letter(last_matrix_column)
    model_last_column = max(7, last_matrix_column)
    alternatives_count = len(alternative_names)
    criteria_count = len(columns)

    _set_title(
        sheet,
        1,
        model_last_column,
        "SYAI — Complete Formula-Driven Calculation Workbook",
    )
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=model_last_column)
    sheet.cell(
        2,
        1,
        "Edit the yellow beta cell to explore the ideal/anti-ideal trade-off. Orange rows "
        "contain weights; blue/red/orange headers identify benefit/cost/target criteria.",
    )
    sheet.cell(2, 1).font = Font(name="Aptos", size=9, italic=True, color="666666")
    sheet.cell(2, 1).alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[2].height = 30

    settings_title_row = 11
    settings_header_row = 12
    settings_data_start = 13
    settings_data_end = settings_data_start + criteria_count - 1

    _set_section_title(sheet, 4, 5, "Model Parameters and Audit Checks")
    parameter_rows = [
        (5, "Beta (β)", float(beta), "Closeness trade-off; must satisfy 0 < β < 1"),
        (6, "Normalization constant (C)", 0.01, "Published fixed constant"),
        (7, "Alternatives", alternatives_count, "Number of decision alternatives"),
        (8, "Criteria", criteria_count, "Number of decision criteria"),
    ]
    for row, label, value, description in parameter_rows:
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
        sheet.cell(row, 3, description)
    sheet.cell(9, 1, "Weight sum check")
    sheet.cell(9, 2, f"=SUM(B{settings_data_start}:B{settings_data_end})")
    sheet.cell(9, 3, '="Expected 1.0000; current "&TEXT(B9,"0.0000")')
    _style_grid(sheet, 5, 9, 1, 5)
    sheet.cell(5, 2).fill = _INPUT_FILL
    sheet.cell(5, 2).number_format = "0.00"
    sheet.cell(6, 2).fill = _FORMULA_FILL
    sheet.cell(6, 2).number_format = "0.00"
    sheet.cell(9, 2).fill = _SUMMARY_FILL
    sheet.cell(9, 2).number_format = _WEIGHT_FORMAT
    sheet.cell(5, 2).comment = Comment(
        "Change beta to recalculate all closeness scores, ranks, and the summary chart. "
        "Valid values are strictly between 0 and 1.",
        "User",
    )
    sheet.cell(6, 2).comment = Comment(
        "C = 0.01 is the fixed SYAI normalization constant.", "User"
    )
    beta_validation = DataValidation(
        type="custom",
        formula1="AND(ISNUMBER(B5),B5>0,B5<1)",
        allow_blank=False,
    )
    beta_validation.error = "Beta must be a number strictly between 0 and 1."
    beta_validation.errorTitle = "Invalid beta"
    beta_validation.prompt = "Enter a beta value strictly between 0 and 1."
    beta_validation.promptTitle = "SYAI beta"
    beta_validation.showErrorMessage = True
    beta_validation.showInputMessage = True
    sheet.add_data_validation(beta_validation)
    beta_validation.add(sheet["B5"])

    raw_title_row = settings_data_end + 3
    raw_weights_row = raw_title_row + 1
    raw_header_row = raw_title_row + 2
    raw_data_start = raw_title_row + 3
    raw_data_end = raw_data_start + alternatives_count - 1
    raw_min_row = raw_data_end + 1
    raw_max_row = raw_data_end + 2
    raw_range_row = raw_data_end + 3
    raw_reference_row = raw_data_end + 4

    _set_section_title(sheet, settings_title_row, 5, "Criterion Settings")
    settings_headers = [
        "Criterion",
        "Normalized Weight",
        "Preference",
        "Ideal point x*",
        "Raw range R_j",
    ]
    for column_index, value in enumerate(settings_headers, start=1):
        sheet.cell(settings_header_row, column_index, value)
    _style_grid(sheet, settings_header_row, settings_data_end, 1, 5)
    for cell in sheet[settings_header_row][:5]:
        cell.fill = _SECTION_FILL
        cell.font = _HEADER_FONT

    for criterion_index, criterion in enumerate(columns):
        row = settings_data_start + criterion_index
        matrix_letter = get_column_letter(criterion_index + 2)
        preference = preferences[criterion]
        _set_text(sheet, row, 1, criterion)
        sheet.cell(row, 2, float(weights[criterion]))
        sheet.cell(row, 2).number_format = _WEIGHT_FORMAT
        sheet.cell(row, 2).fill = _INPUT_FILL
        sheet.cell(row, 3, _preference_label(preference))
        sheet.cell(row, 3).fill = _preference_fill(preference)
        if preference.kind is CriterionType.TARGET:
            sheet.cell(row, 4, float(preference.target_value))
            sheet.cell(row, 4).fill = _INPUT_FILL
        elif preference.kind is CriterionType.BENEFIT:
            sheet.cell(row, 4, f"={matrix_letter}${raw_max_row}")
            sheet.cell(row, 4).fill = _FORMULA_FILL
        else:
            sheet.cell(row, 4, f"={matrix_letter}${raw_min_row}")
            sheet.cell(row, 4).fill = _FORMULA_FILL
        sheet.cell(row, 4).number_format = _RAW_NUMBER_FORMAT
        sheet.cell(row, 5, f"={matrix_letter}${raw_range_row}")
        sheet.cell(row, 5).fill = _FORMULA_FILL
        sheet.cell(row, 5).number_format = _RAW_NUMBER_FORMAT

    weight_validation = DataValidation(
        type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=False
    )
    weight_validation.error = "Weights must be non-negative numbers."
    weight_validation.errorTitle = "Invalid weight"
    weight_validation.showErrorMessage = True
    sheet.add_data_validation(weight_validation)
    weight_validation.add(f"B{settings_data_start}:B{settings_data_end}")

    _set_section_title(
        sheet, raw_title_row, last_matrix_column, "Step 0 — Original Decision Matrix"
    )
    _write_live_weights_row(sheet, raw_weights_row, columns, settings_data_start)
    sheet.cell(raw_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, raw_header_row, column_index, criterion)
    _style_grid(
        sheet,
        raw_header_row,
        raw_reference_row,
        1,
        last_matrix_column,
        number_format=_RAW_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, raw_header_row, columns, preferences)
    for row_offset, (alternative, values) in enumerate(frame.iterrows()):
        row = raw_data_start + row_offset
        _set_text(sheet, row, 1, alternative)
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index, value in enumerate(values, start=2):
            sheet.cell(row, column_index, float(value))
            sheet.cell(row, column_index).fill = _WHITE_FILL
            sheet.cell(row, column_index).number_format = _RAW_NUMBER_FORMAT

    summary_rows = [
        (raw_min_row, "Minimum"),
        (raw_max_row, "Maximum"),
        (raw_range_row, "Range R_j"),
        (raw_reference_row, "Ideal point x*"),
    ]
    for summary_row, label in summary_rows:
        sheet.cell(summary_row, 1, label)
        for column_index in range(1, last_matrix_column + 1):
            sheet.cell(summary_row, column_index).fill = _SUMMARY_FILL
            sheet.cell(summary_row, column_index).font = _HEADER_FONT
    for column_index in range(2, last_matrix_column + 1):
        letter = get_column_letter(column_index)
        sheet.cell(
            raw_min_row,
            column_index,
            f"=MIN({letter}${raw_data_start}:{letter}${raw_data_end})",
        )
        sheet.cell(
            raw_max_row,
            column_index,
            f"=MAX({letter}${raw_data_start}:{letter}${raw_data_end})",
        )
        sheet.cell(
            raw_range_row,
            column_index,
            f"={letter}${raw_max_row}-{letter}${raw_min_row}",
        )
        settings_row = settings_data_start + column_index - 2
        sheet.cell(raw_reference_row, column_index, f"=$D${settings_row}")

    normalized_title_row = raw_reference_row + 3
    normalized_weights_row = normalized_title_row + 1
    normalized_header_row = normalized_title_row + 2
    normalized_data_start = normalized_title_row + 3
    normalized_data_end = normalized_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        normalized_title_row,
        last_matrix_column,
        "Step 1 — Normalized Decision Matrix (N_ij)",
    )
    _write_live_weights_row(sheet, normalized_weights_row, columns, settings_data_start)
    sheet.cell(normalized_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, normalized_header_row, column_index, criterion)
    _style_grid(
        sheet,
        normalized_header_row,
        normalized_data_end,
        1,
        last_matrix_column,
        number_format=_CALC_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, normalized_header_row, columns, preferences)
    for row_offset, _alternative in enumerate(alternative_names):
        row = normalized_data_start + row_offset
        raw_row = raw_data_start + row_offset
        sheet.cell(row, 1, f"=A{raw_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index in range(2, last_matrix_column + 1):
            letter = get_column_letter(column_index)
            settings_row = settings_data_start + column_index - 2
            formula = (
                f"=IF({letter}${raw_max_row}={letter}${raw_min_row},1,"
                f"$B$6+(1-$B$6)*(1-(ABS({letter}{raw_row}-$D${settings_row})/"
                f"MAX($E${settings_row},1E-9))))"
            )
            sheet.cell(row, column_index, formula)
    sheet.cell(normalized_data_start, 2).comment = Comment(
        "N_ij = C + (1-C) × (1 - |x_ij-x*| / R_j). Constant columns return 1.",
        "User",
    )

    weighted_title_row = normalized_data_end + 3
    weighted_weights_row = weighted_title_row + 1
    weighted_header_row = weighted_title_row + 2
    weighted_data_start = weighted_title_row + 3
    weighted_data_end = weighted_data_start + alternatives_count - 1
    ideal_row = weighted_data_end + 1
    anti_ideal_row = weighted_data_end + 2

    _set_section_title(
        sheet,
        weighted_title_row,
        last_matrix_column,
        "Step 2 & 3 — Weighted Matrix and Yielded-Ideal Solutions",
    )
    _write_live_weights_row(sheet, weighted_weights_row, columns, settings_data_start)
    sheet.cell(weighted_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, weighted_header_row, column_index, criterion)
    _style_grid(
        sheet,
        weighted_header_row,
        anti_ideal_row,
        1,
        last_matrix_column,
        number_format=_CALC_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, weighted_header_row, columns, preferences)
    for row_offset, _alternative in enumerate(alternative_names):
        row = weighted_data_start + row_offset
        normalized_row = normalized_data_start + row_offset
        sheet.cell(row, 1, f"=A{normalized_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index in range(2, last_matrix_column + 1):
            letter = get_column_letter(column_index)
            sheet.cell(
                row,
                column_index,
                f"={letter}{normalized_row}*{letter}${weighted_weights_row}",
            )
    for solution_row, label, function_name in [
        (ideal_row, "A+ (Yielded-Ideal)", "MAX"),
        (anti_ideal_row, "A- (Anti-Ideal)", "MIN"),
    ]:
        sheet.cell(solution_row, 1, label)
        for column_index in range(1, last_matrix_column + 1):
            sheet.cell(solution_row, column_index).fill = _SUMMARY_FILL
            sheet.cell(solution_row, column_index).font = _HEADER_FONT
        for column_index in range(2, last_matrix_column + 1):
            letter = get_column_letter(column_index)
            sheet.cell(
                solution_row,
                column_index,
                f"={function_name}({letter}${weighted_data_start}:"
                f"{letter}${weighted_data_end})",
            )

    distances_title_row = anti_ideal_row + 3
    distances_header_row = distances_title_row + 1
    distances_data_start = distances_title_row + 2
    distances_data_end = distances_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        distances_title_row,
        6,
        "Step 4 & 5 — Distances, Closeness Score, Competition Rank, and Sort Order",
    )
    distance_headers = [
        "Alternative",
        "D+ (Ideal)",
        "D- (Anti-Ideal)",
        "Closeness Score (D_i)",
        "Rank",
        "Sort Order",
    ]
    for column_index, header in enumerate(distance_headers, start=1):
        sheet.cell(distances_header_row, column_index, header)
        sheet.cell(distances_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(distances_header_row, column_index).font = _HEADER_FONT
    _style_grid(
        sheet,
        distances_header_row,
        distances_data_end,
        1,
        6,
        number_format=_CALC_NUMBER_FORMAT,
    )
    sheet.row_dimensions[distances_header_row].height = 31

    for row_offset, _alternative in enumerate(alternative_names):
        row = distances_data_start + row_offset
        weighted_row = weighted_data_start + row_offset
        sheet.cell(row, 1, f"=A{weighted_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        ideal_terms = []
        anti_ideal_terms = []
        for column_index in range(2, last_matrix_column + 1):
            letter = get_column_letter(column_index)
            ideal_terms.append(f"ABS({letter}{weighted_row}-${letter}${ideal_row})")
            anti_ideal_terms.append(
                f"ABS({letter}{weighted_row}-${letter}${anti_ideal_row})"
            )
        sheet.cell(
            row,
            2,
            "=" + "+".join(ideal_terms),
        )
        sheet.cell(
            row,
            3,
            "=" + "+".join(anti_ideal_terms),
        )
        sheet.cell(
            row,
            4,
            f"=IF($B$5*B{row}+(1-$B$5)*C{row}<1E-9,1,"
            f"((1-$B$5)*C{row})/($B$5*B{row}+(1-$B$5)*C{row}))",
        )
        sheet.cell(
            row,
            5,
            f"=RANK(D{row},$D${distances_data_start}:$D${distances_data_end},0)",
        )
        sheet.cell(
            row,
            6,
            f"=E{row}+COUNTIF($D${distances_data_start}:D{row},D{row})-1",
        )
        sheet.cell(row, 5).number_format = "0"
        sheet.cell(row, 6).number_format = "0"
    sheet.cell(distances_data_start, 4).comment = Comment(
        "Higher closeness scores are preferred. Beta is read from B5; ties share a rank.",
        "User",
    )
    sheet.conditional_formatting.add(
        f"D{distances_data_start}:D{distances_data_end}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )
    sheet.conditional_formatting.add(
        f"A{distances_data_start}:F{distances_data_end}",
        FormulaRule(
            formula=[f"$E{distances_data_start}=1"],
            fill=_WINNER_FILL,
        ),
    )

    ranking_title_row = distances_data_end + 3
    ranking_header_row = ranking_title_row + 1
    ranking_data_start = ranking_title_row + 2
    ranking_data_end = ranking_data_start + alternatives_count - 1

    _set_section_title(sheet, ranking_title_row, 5, "Final Ranking — Sorted by Score")
    ranking_headers = [
        "Alternative",
        "D+ (Ideal)",
        "D- (Anti-Ideal)",
        "Closeness Score (D_i)",
        "Rank",
    ]
    for column_index, header in enumerate(ranking_headers, start=1):
        sheet.cell(ranking_header_row, column_index, header)
        sheet.cell(ranking_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(ranking_header_row, column_index).font = _HEADER_FONT
    _style_grid(
        sheet,
        ranking_header_row,
        ranking_data_end,
        1,
        5,
        number_format=_CALC_NUMBER_FORMAT,
    )
    sheet.row_dimensions[ranking_header_row].height = 31

    source_columns = ["A", "B", "C", "D", "E"]
    for sort_position in range(1, alternatives_count + 1):
        row = ranking_data_start + sort_position - 1
        match_formula = (
            f"MATCH({sort_position},$F${distances_data_start}:"
            f"$F${distances_data_end},0)"
        )
        for column_index, source_letter in enumerate(source_columns, start=1):
            sheet.cell(
                row,
                column_index,
                f"=INDEX(${source_letter}${distances_data_start}:"
                f"${source_letter}${distances_data_end},{match_formula})",
            )
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        sheet.cell(row, 5).number_format = "0"
    sheet.conditional_formatting.add(
        f"D{ranking_data_start}:D{ranking_data_end}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )
    sheet.conditional_formatting.add(
        f"A{ranking_data_start}:E{ranking_data_end}",
        FormulaRule(formula=[f"$E{ranking_data_start}=1"], fill=_WINNER_FILL),
    )

    sheet.freeze_panes = "B1"
    sheet.print_area = f"A1:{get_column_letter(model_last_column)}{ranking_data_end}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.oddFooter.center.text = "SYAI complete formula model"
    sheet.oddFooter.right.text = "Page &P of &N"

    sheet.column_dimensions["A"].width = max(
        22, min(36, max(len(value) for value in alternative_names) + 3)
    )
    for column_index, criterion in enumerate(columns, start=2):
        letter = get_column_letter(column_index)
        raw_value_width = max(len(f"{float(value):.9f}") for value in frame[criterion])
        sheet.column_dimensions[letter].width = max(
            16,
            min(30, max(len(criterion) + 3, raw_value_width + 2)),
        )
    for column_index in range(last_matrix_column + 1, 7):
        sheet.column_dimensions[get_column_letter(column_index)].width = 17
    for row in range(1, ranking_data_end + 1):
        if sheet.row_dimensions[row].height is None:
            sheet.row_dimensions[row].height = 19
    _autofit_column_widths(sheet)

    return {
        "ranking_header_row": ranking_header_row,
        "ranking_data_start": ranking_data_start,
        "ranking_data_end": ranking_data_end,
    }


def _build_summary_sheet(
    workbook: Workbook,
    formula_positions: Mapping[str, int],
    alternatives_count: int,
) -> None:
    sheet = workbook.create_sheet("Decision Summary")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 90
    _set_title(sheet, 1, 14, "SYAI Decision Summary")
    sheet.merge_cells("A2:N2")
    sheet["A2"] = (
        "A live executive view of the SYAI ranking. Change beta on the SYAI sheet and "
        "Excel will refresh these KPIs, the ranking table, and the chart."
    )
    sheet["A2"].font = Font(name="Aptos", size=9, italic=True, color="666666")
    sheet["A2"].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[2].height = 29

    ranking_start = formula_positions["ranking_data_start"]
    cards = [
        ("A", "B", "Winner", f"='SYAI'!A{ranking_start}", "@"),
        ("C", "D", "Winning score", f"='SYAI'!D{ranking_start}", "0.0%"),
        ("E", "F", "Beta (β)", "='SYAI'!B5", "0.00"),
        ("G", "H", "Alternatives", "='SYAI'!B7", "0"),
    ]
    for start_letter, end_letter, label, formula, number_format in cards:
        sheet.merge_cells(f"{start_letter}4:{end_letter}4")
        sheet.merge_cells(f"{start_letter}5:{end_letter}6")
        label_cell = sheet[f"{start_letter}4"]
        value_cell = sheet[f"{start_letter}5"]
        label_cell.value = label
        value_cell.value = formula
        label_cell.fill = _SECTION_FILL
        value_cell.fill = _SUMMARY_FILL
        label_cell.font = _HEADER_FONT
        value_cell.font = Font(name="Aptos Display", size=16, bold=True, color="0B4F6C")
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = number_format
        for row in range(4, 7):
            for column in range(
                sheet[start_letter + "1"].column, sheet[end_letter + "1"].column + 1
            ):
                sheet.cell(row, column).border = _GRID_BORDER

    summary_header_row = 9
    summary_data_start = 10
    summary_data_end = summary_data_start + alternatives_count - 1
    headers = [
        "Alternative",
        "D+ (Ideal)",
        "D- (Anti-Ideal)",
        "Closeness Score (%)",
        "Rank",
    ]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(summary_header_row, column_index, header)
        sheet.cell(summary_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(summary_header_row, column_index).font = _HEADER_FONT
    _style_grid(
        sheet,
        summary_header_row,
        summary_data_end,
        1,
        5,
        number_format=_CALC_NUMBER_FORMAT,
    )
    for row_offset in range(alternatives_count):
        row = summary_data_start + row_offset
        source_row = ranking_start + row_offset
        for column_index, source_letter in enumerate(["A", "B", "C", "D", "E"], start=1):
            formula = f"='SYAI'!{source_letter}{source_row}"
            if source_letter == "D":
                formula += "*100"
            sheet.cell(row, column_index, formula)
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        sheet.cell(row, 4).number_format = "0.0"
        sheet.cell(row, 5).number_format = "0"
    sheet.conditional_formatting.add(
        f"D{summary_data_start}:D{summary_data_end}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )
    sheet.conditional_formatting.add(
        f"A{summary_data_start}:E{summary_data_end}",
        FormulaRule(formula=[f"$E{summary_data_start}=1"], fill=_WINNER_FILL),
    )

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "SYAI closeness score by alternative (%)"
    chart.x_axis.title = "Alternative"
    chart.y_axis.title = "Closeness score (%)"
    chart.height = 7.2
    chart.width = 14.2
    chart.legend = None
    data = Reference(
        sheet,
        min_col=4,
        min_row=summary_header_row,
        max_row=summary_data_end,
    )
    categories = Reference(
        sheet,
        min_col=1,
        min_row=summary_data_start,
        max_row=summary_data_end,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.numFmt = "0.0"
    chart.y_axis.numFmt = "0"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 100
    sheet.add_chart(chart, "G9")

    notes_row = max(summary_data_end + 3, 27)
    _set_section_title(sheet, notes_row, 8, "How to read this summary")
    notes = [
        "Higher closeness scores indicate stronger alternatives.",
        "Beta controls the balance between closeness to A+ and distance from A-.",
        "Tied scores share the same competition rank; exact ties display in input order.",
        "Verified Values preserves the canonical calculation at download time.",
    ]
    for row_offset, note in enumerate(notes, start=1):
        sheet.merge_cells(
            start_row=notes_row + row_offset,
            start_column=1,
            end_row=notes_row + row_offset,
            end_column=8,
        )
        sheet.cell(notes_row + row_offset, 1, f"• {note}")
        sheet.cell(notes_row + row_offset, 1).font = _BODY_FONT
        sheet.cell(notes_row + row_offset, 1).alignment = Alignment(vertical="center")

    widths = {
        "A": 28,
        "B": 18,
        "C": 18,
        "D": 20,
        "E": 12,
        "F": 3,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 14,
        "L": 14,
        "M": 14,
        "N": 14,
    }
    for letter, width in widths.items():
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A10"
    sheet.print_area = f"A1:N{notes_row + len(notes)}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def _build_verified_values_sheet(
    workbook: Workbook,
    frame: pd.DataFrame,
    weights: Mapping[str, float],
    preferences: Mapping[str, CriterionPreference],
    results: pd.DataFrame,
    steps: Mapping[str, Any],
    *,
    beta: float,
) -> None:
    sheet = workbook.create_sheet("Verified Values")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 90
    max_columns = max(len(frame.columns) + 1, 6)
    _set_title(sheet, 1, max_columns, "SYAI — Verified Numerical Values")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_columns)
    sheet.cell(
        2,
        1,
        "Static values from the canonical Python implementation at download time. Use "
        "them to audit the live formulas on the SYAI sheet.",
    )
    sheet.cell(2, 1).font = Font(name="Aptos", size=9, italic=True, color="666666")
    sheet.cell(2, 1).alignment = Alignment(wrap_text=True)

    _set_section_title(sheet, 4, 4, "Parameters")
    parameter_values = [
        ["Beta (β)", float(beta), "Constant C", 0.01],
        ["Alternatives", len(frame), "Criteria", len(frame.columns)],
    ]
    for row_offset, values in enumerate(parameter_values, start=5):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row_offset, column_index, value)
    _style_grid(sheet, 5, 6, 1, 4)
    sheet.cell(5, 2).number_format = "0.00"
    sheet.cell(5, 4).number_format = "0.00"

    settings_title_row = 8
    settings_header_row = 9
    settings_data_start = 10
    _set_section_title(sheet, settings_title_row, 4, "Criterion Settings")
    settings_headers = ["Criterion", "Weight", "Preference", "Ideal point x*"]
    for column_index, header in enumerate(settings_headers, start=1):
        sheet.cell(settings_header_row, column_index, header)
        sheet.cell(settings_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(settings_header_row, column_index).font = _HEADER_FONT
    for row_offset, criterion in enumerate(frame.columns, start=settings_data_start):
        name = str(criterion)
        preference = preferences[name]
        _set_text(sheet, row_offset, 1, name)
        sheet.cell(row_offset, 2, float(weights[name]))
        sheet.cell(row_offset, 2).number_format = _WEIGHT_FORMAT
        _set_text(sheet, row_offset, 3, _preference_label(preference))
        sheet.cell(row_offset, 3).fill = _preference_fill(preference)
        reference = (
            float(preference.target_value)
            if preference.kind is CriterionType.TARGET
            else float(frame[criterion].max())
            if preference.kind is CriterionType.BENEFIT
            else float(frame[criterion].min())
        )
        sheet.cell(row_offset, 4, reference)
        sheet.cell(row_offset, 4).number_format = _RAW_NUMBER_FORMAT
    settings_data_end = settings_data_start + len(frame.columns) - 1
    _style_grid(
        sheet,
        settings_header_row,
        settings_data_end,
        1,
        4,
        number_format=_CALC_NUMBER_FORMAT,
    )

    next_row = settings_data_end + 3
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 0 — Original Decision Matrix",
        frame,
        matrix_preferences=preferences,
    )
    normalized = steps["Step 1: Normalized Decision Matrix"].reindex(frame.index)
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 1 — Normalized Decision Matrix",
        normalized,
        matrix_preferences=preferences,
    )
    weighted = steps["Step 2: Weighted Normalized Matrix"].reindex(frame.index)
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 2 — Weighted Normalized Matrix",
        weighted,
        matrix_preferences=preferences,
    )

    ideal_values = pd.DataFrame(
        [
            steps["Step 3: Ideal Solutions"]["A+ (Yielded-Ideal Solution)"],
            steps["Step 3: Ideal Solutions"]["A- (Anti-Ideal Solution)"],
        ],
        index=["A+ (Yielded-Ideal)", "A- (Anti-Ideal)"],
    )
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 3 — Ideal Solutions",
        ideal_values,
        header_label="Solution",
        matrix_preferences=preferences,
    )

    distance_values = pd.concat(
        [
            steps["Step 4: Distances to Ideal Solutions"].reindex(frame.index),
            steps["Step 5: Final Closeness Score"].reindex(frame.index),
            results.reindex(frame.index)[["Rank"]],
        ],
        axis=1,
    )
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 4 & 5 — Distances, Closeness Score, and Rank",
        distance_values,
    )
    rank_column = distance_values.columns.get_loc("Rank") + 2
    distance_data_start = next_row - len(distance_values) - 1
    for row in range(distance_data_start, distance_data_start + len(distance_values)):
        sheet.cell(row, rank_column).number_format = "0"

    final_values = results[
        [
            "D+ (Dist to Ideal)",
            "D- (Dist to Anti-Ideal)",
            "Closeness Score (D_i)",
            "Rank",
        ]
    ]
    final_start = next_row
    _write_indexed_dataframe(
        sheet,
        final_start,
        "Final Ranking — Sorted by Rank",
        final_values,
    )
    final_data_start = final_start + 2
    for row in range(final_data_start, final_data_start + len(results)):
        sheet.cell(row, 5).number_format = "0"

    sheet.column_dimensions["A"].width = max(
        24,
        min(
            38,
            max(
                max(len(str(value)) for value in frame.index),
                max(len(str(value)) for value in frame.columns),
            )
            + 3,
        ),
    )
    for column_index in range(2, max_columns + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 22
    _autofit_column_widths(sheet)


def _build_formula_guide_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Formula Guide")
    sheet.sheet_view.showGridLines = False
    _set_title(sheet, 1, 5, "SYAI Formula Guide and Audit Trail")
    sheet.merge_cells("A2:E2")
    sheet["A2"] = (
        "The SYAI sheet contains live Excel formulas. Verified Values preserves the "
        "canonical numerical snapshot generated at download time."
    )
    sheet["A2"].alignment = Alignment(wrap_text=True)
    sheet["A2"].font = Font(name="Aptos", size=9, italic=True, color="666666")

    headers = ["Step", "Symbol", "Mathematical formula", "Meaning", "Excel implementation"]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(4, column_index, header)
        sheet.cell(4, column_index).fill = _SECTION_FILL
        sheet.cell(4, column_index).font = _HEADER_FONT

    rows = [
        (
            "0",
            "x*_j",
            "max(x_j), min(x_j), or target T_j",
            "Ideal point for benefit, cost, or target criteria.",
            "MAX / MIN / target input",
        ),
        (
            "1",
            "N_ij",
            "C + (1-C)[1 - |x_ij-x*_j| / R_j]",
            "Normalize around the ideal point, with C = 0.01.",
            "IF + ABS + MAX; constant columns return 1",
        ),
        (
            "2",
            "v_ij",
            "N_ij × w_j",
            "Apply the normalized criterion weight.",
            "Normalized cell × live weight cell",
        ),
        (
            "3",
            "A+_j",
            "max_i(v_ij)",
            "Yielded-ideal solution for each criterion.",
            "MAX(weighted criterion range)",
        ),
        (
            "3",
            "A-_j",
            "min_i(v_ij)",
            "Anti-ideal solution for each criterion.",
            "MIN(weighted criterion range)",
        ),
        (
            "4",
            "D+_i",
            "Σ_j |v_ij-A+_j|",
            "Manhattan distance to the yielded-ideal solution.",
            "Sum of criterion-level ABS terms",
        ),
        (
            "4",
            "D-_i",
            "Σ_j |v_ij-A-_j|",
            "Manhattan distance to the anti-ideal solution.",
            "Sum of criterion-level ABS terms",
        ),
        (
            "5",
            "D_i",
            "(1-β)D-_i / [βD+_i + (1-β)D-_i]",
            "Final closeness score; higher values are preferred.",
            "IF guard + live beta reference",
        ),
        (
            "6",
            "Rank_i",
            "descending competition rank of D_i",
            "Equal scores share a rank and skipped positions remain skipped.",
            "RANK(score, score range, 0)",
        ),
    ]
    for row_index, values in enumerate(rows, start=5):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row_index, column_index, value)
    _style_grid(sheet, 4, 4 + len(rows), 1, 5)
    for row in range(5, 5 + len(rows)):
        sheet.row_dimensions[row].height = 42

    audit_row = 6 + len(rows)
    _set_section_title(sheet, audit_row, 5, "Workbook Audit Notes")
    audit_notes = [
        "Yellow cells are intended inputs; gray cells are formulas or fixed method values.",
        "Changing beta or weights updates the live model, summary table, chart, and ranks.",
        "Verified Values remains a static reference to the calculation at download time.",
        "The Sort Order helper only makes the display order unique; it does not change tied ranks.",
    ]
    for row_offset, note in enumerate(audit_notes, start=1):
        sheet.cell(audit_row + row_offset, 1, f"• {note}")
        sheet.merge_cells(
            start_row=audit_row + row_offset,
            start_column=1,
            end_row=audit_row + row_offset,
            end_column=5,
        )
        sheet.cell(audit_row + row_offset, 1).font = _BODY_FONT

    source_row = audit_row + len(audit_notes) + 2
    _set_section_title(sheet, source_row, 5, "Method Reference")
    sheet.cell(source_row + 1, 1, "SYAI research paper")
    source_url = "https://www.ejpam.com/index.php/ejpam/article/view/6560/2443"
    sheet.cell(source_row + 1, 2, source_url)
    sheet.merge_cells(
        start_row=source_row + 1,
        start_column=2,
        end_row=source_row + 1,
        end_column=5,
    )
    sheet.cell(source_row + 1, 2).hyperlink = source_url
    sheet.cell(source_row + 1, 2).style = "Hyperlink"

    widths = {"A": 11, "B": 18, "C": 48, "D": 54, "E": 42}
    for letter, width in widths.items():
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A5"


def build_syai_excel_workbook(
    matrix: pd.DataFrame,
    weights: Mapping[str, Any],
    directions: Mapping[str, Any],
    *,
    beta: float = 0.5,
) -> bytes:
    """Return a complete SYAI workbook with live formulas and verified values."""

    frame = validate_crisp_matrix(matrix)
    frame = frame.copy()
    frame.columns = [str(column) for column in frame.columns]
    normalized_weights = validate_weights(weights, frame.columns, normalize=True)
    preferences = validate_method_capabilities("SYAI", frame.columns, directions)
    results, steps = calculate_syai(
        frame,
        normalized_weights,
        directions,
        beta=float(beta),
        return_steps=True,
    )

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "MCDM Calculator"
    workbook.properties.title = "Complete SYAI Formula-Driven Calculation"
    workbook.properties.subject = "Auditable SYAI decision model"
    workbook.properties.version = SYAI_EXCEL_EXPORT_REVISION
    workbook.properties.description = (
        "Live Excel formulas, a decision summary, and canonical numerical values for "
        f"every SYAI calculation stage. Export revision {SYAI_EXCEL_EXPORT_REVISION}."
    )
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    formula_positions = _build_formula_sheet(
        workbook,
        frame,
        normalized_weights,
        preferences,
        beta=float(beta),
    )
    _build_summary_sheet(workbook, formula_positions, len(frame))
    _build_verified_values_sheet(
        workbook,
        frame,
        normalized_weights,
        preferences,
        results,
        steps,
        beta=float(beta),
    )
    _build_formula_guide_sheet(workbook)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


__all__ = [
    "SYAI_EXCEL_EXPORT_FILENAME",
    "SYAI_EXCEL_EXPORT_REVISION",
    "build_syai_excel_workbook",
]
