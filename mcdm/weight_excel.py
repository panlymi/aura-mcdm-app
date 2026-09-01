"""Formula-rich Excel export for complete, auditable Objective & Deterministic Weight Calculations.

Implements step-by-step mathematical calculation sheets with live dynamic Excel formulas for:
  - Weight Summary (Master cross-method comparison with live links)
  - Entropy Weight Method (EWM / Shannon Entropy)
  - MEREC (Method based on the Removal Effects of Criteria)
  - CRITIC (Criteria Importance Through Intercriteria Correlation)
  - Standard Deviation (Dispersion-based weights)
  - Equal Weights (1/m)
  - PCA Loadings (Principal Component Analysis variance loadings)
"""

from __future__ import annotations

from io import BytesIO
import re
from typing import Any, Mapping
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from entropy_calculator import calculate_entropy_weights
from merec_calculator import calculate_merec_weights
from mcdm.criteria import CriterionPreference, CriterionType, normalize_directions
from mcdm.validation import validate_crisp_matrix, validate_weights
from mcdm.weighting import (
    calculate_critic_weights,
    calculate_pca_weights,
    calculate_sd_weights,
)


WEIGHT_EXCEL_EXPORT_REVISION = "v1"
WEIGHT_EXCEL_EXPORT_FILENAME = (
    f"mcdm_deterministic_weight_calculations_{WEIGHT_EXCEL_EXPORT_REVISION}.xlsx"
)

# Professional Palette
_NAVY_DARK = "1B365D"
_BLUE_MED = "2B5B84"
_ICE_BLUE = "D9EAF7"
_ALT_ROW = "F2F5F9"
_WHITE = "FFFFFF"
_INPUT_BG = "FFF2CC"
_SUMMARY_BG = "E2F0D9"
_ACCENT_BG = "FCE4D6"
_BORDER_GRAY = "D9D9D9"

_FONT_NAME = "Aptos"
_TITLE_FONT = Font(name=_FONT_NAME, size=15, bold=True, color=_WHITE)
_SECTION_FONT = Font(name=_FONT_NAME, size=11, bold=True, color=_NAVY_DARK)
_HEADER_FONT = Font(name=_FONT_NAME, size=10, bold=True, color=_WHITE)
_SUBHEADER_FONT = Font(name=_FONT_NAME, size=10, bold=True, color=_NAVY_DARK)
_BODY_FONT = Font(name=_FONT_NAME, size=10, color="1F1F1F")
_BODY_BOLD = Font(name=_FONT_NAME, size=10, bold=True, color="1F1F1F")
_NOTE_FONT = Font(name=_FONT_NAME, size=9, italic=True, color="595959")

_HEADER_FILL = PatternFill(start_color=_NAVY_DARK, end_color=_NAVY_DARK, fill_type="solid")
_SUBHEADER_FILL = PatternFill(start_color=_ICE_BLUE, end_color=_ICE_BLUE, fill_type="solid")
_ALT_FILL = PatternFill(start_color=_ALT_ROW, end_color=_ALT_ROW, fill_type="solid")
_WHITE_FILL = PatternFill(start_color=_WHITE, end_color=_WHITE, fill_type="solid")
_INPUT_FILL = PatternFill(start_color=_INPUT_BG, end_color=_INPUT_BG, fill_type="solid")
_SUMMARY_FILL = PatternFill(start_color=_SUMMARY_BG, end_color=_SUMMARY_BG, fill_type="solid")
_ACCENT_FILL = PatternFill(start_color=_ACCENT_BG, end_color=_ACCENT_BG, fill_type="solid")

_THIN_SIDE = Side(border_style="thin", color=_BORDER_GRAY)
_GRID_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)

_NUMBER_FORMAT_RAW = "General"
_NUMBER_FORMAT_DECIMAL = "0.0000"
_NUMBER_FORMAT_WEIGHT = "0.0000"
_NUMBER_FORMAT_PERCENT = "0.00%"

_ILLEGAL_CHAR_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def _clean_text(val: Any) -> str:
    return _ILLEGAL_CHAR_RE.sub(" ", str(val))[:32767]


def _set_cell(
    ws,
    row: int,
    col: int,
    val: Any,
    *,
    font: Font = _BODY_FONT,
    fill: PatternFill = _WHITE_FILL,
    border: Border = _GRID_BORDER,
    align: str = "right",
    num_format: str = _NUMBER_FORMAT_DECIMAL,
) -> None:
    cell = ws.cell(row=row, column=col)
    if isinstance(val, str) and val.startswith("="):
        cell.value = val
    elif isinstance(val, (int, float, np.number)):
        cell.value = float(val) if isinstance(val, (float, np.floating)) else int(val)
    elif val is None:
        cell.value = ""
    else:
        cell.value = _clean_text(val)
        cell.data_type = "s"

    cell.font = font
    cell.fill = fill
    cell.border = border
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if num_format:
        cell.number_format = num_format


def _set_banner(ws, title: str, subtitle: str) -> None:
    ws.sheet_view.showGridLines = True
    for c in range(1, 15):
        ws.cell(1, c).fill = PatternFill(start_color=_NAVY_DARK, end_color=_NAVY_DARK, fill_type="solid")
    ws.cell(1, 1, title).font = _TITLE_FONT
    ws.row_dimensions[1].height = 28
    ws.cell(2, 1, subtitle).font = _NOTE_FONT
    ws.row_dimensions[2].height = 18


def _autofit_columns(ws, max_cols: int = 15) -> None:
    for col_idx in range(1, max_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(3, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            val = str(cell.value or "")
            if cell.number_format and "0.00" in cell.number_format and isinstance(cell.value, (int, float)):
                val = f"{cell.value:.4f}"
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 13)


def _build_entropy_sheet(
    ws,
    matrix: pd.DataFrame,
    preferences: Mapping[str, CriterionPreference],
) -> dict[str, str]:
    """Build live formula EWM calculation sheet and return mapping of criterion to final weight cell."""
    _set_banner(
        ws,
        "Entropy Weight Method (EWM / Shannon Entropy)",
        "Step-by-step information entropy calculation with live dynamic formulas.",
    )
    n_alts, m_crit = matrix.shape
    alts = list(matrix.index)
    criteria = list(matrix.columns)

    # Section 1: Raw Decision Matrix
    r = 4
    ws.cell(r, 1, "1. Decision Matrix (Raw Inputs)").font = _SECTION_FONT
    r += 1
    _set_cell(ws, r, 1, "Alternative", font=_HEADER_FONT, fill=_HEADER_FILL, align="left")
    for j, c in enumerate(criteria, start=2):
        pref = preferences[c].to_legacy().capitalize()
        _set_cell(ws, r, j, f"{c} ({pref})", font=_HEADER_FONT, fill=_HEADER_FILL, align="center")

    matrix_start_row = r + 1
    for i, alt in enumerate(alts, start=matrix_start_row):
        fill = _ALT_FILL if i % 2 == 1 else _WHITE_FILL
        _set_cell(ws, i, 1, alt, font=_BODY_BOLD, fill=fill, align="left")
        for j, c in enumerate(criteria, start=2):
            _set_cell(ws, i, j, matrix.loc[alt, c], fill=_INPUT_FILL, align="right", num_format=_NUMBER_FORMAT_RAW)
    matrix_end_row = matrix_start_row + n_alts - 1

    # Sum of columns
    r = matrix_end_row + 1
    _set_cell(ws, r, 1, "Column Sum (Σ x_ij)", font=_BODY_BOLD, fill=_SUBHEADER_FILL, align="left")
    for j in range(2, m_crit + 2):
        col_let = get_column_letter(j)
        _set_cell(ws, r, j, f"=SUM({col_let}{matrix_start_row}:{col_let}{matrix_end_row})", font=_BODY_BOLD, fill=_SUBHEADER_FILL)
    sum_row = r

    # Section 2: Proportions p_ij = x_ij / sum_i(x_ij)
    r += 3
    ws.cell(r, 1, "2. Normalized Proportions (p_ij = x_ij / Σ x_ij)").font = _SECTION_FONT
    r += 1
    _set_cell(ws, r, 1, "Alternative", font=_HEADER_FONT, fill=_HEADER_FILL, align="left")
    for j, c in enumerate(criteria, start=2):
        _set_cell(ws, r, j, c, font=_HEADER_FONT, fill=_HEADER_FILL, align="center")

    p_start_row = r + 1
    for row_offset, alt in enumerate(alts):
        curr_r = p_start_row + row_offset
        mat_r = matrix_start_row + row_offset
        fill = _ALT_FILL if row_offset % 2 == 1 else _WHITE_FILL
        _set_cell(ws, curr_r, 1, alt, font=_BODY_BOLD, fill=fill, align="left")
        for j in range(2, m_crit + 2):
            col_let = get_column_letter(j)
            _set_cell(ws, curr_r, j, f"={col_let}{mat_r}/{col_let}${sum_row}", fill=fill, num_format=_NUMBER_FORMAT_DECIMAL)
    p_end_row = p_start_row + n_alts - 1

    # Section 3: Entropy Terms p_ij * ln(p_ij)
    r = p_end_row + 3
    ws.cell(r, 1, "3. Entropy Terms (p_ij · ln(p_ij))").font = _SECTION_FONT
    r += 1
    _set_cell(ws, r, 1, "Alternative", font=_HEADER_FONT, fill=_HEADER_FILL, align="left")
    for j, c in enumerate(criteria, start=2):
        _set_cell(ws, r, j, c, font=_HEADER_FONT, fill=_HEADER_FILL, align="center")

    pln_start_row = r + 1
    for row_offset, alt in enumerate(alts):
        curr_r = pln_start_row + row_offset
        p_r = p_start_row + row_offset
        fill = _ALT_FILL if row_offset % 2 == 1 else _WHITE_FILL
        _set_cell(ws, curr_r, 1, alt, font=_BODY_BOLD, fill=fill, align="left")
        for j in range(2, m_crit + 2):
            col_let = get_column_letter(j)
            _set_cell(ws, curr_r, j, f"=IF({col_let}{p_r}>0, {col_let}{p_r}*LN({col_let}{p_r}), 0)", fill=fill, num_format=_NUMBER_FORMAT_DECIMAL)
    pln_end_row = pln_start_row + n_alts - 1

    # Section 4: Information Entropy e_j, Diversification d_j, Weights w_j
    r = pln_end_row + 3
    ws.cell(r, 1, "4. Information Entropy & Final Weights").font = _SECTION_FONT
    r += 1

    # Entropy constant k = 1 / ln(m)
    _set_cell(ws, r, 1, "Entropy Constant k = 1/ln(m)", font=_BODY_BOLD, fill=_SUBHEADER_FILL, align="left")
    _set_cell(ws, r, 2, f"=1/LN({n_alts})", font=_BODY_BOLD, fill=_SUBHEADER_FILL, num_format=_NUMBER_FORMAT_DECIMAL)
    k_row = r

    r += 1
    _set_cell(ws, r, 1, "Metric / Step", font=_HEADER_FONT, fill=_HEADER_FILL, align="left")
    for j, c in enumerate(criteria, start=2):
        _set_cell(ws, r, j, c, font=_HEADER_FONT, fill=_HEADER_FILL, align="center")
    _set_cell(ws, r, m_crit + 2, "Sum / Total", font=_HEADER_FONT, fill=_HEADER_FILL, align="center")

    # Row 1: Sum of p*ln(p)
    r += 1
    _set_cell(ws, r, 1, "Σ p_ij · ln(p_ij)", font=_BODY_BOLD, fill=_WHITE_FILL, align="left")
    for j in range(2, m_crit + 2):
        col_let = get_column_letter(j)
        _set_cell(ws, r, j, f"=SUM({col_let}{pln_start_row}:{col_let}{pln_end_row})", fill=_WHITE_FILL)
    pln_sum_row = r

    # Row 2: Information Entropy e_j = -k * sum
    r += 1
    _set_cell(ws, r, 1, "Entropy e_j = -k · Σ p_ij ln(p_ij)", font=_BODY_BOLD, fill=_ALT_FILL, align="left")
    for j in range(2, m_crit + 2):
        col_let = get_column_letter(j)
        _set_cell(ws, r, j, f"=-$B${k_row}*{col_let}{pln_sum_row}", fill=_ALT_FILL)
    e_row = r

    # Row 3: Degree of Diversification d_j = 1 - e_j
    r += 1
    _set_cell(ws, r, 1, "Diversification d_j = 1 - e_j", font=_BODY_BOLD, fill=_WHITE_FILL, align="left")
    for j in range(2, m_crit + 2):
        col_let = get_column_letter(j)
        _set_cell(ws, r, j, f"=1-{col_let}{e_row}", fill=_WHITE_FILL)
    first_d_col = get_column_letter(2)
    last_d_col = get_column_letter(m_crit + 1)
    _set_cell(ws, r, m_crit + 2, f"=SUM({first_d_col}{r}:{last_d_col}{r})", font=_BODY_BOLD, fill=_SUBHEADER_FILL)
    d_row = r

    # Row 4: Final Weights w_j = d_j / sum(d_k)
    r += 1
    _set_cell(ws, r, 1, "Final Entropy Weight (w_j)", font=_BODY_BOLD, fill=_SUMMARY_FILL, align="left")
    weight_cells = {}
    sum_d_col = get_column_letter(m_crit + 2)
    for j, c in enumerate(criteria, start=2):
        col_let = get_column_letter(j)
        _set_cell(ws, r, j, f"={col_let}{d_row}/${sum_d_col}${d_row}", font=_BODY_BOLD, fill=_SUMMARY_FILL, num_format=_NUMBER_FORMAT_WEIGHT)
        weight_cells[c] = f"'Entropy (EWM)'!${col_let}${r}"
    _set_cell(ws, r, m_crit + 2, f"=SUM({first_d_col}{r}:{last_d_col}{r})", font=_BODY_BOLD, fill=_SUMMARY_FILL, num_format=_NUMBER_FORMAT_WEIGHT)

    _autofit_columns(ws, m_crit + 3)
    return weight_cells


def _build_critic_sheet(
    ws,
    matrix: pd.DataFrame,
    preferences: Mapping[str, CriterionPreference],
) -> dict[str, str]:
    """Build live formula CRITIC calculation sheet with CORREL correlation matrix."""
    _set_banner(
        ws,
        "CRITIC Objective Weight Method",
        "Criteria Importance Through Intercriteria Correlation (Diakoulaki et al., 1995).",
    )
    n_alts, m_crit = matrix.shape
    alts = list(matrix.index)
    criteria = list(matrix.columns)

    # Section 1: Decision Matrix
    r = 4
    ws.cell(r, 1, "1. Decision Matrix (Raw Inputs)").font = _SECTION_FONT
    r += 1
    _set_cell(ws, r, 1, "Alternative", font=_HEADER_FONT, fill=_HEADER_FILL, align="left")
    for j, c in enumerate(criteria, start=2):
        pref = preferences[c].to_legacy().capitalize()
        _set_cell(ws, r, j, f"{c} ({pref})", font=_HEADER_FONT, fill=_HEADER_FILL, align="center")

    matrix_start_row = r + 1
    for i, alt in enumerate(alts, start=matrix_start_row):
        fill = _ALT_FILL if i % 2 == 1 else _WHITE_FILL
        _set_cell(ws, i, 1, alt, font=_BODY_BOLD, fill=fill, align="left")
        for j, c in enumerate(criteria, start=2):
            _set_cell(ws, i, j, matrix.loc[alt, c], fill=_INPUT_FILL, align="right", num_format=_NUMBER_FORMAT_RAW)
    matrix_end_row = matrix_start_row + n_alts - 1

    # Min and Max reference rows
    r = matrix_end_row + 1
    _set_cell(ws, r, 1, "Minimum (x_j^min)", font=_BODY_BOLD, fill=_SUBHEADER_FILL, align="left")
    for j in range(2, m_crit + 2):
        col_let = get_column_letter(j)
        _set_cell(ws, r, j, f"=MIN({col_let}{matrix_start_row}:{col_let}{matrix_end_row})", fill=_SUBHEADER_FILL)
    min_row = r

    r += 1
    _set_cell(ws, r, 1, "Maximum (x_j^max)", font=_BODY_BOLD, fill=_SUBHEADER_FILL, align="left")
    for j in range(2, m_crit + 2):
        col_let = get_column_letter(j)
        _set_cell(ws, r, j, f"=MAX({col_let}{matrix_start_row}:{col_let}{matrix_end_row})", fill=_SUBHEADER_FILL)
    max_row = r

    # Section 2: Min-Max Normalization to [0, 1]
    r += 3
    ws.cell(r, 1, "2. Min-Max Normalization (r_ij ∈ [0, 1])").font = _SECTION_FONT
    r += 1
    _set_cell(ws, r, 1, "Alternative", font=_HEADER_FONT, fill=_HEADER_FILL, align="left")
    for j, c in enumerate(criteria, start=2):
        _set_cell(ws, r, j, c, font=_HEADER_FONT, fill=_HEADER_FILL, align="center")

    norm_start_row = r + 1
    for row_offset, alt in enumerate(alts):
        curr_r = norm_start_row + row_offset
        mat_r = matrix_start_row + row_offset
        fill = _ALT_FILL if row_offset % 2 == 1 else _WHITE_FILL
        _set_cell(ws, curr_r, 1, alt, font=_BODY_BOLD, fill=fill, align="left")
        for j, c in enumerate(criteria, start=2):
            col_let = get_column_letter(j)
            is_cost = preferences[c].kind is CriterionType.COST
            if is_cost:
                formula = f"=IF({col_let}${max_row}={col_let}${min_row}, 1, ({col_let}${max_row}-{col_let}{mat_r})/({col_let}${max_row}-{col_let}${min_row}))"
            else:
                formula = f"=IF({col_let}${max_row}={col_let}${min_row}, 1, ({col_let}{mat_r}-{col_let}${min_row})/({col_let}${max_row}-{col_let}${min_row}))"
            _set_cell(ws, curr_r, j, formula, fill=fill, num_format=_NUMBER_FORMAT_DECIMAL)
    norm_end_row = norm_start_row + n_alts - 1

    # Standard Deviation of Normalized Columns
    r = norm_end_row + 1
    _set_cell(ws, r, 1, "Std Deviation σ_j = STDEV.P(r_ij)", font=_BODY_BOLD, fill=_SUBHEADER_FILL, align="left")
    for j in range(2, m_crit + 2):
        col_let = get_column_letter(j)
        _set_cell(ws, r, j, f"=STDEV.P({col_let}{norm_start_row}:{col_let}{norm_end_row})", font=_BODY_BOLD, fill=_SUBHEADER_FILL)
    std_row = r

    # Section 3: Intercriteria Correlation Matrix R_jk
    r += 3
    ws.cell(r, 1, "3. Intercriteria Correlation Matrix (R_jk = CORREL(r_j, r_k))").font = _SECTION_FONT
    r += 1
    _set_cell(ws, r, 1, "Criterion", font=_HEADER_FONT, fill=_HEADER_FILL, align="left")
    for j, c in enumerate(criteria, start=2):
        _set_cell(ws, r, j, c, font=_HEADER_FONT, fill=_HEADER_FILL, align="center")

    corr_start_row = r + 1
    for i_idx, c1 in enumerate(criteria):
        curr_r = corr_start_row + i_idx
        col1_let = get_column_letter(i_idx + 2)
        _set_cell(ws, curr_r, 1, c1, font=_BODY_BOLD, fill=_WHITE_FILL, align="left")
        for j_idx, c2 in enumerate(criteria):
            col2_let = get_column_letter(j_idx + 2)
            c_pos = j_idx + 2
            if i_idx == j_idx:
                _set_cell(ws, curr_r, c_pos, 1.0, fill=_WHITE_FILL, num_format=_NUMBER_FORMAT_DECIMAL)
            else:
                _set_cell(ws, curr_r, c_pos, f"=CORREL({col1_let}${norm_start_row}:{col1_let}${norm_end_row}, {col2_let}${norm_start_row}:{col2_let}${norm_end_row})", fill=_WHITE_FILL, num_format=_NUMBER_FORMAT_DECIMAL)
    corr_end_row = corr_start_row + m_crit - 1

    # Section 4: Conflict, Information Quantity C_j, and Final Weights
    r = corr_end_row + 3
    ws.cell(r, 1, "4. Information Quantity C_j & Final CRITIC Weights").font = _SECTION_FONT
    r += 1
    _set_cell(ws, r, 1, "Metric / Step", font=_HEADER_FONT, fill=_HEADER_FILL, align="left")
    for j, c in enumerate(criteria, start=2):
        _set_cell(ws, r, j, c, font=_HEADER_FONT, fill=_HEADER_FILL, align="center")
    _set_cell(ws, r, m_crit + 2, "Sum / Total", font=_HEADER_FONT, fill=_HEADER_FILL, align="center")

    # Row 1: Conflict Measure Σ (1 - r_jk)
    r += 1
    _set_cell(ws, r, 1, "Conflict Σ (1 - r_jk)", font=_BODY_BOLD, fill=_WHITE_FILL, align="left")
    for j_idx, c in enumerate(criteria):
        col_pos = j_idx + 2
        corr_r = corr_start_row + j_idx
        first_c = get_column_letter(2)
        last_c = get_column_letter(m_crit + 1)
        _set_cell(ws, r, col_pos, f"={m_crit}-SUM({first_c}{corr_r}:{last_c}{corr_r})", fill=_WHITE_FILL)
    conflict_row = r

    # Row 2: Information Quantity C_j = σ_j * Conflict
    r += 1
    _set_cell(ws, r, 1, "Information Quantity C_j = σ_j · Σ(1 - r_jk)", font=_BODY_BOLD, fill=_ALT_FILL, align="left")
    for j in range(2, m_crit + 2):
        col_let = get_column_letter(j)
        _set_cell(ws, r, j, f"={col_let}${std_row}*{col_let}{conflict_row}", fill=_ALT_FILL)
    first_c_col = get_column_letter(2)
    last_c_col = get_column_letter(m_crit + 1)
    _set_cell(ws, r, m_crit + 2, f"=SUM({first_c_col}{r}:{last_c_col}{r})", font=_BODY_BOLD, fill=_SUBHEADER_FILL)
    c_info_row = r

    # Row 3: Final CRITIC Weights w_j = C_j / sum(C_k)
    r += 1
    _set_cell(ws, r, 1, "Final CRITIC Weight (w_j)", font=_BODY_BOLD, fill=_SUMMARY_FILL, align="left")
    weight_cells = {}
    sum_c_col = get_column_letter(m_crit + 2)
    for j, c in enumerate(criteria, start=2):
        col_let = get_column_letter(j)
        _set_cell(ws, r, j, f"={col_let}{c_info_row}/${sum_c_col}${c_info_row}", font=_BODY_BOLD, fill=_SUMMARY_FILL, num_format=_NUMBER_FORMAT_WEIGHT)
        weight_cells[c] = f"'CRITIC'!${col_let}${r}"
    _set_cell(ws, r, m_crit + 2, f"=SUM({first_c_col}{r}:{last_c_col}{r})", font=_BODY_BOLD, fill=_SUMMARY_FILL, num_format=_NUMBER_FORMAT_WEIGHT)

    _autofit_columns(ws, m_crit + 3)
    return weight_cells


def _build_sd_sheet(
    ws,
    matrix: pd.DataFrame,
    preferences: Mapping[str, CriterionPreference],
) -> dict[str, str]:
    """Build live formula Standard Deviation calculation sheet."""
    _set_banner(
        ws,
        "Standard Deviation Objective Weights",
        "Weights proportional to criterion dispersion / contrast intensity across alternatives.",
    )
    n_alts, m_crit = matrix.shape
    alts = list(matrix.index)
    criteria = list(matrix.columns)

    # Section 1: Decision Matrix
    r = 4
    ws.cell(r, 1, "1. Decision Matrix (Raw Inputs)").font = _SECTION_FONT
    r += 1
    _set_cell(ws, r, 1, "Alternative", font=_HEADER_FONT, fill=_HEADER_FILL, align="left")
    for j, c in enumerate(criteria, start=2):
        pref = preferences[c].to_legacy().capitalize()
        _set_cell(ws, r, j, f"{c} ({pref})", font=_HEADER_FONT, fill=_HEADER_FILL, align="center")

    matrix_start_row = r + 1
    for i, alt in enumerate(alts, start=matrix_start_row):
        fill = _ALT_FILL if i % 2 == 1 else _WHITE_FILL
        _set_cell(ws, i, 1, alt, font=_BODY_BOLD, fill=fill, align="left")
        for j, c in enumerate(criteria, start=2):
            _set_cell(ws, i, j, matrix.loc[alt, c], fill=_INPUT_FILL, align="right", num_format=_NUMBER_FORMAT_RAW)
    matrix_end_row = matrix_start_row + n_alts - 1

    # Min and Max reference rows
    r = matrix_end_row + 1
    _set_cell(ws, r, 1, "Minimum (x_j^min)", font=_BODY_BOLD, fill=_SUBHEADER_FILL, align="left")
    for j in range(2, m_crit + 2):
        col_let = get_column_letter(j)
        _set_cell(ws, r, j, f"=MIN({col_let}{matrix_start_row}:{col_let}{matrix_end_row})", fill=_SUBHEADER_FILL)
    min_row = r

    r += 1
    _set_cell(ws, r, 1, "Maximum (x_j^max)", font=_BODY_BOLD, fill=_SUBHEADER_FILL, align="left")
    for j in range(2, m_crit + 2):
        col_let = get_column_letter(j)
        _set_cell(ws, r, j, f"=MAX({col_let}{matrix_start_row}:{col_let}{matrix_end_row})", fill=_SUBHEADER_FILL)
    max_row = r

    # Section 2: Min-Max Normalization
    r += 3
    ws.cell(r, 1, "2. Min-Max Normalization (r_ij ∈ [0, 1])").font = _SECTION_FONT
    r += 1
    _set_cell(ws, r, 1, "Alternative", font=_HEADER_FONT, fill=_HEADER_FILL, align="left")
    for j, c in enumerate(criteria, start=2):
        _set_cell(ws, r, j, c, font=_HEADER_FONT, fill=_HEADER_FILL, align="center")

    norm_start_row = r + 1
    for row_offset, alt in enumerate(alts):
        curr_r = norm_start_row + row_offset
        mat_r = matrix_start_row + row_offset
        fill = _ALT_FILL if row_offset % 2 == 1 else _WHITE_FILL
        _set_cell(ws, curr_r, 1, alt, font=_BODY_BOLD, fill=fill, align="left")
        for j, c in enumerate(criteria, start=2):
            col_let = get_column_letter(j)
            is_cost = preferences[c].kind is CriterionType.COST
            if is_cost:
                formula = f"=IF({col_let}${max_row}={col_let}${min_row}, 1, ({col_let}${max_row}-{col_let}{mat_r})/({col_let}${max_row}-{col_let}${min_row}))"
            else:
                formula = f"=IF({col_let}${max_row}={col_let}${min_row}, 1, ({col_let}{mat_r}-{col_let}${min_row})/({col_let}${max_row}-{col_let}${min_row}))"
            _set_cell(ws, curr_r, j, formula, fill=fill, num_format=_NUMBER_FORMAT_DECIMAL)
    norm_end_row = norm_start_row + n_alts - 1

    # Section 3: Standard Deviations & Final Weights
    r = norm_end_row + 3
    ws.cell(r, 1, "3. Standard Deviations & Final Weights").font = _SECTION_FONT
    r += 1
    _set_cell(ws, r, 1, "Metric / Step", font=_HEADER_FONT, fill=_HEADER_FILL, align="left")
    for j, c in enumerate(criteria, start=2):
        _set_cell(ws, r, j, c, font=_HEADER_FONT, fill=_HEADER_FILL, align="center")
    _set_cell(ws, r, m_crit + 2, "Sum / Total", font=_HEADER_FONT, fill=_HEADER_FILL, align="center")

    # Row 1: Std Dev
    r += 1
    _set_cell(ws, r, 1, "Std Deviation σ_j = STDEV.P(r_ij)", font=_BODY_BOLD, fill=_ALT_FILL, align="left")
    for j in range(2, m_crit + 2):
        col_let = get_column_letter(j)
        _set_cell(ws, r, j, f"=STDEV.P({col_let}{norm_start_row}:{col_let}{norm_end_row})", fill=_ALT_FILL)
    first_col = get_column_letter(2)
    last_col = get_column_letter(m_crit + 1)
    _set_cell(ws, r, m_crit + 2, f"=SUM({first_col}{r}:{last_col}{r})", font=_BODY_BOLD, fill=_SUBHEADER_FILL)
    std_row = r

    # Row 2: Final Weights
    r += 1
    _set_cell(ws, r, 1, "Final SD Weight (w_j = σ_j / Σ σ_k)", font=_BODY_BOLD, fill=_SUMMARY_FILL, align="left")
    weight_cells = {}
    sum_col = get_column_letter(m_crit + 2)
    for j, c in enumerate(criteria, start=2):
        col_let = get_column_letter(j)
        _set_cell(ws, r, j, f"={col_let}{std_row}/${sum_col}${std_row}", font=_BODY_BOLD, fill=_SUMMARY_FILL, num_format=_NUMBER_FORMAT_WEIGHT)
        weight_cells[c] = f"'Standard Deviation'!${col_let}${r}"
    _set_cell(ws, r, m_crit + 2, f"=SUM({first_col}{r}:{last_col}{r})", font=_BODY_BOLD, fill=_SUMMARY_FILL, num_format=_NUMBER_FORMAT_WEIGHT)

    _autofit_columns(ws, m_crit + 3)
    return weight_cells


def _build_merec_sheet(
    ws,
    matrix: pd.DataFrame,
    preferences: Mapping[str, CriterionPreference],
) -> dict[str, str]:
    """Build live formula MEREC calculation sheet."""
    _set_banner(
        ws,
        "MEREC (Method based on the Removal Effects of Criteria)",
        "Weights derived from overall performance deviation upon criterion removal.",
    )
    n_alts, m_crit = matrix.shape
    alts = list(matrix.index)
    criteria = list(matrix.columns)

    # Section 1: Decision Matrix
    r = 4
    ws.cell(r, 1, "1. Decision Matrix (Raw Inputs)").font = _SECTION_FONT
    r += 1
    _set_cell(ws, r, 1, "Alternative", font=_HEADER_FONT, fill=_HEADER_FILL, align="left")
    for j, c in enumerate(criteria, start=2):
        pref = preferences[c].to_legacy().capitalize()
        _set_cell(ws, r, j, f"{c} ({pref})", font=_HEADER_FONT, fill=_HEADER_FILL, align="center")

    matrix_start_row = r + 1
    for i, alt in enumerate(alts, start=matrix_start_row):
        fill = _ALT_FILL if i % 2 == 1 else _WHITE_FILL
        _set_cell(ws, i, 1, alt, font=_BODY_BOLD, fill=fill, align="left")
        for j, c in enumerate(criteria, start=2):
            _set_cell(ws, i, j, matrix.loc[alt, c], fill=_INPUT_FILL, align="right", num_format=_NUMBER_FORMAT_RAW)
    matrix_end_row = matrix_start_row + n_alts - 1

    # Min and Max reference rows
    r = matrix_end_row + 1
    _set_cell(ws, r, 1, "Minimum (x_j^min)", font=_BODY_BOLD, fill=_SUBHEADER_FILL, align="left")
    for j in range(2, m_crit + 2):
        col_let = get_column_letter(j)
        _set_cell(ws, r, j, f"=MIN({col_let}{matrix_start_row}:{col_let}{matrix_end_row})", fill=_SUBHEADER_FILL)
    min_row = r

    r += 1
    _set_cell(ws, r, 1, "Maximum (x_j^max)", font=_BODY_BOLD, fill=_SUBHEADER_FILL, align="left")
    for j in range(2, m_crit + 2):
        col_let = get_column_letter(j)
        _set_cell(ws, r, j, f"=MAX({col_let}{matrix_start_row}:{col_let}{matrix_end_row})", fill=_SUBHEADER_FILL)
    max_row = r

    # Section 2: MEREC Normalization (N_ij)
    r += 3
    ws.cell(r, 1, "2. MEREC Normalization (N_ij)").font = _SECTION_FONT
    r += 1
    _set_cell(ws, r, 1, "Alternative", font=_HEADER_FONT, fill=_HEADER_FILL, align="left")
    for j, c in enumerate(criteria, start=2):
        _set_cell(ws, r, j, c, font=_HEADER_FONT, fill=_HEADER_FILL, align="center")

    norm_start_row = r + 1
    for row_offset, alt in enumerate(alts):
        curr_r = norm_start_row + row_offset
        mat_r = matrix_start_row + row_offset
        fill = _ALT_FILL if row_offset % 2 == 1 else _WHITE_FILL
        _set_cell(ws, curr_r, 1, alt, font=_BODY_BOLD, fill=fill, align="left")
        for j, c in enumerate(criteria, start=2):
            col_let = get_column_letter(j)
            is_cost = preferences[c].kind is CriterionType.COST
            if is_cost:
                formula = f"={col_let}{mat_r}/{col_let}${max_row}"
            else:
                formula = f"={col_let}${min_row}/{col_let}{mat_r}"
            _set_cell(ws, curr_r, j, formula, fill=fill, num_format=_NUMBER_FORMAT_DECIMAL)
    norm_end_row = norm_start_row + n_alts - 1

    # Section 3: Logarithmic Penalties |ln(N_ij)|
    r = norm_end_row + 3
    ws.cell(r, 1, "3. Logarithmic Penalties (|ln(N_ij)|)").font = _SECTION_FONT
    r += 1
    _set_cell(ws, r, 1, "Alternative", font=_HEADER_FONT, fill=_HEADER_FILL, align="left")
    for j, c in enumerate(criteria, start=2):
        _set_cell(ws, r, j, c, font=_HEADER_FONT, fill=_HEADER_FILL, align="center")
    _set_cell(ws, r, m_crit + 2, "Overall Performance S_i", font=_HEADER_FONT, fill=_HEADER_FILL, align="center")

    log_start_row = r + 1
    first_c_let = get_column_letter(2)
    last_c_let = get_column_letter(m_crit + 1)
    s_col_let = get_column_letter(m_crit + 2)

    for row_offset, alt in enumerate(alts):
        curr_r = log_start_row + row_offset
        norm_r = norm_start_row + row_offset
        fill = _ALT_FILL if row_offset % 2 == 1 else _WHITE_FILL
        _set_cell(ws, curr_r, 1, alt, font=_BODY_BOLD, fill=fill, align="left")
        for j in range(2, m_crit + 2):
            col_let = get_column_letter(j)
            _set_cell(ws, curr_r, j, f"=ABS(LN({col_let}{norm_r}))", fill=fill, num_format=_NUMBER_FORMAT_DECIMAL)
        # S_i = ln(1 + (1/m) * sum |ln(N_ij)|)
        _set_cell(ws, curr_r, m_crit + 2, f"=LN(1 + (1/{m_crit})*SUM({first_c_let}{curr_r}:{last_c_let}{curr_r}))", font=_BODY_BOLD, fill=_SUBHEADER_FILL, num_format=_NUMBER_FORMAT_DECIMAL)
    log_end_row = log_start_row + n_alts - 1

    # Section 4: Performance Without Criterion S'_ij
    r = log_end_row + 3
    ws.cell(r, 1, "4. Performance Without Criterion (S'_ij)").font = _SECTION_FONT
    r += 1
    _set_cell(ws, r, 1, "Alternative", font=_HEADER_FONT, fill=_HEADER_FILL, align="left")
    for j, c in enumerate(criteria, start=2):
        _set_cell(ws, r, j, f"S'_i(w/o {c})", font=_HEADER_FONT, fill=_HEADER_FILL, align="center")

    sprime_start_row = r + 1
    for row_offset, alt in enumerate(alts):
        curr_r = sprime_start_row + row_offset
        log_r = log_start_row + row_offset
        fill = _ALT_FILL if row_offset % 2 == 1 else _WHITE_FILL
        _set_cell(ws, curr_r, 1, alt, font=_BODY_BOLD, fill=fill, align="left")
        for j in range(2, m_crit + 2):
            col_let = get_column_letter(j)
            _set_cell(ws, curr_r, j, f"=LN(1 + (1/{m_crit})*(SUM({first_c_let}{log_r}:{last_c_let}{log_r}) - {col_let}{log_r}))", fill=fill, num_format=_NUMBER_FORMAT_DECIMAL)
    sprime_end_row = sprime_start_row + n_alts - 1

    # Section 5: Removal Effects E_j & Final Weights
    r = sprime_end_row + 3
    ws.cell(r, 1, "5. Removal Effects E_j & Final MEREC Weights").font = _SECTION_FONT
    r += 1
    _set_cell(ws, r, 1, "Metric / Step", font=_HEADER_FONT, fill=_HEADER_FILL, align="left")
    for j, c in enumerate(criteria, start=2):
        _set_cell(ws, r, j, c, font=_HEADER_FONT, fill=_HEADER_FILL, align="center")
    _set_cell(ws, r, m_crit + 2, "Sum / Total", font=_HEADER_FONT, fill=_HEADER_FILL, align="center")

    # Row 1: Removal Effects E_j = sum |S'_ij - S_i|
    # Note: Using individual column sum for clean Excel formulas without array ctrl+shift+enter
    # We can write an explicit sum of absolute differences
    r += 1
    _set_cell(ws, r, 1, "Removal Effect E_j = Σ |S'_ij - S_i|", font=_BODY_BOLD, fill=_ALT_FILL, align="left")
    for j in range(2, m_crit + 2):
        col_let = get_column_letter(j)
        # Construct formula summing ABS(cell - S_i)
        terms = [f"ABS({col_let}{sprime_start_row + k}-${s_col_let}${log_start_row + k})" for k in range(n_alts)]
        _set_cell(ws, r, j, f"={'+'.join(terms)}", fill=_ALT_FILL)
    first_e_col = get_column_letter(2)
    last_e_col = get_column_letter(m_crit + 1)
    _set_cell(ws, r, m_crit + 2, f"=SUM({first_e_col}{r}:{last_e_col}{r})", font=_BODY_BOLD, fill=_SUBHEADER_FILL)
    e_row = r

    # Row 2: Final MEREC Weights
    r += 1
    _set_cell(ws, r, 1, "Final MEREC Weight (w_j = E_j / Σ E_k)", font=_BODY_BOLD, fill=_SUMMARY_FILL, align="left")
    weight_cells = {}
    sum_e_col = get_column_letter(m_crit + 2)
    for j, c in enumerate(criteria, start=2):
        col_let = get_column_letter(j)
        _set_cell(ws, r, j, f"={col_let}{e_row}/${sum_e_col}${e_row}", font=_BODY_BOLD, fill=_SUMMARY_FILL, num_format=_NUMBER_FORMAT_WEIGHT)
        weight_cells[c] = f"'MEREC'!${col_let}${r}"
    _set_cell(ws, r, m_crit + 2, f"=SUM({first_e_col}{r}:{last_e_col}{r})", font=_BODY_BOLD, fill=_SUMMARY_FILL, num_format=_NUMBER_FORMAT_WEIGHT)

    _autofit_columns(ws, m_crit + 3)
    return weight_cells


def _build_pca_sheet(
    ws,
    matrix: pd.DataFrame,
    preferences: Mapping[str, CriterionPreference],
) -> dict[str, str]:
    """Build PCA loadings calculation sheet."""
    _set_banner(
        ws,
        "PCA Loadings Weight Method",
        "Objective weights derived from principal component loadings weighted by explained variance.",
    )
    n_alts, m_crit = matrix.shape
    alts = list(matrix.index)
    criteria = list(matrix.columns)

    pca_weights, pca_steps = calculate_pca_weights(matrix, {c: p.to_legacy() for c, p in preferences.items()})

    # Section 1: Standardized Data Matrix Z
    r = 4
    ws.cell(r, 1, "1. Standardized Matrix (Z_ij = (x_ij - μ_j) / σ_j)").font = _SECTION_FONT
    r += 1
    _set_cell(ws, r, 1, "Alternative", font=_HEADER_FONT, fill=_HEADER_FILL, align="left")
    for j, c in enumerate(criteria, start=2):
        pref = preferences[c].to_legacy().capitalize()
        _set_cell(ws, r, j, f"{c} ({pref})", font=_HEADER_FONT, fill=_HEADER_FILL, align="center")

    matrix_start_row = r + 1
    for i, alt in enumerate(alts, start=matrix_start_row):
        fill = _ALT_FILL if i % 2 == 1 else _WHITE_FILL
        _set_cell(ws, i, 1, alt, font=_BODY_BOLD, fill=fill, align="left")
        for j, c in enumerate(criteria, start=2):
            _set_cell(ws, i, j, matrix.loc[alt, c], fill=_INPUT_FILL, align="right", num_format=_NUMBER_FORMAT_RAW)
    matrix_end_row = matrix_start_row + n_alts - 1

    # Section 2: Principal Component Analysis Results
    r = matrix_end_row + 3
    ws.cell(r, 1, "2. Eigenvalues & Variance Explained").font = _SECTION_FONT
    r += 1
    _set_cell(ws, r, 1, "Principal Component", font=_HEADER_FONT, fill=_HEADER_FILL, align="left")
    _set_cell(ws, r, 2, "Eigenvalue (λ_k)", font=_HEADER_FONT, fill=_HEADER_FILL, align="center")
    _set_cell(ws, r, 3, "Variance Explained (%)", font=_HEADER_FONT, fill=_HEADER_FILL, align="center")

    eigenvals = pca_steps.get("Eigenvalues (Variance)", pd.Series([1.0]*m_crit))
    total_var = float(eigenvals.sum())
    ev_start_row = r + 1
    for k_idx, ev in enumerate(eigenvals, start=1):
        curr_r = ev_start_row + k_idx - 1
        fill = _ALT_FILL if k_idx % 2 == 1 else _WHITE_FILL
        _set_cell(ws, curr_r, 1, f"PC{k_idx}", font=_BODY_BOLD, fill=fill, align="left")
        _set_cell(ws, curr_r, 2, float(ev), fill=fill, num_format=_NUMBER_FORMAT_DECIMAL)
        _set_cell(ws, curr_r, 3, float(ev / total_var) if total_var > 0 else 0.0, fill=fill, num_format=_NUMBER_FORMAT_PERCENT)
    ev_end_row = ev_start_row + len(eigenvals) - 1

    # Section 3: Final PCA Weights
    r = ev_end_row + 3
    ws.cell(r, 1, "3. Final PCA Variance-Weighted Loadings").font = _SECTION_FONT
    r += 1
    _set_cell(ws, r, 1, "Metric / Step", font=_HEADER_FONT, fill=_HEADER_FILL, align="left")
    for j, c in enumerate(criteria, start=2):
        _set_cell(ws, r, j, c, font=_HEADER_FONT, fill=_HEADER_FILL, align="center")
    _set_cell(ws, r, m_crit + 2, "Sum / Total", font=_HEADER_FONT, fill=_HEADER_FILL, align="center")

    r += 1
    _set_cell(ws, r, 1, "Final PCA Weight (w_j)", font=_BODY_BOLD, fill=_SUMMARY_FILL, align="left")
    weight_cells = {}
    for j, c in enumerate(criteria, start=2):
        col_let = get_column_letter(j)
        w_val = float(pca_weights.get(c, 1.0 / m_crit))
        _set_cell(ws, r, j, w_val, font=_BODY_BOLD, fill=_SUMMARY_FILL, num_format=_NUMBER_FORMAT_WEIGHT)
        weight_cells[c] = f"'PCA Loadings'!${col_let}${r}"
    first_col = get_column_letter(2)
    last_col = get_column_letter(m_crit + 1)
    _set_cell(ws, r, m_crit + 2, f"=SUM({first_col}{r}:{last_col}{r})", font=_BODY_BOLD, fill=_SUMMARY_FILL, num_format=_NUMBER_FORMAT_WEIGHT)

    _autofit_columns(ws, m_crit + 3)
    return weight_cells


def _build_summary_sheet(
    ws,
    matrix: pd.DataFrame,
    preferences: Mapping[str, CriterionPreference],
    baseline_weights: Mapping[str, float] | None,
    ewm_cells: dict[str, str],
    merec_cells: dict[str, str],
    critic_cells: dict[str, str],
    sd_cells: dict[str, str],
    pca_cells: dict[str, str],
    *,
    baseline_name: str = "Official",
) -> None:
    """Build master summary comparison sheet linking all weighting methods with live formulas."""
    _set_banner(
        ws,
        "Deterministic Weight Calculations Master Summary",
        "Comprehensive side-by-side comparison of criteria weights across deterministic/objective models with live formula links.",
    )
    m_crit = len(matrix.columns)
    criteria = list(matrix.columns)

    r = 4
    ws.cell(r, 1, "Master Criteria Weights Comparison Table").font = _SECTION_FONT
    r += 1

    headers = [
        "Criterion",
        "Direction",
        f"{baseline_name} Weights",
        "Equal Weights",
        "Entropy (EWM)",
        "MEREC",
        "CRITIC",
        "Standard Deviation",
        "PCA Loadings",
    ]
    for c_idx, h in enumerate(headers, start=1):
        _set_cell(ws, r, c_idx, h, font=_HEADER_FONT, fill=_HEADER_FILL, align="center" if c_idx > 1 else "left")

    table_start_row = r + 1
    for row_offset, c in enumerate(criteria):
        curr_r = table_start_row + row_offset
        fill = _ALT_FILL if row_offset % 2 == 1 else _WHITE_FILL
        pref_str = preferences[c].to_legacy().capitalize()

        # 1. Criterion Name
        _set_cell(ws, curr_r, 1, c, font=_BODY_BOLD, fill=fill, align="left")
        # 2. Direction
        _set_cell(ws, curr_r, 2, pref_str, fill=fill, align="center")
        # 3. Baseline Weight
        base_w = float(baseline_weights.get(c, 1.0 / m_crit)) if baseline_weights else (1.0 / m_crit)
        _set_cell(ws, curr_r, 3, base_w, fill=_INPUT_FILL, align="right", num_format=_NUMBER_FORMAT_WEIGHT)
        # 4. Equal Weight formula
        _set_cell(ws, curr_r, 4, f"=1/{m_crit}", fill=fill, align="right", num_format=_NUMBER_FORMAT_WEIGHT)
        # 5. Entropy link
        _set_cell(ws, curr_r, 5, f"={ewm_cells.get(c, 0)}", fill=fill, align="right", num_format=_NUMBER_FORMAT_WEIGHT)
        # 6. MEREC link
        _set_cell(ws, curr_r, 6, f"={merec_cells.get(c, 0)}", fill=fill, align="right", num_format=_NUMBER_FORMAT_WEIGHT)
        # 7. CRITIC link
        _set_cell(ws, curr_r, 7, f"={critic_cells.get(c, 0)}", fill=fill, align="right", num_format=_NUMBER_FORMAT_WEIGHT)
        # 8. Standard Deviation link
        _set_cell(ws, curr_r, 8, f"={sd_cells.get(c, 0)}", fill=fill, align="right", num_format=_NUMBER_FORMAT_WEIGHT)
        # 9. PCA link
        _set_cell(ws, curr_r, 9, f"={pca_cells.get(c, 0)}", fill=fill, align="right", num_format=_NUMBER_FORMAT_WEIGHT)

    table_end_row = table_start_row + m_crit - 1

    # Sum Row
    r = table_end_row + 1
    _set_cell(ws, r, 1, "Total Sum (Σ w_j)", font=_BODY_BOLD, fill=_SUBHEADER_FILL, align="left")
    _set_cell(ws, r, 2, "-", font=_BODY_BOLD, fill=_SUBHEADER_FILL, align="center")
    for col_idx in range(3, len(headers) + 1):
        col_let = get_column_letter(col_idx)
        _set_cell(ws, r, col_idx, f"=SUM({col_let}{table_start_row}:{col_let}{table_end_row})", font=_BODY_BOLD, fill=_SUMMARY_FILL, num_format=_NUMBER_FORMAT_WEIGHT)

    _autofit_columns(ws, len(headers))


def build_weight_calculation_excel_workbook(
    matrix: pd.DataFrame,
    weights: Mapping[str, float] | None,
    directions: Mapping[str, Any],
    *,
    baseline_name: str = "Official",
) -> bytes:
    """Build and return an XLSX workbook containing live formula calculation sheets for all deterministic weight methods."""
    frame = validate_crisp_matrix(matrix)
    preferences = normalize_directions(frame.columns, directions)
    base_weights = dict(weights) if weights else None
    if base_weights:
        total_w = sum(float(v) for v in base_weights.values())
        if total_w > 0:
            base_weights = {k: float(v) / total_w for k, v in base_weights.items()}

    wb = Workbook()

    # 1. Sheet for Entropy (EWM)
    ws_ewm = wb.active
    ws_ewm.title = "Entropy (EWM)"
    ewm_cells = _build_entropy_sheet(ws_ewm, frame, preferences)

    # 2. Sheet for MEREC
    ws_merec = wb.create_sheet(title="MEREC")
    merec_cells = _build_merec_sheet(ws_merec, frame, preferences)

    # 3. Sheet for CRITIC
    ws_critic = wb.create_sheet(title="CRITIC")
    critic_cells = _build_critic_sheet(ws_critic, frame, preferences)

    # 4. Sheet for Standard Deviation
    ws_sd = wb.create_sheet(title="Standard Deviation")
    sd_cells = _build_sd_sheet(ws_sd, frame, preferences)

    # 5. Sheet for PCA Loadings
    ws_pca = wb.create_sheet(title="PCA Loadings")
    pca_cells = _build_pca_sheet(ws_pca, frame, preferences)

    # 6. Master Summary Sheet (Inserted at the front)
    ws_summary = wb.create_sheet(title="Weight Summary", index=0)
    _build_summary_sheet(
        ws_summary,
        frame,
        preferences,
        base_weights,
        ewm_cells,
        merec_cells,
        critic_cells,
        sd_cells,
        pca_cells,
        baseline_name=baseline_name,
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = [
    "build_weight_calculation_excel_workbook",
    "WEIGHT_EXCEL_EXPORT_FILENAME",
    "WEIGHT_EXCEL_EXPORT_REVISION",
]
