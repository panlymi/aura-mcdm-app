"""Formula-rich Excel export for complete, auditable MOORA calculations."""

from __future__ import annotations

from io import BytesIO
from typing import Any, Mapping

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from moora_calculator import calculate_moora

from .criteria import CriterionPreference, CriterionType, validate_method_capabilities
from .syai_excel import (
    _ALTERNATIVE_FILL,
    _BODY_FONT,
    _CALC_NUMBER_FORMAT,
    _FORMULA_FILL,
    _GRID_BORDER,
    _HEADER_FONT,
    _INPUT_FILL,
    _RAW_NUMBER_FORMAT,
    _SECTION_FILL,
    _SUMMARY_FILL,
    _WEIGHT_FORMAT,
    _WINNER_FILL,
    _WHITE_FILL,
    _autofit_column_widths,
    _preference_fill,
    _preference_label,
    _set_section_title,
    _set_text,
    _set_title,
    _style_grid,
    _style_matrix_header,
    _write_indexed_dataframe,
)
from .validation import validate_crisp_matrix, validate_weights


_NUMERICAL_GUARD = 1e-9

# Included in Streamlit's cache key so workbook changes invalidate old bytes.
MOORA_EXCEL_EXPORT_REVISION = "v3"
MOORA_EXCEL_EXPORT_FILENAME = (
    f"moora_complete_formula_calculation_{MOORA_EXCEL_EXPORT_REVISION}.xlsx"
)


def _write_normalized_live_weights_row(
    sheet,
    row: int,
    columns: list[str],
    settings_data_start: int,
    *,
    weight_sum_row: int,
) -> None:
    """Write effective weights that mirror the app's automatic normalization."""

    sheet.cell(row, 1, "Effective weight")
    for offset, _criterion in enumerate(columns, start=2):
        settings_row = settings_data_start + offset - 2
        sheet.cell(
            row,
            offset,
            f"=IF($B${weight_sum_row}<=0,0,$B${settings_row}/$B${weight_sum_row})",
        )
    for cell in sheet[row][: len(columns) + 1]:
        cell.fill = _INPUT_FILL
        cell.font = _HEADER_FONT
        cell.border = _GRID_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.column > 1:
            cell.number_format = _WEIGHT_FORMAT


def _build_formula_sheet(
    workbook: Workbook,
    frame: pd.DataFrame,
    weights: Mapping[str, float],
    preferences: Mapping[str, CriterionPreference],
) -> dict[str, int]:
    sheet = workbook.create_sheet("MOORA")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 82

    columns = [str(column) for column in frame.columns]
    alternatives = [str(index) for index in frame.index]
    criteria_count = len(columns)
    alternatives_count = len(alternatives)
    last_matrix_column = criteria_count + 1
    model_last_column = max(7, last_matrix_column)

    _set_title(
        sheet,
        1,
        model_last_column,
        "MOORA — Complete Formula-Driven Calculation Workbook",
    )
    sheet.merge_cells(
        start_row=2, start_column=1, end_row=2, end_column=model_last_column
    )
    sheet.cell(
        2,
        1,
        "Edit the criterion weights and matrix inputs to explore the decision model. "
        "Every vector normalization, weighted value, benefit/cost sum, assessment value (y_i), "
        "rank, KPI, and chart is linked to the live model.",
    )
    sheet.cell(2, 1).font = Font(name="Aptos", size=9, italic=True, color="666666")
    sheet.cell(2, 1).alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[2].height = 30

    settings_title_row = 12
    settings_header_row = 13
    settings_data_start = 14
    settings_data_end = settings_data_start + criteria_count - 1

    _set_section_title(sheet, 4, 5, "Model Parameters and Audit Checks")
    parameter_rows = [
        (5, "Alternatives", alternatives_count, "Number of decision alternatives"),
        (6, "Criteria", criteria_count, "Number of decision criteria"),
        (
            7,
            "Benefit criteria",
            sum(1 for p in preferences.values() if p.kind is CriterionType.BENEFIT),
            "Number of criteria to maximize",
        ),
        (
            8,
            "Cost criteria",
            sum(1 for p in preferences.values() if p.kind is CriterionType.COST),
            "Number of criteria to minimize",
        ),
    ]
    for row, label, value, description in parameter_rows:
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
        sheet.cell(row, 3, description)
    sheet.cell(9, 1, "Weight sum check")
    sheet.cell(9, 2, f"=SUM(B{settings_data_start}:B{settings_data_end})")
    sheet.cell(
        9,
        3,
        (
            '=IF(B9<=0,"ERROR: enter at least one positive weight",'
            f'IF(ABS(B9-1)<={_NUMERICAL_GUARD:.12g},'
            '"OK — weights already sum to 1",'
            '"OK — live formulas normalize weights to 1"))'
        ),
    )
    _style_grid(sheet, 5, 9, 1, 5)
    sheet.cell(5, 2).fill = _FORMULA_FILL
    sheet.cell(6, 2).fill = _FORMULA_FILL
    sheet.cell(7, 2).fill = _FORMULA_FILL
    sheet.cell(8, 2).fill = _FORMULA_FILL
    sheet.cell(9, 2).fill = _SUMMARY_FILL
    sheet.cell(9, 2).number_format = _WEIGHT_FORMAT

    _set_section_title(sheet, settings_title_row, 5, "Criterion Settings and Denominators")
    headers = [
        "Criterion",
        "Weight (w_j)",
        "Preference",
        "Denominator √(∑x²)",
        "Audit Note",
    ]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(settings_header_row, column_index, header)
        sheet.cell(settings_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(settings_header_row, column_index).font = _HEADER_FONT
    sheet.row_dimensions[settings_header_row].height = 24

    matrix_title_row = settings_data_end + 3
    raw_weights_row = matrix_title_row + 1
    raw_header_row = matrix_title_row + 2
    raw_data_start = matrix_title_row + 3
    raw_data_end = raw_data_start + alternatives_count - 1
    matrix_denominator_row = raw_data_end + 1

    weight_validation = DataValidation(
        type="decimal",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=False,
    )
    weight_validation.promptTitle = "Non-negative criterion weight"
    weight_validation.prompt = (
        "Enter any non-negative weight. Live effective weights are normalized automatically."
    )
    weight_validation.errorTitle = "Invalid weight"
    weight_validation.error = "Weights must be numeric and non-negative."
    weight_validation.errorStyle = "stop"
    weight_validation.showInputMessage = True
    weight_validation.showErrorMessage = True
    sheet.add_data_validation(weight_validation)
    weight_validation.add(f"B{settings_data_start}:B{settings_data_end}")

    for row_offset, criterion in enumerate(columns):
        row = settings_data_start + row_offset
        preference = preferences[criterion]
        _set_text(sheet, row, 1, criterion)
        sheet.cell(row, 2, float(weights[criterion]))
        sheet.cell(row, 2).number_format = _WEIGHT_FORMAT
        sheet.cell(row, 2).fill = _INPUT_FILL
        sheet.cell(row, 2).comment = Comment(
            "Criterion weight. You can edit this cell; all MOORA scores and ranks will update.",
            "User",
        )
        _set_text(sheet, row, 3, _preference_label(preference))
        sheet.cell(row, 3).fill = _preference_fill(preference)
        sheet.cell(row, 3).comment = Comment(
            "Benefit criteria contribute positively to y_i; Cost criteria are subtracted.",
            "User",
        )
        matrix_column_letter = get_column_letter(row_offset + 2)
        sheet.cell(row, 4, f"={matrix_column_letter}${matrix_denominator_row}")
        sheet.cell(row, 4).fill = _FORMULA_FILL
        sheet.cell(row, 4).number_format = _CALC_NUMBER_FORMAT
        sheet.cell(
            row,
            5,
            f'=IF($B$9<=0,"Invalid: total weight must be positive",'
            f'"Normalized weight: "&TEXT(B{row}/$B$9,"0.0000"))',
        )
        sheet.cell(row, 5).font = Font(
            name="Aptos", size=9, italic=True, color="555555"
        )
    _style_grid(
        sheet,
        settings_data_start,
        settings_data_end,
        1,
        5,
        number_format=_CALC_NUMBER_FORMAT,
    )

    _set_section_title(
        sheet,
        matrix_title_row,
        last_matrix_column,
        "Step 1 — Original Decision Matrix and Vector Denominators",
    )
    _write_normalized_live_weights_row(
        sheet,
        raw_weights_row,
        columns,
        settings_data_start,
        weight_sum_row=9,
    )
    sheet.cell(raw_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, raw_header_row, column_index, criterion)
    _style_grid(
        sheet,
        raw_header_row,
        raw_data_end,
        1,
        last_matrix_column,
        number_format=_RAW_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, raw_header_row, columns, preferences)
    for row_offset, (alternative, row_values) in enumerate(frame.iterrows()):
        row = raw_data_start + row_offset
        _set_text(sheet, row, 1, str(alternative))
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index, value in enumerate(row_values, start=2):
            sheet.cell(row, column_index, float(value))
            sheet.cell(row, column_index).number_format = _RAW_NUMBER_FORMAT
            sheet.cell(row, column_index).fill = _WHITE_FILL

    sheet.cell(matrix_denominator_row, 1, "Denominator √(∑x²)")
    sheet.cell(matrix_denominator_row, 1).font = _HEADER_FONT
    sheet.cell(matrix_denominator_row, 1).fill = _SECTION_FILL
    for column_index in range(2, last_matrix_column + 1):
        column_letter = get_column_letter(column_index)
        formula = f"=SQRT(SUMSQ({column_letter}{raw_data_start}:{column_letter}{raw_data_end}))"
        sheet.cell(matrix_denominator_row, column_index, formula)
        sheet.cell(matrix_denominator_row, column_index).fill = _SUMMARY_FILL
        sheet.cell(matrix_denominator_row, column_index).number_format = _CALC_NUMBER_FORMAT
    _style_grid(
        sheet,
        matrix_denominator_row,
        matrix_denominator_row,
        1,
        last_matrix_column,
        number_format=_CALC_NUMBER_FORMAT,
    )
    for column_index in range(1, last_matrix_column + 1):
        sheet.cell(matrix_denominator_row, column_index).border = _GRID_BORDER

    norm_title_row = matrix_denominator_row + 3
    norm_weights_row = norm_title_row + 1
    norm_header_row = norm_title_row + 2
    norm_data_start = norm_title_row + 3
    norm_data_end = norm_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        norm_title_row,
        last_matrix_column,
        "Step 2 — Vector (Ratio) Normalized Decision Matrix (x*_ij)",
    )
    _write_normalized_live_weights_row(
        sheet,
        norm_weights_row,
        columns,
        settings_data_start,
        weight_sum_row=9,
    )
    sheet.cell(norm_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, norm_header_row, column_index, criterion)
    _style_grid(
        sheet,
        norm_header_row,
        norm_data_end,
        1,
        last_matrix_column,
        number_format=_CALC_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, norm_header_row, columns, preferences)
    for row_offset in range(alternatives_count):
        row = norm_data_start + row_offset
        matrix_row = raw_data_start + row_offset
        _set_text(sheet, row, 1, alternatives[row_offset])
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index in range(2, last_matrix_column + 1):
            letter = get_column_letter(column_index)
            sheet.cell(
                row,
                column_index,
                f"=IF(ABS({letter}${matrix_denominator_row})<={_NUMERICAL_GUARD:.12g},0,"
                f"{letter}{matrix_row}/{letter}${matrix_denominator_row})",
            )
            sheet.cell(row, column_index).fill = _FORMULA_FILL
            sheet.cell(row, column_index).number_format = _CALC_NUMBER_FORMAT

    weighted_title_row = norm_data_end + 3
    weighted_weights_row = weighted_title_row + 1
    weighted_header_row = weighted_title_row + 2
    weighted_data_start = weighted_title_row + 3
    weighted_data_end = weighted_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        weighted_title_row,
        last_matrix_column,
        "Step 3 — Weighted Normalized Decision Matrix (v_ij = w_j × x*_ij)",
    )
    _write_normalized_live_weights_row(
        sheet,
        weighted_weights_row,
        columns,
        settings_data_start,
        weight_sum_row=9,
    )
    sheet.cell(weighted_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, weighted_header_row, column_index, criterion)
    _style_grid(
        sheet,
        weighted_header_row,
        weighted_data_end,
        1,
        last_matrix_column,
        number_format=_CALC_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, weighted_header_row, columns, preferences)
    for row_offset in range(alternatives_count):
        row = weighted_data_start + row_offset
        norm_row = norm_data_start + row_offset
        _set_text(sheet, row, 1, alternatives[row_offset])
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index in range(2, last_matrix_column + 1):
            letter = get_column_letter(column_index)
            sheet.cell(
                row,
                column_index,
                f"={letter}{norm_row}*{letter}${weighted_weights_row}",
            )
            sheet.cell(row, column_index).fill = _FORMULA_FILL
            sheet.cell(row, column_index).number_format = _CALC_NUMBER_FORMAT

    results_title_row = weighted_data_end + 3
    results_header_row = results_title_row + 1
    results_data_start = results_header_row + 1
    results_data_end = results_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        results_title_row,
        6,
        "Step 4 — Benefit/Cost Sums, Assessment Value, Rank, and Tie-Safe Sort Order",
    )
    result_headers = [
        "Alternative",
        "Sum (Maximize)",
        "Sum (Minimize)",
        "Assessment Value (y_i)",
        "Rank",
        "Sort Order",
    ]
    for column_index, header in enumerate(result_headers, start=1):
        sheet.cell(results_header_row, column_index, header)
        sheet.cell(results_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(results_header_row, column_index).font = _HEADER_FONT
    sheet.row_dimensions[results_header_row].height = 24

    max_col_indices = [
        col_idx
        for col_idx, col in enumerate(columns, start=2)
        if preferences[col].kind is CriterionType.BENEFIT
    ]
    min_col_indices = [
        col_idx
        for col_idx, col in enumerate(columns, start=2)
        if preferences[col].kind is CriterionType.COST
    ]

    for row_offset in range(alternatives_count):
        row = results_data_start + row_offset
        weighted_row = weighted_data_start + row_offset
        _set_text(sheet, row, 1, alternatives[row_offset])
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL

        if max_col_indices:
            max_cells = [f"{get_column_letter(c)}{weighted_row}" for c in max_col_indices]
            max_formula = f"={max_cells[0]}" if len(max_cells) == 1 else f"=SUM({','.join(max_cells)})"
        else:
            max_formula = "=0"

        if min_col_indices:
            min_cells = [f"{get_column_letter(c)}{weighted_row}" for c in min_col_indices]
            min_formula = f"={min_cells[0]}" if len(min_cells) == 1 else f"=SUM({','.join(min_cells)})"
        else:
            min_formula = "=0"

        sheet.cell(row, 2, max_formula)
        sheet.cell(row, 3, min_formula)
        sheet.cell(row, 4, f"=B{row}-C{row}")
        sheet.cell(
            row,
            5,
            f"=RANK(D{row},$D${results_data_start}:$D${results_data_end},0)",
        )
        sheet.cell(row, 2).fill = _FORMULA_FILL
        sheet.cell(row, 3).fill = _FORMULA_FILL
        sheet.cell(row, 4).fill = _SUMMARY_FILL
        sheet.cell(row, 5).fill = _WHITE_FILL
        sheet.cell(row, 5).number_format = "0"
    _style_grid(
        sheet,
        results_data_start,
        results_data_end,
        1,
        6,
        number_format=_CALC_NUMBER_FORMAT,
    )
    for row in range(results_data_start, results_data_end + 1):
        sheet.cell(row, 5).number_format = "0"
        sheet.cell(
            row,
            6,
            f"=E{row}+COUNTIF($D${results_data_start}:D{row},D{row})-1",
        )
        sheet.cell(row, 6).number_format = "0"

    ranking_title_row = results_data_end + 3
    ranking_header_row = ranking_title_row + 1
    ranking_data_start = ranking_header_row + 1
    ranking_data_end = ranking_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        ranking_title_row,
        5,
        "Step 5 — Final Ranking (Sorted by Score)",
    )
    for column_index, header in enumerate(result_headers[:5], start=1):
        sheet.cell(ranking_header_row, column_index, header)
        sheet.cell(ranking_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(ranking_header_row, column_index).font = _HEADER_FONT
    _style_grid(
        sheet,
        ranking_data_start,
        ranking_data_end,
        1,
        5,
        number_format=_CALC_NUMBER_FORMAT,
    )
    sheet.row_dimensions[ranking_header_row].height = 31
    for sort_position in range(1, alternatives_count + 1):
        row = ranking_data_start + sort_position - 1
        match_formula = (
            f"MATCH({sort_position},$F${results_data_start}:$F${results_data_end},0)"
        )
        for column_index, source_letter in enumerate(["A", "B", "C", "D", "E"], start=1):
            sheet.cell(
                row,
                column_index,
                f"=INDEX(${source_letter}${results_data_start}:"
                f"${source_letter}${results_data_end},{match_formula})",
            )
            if column_index > 1:
                sheet.cell(row, column_index).fill = _WHITE_FILL
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        sheet.cell(row, 5).number_format = "0"
    sheet.conditional_formatting.add(
        f"D{ranking_data_start}:D{ranking_data_end}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )
    sheet.conditional_formatting.add(
        f"A{ranking_data_start}:E{ranking_data_end}",
        FormulaRule(formula=[f"$E{ranking_data_start}=1"], fill=_WINNER_FILL),
    )

    sheet.freeze_panes = "B1"
    sheet.print_area = f"A1:{get_column_letter(model_last_column)}{ranking_data_end}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.oddFooter.center.text = "MOORA complete formula model"
    sheet.oddFooter.right.text = "Page &P of &N"
    sheet.column_dimensions["A"].width = max(
        25, min(38, max(len(value) for value in alternatives) + 3)
    )
    for column_index, criterion in enumerate(columns, start=2):
        letter = get_column_letter(column_index)
        raw_value_width = max(len(f"{float(value):.9f}") for value in frame[criterion])
        sheet.column_dimensions[letter].width = max(
            16, min(30, max(len(criterion) + 3, raw_value_width + 2))
        )
    for row in range(1, ranking_data_end + 1):
        if sheet.row_dimensions[row].height is None:
            sheet.row_dimensions[row].height = 19
    _autofit_column_widths(sheet)

    return {
        "results_data_start": results_data_start,
        "results_data_end": results_data_end,
        "ranking_data_start": ranking_data_start,
        "ranking_data_end": ranking_data_end,
    }


def _build_summary_sheet(
    workbook: Workbook,
    formula_positions: Mapping[str, int],
    alternatives_count: int,
) -> None:
    sheet = workbook.create_sheet("Decision Summary")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 90
    _set_title(sheet, 1, 14, "MOORA Decision Summary")
    sheet.merge_cells("A2:N2")
    sheet["A2"] = (
        "A live executive summary of the MOORA Multi-Objective Optimization results. "
        "Change weights or matrix values on the MOORA sheet to refresh the KPIs, table, and chart."
    )
    sheet["A2"].font = Font(name="Aptos", size=9, italic=True, color="666666")
    sheet["A2"].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[2].height = 29

    ranking_start = formula_positions["ranking_data_start"]
    cards = [
        (
            "A",
            "B",
            "First Rank-1 Alternative",
            f"='MOORA'!A{ranking_start}",
            "@",
        ),
        ("C", "D", "Winning Assessment Value", f"='MOORA'!D{ranking_start}", "0.0000"),
        ("E", "F", "Alternatives", "='MOORA'!B5", "0"),
        ("G", "H", "Total Criteria", "='MOORA'!B6", "0"),
        ("I", "J", "Benefit Criteria", "='MOORA'!B7", "0"),
    ]
    for start_letter, end_letter, label, formula, number_format in cards:
        sheet.merge_cells(f"{start_letter}4:{end_letter}4")
        sheet.merge_cells(f"{start_letter}5:{end_letter}6")
        label_cell = sheet[f"{start_letter}4"]
        value_cell = sheet[f"{start_letter}5"]
        label_cell.value = label
        value_cell.value = formula
        label_cell.fill = _SECTION_FILL
        value_cell.fill = _SUMMARY_FILL
        label_cell.font = _HEADER_FONT
        value_cell.font = Font(name="Aptos Display", size=15, bold=True, color="0B4F6C")
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        value_cell.number_format = number_format
        for row in range(4, 7):
            for column in range(
                sheet[start_letter + "1"].column,
                sheet[end_letter + "1"].column + 1,
            ):
                sheet.cell(row, column).border = _GRID_BORDER

    summary_header_row = 9
    summary_data_start = 10
    summary_data_end = summary_data_start + alternatives_count - 1
    headers = [
        "Alternative",
        "Sum (Maximize)",
        "Sum (Minimize)",
        "Assessment Value (y_i)",
        "Rank",
    ]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(summary_header_row, column_index, header)
        sheet.cell(summary_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(summary_header_row, column_index).font = _HEADER_FONT
    _style_grid(
        sheet,
        summary_header_row,
        summary_data_end,
        1,
        5,
        number_format=_CALC_NUMBER_FORMAT,
    )
    for row_offset in range(alternatives_count):
        row = summary_data_start + row_offset
        source_row = ranking_start + row_offset
        for column_index, source_letter in enumerate(["A", "B", "C", "D", "E"], start=1):
            sheet.cell(row, column_index, f"='MOORA'!{source_letter}{source_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        sheet.cell(row, 5).number_format = "0"
    sheet.conditional_formatting.add(
        f"D{summary_data_start}:D{summary_data_end}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )
    sheet.conditional_formatting.add(
        f"A{summary_data_start}:E{summary_data_end}",
        FormulaRule(formula=[f"$E{summary_data_start}=1"], fill=_WINNER_FILL),
    )

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "MOORA Assessment Value (y_i) by alternative"
    chart.x_axis.title = "Alternative"
    chart.y_axis.title = "Assessment Value (y_i)"
    chart.height = 7.2
    chart.width = 14.2
    chart.legend = None
    data = Reference(
        sheet, min_col=4, min_row=summary_header_row, max_row=summary_data_end
    )
    categories = Reference(
        sheet, min_col=1, min_row=summary_data_start, max_row=summary_data_end
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.numFmt = "0.000"
    chart.y_axis.numFmt = "0.00"
    sheet.add_chart(chart, "G9")

    notes_row = max(summary_data_end + 3, 27)
    _set_section_title(sheet, notes_row, 10, "How to read this summary")
    notes = [
        "Higher MOORA assessment values (y_i) indicate superior overall performance.",
        "Vector normalization divides each cell by the square root of the sum of squares of its criterion column.",
        "Benefit criteria are summed positively; Cost criteria are subtracted.",
        "Tied assessment values share a competition rank; sort order only controls display order.",
        "The Verified Values sheet preserves the canonical calculation at download time.",
    ]
    for row_offset, note in enumerate(notes, start=1):
        sheet.merge_cells(
            start_row=notes_row + row_offset,
            start_column=1,
            end_row=notes_row + row_offset,
            end_column=10,
        )
        sheet.cell(notes_row + row_offset, 1, f"• {note}")
        sheet.cell(notes_row + row_offset, 1).font = _BODY_FONT
        sheet.cell(notes_row + row_offset, 1).alignment = Alignment(vertical="center")

    widths = {
        "A": 28,
        "B": 17,
        "C": 17,
        "D": 22,
        "E": 12,
        "F": 3,
        "G": 14,
        "H": 14,
        "I": 18,
        "J": 18,
        "K": 14,
        "L": 14,
        "M": 14,
        "N": 14,
    }
    for letter, width in widths.items():
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A10"
    sheet.print_area = f"A1:N{notes_row + len(notes)}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def _build_verified_values_sheet(
    workbook: Workbook,
    frame: pd.DataFrame,
    weights: Mapping[str, float],
    preferences: Mapping[str, CriterionPreference],
    results: pd.DataFrame,
    steps: Mapping[str, Any],
) -> None:
    sheet = workbook.create_sheet("Verified Values")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 88
    max_columns = max(len(frame.columns) + 1, 6)
    _set_title(sheet, 1, max_columns, "MOORA — Verified Numerical Values")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_columns)
    sheet.cell(
        2,
        1,
        "Static values from the canonical Python implementation at download time. "
        "Use them to reconcile every live Excel calculation stage.",
    )
    sheet.cell(2, 1).font = Font(name="Aptos", size=9, italic=True, color="666666")
    sheet.cell(2, 1).alignment = Alignment(wrap_text=True)

    _set_section_title(sheet, 4, 6, "Parameters")
    parameter_values = [
        ["Alternatives", len(frame), "Criteria", len(frame.columns), "Benefit Criteria", sum(1 for p in preferences.values() if p.kind is CriterionType.BENEFIT)],
        ["Cost Criteria", sum(1 for p in preferences.values() if p.kind is CriterionType.COST), None, None, None, None],
    ]
    for row_offset, values in enumerate(parameter_values, start=5):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row_offset, column_index, value)
    _style_grid(sheet, 5, 6, 1, 6)

    settings_title_row = 8
    settings_header_row = 9
    settings_data_start = 10
    _set_section_title(sheet, settings_title_row, 4, "Criterion Settings")
    settings_headers = ["Criterion", "Weight", "Preference", "Denominator √(∑x²)"]
    for column_index, header in enumerate(settings_headers, start=1):
        sheet.cell(settings_header_row, column_index, header)
        sheet.cell(settings_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(settings_header_row, column_index).font = _HEADER_FONT
    for row_offset, criterion in enumerate(frame.columns, start=settings_data_start):
        name = str(criterion)
        preference = preferences[name]
        _set_text(sheet, row_offset, 1, name)
        sheet.cell(row_offset, 2, float(weights[name]))
        sheet.cell(row_offset, 2).number_format = _WEIGHT_FORMAT
        _set_text(sheet, row_offset, 3, _preference_label(preference))
        sheet.cell(row_offset, 3).fill = _preference_fill(preference)
        denominator = float((frame[criterion] ** 2).sum() ** 0.5)
        sheet.cell(row_offset, 4, denominator)
        sheet.cell(row_offset, 4).number_format = _CALC_NUMBER_FORMAT
    settings_data_end = settings_data_start + len(frame.columns) - 1
    _style_grid(
        sheet,
        settings_header_row,
        settings_data_end,
        1,
        4,
        number_format=_CALC_NUMBER_FORMAT,
    )

    next_row = settings_data_end + 3
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 1 — Original Decision Matrix",
        frame,
        matrix_preferences=preferences,
    )
    normalized = steps["Step 2: Ratio Normalized Matrix ($x^*_{ij}$)"].reindex(frame.index)
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 2 — Vector (Ratio) Normalized Decision Matrix",
        normalized,
        matrix_preferences=preferences,
    )
    weighted = steps["Step 3: Weighted Normalized Matrix ($v_{ij}$)"].reindex(frame.index)
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 3 — Weighted Normalized Decision Matrix",
        weighted,
        matrix_preferences=preferences,
    )
    step4_df = steps["Step 4: Normalized Assessment Value ($y_i$)"].reindex(frame.index)
    step4_with_rank = pd.concat(
        [step4_df, results.reindex(frame.index)[["Rank"]]], axis=1
    )
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 4 — Benefit Sum, Cost Sum, Assessment Value (y_i), and Rank",
        step4_with_rank,
    )
    step4_data_start = next_row - len(step4_with_rank) - 1
    for row in range(step4_data_start, step4_data_start + len(step4_with_rank)):
        sheet.cell(row, 5).number_format = "0"

    final_values = results[
        ["y_i (Assessment Value)", "Rank"]
    ]
    final_start = next_row
    _write_indexed_dataframe(
        sheet,
        final_start,
        "Final Ranking — Sorted by Rank",
        final_values,
    )
    final_data_start = final_start + 2
    for row in range(final_data_start, final_data_start + len(results)):
        sheet.cell(row, 3).number_format = "0"

    sheet.column_dimensions["A"].width = max(
        26,
        min(
            40,
            max(
                max(len(str(value)) for value in frame.index),
                max(len(str(value)) for value in frame.columns),
            )
            + 3,
        ),
    )
    for column_index in range(2, max_columns + 1):
        letter = get_column_letter(column_index)
        sheet.column_dimensions[letter].width = 18
    _autofit_column_widths(sheet)


def _build_guide_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Formula Guide")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 90
    _set_title(sheet, 1, 10, "MOORA Mathematical Model & Reference Guide")
    sheet.merge_cells("A2:J2")
    sheet["A2"] = (
        "Complete equations and documentation for the Multi-Objective Optimization "
        "on the basis of Ratio Analysis (MOORA) method."
    )
    sheet["A2"].font = Font(name="Aptos", size=9, italic=True, color="666666")
    sheet["A2"].alignment = Alignment(wrap_text=True)

    sections = [
        (
            4,
            "1. Vector (Ratio) Normalization",
            [
                "Each decision value is divided by the square root of the sum of squared values for that criterion across all alternatives:",
                "x*_ij = x_ij / √(∑_{k=1}^m x_kj^2)",
                "This produces a dimensionless ratio, preserves the sign of each value, and keeps absolute normalized magnitudes at or below one.",
            ],
        ),
        (
            10,
            "2. Weighted Normalization",
            [
                "Each normalized value is multiplied by its criterion weight w_j:",
                "v_ij = w_j × x*_ij",
                "where ∑_{j=1}^n w_j = 1.",
            ],
        ),
        (
            16,
            "3. Multi-Objective Optimization (Ratio System)",
            [
                "The normalized assessment value (y_i) subtracts the sum of weighted cost criteria from the sum of weighted benefit criteria:",
                "y_i = ∑_{j ∈ B} v_ij - ∑_{j ∈ C} v_ij",
                "where B is the set of benefit criteria (to maximize) and C is the set of cost criteria (to minimize).",
            ],
        ),
        (
            22,
            "4. Ranking and Competition Ties",
            [
                "Alternatives are ranked in descending order of y_i (the highest assessment value receives Rank 1):",
                "Rank(A_i) = 1 + count of alternatives with higher y_k.",
                "Tied scores receive identical competition ranks (e.g. 1, 2, 2, 4).",
            ],
        ),
        (
            28,
            "5. References and Primary Literature",
            [
                "• Brauers, W. K. M., & Zavadskas, E. K. (2006). The MOORA method and its application to privatization in a transition economy. Control and Cybernetics, 35(2), 445–469.",
                "https://pldml.icm.edu.pl/pldml/element/bwmeta1.element.bwnjournal-article-ccv35i2p445bwm",
                "• Brauers, W. K. M., & Zavadskas, E. K. (2009). Multi-objective Optimization with Discrete Alternatives on the Basis of Ratio Analysis. Intellectual Economics, 2(6), 24–31.",
                "https://ojs.mruni.eu/ojs/intellectual-economics/article/view/1193",
            ],
        ),
    ]

    for start_row, heading, paragraphs in sections:
        _set_section_title(sheet, start_row, 10, heading)
        current_row = start_row + 1
        for paragraph in paragraphs:
            sheet.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=10,
            )
            cell = sheet.cell(current_row, 1, paragraph)
            if paragraph.startswith("x*_ij =") or paragraph.startswith("v_ij =") or paragraph.startswith("y_i =") or paragraph.startswith("Rank("):
                cell.font = Font(name="Courier New", size=10, bold=True, color="002060")
                cell.fill = _SECTION_FILL
                sheet.row_dimensions[current_row].height = 24
            else:
                cell.font = _BODY_FONT
                sheet.row_dimensions[current_row].height = 20
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            current_row += 1

    widths = {
        "A": 18,
        "B": 18,
        "C": 18,
        "D": 18,
        "E": 18,
        "F": 18,
        "G": 18,
        "H": 18,
        "I": 18,
        "J": 18,
    }
    for letter, width in widths.items():
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A1"
    sheet.print_area = "A1:J35"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def build_moora_excel_workbook(
    data: pd.DataFrame,
    weights: Mapping[str, Any],
    directions: Mapping[str, Any],
) -> bytes:
    """Return a complete formula-driven MOORA workbook as XLSX bytes."""

    frame = validate_crisp_matrix(data)
    columns = [str(column) for column in frame.columns]
    preferences = validate_method_capabilities("MOORA", columns, directions)
    normalized_weights = validate_weights(weights, columns, normalize=True)
    results, steps = calculate_moora(
        frame,
        normalized_weights,
        {column: preferences[column].to_legacy() for column in columns},
        return_steps=True,
    )

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.properties.version = MOORA_EXCEL_EXPORT_REVISION
    workbook.properties.creator = "AURA MCDM App"
    workbook.properties.title = "Complete MOORA calculation workbook"
    workbook.properties.subject = "Auditable MOORA decision model"
    workbook.properties.keywords = "MOORA, MCDM, decision analysis, formulas"

    positions = _build_formula_sheet(
        workbook,
        frame,
        normalized_weights,
        preferences,
    )
    _build_summary_sheet(workbook, positions, len(frame))
    _build_verified_values_sheet(
        workbook,
        frame,
        normalized_weights,
        preferences,
        results,
        steps,
    )
    _build_guide_sheet(workbook)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


__all__ = [
    "MOORA_EXCEL_EXPORT_FILENAME",
    "MOORA_EXCEL_EXPORT_REVISION",
    "build_moora_excel_workbook",
]
