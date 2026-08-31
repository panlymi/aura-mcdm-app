"""Formula-rich Excel export for complete, auditable ARIE calculations."""

from __future__ import annotations

from io import BytesIO
from typing import Any, Mapping

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from arie_calculator import calculate_arie

from .criteria import CriterionPreference, CriterionType, validate_method_capabilities
from .syai_excel import (
    _ALTERNATIVE_FILL,
    _BODY_FONT,
    _CALC_NUMBER_FORMAT,
    _FORMULA_FILL,
    _GRID_BORDER,
    _HEADER_FONT,
    _INPUT_FILL,
    _RAW_NUMBER_FORMAT,
    _SECTION_FILL,
    _SUMMARY_FILL,
    _WEIGHT_FORMAT,
    _WINNER_FILL,
    _WHITE_FILL,
    _autofit_column_widths,
    _preference_fill,
    _preference_label,
    _set_section_title,
    _set_text,
    _set_title,
    _style_grid,
    _style_matrix_header,
    _write_indexed_dataframe,
    _write_live_weights_row,
)
from .validation import validate_crisp_matrix, validate_weights


_EPSILON = 1e-9

# Included in Streamlit's cache key so workbook changes invalidate old bytes.
ARIE_EXCEL_EXPORT_REVISION = "v2"
ARIE_EXCEL_EXPORT_FILENAME = (
    f"arie_complete_formula_calculation_{ARIE_EXCEL_EXPORT_REVISION}.xlsx"
)


def _build_formula_sheet(
    workbook: Workbook,
    frame: pd.DataFrame,
    weights: Mapping[str, float],
    preferences: Mapping[str, CriterionPreference],
    *,
    gamma: float,
    kappa: float,
) -> dict[str, int]:
    sheet = workbook.create_sheet("ARIE")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 82

    columns = [str(column) for column in frame.columns]
    alternative_names = [str(index) for index in frame.index]
    last_matrix_column = len(columns) + 1
    last_matrix_letter = get_column_letter(last_matrix_column)
    similarity_total_column = last_matrix_column + 1
    similarity_total_letter = get_column_letter(similarity_total_column)
    model_last_column = max(7, similarity_total_column)
    alternatives_count = len(alternative_names)
    criteria_count = len(columns)

    _set_title(
        sheet,
        1,
        model_last_column,
        "ARIE — Complete Formula-Driven Calculation Workbook",
    )
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=model_last_column)
    sheet.cell(
        2,
        1,
        "Edit the yellow gamma and kappa cells to explore sensitivity and balance. "
        "The similarity-contribution matrices expose how every criterion shapes the result.",
    )
    sheet.cell(2, 1).font = Font(name="Aptos", size=9, italic=True, color="666666")
    sheet.cell(2, 1).alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[2].height = 30

    settings_title_row = 12
    settings_header_row = 13
    settings_data_start = 14
    settings_data_end = settings_data_start + criteria_count - 1

    _set_section_title(sheet, 4, 5, "Model Parameters and Audit Checks")
    parameter_rows = [
        (5, "Gamma (γ)", float(gamma), "Similarity sensitivity; must be greater than 0"),
        (6, "Kappa (κ)", float(kappa), "Ideal/anti-ideal balance; valid range is 0 to 1"),
        (7, "Numerical epsilon", _EPSILON, "Guard used by the canonical ARIE calculator"),
        (8, "Alternatives", alternatives_count, "Number of decision alternatives"),
        (9, "Criteria", criteria_count, "Number of decision criteria"),
    ]
    for row, label, value, description in parameter_rows:
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
        sheet.cell(row, 3, description)
    sheet.cell(10, 1, "Weight sum check")
    sheet.cell(10, 2, f"=SUM(B{settings_data_start}:B{settings_data_end})")
    sheet.cell(10, 3, '="Expected 1.0000; current "&TEXT(B10,"0.0000")')
    _style_grid(sheet, 5, 10, 1, 5)
    for row in (5, 6):
        sheet.cell(row, 2).fill = _INPUT_FILL
        sheet.cell(row, 2).number_format = "0.00"
    sheet.cell(7, 2).fill = _FORMULA_FILL
    sheet.cell(7, 2).number_format = "0.0E+00"
    sheet.cell(10, 2).fill = _SUMMARY_FILL
    sheet.cell(10, 2).number_format = _WEIGHT_FORMAT
    sheet.cell(5, 2).comment = Comment(
        "Gamma controls nonlinearity. Values above 1 sharpen similarity differences; "
        "values below 1 soften them.",
        "User",
    )
    sheet.cell(6, 2).comment = Comment(
        "Kappa balances similarity to the ideal against similarity to the anti-ideal. "
        "The valid range is 0 through 1.",
        "User",
    )
    sheet.cell(7, 2).comment = Comment(
        "This 1E-9 guard matches the canonical Python calculation and should normally "
        "remain unchanged.",
        "User",
    )

    gamma_validation = DataValidation(
        type="custom", formula1="AND(ISNUMBER(B5),B5>0)", allow_blank=False
    )
    gamma_validation.error = "Gamma must be a number greater than zero."
    gamma_validation.errorTitle = "Invalid gamma"
    gamma_validation.prompt = "Enter a positive gamma value."
    gamma_validation.promptTitle = "ARIE gamma"
    gamma_validation.showErrorMessage = True
    gamma_validation.showInputMessage = True
    sheet.add_data_validation(gamma_validation)
    gamma_validation.add(sheet["B5"])

    kappa_validation = DataValidation(
        type="decimal",
        operator="between",
        formula1="0",
        formula2="1",
        allow_blank=False,
    )
    kappa_validation.error = "Kappa must be between 0 and 1."
    kappa_validation.errorTitle = "Invalid kappa"
    kappa_validation.prompt = "Enter a kappa value from 0 through 1."
    kappa_validation.promptTitle = "ARIE kappa"
    kappa_validation.showErrorMessage = True
    kappa_validation.showInputMessage = True
    sheet.add_data_validation(kappa_validation)
    kappa_validation.add(sheet["B6"])

    raw_title_row = settings_data_end + 3
    raw_weights_row = raw_title_row + 1
    raw_header_row = raw_title_row + 2
    raw_data_start = raw_title_row + 3
    raw_data_end = raw_data_start + alternatives_count - 1
    raw_min_row = raw_data_end + 1
    raw_max_row = raw_data_end + 2
    raw_reference_row = raw_data_end + 3
    raw_target_scale_row = raw_data_end + 4

    _set_section_title(sheet, settings_title_row, 5, "Criterion Settings")
    settings_headers = [
        "Criterion",
        "Normalized Weight",
        "Preference",
        "Reference x*",
        "Target deviation scale",
    ]
    for column_index, value in enumerate(settings_headers, start=1):
        sheet.cell(settings_header_row, column_index, value)
    _style_grid(sheet, settings_header_row, settings_data_end, 1, 5)
    for cell in sheet[settings_header_row][:5]:
        cell.fill = _SECTION_FILL
        cell.font = _HEADER_FONT

    for criterion_index, criterion in enumerate(columns):
        row = settings_data_start + criterion_index
        matrix_letter = get_column_letter(criterion_index + 2)
        preference = preferences[criterion]
        _set_text(sheet, row, 1, criterion)
        sheet.cell(row, 2, float(weights[criterion]))
        sheet.cell(row, 2).number_format = _WEIGHT_FORMAT
        sheet.cell(row, 2).fill = _INPUT_FILL
        sheet.cell(row, 3, _preference_label(preference))
        sheet.cell(row, 3).fill = _preference_fill(preference)
        if preference.kind is CriterionType.TARGET:
            sheet.cell(row, 4, float(preference.target_value))
            sheet.cell(row, 4).fill = _INPUT_FILL
        elif preference.kind is CriterionType.BENEFIT:
            sheet.cell(row, 4, f"={matrix_letter}${raw_max_row}")
            sheet.cell(row, 4).fill = _FORMULA_FILL
        else:
            sheet.cell(row, 4, f"={matrix_letter}${raw_min_row}")
            sheet.cell(row, 4).fill = _FORMULA_FILL
        sheet.cell(row, 4).number_format = _RAW_NUMBER_FORMAT
        sheet.cell(row, 5, f"={matrix_letter}${raw_target_scale_row}")
        sheet.cell(row, 5).fill = _FORMULA_FILL
        sheet.cell(row, 5).number_format = _RAW_NUMBER_FORMAT

    weight_validation = DataValidation(
        type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=False
    )
    weight_validation.error = "Weights must be non-negative numbers."
    weight_validation.errorTitle = "Invalid weight"
    weight_validation.showErrorMessage = True
    sheet.add_data_validation(weight_validation)
    weight_validation.add(f"B{settings_data_start}:B{settings_data_end}")

    _set_section_title(
        sheet, raw_title_row, last_matrix_column, "Step 1 — Original Decision Matrix"
    )
    _write_live_weights_row(sheet, raw_weights_row, columns, settings_data_start)
    sheet.cell(raw_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, raw_header_row, column_index, criterion)
    _style_grid(
        sheet,
        raw_header_row,
        raw_target_scale_row,
        1,
        last_matrix_column,
        number_format=_RAW_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, raw_header_row, columns, preferences)
    for row_offset, (alternative, values) in enumerate(frame.iterrows()):
        row = raw_data_start + row_offset
        _set_text(sheet, row, 1, alternative)
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index, value in enumerate(values, start=2):
            sheet.cell(row, column_index, float(value))
            sheet.cell(row, column_index).fill = _WHITE_FILL
            sheet.cell(row, column_index).number_format = _RAW_NUMBER_FORMAT

    raw_summary_rows = [
        (raw_min_row, "Minimum"),
        (raw_max_row, "Maximum"),
        (raw_reference_row, "Reference x*"),
        (raw_target_scale_row, "Target deviation scale"),
    ]
    for summary_row, label in raw_summary_rows:
        sheet.cell(summary_row, 1, label)
        for column_index in range(1, last_matrix_column + 1):
            sheet.cell(summary_row, column_index).fill = _SUMMARY_FILL
            sheet.cell(summary_row, column_index).font = _HEADER_FONT
    for column_index in range(2, last_matrix_column + 1):
        letter = get_column_letter(column_index)
        settings_row = settings_data_start + column_index - 2
        preference = preferences[columns[column_index - 2]]
        sheet.cell(
            raw_min_row,
            column_index,
            f"=MIN({letter}${raw_data_start}:{letter}${raw_data_end})",
        )
        sheet.cell(
            raw_max_row,
            column_index,
            f"=MAX({letter}${raw_data_start}:{letter}${raw_data_end})",
        )
        sheet.cell(raw_reference_row, column_index, f"=$D${settings_row}")
        if preference.kind is CriterionType.TARGET:
            sheet.cell(
                raw_target_scale_row,
                column_index,
                f"=MAX(ABS({letter}${raw_max_row}-{letter}${raw_reference_row}),"
                f"ABS({letter}${raw_min_row}-{letter}${raw_reference_row}))",
            )
        else:
            sheet.cell(raw_target_scale_row, column_index, 0.0)

    normalized_title_row = raw_target_scale_row + 3
    normalized_weights_row = normalized_title_row + 1
    normalized_header_row = normalized_title_row + 2
    normalized_data_start = normalized_title_row + 3
    normalized_data_end = normalized_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        normalized_title_row,
        last_matrix_column,
        "Step 2 — Normalized Decision Matrix (r_ij)",
    )
    _write_live_weights_row(sheet, normalized_weights_row, columns, settings_data_start)
    sheet.cell(normalized_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, normalized_header_row, column_index, criterion)
    _style_grid(
        sheet,
        normalized_header_row,
        normalized_data_end,
        1,
        last_matrix_column,
        number_format=_CALC_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, normalized_header_row, columns, preferences)
    for row_offset, _alternative in enumerate(alternative_names):
        row = normalized_data_start + row_offset
        raw_row = raw_data_start + row_offset
        sheet.cell(row, 1, f"=A{raw_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index in range(2, last_matrix_column + 1):
            letter = get_column_letter(column_index)
            criterion = columns[column_index - 2]
            preference = preferences[criterion]
            if preference.kind is CriterionType.BENEFIT:
                formula = (
                    f"=IF(ABS({letter}${raw_max_row})<$B$7,0,"
                    f"{letter}{raw_row}/{letter}${raw_max_row})"
                )
            elif preference.kind is CriterionType.COST:
                formula = (
                    f"={letter}${raw_min_row}/({letter}{raw_row}+$B$7)"
                )
            else:
                formula = (
                    f"=IF(ABS({letter}${raw_target_scale_row})<$B$7,1,"
                    f"1-ABS({letter}{raw_row}-{letter}${raw_reference_row})/"
                    f"{letter}${raw_target_scale_row})"
                )
            sheet.cell(row, column_index, formula)
    sheet.cell(normalized_data_start, 2).comment = Comment(
        "Benefit, cost, and target criteria use their native ARIE normalization rules. "
        "The formulas reference the visible summary rows above.",
        "User",
    )

    weighted_title_row = normalized_data_end + 3
    weighted_weights_row = weighted_title_row + 1
    weighted_header_row = weighted_title_row + 2
    weighted_data_start = weighted_title_row + 3
    weighted_data_end = weighted_data_start + alternatives_count - 1
    ideal_row = weighted_data_end + 1
    anti_ideal_row = weighted_data_end + 2

    _set_section_title(
        sheet,
        weighted_title_row,
        last_matrix_column,
        "Step 3 & 4a — Weighted Matrix and Ideal Solutions",
    )
    _write_live_weights_row(sheet, weighted_weights_row, columns, settings_data_start)
    sheet.cell(weighted_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, weighted_header_row, column_index, criterion)
    _style_grid(
        sheet,
        weighted_header_row,
        anti_ideal_row,
        1,
        last_matrix_column,
        number_format=_CALC_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, weighted_header_row, columns, preferences)
    for row_offset, _alternative in enumerate(alternative_names):
        row = weighted_data_start + row_offset
        normalized_row = normalized_data_start + row_offset
        sheet.cell(row, 1, f"=A{normalized_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index in range(2, last_matrix_column + 1):
            letter = get_column_letter(column_index)
            sheet.cell(
                row,
                column_index,
                f"={letter}{normalized_row}*{letter}${weighted_weights_row}",
            )
    for solution_row, label, function_name in [
        (ideal_row, "Ideal solution (v_max)", "MAX"),
        (anti_ideal_row, "Anti-ideal solution (v_min)", "MIN"),
    ]:
        sheet.cell(solution_row, 1, label)
        for column_index in range(1, last_matrix_column + 1):
            sheet.cell(solution_row, column_index).fill = _SUMMARY_FILL
            sheet.cell(solution_row, column_index).font = _HEADER_FONT
        for column_index in range(2, last_matrix_column + 1):
            letter = get_column_letter(column_index)
            sheet.cell(
                solution_row,
                column_index,
                f"={function_name}({letter}${weighted_data_start}:"
                f"{letter}${weighted_data_end})",
            )

    best_title_row = anti_ideal_row + 3
    best_header_row = best_title_row + 1
    best_data_start = best_title_row + 2
    best_data_end = best_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        best_title_row,
        similarity_total_column,
        "Step 4b — Similarity-to-Ideal Contributions",
    )
    sheet.cell(best_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, best_header_row, column_index, criterion)
    sheet.cell(best_header_row, similarity_total_column, "Sim_best")
    _style_grid(
        sheet,
        best_header_row,
        best_data_end,
        1,
        similarity_total_column,
        number_format=_CALC_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, best_header_row, columns, preferences)
    sheet.cell(best_header_row, similarity_total_column).fill = _SECTION_FILL
    sheet.cell(best_header_row, similarity_total_column).font = _HEADER_FONT
    for row_offset, _alternative in enumerate(alternative_names):
        row = best_data_start + row_offset
        weighted_row = weighted_data_start + row_offset
        sheet.cell(row, 1, f"=A{weighted_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index in range(2, last_matrix_column + 1):
            letter = get_column_letter(column_index)
            sheet.cell(
                row,
                column_index,
                f"=POWER({letter}{weighted_row}/({letter}${ideal_row}+$B$7),$B$5)",
            )
        sheet.cell(
            row,
            similarity_total_column,
            f"=SUM(B{row}:{last_matrix_letter}{row})",
        )
        sheet.cell(row, similarity_total_column).fill = _SUMMARY_FILL

    worst_title_row = best_data_end + 3
    worst_header_row = worst_title_row + 1
    worst_data_start = worst_title_row + 2
    worst_data_end = worst_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        worst_title_row,
        similarity_total_column,
        "Step 4b — Similarity-to-Anti-Ideal Contributions",
    )
    sheet.cell(worst_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, worst_header_row, column_index, criterion)
    sheet.cell(worst_header_row, similarity_total_column, "Sim_worst")
    _style_grid(
        sheet,
        worst_header_row,
        worst_data_end,
        1,
        similarity_total_column,
        number_format=_CALC_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, worst_header_row, columns, preferences)
    sheet.cell(worst_header_row, similarity_total_column).fill = _SECTION_FILL
    sheet.cell(worst_header_row, similarity_total_column).font = _HEADER_FONT
    for row_offset, _alternative in enumerate(alternative_names):
        row = worst_data_start + row_offset
        weighted_row = weighted_data_start + row_offset
        sheet.cell(row, 1, f"=A{weighted_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index in range(2, last_matrix_column + 1):
            letter = get_column_letter(column_index)
            sheet.cell(
                row,
                column_index,
                f"=POWER({letter}${anti_ideal_row}/({letter}{weighted_row}+$B$7),$B$5)",
            )
        sheet.cell(
            row,
            similarity_total_column,
            f"=SUM(B{row}:{last_matrix_letter}{row})",
        )
        sheet.cell(row, similarity_total_column).fill = _SUMMARY_FILL

    results_title_row = worst_data_end + 3
    results_header_row = results_title_row + 1
    results_data_start = results_title_row + 2
    results_data_end = results_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        results_title_row,
        6,
        "Step 5 — Relative Closeness, Competition Rank, and Sort Order",
    )
    result_headers = [
        "Alternative",
        "Sim_best",
        "Sim_worst",
        "Relative Closeness (RC_i)",
        "Rank",
        "Sort Order",
    ]
    for column_index, header in enumerate(result_headers, start=1):
        sheet.cell(results_header_row, column_index, header)
        sheet.cell(results_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(results_header_row, column_index).font = _HEADER_FONT
    _style_grid(
        sheet,
        results_header_row,
        results_data_end,
        1,
        6,
        number_format=_CALC_NUMBER_FORMAT,
    )
    sheet.row_dimensions[results_header_row].height = 31
    for row_offset, _alternative in enumerate(alternative_names):
        row = results_data_start + row_offset
        best_row = best_data_start + row_offset
        worst_row = worst_data_start + row_offset
        sheet.cell(row, 1, f"=A{best_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        sheet.cell(row, 2, f"={similarity_total_letter}{best_row}")
        sheet.cell(row, 3, f"={similarity_total_letter}{worst_row}")
        sheet.cell(
            row,
            4,
            f"=($B$6*B{row})/($B$6*B{row}+(1-$B$6)*C{row}+$B$7)",
        )
        sheet.cell(
            row,
            5,
            f"=RANK(D{row},$D${results_data_start}:$D${results_data_end},0)",
        )
        sheet.cell(
            row,
            6,
            f"=E{row}+COUNTIF($D${results_data_start}:D{row},D{row})-1",
        )
        sheet.cell(row, 5).number_format = "0"
        sheet.cell(row, 6).number_format = "0"
    sheet.cell(results_data_start, 4).comment = Comment(
        "Higher relative-closeness values are preferred. Gamma is read from B5 and "
        "kappa from B6; tied values share a competition rank.",
        "User",
    )
    sheet.conditional_formatting.add(
        f"D{results_data_start}:D{results_data_end}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )
    sheet.conditional_formatting.add(
        f"A{results_data_start}:F{results_data_end}",
        FormulaRule(formula=[f"$E{results_data_start}=1"], fill=_WINNER_FILL),
    )

    ranking_title_row = results_data_end + 3
    ranking_header_row = ranking_title_row + 1
    ranking_data_start = ranking_title_row + 2
    ranking_data_end = ranking_data_start + alternatives_count - 1

    _set_section_title(sheet, ranking_title_row, 5, "Final Ranking — Sorted by Score")
    ranking_headers = [
        "Alternative",
        "Sim_best",
        "Sim_worst",
        "Relative Closeness (RC_i)",
        "Rank",
    ]
    for column_index, header in enumerate(ranking_headers, start=1):
        sheet.cell(ranking_header_row, column_index, header)
        sheet.cell(ranking_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(ranking_header_row, column_index).font = _HEADER_FONT
    _style_grid(
        sheet,
        ranking_header_row,
        ranking_data_end,
        1,
        5,
        number_format=_CALC_NUMBER_FORMAT,
    )
    sheet.row_dimensions[ranking_header_row].height = 31
    for sort_position in range(1, alternatives_count + 1):
        row = ranking_data_start + sort_position - 1
        match_formula = (
            f"MATCH({sort_position},$F${results_data_start}:$F${results_data_end},0)"
        )
        for column_index, source_letter in enumerate(["A", "B", "C", "D", "E"], start=1):
            sheet.cell(
                row,
                column_index,
                f"=INDEX(${source_letter}${results_data_start}:"
                f"${source_letter}${results_data_end},{match_formula})",
            )
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        sheet.cell(row, 5).number_format = "0"
    sheet.conditional_formatting.add(
        f"D{ranking_data_start}:D{ranking_data_end}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )
    sheet.conditional_formatting.add(
        f"A{ranking_data_start}:E{ranking_data_end}",
        FormulaRule(formula=[f"$E{ranking_data_start}=1"], fill=_WINNER_FILL),
    )

    sheet.freeze_panes = "B1"
    sheet.print_area = f"A1:{get_column_letter(model_last_column)}{ranking_data_end}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.oddFooter.center.text = "ARIE complete formula model"
    sheet.oddFooter.right.text = "Page &P of &N"
    sheet.column_dimensions["A"].width = max(
        25, min(38, max(len(value) for value in alternative_names) + 3)
    )
    for column_index, criterion in enumerate(columns, start=2):
        letter = get_column_letter(column_index)
        raw_value_width = max(len(f"{float(value):.9f}") for value in frame[criterion])
        sheet.column_dimensions[letter].width = max(
            16,
            min(30, max(len(criterion) + 3, raw_value_width + 2)),
        )
    sheet.column_dimensions[similarity_total_letter].width = 18
    for row in range(1, ranking_data_end + 1):
        if sheet.row_dimensions[row].height is None:
            sheet.row_dimensions[row].height = 19
    _autofit_column_widths(sheet)

    return {
        "ranking_data_start": ranking_data_start,
        "ranking_data_end": ranking_data_end,
    }


def _build_summary_sheet(
    workbook: Workbook,
    formula_positions: Mapping[str, int],
    alternatives_count: int,
) -> None:
    sheet = workbook.create_sheet("Decision Summary")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 90
    _set_title(sheet, 1, 14, "ARIE Decision Summary")
    sheet.merge_cells("A2:N2")
    sheet["A2"] = (
        "A live view of ARIE's ideal-evaluation ranking. Change gamma or kappa on the "
        "ARIE sheet and Excel will refresh the KPIs, table, and chart."
    )
    sheet["A2"].font = Font(name="Aptos", size=9, italic=True, color="666666")
    sheet["A2"].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[2].height = 29

    ranking_start = formula_positions["ranking_data_start"]
    cards = [
        ("A", "B", "Winner", f"='ARIE'!A{ranking_start}", "@"),
        ("C", "D", "Winning closeness", f"='ARIE'!D{ranking_start}", "0.0%"),
        ("E", "F", "Gamma (γ)", "='ARIE'!B5", "0.00"),
        ("G", "H", "Kappa (κ)", "='ARIE'!B6", "0.00"),
        (
            "I",
            "J",
            "Balance stance",
            '=IF(\'ARIE\'!B6>0.5,"Ideal emphasis",IF(\'ARIE\'!B6<0.5,'
            '"Anti-ideal emphasis","Balanced"))',
            "@",
        ),
    ]
    for start_letter, end_letter, label, formula, number_format in cards:
        sheet.merge_cells(f"{start_letter}4:{end_letter}4")
        sheet.merge_cells(f"{start_letter}5:{end_letter}6")
        label_cell = sheet[f"{start_letter}4"]
        value_cell = sheet[f"{start_letter}5"]
        label_cell.value = label
        value_cell.value = formula
        label_cell.fill = _SECTION_FILL
        value_cell.fill = _SUMMARY_FILL
        label_cell.font = _HEADER_FONT
        value_cell.font = Font(name="Aptos Display", size=15, bold=True, color="0B4F6C")
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        value_cell.number_format = number_format
        for row in range(4, 7):
            for column in range(
                sheet[start_letter + "1"].column, sheet[end_letter + "1"].column + 1
            ):
                sheet.cell(row, column).border = _GRID_BORDER

    summary_header_row = 9
    summary_data_start = 10
    summary_data_end = summary_data_start + alternatives_count - 1
    headers = [
        "Alternative",
        "Sim_best",
        "Sim_worst",
        "Relative Closeness (%)",
        "Rank",
    ]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(summary_header_row, column_index, header)
        sheet.cell(summary_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(summary_header_row, column_index).font = _HEADER_FONT
    _style_grid(
        sheet,
        summary_header_row,
        summary_data_end,
        1,
        5,
        number_format=_CALC_NUMBER_FORMAT,
    )
    for row_offset in range(alternatives_count):
        row = summary_data_start + row_offset
        source_row = ranking_start + row_offset
        for column_index, source_letter in enumerate(["A", "B", "C", "D", "E"], start=1):
            formula = f"='ARIE'!{source_letter}{source_row}"
            if source_letter == "D":
                formula += "*100"
            sheet.cell(row, column_index, formula)
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        sheet.cell(row, 4).number_format = "0.0"
        sheet.cell(row, 5).number_format = "0"
    sheet.conditional_formatting.add(
        f"D{summary_data_start}:D{summary_data_end}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )
    sheet.conditional_formatting.add(
        f"A{summary_data_start}:E{summary_data_end}",
        FormulaRule(formula=[f"$E{summary_data_start}=1"], fill=_WINNER_FILL),
    )

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "ARIE relative closeness by alternative (%)"
    chart.x_axis.title = "Alternative"
    chart.y_axis.title = "Relative closeness (%)"
    chart.height = 7.2
    chart.width = 14.2
    chart.legend = None
    data = Reference(
        sheet,
        min_col=4,
        min_row=summary_header_row,
        max_row=summary_data_end,
    )
    categories = Reference(
        sheet,
        min_col=1,
        min_row=summary_data_start,
        max_row=summary_data_end,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.numFmt = "0.0"
    chart.y_axis.numFmt = "0"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 100
    sheet.add_chart(chart, "G9")

    notes_row = max(summary_data_end + 3, 27)
    _set_section_title(sheet, notes_row, 10, "How to read this summary")
    notes = [
        "Higher relative closeness indicates a stronger alternative.",
        "Gamma sharpens or softens the similarity contribution of each criterion.",
        "Kappa above 0.5 emphasizes the ideal; below 0.5 emphasizes anti-ideal avoidance.",
        "Tied scores share a competition rank; exact ties display in input order.",
        "Verified Values preserves the canonical calculation at download time.",
    ]
    for row_offset, note in enumerate(notes, start=1):
        sheet.merge_cells(
            start_row=notes_row + row_offset,
            start_column=1,
            end_row=notes_row + row_offset,
            end_column=10,
        )
        sheet.cell(notes_row + row_offset, 1, f"• {note}")
        sheet.cell(notes_row + row_offset, 1).font = _BODY_FONT
        sheet.cell(notes_row + row_offset, 1).alignment = Alignment(vertical="center")

    widths = {
        "A": 28,
        "B": 17,
        "C": 17,
        "D": 23,
        "E": 12,
        "F": 3,
        "G": 14,
        "H": 14,
        "I": 18,
        "J": 18,
        "K": 14,
        "L": 14,
        "M": 14,
        "N": 14,
    }
    for letter, width in widths.items():
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A10"
    sheet.print_area = f"A1:N{notes_row + len(notes)}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def _build_verified_values_sheet(
    workbook: Workbook,
    frame: pd.DataFrame,
    weights: Mapping[str, float],
    preferences: Mapping[str, CriterionPreference],
    results: pd.DataFrame,
    steps: Mapping[str, Any],
    *,
    gamma: float,
    kappa: float,
) -> None:
    sheet = workbook.create_sheet("Verified Values")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 88
    max_columns = max(len(frame.columns) + 1, 6)
    _set_title(sheet, 1, max_columns, "ARIE — Verified Numerical Values")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_columns)
    sheet.cell(
        2,
        1,
        "Static values from the canonical Python implementation at download time, "
        "including criterion-level similarity contributions for a complete audit trail.",
    )
    sheet.cell(2, 1).font = Font(name="Aptos", size=9, italic=True, color="666666")
    sheet.cell(2, 1).alignment = Alignment(wrap_text=True)

    _set_section_title(sheet, 4, 6, "Parameters")
    parameter_values = [
        ["Gamma (γ)", float(gamma), "Kappa (κ)", float(kappa), "Epsilon", _EPSILON],
        ["Alternatives", len(frame), "Criteria", len(frame.columns), None, None],
    ]
    for row_offset, values in enumerate(parameter_values, start=5):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row_offset, column_index, value)
    _style_grid(sheet, 5, 6, 1, 6)
    sheet.cell(5, 2).number_format = "0.00"
    sheet.cell(5, 4).number_format = "0.00"
    sheet.cell(5, 6).number_format = "0.0E+00"

    settings_title_row = 8
    settings_header_row = 9
    settings_data_start = 10
    _set_section_title(sheet, settings_title_row, 4, "Criterion Settings")
    settings_headers = ["Criterion", "Weight", "Preference", "Reference x*"]
    for column_index, header in enumerate(settings_headers, start=1):
        sheet.cell(settings_header_row, column_index, header)
        sheet.cell(settings_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(settings_header_row, column_index).font = _HEADER_FONT
    for row_offset, criterion in enumerate(frame.columns, start=settings_data_start):
        name = str(criterion)
        preference = preferences[name]
        _set_text(sheet, row_offset, 1, name)
        sheet.cell(row_offset, 2, float(weights[name]))
        sheet.cell(row_offset, 2).number_format = _WEIGHT_FORMAT
        _set_text(sheet, row_offset, 3, _preference_label(preference))
        sheet.cell(row_offset, 3).fill = _preference_fill(preference)
        reference = (
            float(preference.target_value)
            if preference.kind is CriterionType.TARGET
            else float(frame[criterion].max())
            if preference.kind is CriterionType.BENEFIT
            else float(frame[criterion].min())
        )
        sheet.cell(row_offset, 4, reference)
        sheet.cell(row_offset, 4).number_format = _RAW_NUMBER_FORMAT
    settings_data_end = settings_data_start + len(frame.columns) - 1
    _style_grid(
        sheet,
        settings_header_row,
        settings_data_end,
        1,
        4,
        number_format=_CALC_NUMBER_FORMAT,
    )

    next_row = settings_data_end + 3
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 1 — Original Decision Matrix",
        frame,
        matrix_preferences=preferences,
    )
    normalized = steps["Step 2: Normalized Decision Matrix"].reindex(frame.index)
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 2 — Normalized Decision Matrix",
        normalized,
        matrix_preferences=preferences,
    )
    weighted = steps["Step 3: Weighted Normalized Matrix"].reindex(frame.index)
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 3 — Weighted Normalized Matrix",
        weighted,
        matrix_preferences=preferences,
    )
    ideal_values = steps["Step 4a: Ideal and Anti-Ideal Solutions"]
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 4a — Ideal and Anti-Ideal Solutions",
        ideal_values,
        header_label="Solution",
        matrix_preferences=preferences,
    )

    ideal = ideal_values.loc["Ideal Solution (v_max)"]
    anti_ideal = ideal_values.loc["Anti-Ideal Solution (v_min)"]
    sim_best_parts = weighted.div(ideal + _EPSILON, axis="columns").pow(float(gamma))
    sim_worst_parts = (anti_ideal / (weighted + _EPSILON)).pow(float(gamma))
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 4b — Similarity-to-Ideal Contributions",
        sim_best_parts,
        matrix_preferences=preferences,
    )
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 4b — Similarity-to-Anti-Ideal Contributions",
        sim_worst_parts,
        matrix_preferences=preferences,
    )

    similarity_totals = steps["Step 4b: Similarity Computations"].reindex(frame.index)
    result_values = pd.concat(
        [
            similarity_totals,
            results.reindex(frame.index)[["Relative Closeness (RC_i)", "Rank"]],
        ],
        axis=1,
    )
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 5 — Similarity Totals, Relative Closeness, and Rank",
        result_values,
    )
    rank_column = result_values.columns.get_loc("Rank") + 2
    result_data_start = next_row - len(result_values) - 1
    for row in range(result_data_start, result_data_start + len(result_values)):
        sheet.cell(row, rank_column).number_format = "0"

    final_values = results[
        ["Sim_best", "Sim_worst", "Relative Closeness (RC_i)", "Rank"]
    ]
    final_start = next_row
    _write_indexed_dataframe(
        sheet,
        final_start,
        "Final Ranking — Sorted by Rank",
        final_values,
    )
    final_data_start = final_start + 2
    for row in range(final_data_start, final_data_start + len(results)):
        sheet.cell(row, 5).number_format = "0"

    sheet.column_dimensions["A"].width = max(
        26,
        min(
            40,
            max(
                max(len(str(value)) for value in frame.index),
                max(len(str(value)) for value in frame.columns),
            )
            + 3,
        ),
    )
    for column_index in range(2, max_columns + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 22
    _autofit_column_widths(sheet)


def _build_formula_guide_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Formula Guide")
    sheet.sheet_view.showGridLines = False
    _set_title(sheet, 1, 5, "ARIE Formula Guide and Audit Trail")
    sheet.merge_cells("A2:E2")
    sheet["A2"] = (
        "The ARIE sheet contains live Excel formulas. Verified Values preserves the "
        "canonical numerical snapshot generated at download time."
    )
    sheet["A2"].alignment = Alignment(wrap_text=True)
    sheet["A2"].font = Font(name="Aptos", size=9, italic=True, color="666666")

    headers = ["Step", "Symbol", "Mathematical formula", "Meaning", "Excel implementation"]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(4, column_index, header)
        sheet.cell(4, column_index).fill = _SECTION_FILL
        sheet.cell(4, column_index).font = _HEADER_FONT

    rows = [
        (
            "2",
            "r_ij benefit",
            "x_ij / x_j^max",
            "Normalize a criterion where larger values are preferred.",
            "IF guard + division by visible maximum",
        ),
        (
            "2",
            "r_ij cost",
            "x_j^min / (x_ij + ε)",
            "Normalize a criterion where smaller positive values are preferred.",
            "Visible minimum / (input + epsilon)",
        ),
        (
            "2",
            "r_ij target",
            "1 - |x_ij-T_j| / max(|x_j^max-T_j|, |x_j^min-T_j|)",
            "Normalize a criterion around its target value.",
            "IF guard + ABS + MAX",
        ),
        (
            "3",
            "v_ij",
            "w_j × r_ij",
            "Apply the normalized criterion weight.",
            "Normalized cell × live weight cell",
        ),
        (
            "4a",
            "v_j^max / v_j^min",
            "max_i(v_ij) / min_i(v_ij)",
            "Create ideal and anti-ideal weighted benchmarks.",
            "MAX / MIN of weighted criterion range",
        ),
        (
            "4b",
            "s_ij^best",
            "[v_ij / (v_j^max + ε)]^γ",
            "Criterion contribution to similarity with the ideal.",
            "POWER with live gamma reference",
        ),
        (
            "4b",
            "s_ij^worst",
            "[v_j^min / (v_ij + ε)]^γ",
            "Criterion contribution to similarity with the anti-ideal.",
            "POWER with live gamma reference",
        ),
        (
            "4b",
            "Sim_i",
            "Σ_j s_ij",
            "Aggregate ideal or anti-ideal similarity contributions.",
            "SUM across criterion contributions",
        ),
        (
            "5",
            "RC_i",
            "κ Sim_i^best / [κ Sim_i^best + (1-κ) Sim_i^worst + ε]",
            "Final relative closeness; higher values are preferred.",
            "Live kappa reference plus epsilon guard",
        ),
        (
            "5",
            "Rank_i",
            "descending competition rank of RC_i",
            "Equal scores share a rank and skipped positions remain skipped.",
            "RANK(score, score range, 0)",
        ),
    ]
    for row_index, values in enumerate(rows, start=5):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row_index, column_index, value)
    _style_grid(sheet, 4, 4 + len(rows), 1, 5)
    for row in range(5, 5 + len(rows)):
        sheet.row_dimensions[row].height = 43

    audit_row = 6 + len(rows)
    _set_section_title(sheet, audit_row, 5, "Workbook Audit Notes")
    audit_notes = [
        "Yellow cells are intended inputs; gray cells are formulas or fixed method values.",
        "Changing gamma, kappa, target references, or weights updates all live outputs.",
        "Similarity contribution matrices make each criterion's influence directly traceable.",
        "Verified Values remains a static reference to the calculation at download time.",
        "Sort Order only makes display order unique; it does not change tied ranks.",
    ]
    for row_offset, note in enumerate(audit_notes, start=1):
        sheet.cell(audit_row + row_offset, 1, f"• {note}")
        sheet.merge_cells(
            start_row=audit_row + row_offset,
            start_column=1,
            end_row=audit_row + row_offset,
            end_column=5,
        )
        sheet.cell(audit_row + row_offset, 1).font = _BODY_FONT

    source_row = audit_row + len(audit_notes) + 2
    _set_section_title(sheet, source_row, 5, "Method Reference")
    article_url = "https://www.ejpam.com/index.php/ejpam/article/view/6578"
    doi_url = "https://doi.org/10.29020/nybg.ejpam.v18i4.6578"
    sheet.cell(source_row + 1, 1, "ARIE article")
    sheet.cell(source_row + 1, 2, article_url)
    sheet.merge_cells(
        start_row=source_row + 1,
        start_column=2,
        end_row=source_row + 1,
        end_column=5,
    )
    sheet.cell(source_row + 1, 2).hyperlink = article_url
    sheet.cell(source_row + 1, 2).style = "Hyperlink"
    sheet.cell(source_row + 2, 1, "DOI")
    sheet.cell(source_row + 2, 2, doi_url)
    sheet.merge_cells(
        start_row=source_row + 2,
        start_column=2,
        end_row=source_row + 2,
        end_column=5,
    )
    sheet.cell(source_row + 2, 2).hyperlink = doi_url
    sheet.cell(source_row + 2, 2).style = "Hyperlink"

    widths = {"A": 13, "B": 20, "C": 55, "D": 55, "E": 44}
    for letter, width in widths.items():
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A5"


def build_arie_excel_workbook(
    matrix: pd.DataFrame,
    weights: Mapping[str, Any],
    directions: Mapping[str, Any],
    *,
    gamma: float = 1.0,
    kappa: float = 0.5,
) -> bytes:
    """Return a complete ARIE workbook with live formulas and verified values."""

    frame = validate_crisp_matrix(matrix)
    frame = frame.copy()
    frame.columns = [str(column) for column in frame.columns]
    normalized_weights = validate_weights(weights, frame.columns, normalize=True)
    preferences = validate_method_capabilities("ARIE", frame.columns, directions)
    results, steps = calculate_arie(
        frame,
        normalized_weights,
        directions,
        gamma=float(gamma),
        kappa=float(kappa),
        return_steps=True,
    )

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "MCDM Calculator"
    workbook.properties.title = "Complete ARIE Formula-Driven Calculation"
    workbook.properties.subject = "Auditable ARIE decision model"
    workbook.properties.version = ARIE_EXCEL_EXPORT_REVISION
    workbook.properties.description = (
        "Live Excel formulas, a decision summary, and canonical numerical values for "
        f"every ARIE calculation stage. Export revision {ARIE_EXCEL_EXPORT_REVISION}."
    )
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    formula_positions = _build_formula_sheet(
        workbook,
        frame,
        normalized_weights,
        preferences,
        gamma=float(gamma),
        kappa=float(kappa),
    )
    _build_summary_sheet(workbook, formula_positions, len(frame))
    _build_verified_values_sheet(
        workbook,
        frame,
        normalized_weights,
        preferences,
        results,
        steps,
        gamma=float(gamma),
        kappa=float(kappa),
    )
    _build_formula_guide_sheet(workbook)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


__all__ = [
    "ARIE_EXCEL_EXPORT_FILENAME",
    "ARIE_EXCEL_EXPORT_REVISION",
    "build_arie_excel_workbook",
]
