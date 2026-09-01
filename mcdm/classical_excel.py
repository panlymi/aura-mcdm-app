"""Shared formula-rich Excel exports for classical MCDM methods.

The app calculators remain the canonical numerical implementation.  These
builders mirror their calculations with transparent Excel formulas and include
static verified-value snapshots for reproducibility.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any, Mapping, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from saw_calculator import calculate_saw
from topsis_calculator import calculate_topsis
from vikor_calculator import calculate_vikor

from .criteria import CriterionPreference, CriterionType, validate_method_capabilities
from .syai_excel import (
    _ALTERNATIVE_FILL,
    _BODY_FONT,
    _CALC_NUMBER_FORMAT,
    _FORMULA_FILL,
    _GRID_BORDER,
    _HEADER_FONT,
    _INPUT_FILL,
    _RAW_NUMBER_FORMAT,
    _SECTION_FILL,
    _SUMMARY_FILL,
    _WEIGHT_FORMAT,
    _WINNER_FILL,
    _WHITE_FILL,
    _autofit_column_widths,
    _preference_fill,
    _preference_label,
    _set_section_title,
    _set_text,
    _set_title,
    _style_grid,
    _style_matrix_header,
    _write_indexed_dataframe,
)
from .validation import validate_crisp_matrix, validate_method_matrix, validate_weights


_NUMERICAL_GUARD = 1e-9
_ERROR_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")

SAW_EXCEL_EXPORT_REVISION = "v2"
SAW_EXCEL_EXPORT_FILENAME = (
    f"saw_complete_formula_calculation_{SAW_EXCEL_EXPORT_REVISION}.xlsx"
)
TOPSIS_EXCEL_EXPORT_REVISION = "v2"
TOPSIS_EXCEL_EXPORT_FILENAME = (
    f"topsis_complete_formula_calculation_{TOPSIS_EXCEL_EXPORT_REVISION}.xlsx"
)
VIKOR_EXCEL_EXPORT_REVISION = "v2"
VIKOR_EXCEL_EXPORT_FILENAME = (
    f"vikor_complete_formula_calculation_{VIKOR_EXCEL_EXPORT_REVISION}.xlsx"
)


@dataclass(frozen=True)
class _FormulaLayout:
    method: str
    columns: list[str]
    alternatives: list[str]
    preferences: Mapping[str, CriterionPreference]
    settings_data_start: int
    settings_data_end: int
    raw_title_row: int
    raw_weights_row: int
    raw_header_row: int
    raw_data_start: int
    raw_data_end: int
    last_matrix_column: int
    model_last_column: int
    weight_sum_row: int


@dataclass(frozen=True)
class _RankingLayout:
    header_row: int
    data_start: int
    data_end: int
    score_column: int
    rank_column: int


def _legacy_directions(
    columns: Sequence[str], preferences: Mapping[str, CriterionPreference]
) -> dict[str, str | dict[str, float | str]]:
    return {column: preferences[column].to_legacy() for column in columns}


def _formula_guard() -> str:
    return f"{_NUMERICAL_GUARD:.12g}"


def _write_effective_weights_row(
    sheet,
    row: int,
    columns: Sequence[str],
    settings_data_start: int,
) -> None:
    sheet.cell(row, 1, "Effective weight")
    for offset, _criterion in enumerate(columns, start=2):
        settings_row = settings_data_start + offset - 2
        sheet.cell(row, offset, f"=$C${settings_row}")
    for cell in sheet[row][: len(columns) + 1]:
        cell.fill = _INPUT_FILL
        cell.font = _HEADER_FONT
        cell.border = _GRID_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.column > 1:
            cell.number_format = _WEIGHT_FORMAT


def _build_base_formula_sheet(
    workbook: Workbook,
    *,
    method: str,
    frame: pd.DataFrame,
    weights: Mapping[str, float],
    preferences: Mapping[str, CriterionPreference],
    subtitle: str,
    extra_setting_headers: Sequence[str],
    minimum_columns: int = 7,
    parameter_overrides: Sequence[tuple[str, Any, str]] = (),
) -> tuple[Any, _FormulaLayout]:
    sheet = workbook.create_sheet(method)
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 82

    columns = [str(column) for column in frame.columns]
    alternatives = [str(index) for index in frame.index]
    criteria_count = len(columns)
    alternatives_count = len(alternatives)
    last_matrix_column = criteria_count + 1
    settings_last_column = 4 + len(extra_setting_headers)
    model_last_column = max(minimum_columns, last_matrix_column, settings_last_column)

    _set_title(
        sheet,
        1,
        model_last_column,
        f"{method} — Complete Formula-Driven Calculation Workbook",
    )
    sheet.merge_cells(
        start_row=2, start_column=1, end_row=2, end_column=model_last_column
    )
    sheet.cell(2, 1, subtitle)
    sheet.cell(2, 1).font = Font(name="Aptos", size=9, italic=True, color="666666")
    sheet.cell(2, 1).alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[2].height = 30

    settings_title_row = 12
    settings_header_row = 13
    settings_data_start = 14
    settings_data_end = settings_data_start + criteria_count - 1
    raw_title_row = settings_data_end + 3
    raw_weights_row = raw_title_row + 1
    raw_header_row = raw_title_row + 2
    raw_data_start = raw_title_row + 3
    raw_data_end = raw_data_start + alternatives_count - 1
    weight_sum_row = 9

    _set_section_title(sheet, 4, max(5, settings_last_column), "Model Parameters and Audit Checks")
    parameter_rows = [
        (5, "Alternatives", alternatives_count, "Number of decision alternatives"),
        (6, "Criteria", criteria_count, "Number of decision criteria"),
        (
            7,
            "Benefit criteria",
            sum(1 for item in preferences.values() if item.kind is CriterionType.BENEFIT),
            "Criteria to maximize",
        ),
        (
            8,
            "Cost criteria",
            sum(1 for item in preferences.values() if item.kind is CriterionType.COST),
            "Criteria to minimize",
        ),
    ]
    for row, label, value, description in parameter_rows:
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
        sheet.cell(row, 3, description)
    for label, value, description in parameter_overrides:
        for row in range(5, 9):
            if sheet.cell(row, 1).value == label:
                sheet.cell(row, 2, value)
                sheet.cell(row, 3, description)
                break

    sheet.cell(weight_sum_row, 1, "Entered weight sum")
    sheet.cell(
        weight_sum_row,
        2,
        f"=SUM(B{settings_data_start}:B{settings_data_end})",
    )
    sheet.cell(
        weight_sum_row,
        3,
        (
            f'=IF(B{weight_sum_row}<=0,"ERROR: enter at least one positive weight",'
            f'IF(ABS(B{weight_sum_row}-1)<={_formula_guard()},'
            '"OK — weights already sum to 1",'
            '"OK — effective weights are normalized to 1"))'
        ),
    )
    _style_grid(sheet, 5, weight_sum_row, 1, max(5, settings_last_column))
    for row in range(5, 9):
        sheet.cell(row, 2).fill = _FORMULA_FILL
    sheet.cell(weight_sum_row, 2).fill = _SUMMARY_FILL
    sheet.cell(weight_sum_row, 2).number_format = _WEIGHT_FORMAT
    sheet.cell(weight_sum_row, 3).alignment = Alignment(wrap_text=True)
    sheet.conditional_formatting.add(
        f"B{weight_sum_row}",
        CellIsRule(operator="lessThanOrEqual", formula=["0"], fill=_ERROR_FILL),
    )

    _set_section_title(
        sheet,
        settings_title_row,
        settings_last_column,
        "Criterion Settings and Live References",
    )
    settings_headers = [
        "Criterion",
        "Entered Weight",
        "Effective Weight",
        "Preference",
        *extra_setting_headers,
    ]
    for column_index, header in enumerate(settings_headers, start=1):
        sheet.cell(settings_header_row, column_index, header)
        sheet.cell(settings_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(settings_header_row, column_index).font = _HEADER_FONT
        sheet.cell(settings_header_row, column_index).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    sheet.row_dimensions[settings_header_row].height = 31

    weight_validation = DataValidation(
        type="decimal",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=False,
    )
    weight_validation.promptTitle = "Non-negative criterion weight"
    weight_validation.prompt = (
        "Enter any non-negative weight. Effective weights are normalized automatically."
    )
    weight_validation.errorTitle = "Invalid weight"
    weight_validation.error = "Weights must be numeric and non-negative."
    weight_validation.errorStyle = "stop"
    weight_validation.showInputMessage = True
    weight_validation.showErrorMessage = True
    sheet.add_data_validation(weight_validation)
    weight_validation.add(f"B{settings_data_start}:B{settings_data_end}")

    for row_offset, criterion in enumerate(columns):
        row = settings_data_start + row_offset
        preference = preferences[criterion]
        _set_text(sheet, row, 1, criterion)
        sheet.cell(row, 2, float(weights[criterion]))
        sheet.cell(row, 2).number_format = _WEIGHT_FORMAT
        sheet.cell(row, 2).fill = _INPUT_FILL
        sheet.cell(row, 2).comment = Comment(
            "Editable criterion weight. Effective weights in column C are normalized to sum to one.",
            "User",
        )
        sheet.cell(
            row,
            3,
            f"=IF($B${weight_sum_row}<=0,0,B{row}/$B${weight_sum_row})",
        )
        sheet.cell(row, 3).number_format = _WEIGHT_FORMAT
        sheet.cell(row, 3).fill = _FORMULA_FILL
        _set_text(sheet, row, 4, _preference_label(preference))
        sheet.cell(row, 4).fill = _preference_fill(preference)
        sheet.cell(row, 4).comment = Comment(
            "Criterion direction is fixed by the calculation submitted in the app.",
            "User",
        )
    _style_grid(
        sheet,
        settings_data_start,
        settings_data_end,
        1,
        settings_last_column,
        number_format=_CALC_NUMBER_FORMAT,
    )
    for row in range(settings_data_start, settings_data_end + 1):
        sheet.cell(row, 2).number_format = _WEIGHT_FORMAT
        sheet.cell(row, 3).number_format = _WEIGHT_FORMAT

    _set_section_title(
        sheet,
        raw_title_row,
        last_matrix_column,
        "Step 1 — Original Decision Matrix",
    )
    _write_effective_weights_row(sheet, raw_weights_row, columns, settings_data_start)
    sheet.cell(raw_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, raw_header_row, column_index, criterion)
    _style_grid(
        sheet,
        raw_header_row,
        raw_data_end,
        1,
        last_matrix_column,
        number_format=_RAW_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, raw_header_row, columns, preferences)
    for row_offset, (alternative, row_values) in enumerate(frame.iterrows()):
        row = raw_data_start + row_offset
        _set_text(sheet, row, 1, str(alternative))
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index, value in enumerate(row_values, start=2):
            sheet.cell(row, column_index, float(value))
            sheet.cell(row, column_index).number_format = _RAW_NUMBER_FORMAT
            sheet.cell(row, column_index).fill = _WHITE_FILL
    sheet.cell(raw_data_start, 2).comment = Comment(
        "Editable decision-matrix input. Formula sections below update when these values change.",
        "User",
    )

    layout = _FormulaLayout(
        method=method,
        columns=columns,
        alternatives=alternatives,
        preferences=preferences,
        settings_data_start=settings_data_start,
        settings_data_end=settings_data_end,
        raw_title_row=raw_title_row,
        raw_weights_row=raw_weights_row,
        raw_header_row=raw_header_row,
        raw_data_start=raw_data_start,
        raw_data_end=raw_data_end,
        last_matrix_column=last_matrix_column,
        model_last_column=model_last_column,
        weight_sum_row=weight_sum_row,
    )
    return sheet, layout


def _write_formula_matrix_section(
    sheet,
    *,
    start_row: int,
    title: str,
    layout: _FormulaLayout,
) -> tuple[int, int, int, int, int]:
    title_row = start_row
    weights_row = start_row + 1
    header_row = start_row + 2
    data_start = start_row + 3
    data_end = data_start + len(layout.alternatives) - 1
    _set_section_title(sheet, title_row, layout.last_matrix_column, title)
    _write_effective_weights_row(
        sheet,
        weights_row,
        layout.columns,
        layout.settings_data_start,
    )
    sheet.cell(header_row, 1, "Alternative")
    for column_index, criterion in enumerate(layout.columns, start=2):
        _set_text(sheet, header_row, column_index, criterion)
    _style_grid(
        sheet,
        header_row,
        data_end,
        1,
        layout.last_matrix_column,
        number_format=_CALC_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, header_row, layout.columns, layout.preferences)
    for row_offset in range(len(layout.alternatives)):
        row = data_start + row_offset
        sheet.cell(row, 1, f"=A{layout.raw_data_start + row_offset}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index in range(2, layout.last_matrix_column + 1):
            sheet.cell(row, column_index).fill = _FORMULA_FILL
            sheet.cell(row, column_index).number_format = _CALC_NUMBER_FORMAT
    return title_row, weights_row, header_row, data_start, data_end


def _write_sorted_ranking(
    sheet,
    *,
    start_row: int,
    title: str,
    headers: Sequence[str],
    source_letters: Sequence[str],
    results_data_start: int,
    results_data_end: int,
    sort_order_letter: str,
    score_column: int,
    rank_column: int,
    lower_is_better: bool,
) -> _RankingLayout:
    title_row = start_row
    header_row = start_row + 1
    data_start = start_row + 2
    data_end = data_start + (results_data_end - results_data_start)
    _set_section_title(sheet, title_row, len(headers), title)
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(header_row, column_index, header)
        sheet.cell(header_row, column_index).fill = _SECTION_FILL
        sheet.cell(header_row, column_index).font = _HEADER_FONT
        sheet.cell(header_row, column_index).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    _style_grid(
        sheet,
        header_row,
        data_end,
        1,
        len(headers),
        number_format=_CALC_NUMBER_FORMAT,
    )
    sheet.row_dimensions[header_row].height = 31

    for sort_position in range(1, data_end - data_start + 2):
        row = data_start + sort_position - 1
        match_formula = (
            f"MATCH({sort_position},${sort_order_letter}${results_data_start}:"
            f"${sort_order_letter}${results_data_end},0)"
        )
        for column_index, source_letter in enumerate(source_letters, start=1):
            sheet.cell(
                row,
                column_index,
                f"=INDEX(${source_letter}${results_data_start}:"
                f"${source_letter}${results_data_end},{match_formula})",
            )
            if column_index > 1:
                sheet.cell(row, column_index).fill = _WHITE_FILL
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        sheet.cell(row, rank_column).number_format = "0"

    if lower_is_better:
        start_color, end_color = "63BE7B", "F8696B"
    else:
        start_color, end_color = "F8696B", "63BE7B"
    sheet.conditional_formatting.add(
        f"{get_column_letter(score_column)}{data_start}:"
        f"{get_column_letter(score_column)}{data_end}",
        ColorScaleRule(
            start_type="min",
            start_color=start_color,
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color=end_color,
        ),
    )
    rank_letter = get_column_letter(rank_column)
    sheet.conditional_formatting.add(
        f"A{data_start}:{get_column_letter(len(headers))}{data_end}",
        FormulaRule(formula=[f"${rank_letter}{data_start}=1"], fill=_WINNER_FILL),
    )
    return _RankingLayout(
        header_row=header_row,
        data_start=data_start,
        data_end=data_end,
        score_column=score_column,
        rank_column=rank_column,
    )


def _finalize_formula_sheet(sheet, layout: _FormulaLayout, last_row: int) -> None:
    sheet.freeze_panes = "B1"
    sheet.print_area = f"A1:{get_column_letter(layout.model_last_column)}{last_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.oddFooter.center.text = f"{layout.method} complete formula model"
    sheet.oddFooter.right.text = "Page &P of &N"
    sheet.column_dimensions["A"].width = max(
        25,
        min(38, max(len(value) for value in layout.alternatives) + 3),
    )
    for column_index, criterion in enumerate(layout.columns, start=2):
        letter = get_column_letter(column_index)
        sheet.column_dimensions[letter].width = max(16, min(30, len(criterion) + 4))
    for column_index in range(layout.last_matrix_column + 1, layout.model_last_column + 1):
        letter = get_column_letter(column_index)
        sheet.column_dimensions[letter].width = max(
            16, sheet.column_dimensions[letter].width or 0
        )
    for row in range(1, last_row + 1):
        if sheet.row_dimensions[row].height is None:
            sheet.row_dimensions[row].height = 19
    _autofit_column_widths(sheet)


def _build_summary_sheet(
    workbook: Workbook,
    *,
    method: str,
    title: str,
    subtitle: str,
    ranking: _RankingLayout,
    headers: Sequence[str],
    alternatives_count: int,
    cards: Sequence[tuple[str, str, str, str, str]],
    score_label: str,
    lower_is_better: bool,
    notes: Sequence[str],
) -> None:
    sheet = workbook.create_sheet("Decision Summary")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 90
    _set_title(sheet, 1, 14, title)
    sheet.merge_cells("A2:N2")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(name="Aptos", size=9, italic=True, color="666666")
    sheet["A2"].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[2].height = 29

    for start_letter, end_letter, label, formula, number_format in cards:
        sheet.merge_cells(f"{start_letter}4:{end_letter}4")
        sheet.merge_cells(f"{start_letter}5:{end_letter}6")
        label_cell = sheet[f"{start_letter}4"]
        value_cell = sheet[f"{start_letter}5"]
        label_cell.value = label
        value_cell.value = formula
        label_cell.fill = _SECTION_FILL
        value_cell.fill = _SUMMARY_FILL
        label_cell.font = _HEADER_FONT
        value_cell.font = Font(name="Aptos Display", size=15, bold=True, color="0B4F6C")
        label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        value_cell.number_format = number_format
        for row in range(4, 7):
            for column in range(
                sheet[start_letter + "1"].column,
                sheet[end_letter + "1"].column + 1,
            ):
                sheet.cell(row, column).border = _GRID_BORDER

    summary_header_row = 9
    summary_data_start = 10
    summary_data_end = summary_data_start + alternatives_count - 1
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(summary_header_row, column_index, header)
        sheet.cell(summary_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(summary_header_row, column_index).font = _HEADER_FONT
        sheet.cell(summary_header_row, column_index).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    _style_grid(
        sheet,
        summary_header_row,
        summary_data_end,
        1,
        len(headers),
        number_format=_CALC_NUMBER_FORMAT,
    )
    for row_offset in range(alternatives_count):
        row = summary_data_start + row_offset
        source_row = ranking.data_start + row_offset
        for column_index in range(1, len(headers) + 1):
            source_letter = get_column_letter(column_index)
            sheet.cell(row, column_index, f"='{method}'!{source_letter}{source_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        sheet.cell(row, ranking.rank_column).number_format = "0"

    score_letter = get_column_letter(ranking.score_column)
    if lower_is_better:
        start_color, end_color = "63BE7B", "F8696B"
    else:
        start_color, end_color = "F8696B", "63BE7B"
    sheet.conditional_formatting.add(
        f"{score_letter}{summary_data_start}:{score_letter}{summary_data_end}",
        ColorScaleRule(
            start_type="min",
            start_color=start_color,
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color=end_color,
        ),
    )
    rank_letter = get_column_letter(ranking.rank_column)
    sheet.conditional_formatting.add(
        f"A{summary_data_start}:{get_column_letter(len(headers))}{summary_data_end}",
        FormulaRule(formula=[f"${rank_letter}{summary_data_start}=1"], fill=_WINNER_FILL),
    )

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = f"{score_label} by alternative — ranked best to worst"
    chart.x_axis.title = "Alternative"
    chart.y_axis.title = score_label
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.numFmt = "0.00"
    chart.height = 7.2
    chart.width = 14.2
    chart.legend = None
    chart.add_data(
        Reference(
            sheet,
            min_col=ranking.score_column,
            min_row=summary_header_row,
            max_row=summary_data_end,
        ),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(
            sheet,
            min_col=1,
            min_row=summary_data_start,
            max_row=summary_data_end,
        )
    )
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.numFmt = "0.000"
    sheet.add_chart(chart, "G9")

    notes_row = max(summary_data_end + 3, 27)
    _set_section_title(sheet, notes_row, 10, "How to read this summary")
    for row_offset, note in enumerate(notes, start=1):
        row = notes_row + row_offset
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        sheet.cell(row, 1, f"• {note}")
        sheet.cell(row, 1).font = _BODY_FONT
        sheet.cell(row, 1).alignment = Alignment(vertical="center", wrap_text=True)

    widths = {
        "A": 28,
        "B": 17,
        "C": 17,
        "D": 22,
        "E": 12,
        "F": 3,
        "G": 14,
        "H": 14,
        "I": 18,
        "J": 18,
        "K": 14,
        "L": 14,
        "M": 14,
        "N": 14,
    }
    for letter, width in widths.items():
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A10"
    sheet.print_area = f"A1:N{notes_row + len(notes)}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def _begin_verified_sheet(
    workbook: Workbook,
    *,
    method: str,
    frame: pd.DataFrame,
    weights: Mapping[str, float],
    preferences: Mapping[str, CriterionPreference],
    method_parameters: Sequence[tuple[str, Any, str]] = (),
    extra_setting_headers: Sequence[str] = (),
    extra_setting_values: Mapping[str, Sequence[Any]] | None = None,
) -> tuple[Any, int]:
    sheet = workbook.create_sheet("Verified Values")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 88
    max_columns = max(len(frame.columns) + 1, 6, 4 + len(extra_setting_headers))
    _set_title(sheet, 1, max_columns, f"{method} — Verified Numerical Values")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_columns)
    sheet.cell(
        2,
        1,
        "Static values from the canonical Python implementation at download time. "
        "Use them to reconcile every live Excel calculation stage.",
    )
    sheet.cell(2, 1).font = Font(name="Aptos", size=9, italic=True, color="666666")
    sheet.cell(2, 1).alignment = Alignment(wrap_text=True)

    _set_section_title(sheet, 4, max_columns, "Parameters")
    parameter_rows = [
        ("Alternatives", len(frame), "Number of decision alternatives"),
        ("Criteria", len(frame.columns), "Number of decision criteria"),
        (
            "Benefit criteria",
            sum(1 for item in preferences.values() if item.kind is CriterionType.BENEFIT),
            "Criteria to maximize",
        ),
        (
            "Cost criteria",
            sum(1 for item in preferences.values() if item.kind is CriterionType.COST),
            "Criteria to minimize",
        ),
        *method_parameters,
    ]
    parameter_header_row = 5
    for column_index, header in enumerate(["Parameter", "Value", "Meaning"], start=1):
        sheet.cell(parameter_header_row, column_index, header)
        sheet.cell(parameter_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(parameter_header_row, column_index).font = _HEADER_FONT
    for row_offset, (label, value, meaning) in enumerate(parameter_rows, start=1):
        row = parameter_header_row + row_offset
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
        sheet.cell(row, 3, meaning)
    parameter_end = parameter_header_row + len(parameter_rows)
    _style_grid(sheet, parameter_header_row, parameter_end, 1, 3)

    settings_title_row = parameter_end + 2
    settings_header_row = settings_title_row + 1
    settings_data_start = settings_header_row + 1
    headers = ["Criterion", "Weight", "Preference", *extra_setting_headers]
    _set_section_title(sheet, settings_title_row, len(headers), "Criterion Settings")
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(settings_header_row, column_index, header)
        sheet.cell(settings_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(settings_header_row, column_index).font = _HEADER_FONT
    extra_setting_values = extra_setting_values or {}
    for row_offset, criterion in enumerate(frame.columns):
        row = settings_data_start + row_offset
        name = str(criterion)
        preference = preferences[name]
        _set_text(sheet, row, 1, name)
        sheet.cell(row, 2, float(weights[name]))
        sheet.cell(row, 2).number_format = _WEIGHT_FORMAT
        _set_text(sheet, row, 3, _preference_label(preference))
        sheet.cell(row, 3).fill = _preference_fill(preference)
        for extra_offset, value in enumerate(extra_setting_values.get(name, ()), start=4):
            sheet.cell(row, extra_offset, value)
            if isinstance(value, (float, int)):
                sheet.cell(row, extra_offset).number_format = _CALC_NUMBER_FORMAT
    settings_data_end = settings_data_start + len(frame.columns) - 1
    _style_grid(
        sheet,
        settings_header_row,
        settings_data_end,
        1,
        len(headers),
        number_format=_CALC_NUMBER_FORMAT,
    )
    return sheet, settings_data_end + 3


def _finish_verified_sheet(sheet, frame: pd.DataFrame, last_row: int) -> None:
    max_columns = max(len(frame.columns) + 1, sheet.max_column)
    sheet.column_dimensions["A"].width = max(
        26,
        min(
            40,
            max(
                max(len(str(value)) for value in frame.index),
                max(len(str(value)) for value in frame.columns),
            )
            + 3,
        ),
    )
    for column_index in range(2, max_columns + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 20
    sheet.freeze_panes = "A1"
    sheet.print_area = f"A1:{get_column_letter(max_columns)}{last_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    _autofit_column_widths(sheet)


def _build_formula_guide_sheet(
    workbook: Workbook,
    *,
    method: str,
    title: str,
    introduction: str,
    sections: Sequence[tuple[str, Sequence[tuple[str, bool]]]],
) -> None:
    sheet = workbook.create_sheet("Formula Guide")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 90
    _set_title(sheet, 1, 10, title)
    sheet.merge_cells("A2:J2")
    sheet["A2"] = introduction
    sheet["A2"].font = Font(name="Aptos", size=9, italic=True, color="666666")
    sheet["A2"].alignment = Alignment(wrap_text=True)

    row = 4
    for section_index, (heading, lines) in enumerate(sections, start=1):
        _set_section_title(sheet, row, 10, f"{section_index}. {heading}")
        row += 1
        for text, is_formula in lines:
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
            cell = sheet.cell(row, 1)
            _set_text(sheet, row, 1, text)
            if is_formula:
                cell.font = Font(name="Courier New", size=10, bold=True, color="002060")
                cell.fill = _SECTION_FILL
            elif text.startswith("https://"):
                cell.font = Font(name="Aptos", size=9, color="0563C1", underline="single")
                cell.hyperlink = text
            else:
                cell.font = _BODY_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            sheet.row_dimensions[row].height = 30 if len(text) > 140 else 22
            row += 1
        row += 2

    for letter in "ABCDEFGHIJ":
        sheet.column_dimensions[letter].width = 18
    sheet.freeze_panes = "A1"
    sheet.print_area = f"A1:J{row - 1}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.oddFooter.center.text = f"{method} formula guide"


def _new_workbook(*, method: str, revision: str) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.properties.creator = "AURA MCDM App"
    workbook.properties.title = f"Complete {method} calculation workbook"
    workbook.properties.subject = f"Auditable {method} decision model"
    workbook.properties.keywords = f"{method}, MCDM, decision analysis, formulas"
    workbook.properties.version = revision
    return workbook


def _workbook_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_saw_formula_sheet(
    workbook: Workbook,
    frame: pd.DataFrame,
    weights: Mapping[str, float],
    preferences: Mapping[str, CriterionPreference],
) -> dict[str, Any]:
    sheet, layout = _build_base_formula_sheet(
        workbook,
        method="SAW",
        frame=frame,
        weights=weights,
        preferences=preferences,
        subtitle=(
            "Edit the entered criterion weights and matrix inputs to explore the model. "
            "Weights are normalized automatically, and every ratio, weighted value, "
            "score, competition rank, KPI, and chart remains linked to the live inputs."
        ),
        extra_setting_headers=["Normalization Reference"],
    )

    for row_offset, criterion in enumerate(layout.columns):
        settings_row = layout.settings_data_start + row_offset
        column_letter = get_column_letter(row_offset + 2)
        range_ref = (
            f"{column_letter}${layout.raw_data_start}:"
            f"{column_letter}${layout.raw_data_end}"
        )
        aggregate = (
            "MAX"
            if layout.preferences[criterion].kind is CriterionType.BENEFIT
            else "MIN"
        )
        sheet.cell(settings_row, 5, f"={aggregate}({range_ref})")
        sheet.cell(settings_row, 5).fill = _FORMULA_FILL
        sheet.cell(settings_row, 5).number_format = _CALC_NUMBER_FORMAT
        sheet.cell(settings_row, 5).comment = Comment(
            "Benefit criteria use the column maximum; cost criteria use the column minimum.",
            "User",
        )

    (
        _norm_title,
        _norm_weights,
        _norm_header,
        norm_data_start,
        norm_data_end,
    ) = _write_formula_matrix_section(
        sheet,
        start_row=layout.raw_data_end + 3,
        title="Step 2 — Ratio-Normalized Decision Matrix (r_ij)",
        layout=layout,
    )
    for row_offset in range(len(layout.alternatives)):
        row = norm_data_start + row_offset
        raw_row = layout.raw_data_start + row_offset
        for column_index, criterion in enumerate(layout.columns, start=2):
            letter = get_column_letter(column_index)
            settings_row = layout.settings_data_start + column_index - 2
            if layout.preferences[criterion].kind is CriterionType.BENEFIT:
                formula = (
                    f"=IF(ABS($E${settings_row})<={_formula_guard()},0,"
                    f"{letter}{raw_row}/$E${settings_row})"
                )
            else:
                formula = (
                    f"=IF(ABS({letter}{raw_row})<={_formula_guard()},0,"
                    f"$E${settings_row}/{letter}{raw_row})"
                )
            sheet.cell(row, column_index, formula)

    (
        _weighted_title,
        weighted_weights_row,
        _weighted_header,
        weighted_data_start,
        weighted_data_end,
    ) = _write_formula_matrix_section(
        sheet,
        start_row=norm_data_end + 3,
        title="Step 3 — Weighted Normalized Decision Matrix (w_j × r_ij)",
        layout=layout,
    )
    for row_offset in range(len(layout.alternatives)):
        row = weighted_data_start + row_offset
        norm_row = norm_data_start + row_offset
        for column_index in range(2, layout.last_matrix_column + 1):
            letter = get_column_letter(column_index)
            sheet.cell(
                row,
                column_index,
                f"={letter}{norm_row}*{letter}${weighted_weights_row}",
            )

    results_title_row = weighted_data_end + 3
    results_header_row = results_title_row + 1
    results_data_start = results_title_row + 2
    results_data_end = results_data_start + len(layout.alternatives) - 1
    _set_section_title(
        sheet,
        results_title_row,
        4,
        "Step 4 — SAW Score, Competition Rank, and Tie-Safe Sort Order",
    )
    result_headers = ["Alternative", "V_i (SAW Score)", "Rank", "Sort Order"]
    for column_index, header in enumerate(result_headers, start=1):
        sheet.cell(results_header_row, column_index, header)
        sheet.cell(results_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(results_header_row, column_index).font = _HEADER_FONT
    _style_grid(
        sheet,
        results_header_row,
        results_data_end,
        1,
        4,
        number_format=_CALC_NUMBER_FORMAT,
    )
    for row_offset in range(len(layout.alternatives)):
        row = results_data_start + row_offset
        weighted_row = weighted_data_start + row_offset
        sheet.cell(row, 1, f"=A{weighted_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        sheet.cell(
            row,
            2,
            f"=SUM(B{weighted_row}:{get_column_letter(layout.last_matrix_column)}{weighted_row})",
        )
        sheet.cell(
            row,
            3,
            f"=RANK(B{row},$B${results_data_start}:$B${results_data_end},0)",
        )
        sheet.cell(
            row,
            4,
            f"=C{row}+COUNTIF($B${results_data_start}:B{row},B{row})-1",
        )
        sheet.cell(row, 2).fill = _SUMMARY_FILL
        sheet.cell(row, 3).number_format = "0"
        sheet.cell(row, 4).number_format = "0"
    sheet.cell(results_data_start, 2).comment = Comment(
        "Higher SAW scores are preferred. Competition ties share a rank; Sort Order "
        "keeps every tied alternative in the final table.",
        "User",
    )

    ranking = _write_sorted_ranking(
        sheet,
        start_row=results_data_end + 3,
        title="Final Ranking — Sorted by SAW Score",
        headers=["Alternative", "V_i (SAW Score)", "Rank"],
        source_letters=["A", "B", "C"],
        results_data_start=results_data_start,
        results_data_end=results_data_end,
        sort_order_letter="D",
        score_column=2,
        rank_column=3,
        lower_is_better=False,
    )
    _finalize_formula_sheet(sheet, layout, ranking.data_end)

    return {
        "results_data_start": results_data_start,
        "results_data_end": results_data_end,
        "ranking_data_start": ranking.data_start,
        "ranking_data_end": ranking.data_end,
        "ranking": ranking,
        "layout": layout,
        "rank_column_letter": "C",
        "score_column_letter": "B",
    }


def build_saw_excel_workbook(
    data: pd.DataFrame,
    weights: Mapping[str, Any],
    directions: Mapping[str, Any],
) -> bytes:
    """Return a complete formula-driven SAW workbook as XLSX bytes."""

    frame = validate_crisp_matrix(data)
    columns = [str(column) for column in frame.columns]
    preferences = validate_method_capabilities("SAW", columns, directions)
    validate_method_matrix("SAW", frame, directions)
    normalized_weights = validate_weights(weights, columns, normalize=True)
    results, steps = calculate_saw(
        frame,
        normalized_weights,
        _legacy_directions(columns, preferences),
        return_steps=True,
    )

    workbook = _new_workbook(method="SAW", revision=SAW_EXCEL_EXPORT_REVISION)
    pos = _build_saw_formula_sheet(workbook, frame, normalized_weights, preferences)
    ranking = pos["ranking"]
    layout = pos["layout"]

    _build_summary_sheet(
        workbook,
        method="SAW",
        title="SAW Decision Summary",
        subtitle=(
            "A live summary of the Simple Additive Weighting result. Change weights "
            "or matrix values on the SAW sheet to refresh every KPI, row, and chart."
        ),
        ranking=ranking,
        headers=["Alternative", "V_i (SAW Score)", "Rank"],
        alternatives_count=len(frame),
        cards=[
            ("A", "B", "First Rank-1 Alternative", f"='SAW'!A{ranking.data_start}", "@"),
            ("C", "D", "Winning SAW Score", f"='SAW'!B{ranking.data_start}", "0.0000"),
            ("E", "F", "Alternatives", "='SAW'!B5", "0"),
            ("G", "H", "Total Criteria", "='SAW'!B6", "0"),
            ("I", "J", "Benefit Criteria", "='SAW'!B7", "0"),
        ],
        score_label="SAW score (V_i)",
        lower_is_better=False,
        notes=[
            "Higher V_i is preferred; a score is the sum of effective-weighted normalized ratings.",
            "Benefit values use x_ij / max(x_j); cost values use min(x_j) / x_ij.",
            "Entered weights are normalized in the workbook, matching the app calculator.",
            "Tied scores share competition ranks, while the sort helper retains every tied row.",
            "The Verified Values sheet preserves the canonical calculation at download time.",
        ],
    )

    reference_values = {
        criterion: [
            float(frame[criterion].max())
            if preferences[criterion].kind is CriterionType.BENEFIT
            else float(frame[criterion].min())
        ]
        for criterion in columns
    }
    verified, next_row = _begin_verified_sheet(
        workbook,
        method="SAW",
        frame=frame,
        weights=normalized_weights,
        preferences=preferences,
        extra_setting_headers=["Normalization Reference"],
        extra_setting_values=reference_values,
    )
    next_row = _write_indexed_dataframe(
        verified,
        next_row,
        "Step 1 — Original Decision Matrix",
        frame,
        matrix_preferences=preferences,
    )
    next_row = _write_indexed_dataframe(
        verified,
        next_row,
        "Step 2 — Ratio-Normalized Decision Matrix",
        steps["Step 2: Normalized Decision Matrix"].reindex(frame.index),
        matrix_preferences=preferences,
    )
    next_row = _write_indexed_dataframe(
        verified,
        next_row,
        "Step 3 — Weighted Normalized Decision Matrix",
        steps["Step 3: Weighted Normalized Matrix"].reindex(frame.index),
        matrix_preferences=preferences,
    )
    final_values = results[["V_i (SAW Score)", "Rank"]]
    final_start = next_row
    next_row = _write_indexed_dataframe(
        verified,
        final_start,
        "Step 4 — Final SAW Score and Ranking",
        final_values,
    )
    final_data_start = final_start + 2
    for row in range(final_data_start, final_data_start + len(final_values)):
        verified.cell(row, 3).number_format = "0"
    _finish_verified_sheet(verified, frame, next_row - 1)

    _build_formula_guide_sheet(
        workbook,
        method="SAW",
        title="SAW Formula Guide and Audit Trail",
        introduction=(
            "Equations and audit notes for the Simple Additive Weighting implementation "
            "used by this app (benefit and cost criteria)."
        ),
        sections=[
            (
                "Effective criterion weights",
                [
                    (
                        "Entered non-negative weights are normalized so that their effective total equals one.",
                        False,
                    ),
                    ("w_j = entered_weight_j / Σ_k entered_weight_k", True),
                ],
            ),
            (
                "Ratio normalization",
                [
                    ("For a benefit criterion: r_ij = x_ij / max_i(x_ij)", True),
                    ("For a cost criterion: r_ij = min_i(x_ij) / x_ij", True),
                    (
                        "The app requires non-negative benefit columns with at least one positive value and strictly positive cost columns.",
                        False,
                    ),
                ],
            ),
            (
                "Aggregation and ranking",
                [
                    ("V_i = Σ_j w_j r_ij", True),
                    (
                        "Alternatives are ranked from the highest V_i to the lowest. Exact ties receive the same competition rank.",
                        False,
                    ),
                ],
            ),
            (
                "Reference and implementation scope",
                [
                    (
                        "Fishburn, P. C. (1967). Additive Utilities with Incomplete Product Sets: Application to Priorities and Assignments. Operations Research, 15(3), 537–542.",
                        False,
                    ),
                    ("https://doi.org/10.1287/opre.15.3.537", False),
                    (
                        "This workbook documents the app's stated SAW normalization rules and does not silently transform unsupported target criteria.",
                        False,
                    ),
                ],
            ),
        ],
    )
    return _workbook_bytes(workbook)


def build_topsis_excel_workbook(
    data: pd.DataFrame,
    weights: Mapping[str, Any],
    directions: Mapping[str, Any],
) -> bytes:
    """Return a complete formula-driven TOPSIS workbook as XLSX bytes."""

    frame = validate_crisp_matrix(data)
    columns = [str(column) for column in frame.columns]
    preferences = validate_method_capabilities("TOPSIS", columns, directions)
def _build_topsis_formula_sheet(
    workbook: Workbook,
    frame: pd.DataFrame,
    weights: Mapping[str, float],
    preferences: Mapping[str, CriterionPreference],
) -> dict[str, Any]:
    sheet, layout = _build_base_formula_sheet(
        workbook,
        method="TOPSIS",
        frame=frame,
        weights=weights,
        preferences=preferences,
        subtitle=(
            "Edit the entered criterion weights and matrix inputs to explore the model. "
            "The workbook recalculates vector normalization, ideal solutions, Euclidean "
            "distances, closeness, ranks, summary KPIs, and the chart."
        ),
        extra_setting_headers=["Vector Denominator √(Σx²)"],
    )
    for row_offset, _criterion in enumerate(layout.columns):
        settings_row = layout.settings_data_start + row_offset
        letter = get_column_letter(row_offset + 2)
        sheet.cell(
            settings_row,
            5,
            f"=SQRT(SUMSQ({letter}${layout.raw_data_start}:{letter}${layout.raw_data_end}))",
        )
        sheet.cell(settings_row, 5).fill = _FORMULA_FILL
        sheet.cell(settings_row, 5).number_format = _CALC_NUMBER_FORMAT

    (
        _norm_title,
        _norm_weights,
        _norm_header,
        norm_data_start,
        norm_data_end,
    ) = _write_formula_matrix_section(
        sheet,
        start_row=layout.raw_data_end + 3,
        title="Step 2 — Vector-Normalized Decision Matrix (r_ij)",
        layout=layout,
    )
    for row_offset in range(len(layout.alternatives)):
        row = norm_data_start + row_offset
        raw_row = layout.raw_data_start + row_offset
        for column_index in range(2, layout.last_matrix_column + 1):
            letter = get_column_letter(column_index)
            settings_row = layout.settings_data_start + column_index - 2
            sheet.cell(
                row,
                column_index,
                f"=IF(ABS($E${settings_row})<={_formula_guard()},0,"
                f"{letter}{raw_row}/$E${settings_row})",
            )

    (
        _weighted_title,
        weighted_weights_row,
        _weighted_header,
        weighted_data_start,
        weighted_data_end,
    ) = _write_formula_matrix_section(
        sheet,
        start_row=norm_data_end + 3,
        title="Step 3 — Weighted Normalized Decision Matrix (v_ij = w_j × r_ij)",
        layout=layout,
    )
    for row_offset in range(len(layout.alternatives)):
        row = weighted_data_start + row_offset
        norm_row = norm_data_start + row_offset
        for column_index in range(2, layout.last_matrix_column + 1):
            letter = get_column_letter(column_index)
            sheet.cell(
                row,
                column_index,
                f"={letter}{norm_row}*{letter}${weighted_weights_row}",
            )

    ideal_title_row = weighted_data_end + 3
    ideal_header_row = ideal_title_row + 1
    pis_row = ideal_title_row + 2
    nis_row = ideal_title_row + 3
    _set_section_title(
        sheet,
        ideal_title_row,
        layout.last_matrix_column,
        "Step 4 — Positive-Ideal (A+) and Negative-Ideal (A−) Solutions",
    )
    sheet.cell(ideal_header_row, 1, "Solution")
    for column_index, criterion in enumerate(layout.columns, start=2):
        _set_text(sheet, ideal_header_row, column_index, criterion)
    _style_grid(
        sheet,
        ideal_header_row,
        nis_row,
        1,
        layout.last_matrix_column,
        number_format=_CALC_NUMBER_FORMAT,
    )
    _style_matrix_header(sheet, ideal_header_row, layout.columns, layout.preferences)
    sheet.cell(pis_row, 1, "PIS (A+)")
    sheet.cell(nis_row, 1, "NIS (A−)")
    sheet.cell(pis_row, 1).fill = _ALTERNATIVE_FILL
    sheet.cell(nis_row, 1).fill = _ALTERNATIVE_FILL
    for column_index, criterion in enumerate(layout.columns, start=2):
        letter = get_column_letter(column_index)
        weighted_range = (
            f"{letter}${weighted_data_start}:{letter}${weighted_data_end}"
        )
        is_benefit = layout.preferences[criterion].kind is CriterionType.BENEFIT
        sheet.cell(pis_row, column_index, f"={'MAX' if is_benefit else 'MIN'}({weighted_range})")
        sheet.cell(nis_row, column_index, f"={'MIN' if is_benefit else 'MAX'}({weighted_range})")
        sheet.cell(pis_row, column_index).fill = _SUMMARY_FILL
        sheet.cell(nis_row, column_index).fill = _FORMULA_FILL

    results_title_row = nis_row + 3
    results_header_row = results_title_row + 1
    results_data_start = results_title_row + 2
    results_data_end = results_data_start + len(layout.alternatives) - 1
    _set_section_title(
        sheet,
        results_title_row,
        6,
        "Step 5 — Separation, Relative Closeness, Competition Rank, and Sort Order",
    )
    result_headers = [
        "Alternative",
        "D+ (Ideal)",
        "D− (Anti-Ideal)",
        "Relative Closeness (C_i)",
        "Rank",
        "Sort Order",
    ]
    for column_index, header in enumerate(result_headers, start=1):
        sheet.cell(results_header_row, column_index, header)
        sheet.cell(results_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(results_header_row, column_index).font = _HEADER_FONT
        sheet.cell(results_header_row, column_index).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    _style_grid(
        sheet,
        results_header_row,
        results_data_end,
        1,
        6,
        number_format=_CALC_NUMBER_FORMAT,
    )
    for row_offset in range(len(layout.alternatives)):
        row = results_data_start + row_offset
        weighted_row = weighted_data_start + row_offset
        sheet.cell(row, 1, f"=A{weighted_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        plus_terms = [
            f"({get_column_letter(column_index)}{weighted_row}-"
            f"{get_column_letter(column_index)}${pis_row})^2"
            for column_index in range(2, layout.last_matrix_column + 1)
        ]
        minus_terms = [
            f"({get_column_letter(column_index)}{weighted_row}-"
            f"{get_column_letter(column_index)}${nis_row})^2"
            for column_index in range(2, layout.last_matrix_column + 1)
        ]
        sheet.cell(row, 2, f"=SQRT(SUM({','.join(plus_terms)}))")
        sheet.cell(row, 3, f"=SQRT(SUM({','.join(minus_terms)}))")
        sheet.cell(
            row,
            4,
            f"=IF(B{row}+C{row}<={_formula_guard()},0,C{row}/(B{row}+C{row}))",
        )
        sheet.cell(
            row,
            5,
            f"=RANK(D{row},$D${results_data_start}:$D${results_data_end},0)",
        )
        sheet.cell(
            row,
            6,
            f"=E{row}+COUNTIF($D${results_data_start}:D{row},D{row})-1",
        )
        sheet.cell(row, 4).fill = _SUMMARY_FILL
        sheet.cell(row, 5).number_format = "0"
        sheet.cell(row, 6).number_format = "0"
    sheet.cell(results_data_start, 4).comment = Comment(
        "Higher relative closeness is preferred. Exact ties share a competition rank; "
        "Sort Order preserves each tied alternative in the ranking table.",
        "User",
    )

    ranking = _write_sorted_ranking(
        sheet,
        start_row=results_data_end + 3,
        title="Final Ranking — Sorted by Relative Closeness",
        headers=[
            "Alternative",
            "D+ (Ideal)",
            "D− (Anti-Ideal)",
            "Relative Closeness (C_i)",
            "Rank",
        ],
        source_letters=["A", "B", "C", "D", "E"],
        results_data_start=results_data_start,
        results_data_end=results_data_end,
        sort_order_letter="F",
        score_column=4,
        rank_column=5,
        lower_is_better=False,
    )
    _finalize_formula_sheet(sheet, layout, ranking.data_end)

    return {
        "results_data_start": results_data_start,
        "results_data_end": results_data_end,
        "ranking_data_start": ranking.data_start,
        "ranking_data_end": ranking.data_end,
        "ranking": ranking,
        "layout": layout,
        "rank_column_letter": "E",
        "score_column_letter": "D",
    }


def build_topsis_excel_workbook(
    data: pd.DataFrame,
    weights: Mapping[str, Any],
    directions: Mapping[str, Any],
) -> bytes:
    """Return a complete formula-driven TOPSIS workbook as XLSX bytes."""

    frame = validate_crisp_matrix(data)
    columns = [str(column) for column in frame.columns]
    preferences = validate_method_capabilities("TOPSIS", columns, directions)
    normalized_weights = validate_weights(weights, columns, normalize=True)
    results, steps = calculate_topsis(
        frame,
        normalized_weights,
        _legacy_directions(columns, preferences),
        return_steps=True,
    )

    workbook = _new_workbook(method="TOPSIS", revision=TOPSIS_EXCEL_EXPORT_REVISION)
    pos = _build_topsis_formula_sheet(workbook, frame, normalized_weights, preferences)
    ranking = pos["ranking"]
    layout = pos["layout"]

    _build_summary_sheet(
        workbook,
        method="TOPSIS",
        title="TOPSIS Decision Summary",
        subtitle=(
            "A live summary of closeness to the positive ideal and distance from the "
            "negative ideal. Change inputs on the TOPSIS sheet to refresh the result."
        ),
        ranking=ranking,
        headers=[
            "Alternative",
            "D+ (Ideal)",
            "D− (Anti-Ideal)",
            "Relative Closeness (C_i)",
            "Rank",
        ],
        alternatives_count=len(frame),
        cards=[
            (
                "A",
                "B",
                "First Rank-1 Alternative",
                f"='TOPSIS'!A{ranking.data_start}",
                "@",
            ),
            (
                "C",
                "D",
                "Winning Closeness",
                f"='TOPSIS'!D{ranking.data_start}",
                "0.0000",
            ),
            ("E", "F", "Alternatives", "='TOPSIS'!B5", "0"),
            ("G", "H", "Total Criteria", "='TOPSIS'!B6", "0"),
            ("I", "J", "Benefit Criteria", "='TOPSIS'!B7", "0"),
        ],
        score_label="Relative closeness (C_i)",
        lower_is_better=False,
        notes=[
            "Higher relative closeness is preferred; values normally lie between zero and one.",
            "The positive ideal uses the best weighted value per criterion; the negative ideal uses the worst.",
            "D+ measures distance to the positive ideal and D− measures distance to the negative ideal.",
            "Entered weights are normalized in the workbook, matching the app calculator.",
            "Competition ties retain all alternatives through a separate sequential sort helper.",
        ],
    )

    denominator_values = {
        criterion: [float((frame[criterion] ** 2).sum() ** 0.5)]
        for criterion in columns
    }
    verified, next_row = _begin_verified_sheet(
        workbook,
        method="TOPSIS",
        frame=frame,
        weights=normalized_weights,
        preferences=preferences,
        extra_setting_headers=["Vector Denominator √(Σx²)"],
        extra_setting_values=denominator_values,
    )
    next_row = _write_indexed_dataframe(
        verified,
        next_row,
        "Step 1 — Original Decision Matrix",
        frame,
        matrix_preferences=preferences,
    )
    next_row = _write_indexed_dataframe(
        verified,
        next_row,
        "Step 2 — Vector-Normalized Decision Matrix",
        steps["Step 2: Normalized Decision Matrix ($r_{ij}$)"].reindex(frame.index),
        matrix_preferences=preferences,
    )
    next_row = _write_indexed_dataframe(
        verified,
        next_row,
        "Step 3 — Weighted Normalized Decision Matrix",
        steps["Step 3: Weighted Normalized Matrix ($v_{ij}$)"].reindex(frame.index),
        matrix_preferences=preferences,
    )
    next_row = _write_indexed_dataframe(
        verified,
        next_row,
        "Step 4 — Positive-Ideal and Negative-Ideal Solutions",
        steps["Step 4: Ideal and Anti-Ideal Solutions"],
        header_label="Solution",
        matrix_preferences=preferences,
    )
    next_row = _write_indexed_dataframe(
        verified,
        next_row,
        "Step 5 — Separation Measures",
        steps["Step 5: Separation Measures"].reindex(frame.index),
    )
    final_values = results[
        [
            "D+ (Ideal)",
            "D- (Anti-Ideal)",
            "Relative Closeness (C_i)",
            "Rank",
        ]
    ]
    final_start = next_row
    next_row = _write_indexed_dataframe(
        verified,
        final_start,
        "Step 6 — Relative Closeness and Final Ranking",
        final_values,
    )
    for row in range(final_start + 2, final_start + 2 + len(final_values)):
        verified.cell(row, 5).number_format = "0"
    _finish_verified_sheet(verified, frame, next_row - 1)

    _build_formula_guide_sheet(
        workbook,
        method="TOPSIS",
        title="TOPSIS Formula Guide and Audit Trail",
        introduction=(
            "Equations and audit notes for the vector-normalized TOPSIS implementation "
            "used by this app (benefit and cost criteria)."
        ),
        sections=[
            (
                "Vector normalization and weighting",
                [
                    ("r_ij = x_ij / √(Σ_i x_ij²)", True),
                    ("v_ij = w_j r_ij, with Σ_j w_j = 1", True),
                    (
                        "A zero vector denominator produces a zero normalized column, matching the app's numerical guard.",
                        False,
                    ),
                ],
            ),
            (
                "Ideal solutions",
                [
                    (
                        "For benefit criteria, A+ uses max(v_ij) and A− uses min(v_ij); cost criteria reverse those choices.",
                        False,
                    ),
                    ("A+_j = best weighted value; A−_j = worst weighted value", True),
                ],
            ),
            (
                "Distances, closeness, and ranking",
                [
                    ("D+_i = √(Σ_j (v_ij − A+_j)²)", True),
                    ("D−_i = √(Σ_j (v_ij − A−_j)²)", True),
                    ("C_i = D−_i / (D+_i + D−_i)", True),
                    (
                        "Higher C_i is preferred. If both distances are numerically zero, the app assigns C_i = 0.",
                        False,
                    ),
                ],
            ),
            (
                "Primary references",
                [
                    (
                        "Hwang, C.-L., & Yoon, K. (1981). Multiple Attribute Decision Making: Methods and Applications. Springer.",
                        False,
                    ),
                    ("https://doi.org/10.1007/978-3-642-48318-9", False),
                    (
                        "Opricovic, S., & Tzeng, G.-H. (2004). Compromise solution by MCDM methods: A comparative analysis of VIKOR and TOPSIS. European Journal of Operational Research, 156(2), 445–455.",
                        False,
                    ),
                    ("https://doi.org/10.1016/S0377-2217(03)00020-1", False),
                ],
            ),
        ],
    )
    return _workbook_bytes(workbook)


def _build_vikor_formula_sheet(
    workbook: Workbook,
    frame: pd.DataFrame,
    weights: Mapping[str, float],
    preferences: Mapping[str, CriterionPreference],
    v_param: float = 0.5,
) -> dict[str, Any]:
    sheet, layout = _build_base_formula_sheet(
        workbook,
        method="VIKOR",
        frame=frame,
        weights=weights,
        preferences=preferences,
        subtitle=(
            "Edit v, the entered criterion weights, and matrix inputs to explore the "
            "compromise ranking. The live model exposes every best/worst reference, "
            "loss contribution, utility, regret, Q index, rank, KPI, and chart."
        ),
        extra_setting_headers=["Best (f*)", "Worst (f−)"],
        minimum_columns=8,
    )
    sheet["D5"] = "v (majority strategy)"
    sheet["E5"] = float(v_param)
    sheet["E5"].fill = _INPUT_FILL
    sheet["E5"].number_format = "0.00"
    sheet["F5"] = "0 = regret emphasis; 0.5 = consensus; 1 = group utility emphasis"
    sheet["E5"].comment = Comment(
        "Editable VIKOR compromise parameter. It must remain between 0 and 1.",
        "User",
    )
    v_validation = DataValidation(
        type="decimal",
        operator="between",
        formula1="0",
        formula2="1",
        allow_blank=False,
    )
    v_validation.promptTitle = "VIKOR strategy parameter"
    v_validation.prompt = "Enter v between 0 and 1 inclusive."
    v_validation.errorTitle = "Invalid v"
    v_validation.error = "v must be between 0 and 1."
    v_validation.errorStyle = "stop"
    v_validation.showInputMessage = True
    v_validation.showErrorMessage = True
    sheet.add_data_validation(v_validation)
    v_validation.add(sheet["E5"])

    for row_offset, criterion in enumerate(layout.columns):
        settings_row = layout.settings_data_start + row_offset
        letter = get_column_letter(row_offset + 2)
        source_range = (
            f"{letter}${layout.raw_data_start}:{letter}${layout.raw_data_end}"
        )
        is_benefit = layout.preferences[criterion].kind is CriterionType.BENEFIT
        sheet.cell(settings_row, 5, f"={'MAX' if is_benefit else 'MIN'}({source_range})")
        sheet.cell(settings_row, 6, f"={'MIN' if is_benefit else 'MAX'}({source_range})")
        for column_index in (5, 6):
            sheet.cell(settings_row, column_index).fill = _FORMULA_FILL
            sheet.cell(settings_row, column_index).number_format = _CALC_NUMBER_FORMAT

    (
        _distance_title,
        _distance_weights,
        _distance_header,
        distance_data_start,
        distance_data_end,
    ) = _write_formula_matrix_section(
        sheet,
        start_row=layout.raw_data_end + 3,
        title="Step 2 — Normalized Distance from the Criterion Best (d_ij)",
        layout=layout,
    )
    for row_offset in range(len(layout.alternatives)):
        row = distance_data_start + row_offset
        raw_row = layout.raw_data_start + row_offset
        for column_index, criterion in enumerate(layout.columns, start=2):
            letter = get_column_letter(column_index)
            settings_row = layout.settings_data_start + column_index - 2
            if layout.preferences[criterion].kind is CriterionType.BENEFIT:
                numerator = f"$E${settings_row}-{letter}{raw_row}"
                denominator = f"$E${settings_row}-$F${settings_row}"
            else:
                numerator = f"{letter}{raw_row}-$E${settings_row}"
                denominator = f"$F${settings_row}-$E${settings_row}"
            sheet.cell(
                row,
                column_index,
                f"=IF(ABS({denominator})<={_formula_guard()},0,"
                f"({numerator})/({denominator}))",
            )

    (
        _weighted_title,
        weighted_weights_row,
        _weighted_header,
        weighted_data_start,
        weighted_data_end,
    ) = _write_formula_matrix_section(
        sheet,
        start_row=distance_data_end + 3,
        title="Step 3 — Weighted Normalized Distance Matrix (w_j × d_ij)",
        layout=layout,
    )
    for row_offset in range(len(layout.alternatives)):
        row = weighted_data_start + row_offset
        distance_row = distance_data_start + row_offset
        for column_index in range(2, layout.last_matrix_column + 1):
            letter = get_column_letter(column_index)
            sheet.cell(
                row,
                column_index,
                f"={letter}{distance_row}*{letter}${weighted_weights_row}",
            )

    sr_title_row = weighted_data_end + 3
    sr_header_row = sr_title_row + 1
    sr_data_start = sr_title_row + 2
    sr_data_end = sr_data_start + len(layout.alternatives) - 1
    _set_section_title(
        sheet,
        sr_title_row,
        3,
        "Step 4 — Group Utility (S_i) and Individual Regret (R_i)",
    )
    for column_index, header in enumerate(["Alternative", "S_i", "R_i"], start=1):
        sheet.cell(sr_header_row, column_index, header)
        sheet.cell(sr_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(sr_header_row, column_index).font = _HEADER_FONT
    _style_grid(
        sheet,
        sr_header_row,
        sr_data_end,
        1,
        3,
        number_format=_CALC_NUMBER_FORMAT,
    )
    last_matrix_letter = get_column_letter(layout.last_matrix_column)
    for row_offset in range(len(layout.alternatives)):
        row = sr_data_start + row_offset
        weighted_row = weighted_data_start + row_offset
        sheet.cell(row, 1, f"=A{weighted_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        sheet.cell(row, 2, f"=SUM(B{weighted_row}:{last_matrix_letter}{weighted_row})")
        sheet.cell(row, 3, f"=MAX(B{weighted_row}:{last_matrix_letter}{weighted_row})")

    q_parameters_title_row = sr_data_end + 3
    q_parameters_header_row = q_parameters_title_row + 1
    q_parameters_data_start = q_parameters_title_row + 2
    _set_section_title(
        sheet,
        q_parameters_title_row,
        3,
        "Step 5 — VIKOR Index Parameters",
    )
    for column_index, header in enumerate(["Parameter", "Live Value", "Meaning"], start=1):
        sheet.cell(q_parameters_header_row, column_index, header)
        sheet.cell(q_parameters_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(q_parameters_header_row, column_index).font = _HEADER_FONT
    q_parameter_rows = [
        ("v", "=$E$5", "Weight of group utility relative to individual regret"),
        ("S*", f"=MIN(B{sr_data_start}:B{sr_data_end})", "Best (minimum) S_i"),
        ("S−", f"=MAX(B{sr_data_start}:B{sr_data_end})", "Worst (maximum) S_i"),
        ("R*", f"=MIN(C{sr_data_start}:C{sr_data_end})", "Best (minimum) R_i"),
        ("R−", f"=MAX(C{sr_data_start}:C{sr_data_end})", "Worst (maximum) R_i"),
    ]
    q_parameter_cells: dict[str, int] = {}
    for row_offset, (label, formula, meaning) in enumerate(q_parameter_rows):
        row = q_parameters_data_start + row_offset
        q_parameter_cells[label] = row
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, formula)
        sheet.cell(row, 2).fill = _FORMULA_FILL
        sheet.cell(row, 2).number_format = _CALC_NUMBER_FORMAT
        sheet.cell(row, 3, meaning)
    q_parameters_data_end = q_parameters_data_start + len(q_parameter_rows) - 1
    _style_grid(
        sheet,
        q_parameters_header_row,
        q_parameters_data_end,
        1,
        3,
        number_format=_CALC_NUMBER_FORMAT,
    )

    results_title_row = q_parameters_data_end + 3
    results_header_row = results_title_row + 1
    results_data_start = results_title_row + 2
    results_data_end = results_data_start + len(layout.alternatives) - 1
    _set_section_title(
        sheet,
        results_title_row,
        8,
        "Step 6 — Normalized Utility/Regret Terms, Q Index, Rank, and Sort Order",
    )
    result_headers = [
        "Alternative",
        "S_i (Utility)",
        "R_i (Regret)",
        "Normalized S Term",
        "Normalized R Term",
        "Q_i (VIKOR Index)",
        "Rank",
        "Sort Order",
    ]
    for column_index, header in enumerate(result_headers, start=1):
        sheet.cell(results_header_row, column_index, header)
        sheet.cell(results_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(results_header_row, column_index).font = _HEADER_FONT
        sheet.cell(results_header_row, column_index).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    _style_grid(
        sheet,
        results_header_row,
        results_data_end,
        1,
        8,
        number_format=_CALC_NUMBER_FORMAT,
    )
    s_star_row = q_parameter_cells["S*"]
    s_minus_row = q_parameter_cells["S−"]
    r_star_row = q_parameter_cells["R*"]
    r_minus_row = q_parameter_cells["R−"]
    v_row = q_parameter_cells["v"]
    for row_offset in range(len(layout.alternatives)):
        row = results_data_start + row_offset
        sr_row = sr_data_start + row_offset
        sheet.cell(row, 1, f"=A{sr_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        sheet.cell(row, 2, f"=B{sr_row}")
        sheet.cell(row, 3, f"=C{sr_row}")
        sheet.cell(
            row,
            4,
            f"=IF(ABS($B${s_minus_row}-$B${s_star_row})<={_formula_guard()},0,"
            f"(B{row}-$B${s_star_row})/($B${s_minus_row}-$B${s_star_row}))",
        )
        sheet.cell(
            row,
            5,
            f"=IF(ABS($B${r_minus_row}-$B${r_star_row})<={_formula_guard()},0,"
            f"(C{row}-$B${r_star_row})/($B${r_minus_row}-$B${r_star_row}))",
        )
        sheet.cell(row, 6, f"=$B${v_row}*D{row}+(1-$B${v_row})*E{row}")
        sheet.cell(
            row,
            7,
            f"=RANK(F{row},$F${results_data_start}:$F${results_data_end},1)",
        )
        sheet.cell(
            row,
            8,
            f"=G{row}+COUNTIF($F${results_data_start}:F{row},F{row})-1",
        )
        sheet.cell(row, 6).fill = _SUMMARY_FILL
        sheet.cell(row, 7).number_format = "0"
        sheet.cell(row, 8).number_format = "0"
    sheet.cell(results_data_start, 6).comment = Comment(
        "Lower Q_i is preferred. Competition ties share a rank; Sort Order ensures "
        "that each tied alternative appears once in the final ranking.",
        "User",
    )

    ranking = _write_sorted_ranking(
        sheet,
        start_row=results_data_end + 3,
        title="Step 7 — Final Ranking Sorted by VIKOR Q Index",
        headers=[
            "Alternative",
            "S_i (Utility)",
            "R_i (Regret)",
            "Q_i (VIKOR Index)",
            "Rank",
        ],
        source_letters=["A", "B", "C", "F", "G"],
        results_data_start=results_data_start,
        results_data_end=results_data_end,
        sort_order_letter="H",
        score_column=4,
        rank_column=5,
        lower_is_better=True,
    )
    _finalize_formula_sheet(sheet, layout, ranking.data_end)

    return {
        "results_data_start": results_data_start,
        "results_data_end": results_data_end,
        "ranking_data_start": ranking.data_start,
        "ranking_data_end": ranking.data_end,
        "ranking": ranking,
        "layout": layout,
        "rank_column_letter": "G",
        "score_column_letter": "F",
    }


def build_vikor_excel_workbook(
    data: pd.DataFrame,
    weights: Mapping[str, Any],
    directions: Mapping[str, Any],
    v_param: float = 0.5,
) -> bytes:
    """Return a complete formula-driven VIKOR workbook as XLSX bytes."""

    frame = validate_crisp_matrix(data)
    columns = [str(column) for column in frame.columns]
    preferences = validate_method_capabilities("VIKOR", columns, directions)
    normalized_weights = validate_weights(weights, columns, normalize=True)
    results, steps = calculate_vikor(
        frame,
        normalized_weights,
        _legacy_directions(columns, preferences),
        v_param=float(v_param),
        return_steps=True,
    )

    workbook = _new_workbook(method="VIKOR", revision=VIKOR_EXCEL_EXPORT_REVISION)
    pos = _build_vikor_formula_sheet(workbook, frame, normalized_weights, preferences, v_param=float(v_param))
    ranking = pos["ranking"]
    layout = pos["layout"]

    _build_summary_sheet(
        workbook,
        method="VIKOR",
        title="VIKOR Decision Summary",
        subtitle=(
            "A live summary of the compromise ranking. Lower Q_i values are preferred; "
            "change v, weights, or matrix values on the VIKOR sheet to refresh the model."
        ),
        ranking=ranking,
        headers=[
            "Alternative",
            "S_i (Utility)",
            "R_i (Regret)",
            "Q_i (VIKOR Index)",
            "Rank",
        ],
        alternatives_count=len(frame),
        cards=[
            (
                "A",
                "B",
                "First Rank-1 Alternative",
                f"='VIKOR'!A{ranking.data_start}",
                "@",
            ),
            (
                "C",
                "D",
                "Winning Q Index",
                f"='VIKOR'!D{ranking.data_start}",
                "0.0000",
            ),
            ("E", "F", "v (Majority Weight)", "='VIKOR'!E5", "0.00"),
            ("G", "H", "Alternatives", "='VIKOR'!B5", "0"),
            ("I", "J", "Total Criteria", "='VIKOR'!B6", "0"),
        ],
        score_label="VIKOR index (Q_i; lower is better)",
        lower_is_better=True,
        notes=[
            "Lower Q_i is preferred; Q_i combines normalized group utility and individual regret.",
            "S_i is the sum of weighted losses, while R_i is the largest single weighted loss.",
            "v = 0.5 balances both strategies; larger v emphasizes group utility and smaller v emphasizes regret.",
            "Entered weights are normalized in the workbook, matching the app calculator.",
            "This app reports the Q ranking; it does not infer a separate VIKOR compromise set from acceptance conditions.",
        ],
    )

    best_worst = steps["Step 2: Best (f*) and Worst (f-) Values"]
    extra_settings = {
        criterion: [
            float(best_worst.loc["Best (f*)", criterion]),
            float(best_worst.loc["Worst (f-)", criterion]),
        ]
        for criterion in columns
    }
    verified, next_row = _begin_verified_sheet(
        workbook,
        method="VIKOR",
        frame=frame,
        weights=normalized_weights,
        preferences=preferences,
        method_parameters=[
            ("v", float(v_param), "Weight of group utility relative to individual regret")
        ],
        extra_setting_headers=["Best (f*)", "Worst (f−)"],
        extra_setting_values=extra_settings,
    )
    next_row = _write_indexed_dataframe(
        verified,
        next_row,
        "Step 1 — Original Decision Matrix",
        frame,
        matrix_preferences=preferences,
    )
    next_row = _write_indexed_dataframe(
        verified,
        next_row,
        "Step 2 — Best and Worst Criterion Values",
        best_worst,
        header_label="Solution",
        matrix_preferences=preferences,
    )

    normalized_distance = pd.DataFrame(index=frame.index, columns=frame.columns, dtype=float)
    for criterion in frame.columns:
        name = str(criterion)
        best = float(best_worst.loc["Best (f*)", criterion])
        worst = float(best_worst.loc["Worst (f-)", criterion])
        denominator = abs(best - worst)
        if denominator <= _NUMERICAL_GUARD:
            normalized_distance[criterion] = 0.0
        elif preferences[name].kind is CriterionType.BENEFIT:
            normalized_distance[criterion] = (best - frame[criterion]) / (best - worst)
        else:
            normalized_distance[criterion] = (frame[criterion] - best) / (worst - best)
    next_row = _write_indexed_dataframe(
        verified,
        next_row,
        "Step 3 — Normalized Distance from the Criterion Best",
        normalized_distance,
        matrix_preferences=preferences,
    )
    next_row = _write_indexed_dataframe(
        verified,
        next_row,
        "Step 4 — Weighted Normalized Distance Matrix",
        steps["Step 3: Weighted Normalized Distance Matrix"].reindex(frame.index),
        matrix_preferences=preferences,
    )
    next_row = _write_indexed_dataframe(
        verified,
        next_row,
        "Step 5 — Group Utility and Individual Regret",
        steps["Step 4: Utility (S_i) and Regret (R_i) Measures"].reindex(frame.index),
    )
    q_parameters = steps["Step 5: VIKOR Index (Q_i) Parameters"]
    q_parameter_frame = pd.DataFrame(
        {"Value": [float(value) for value in q_parameters.values()]},
        index=list(q_parameters.keys()),
    )
    next_row = _write_indexed_dataframe(
        verified,
        next_row,
        "Step 6 — VIKOR Index Parameters",
        q_parameter_frame,
        header_label="Parameter",
    )
    final_values = results[
        ["S_i (Utility)", "R_i (Regret)", "Q_i (VIKOR Index)", "Rank"]
    ]
    final_start = next_row
    next_row = _write_indexed_dataframe(
        verified,
        final_start,
        "Step 7 — Final VIKOR Index and Ranking",
        final_values,
    )
    for row in range(final_start + 2, final_start + 2 + len(final_values)):
        verified.cell(row, 5).number_format = "0"
    _finish_verified_sheet(verified, frame, next_row - 1)

    _build_formula_guide_sheet(
        workbook,
        method="VIKOR",
        title="VIKOR Formula Guide and Audit Trail",
        introduction=(
            "Equations and audit notes for the VIKOR compromise-index implementation "
            "used by this app (benefit and cost criteria)."
        ),
        sections=[
            (
                "Criterion best, worst, and normalized loss",
                [
                    (
                        "Benefit criteria use f*_j = max_i(f_ij) and f−_j = min_i(f_ij); cost criteria reverse these definitions.",
                        False,
                    ),
                    ("Benefit: d_ij = (f*_j − f_ij) / (f*_j − f−_j)", True),
                    ("Cost: d_ij = (f_ij − f*_j) / (f−_j − f*_j)", True),
                    (
                        "If a criterion has no range, its normalized loss is zero for every alternative.",
                        False,
                    ),
                ],
            ),
            (
                "Group utility and individual regret",
                [
                    ("S_i = Σ_j w_j d_ij", True),
                    ("R_i = max_j(w_j d_ij)", True),
                    (
                        "Lower S_i and R_i indicate smaller aggregate and worst-criterion loss from the ideal.",
                        False,
                    ),
                ],
            ),
            (
                "Compromise index and ranking",
                [
                    (
                        "Q_i = v(S_i − S*)/(S− − S*) + (1 − v)(R_i − R*)/(R− − R*)",
                        True,
                    ),
                    ("S* = min_i S_i; S− = max_i S_i; R* = min_i R_i; R− = max_i R_i", True),
                    (
                        "Lower Q_i is preferred. A zero S or R range contributes zero, matching the app's numerical guard.",
                        False,
                    ),
                    (
                        "The workbook mirrors the app's Q ranking only; acceptable-advantage and stability conditions for a compromise set are not computed.",
                        False,
                    ),
                ],
            ),
            (
                "Primary reference",
                [
                    (
                        "Opricovic, S., & Tzeng, G.-H. (2004). Compromise solution by MCDM methods: A comparative analysis of VIKOR and TOPSIS. European Journal of Operational Research, 156(2), 445–455.",
                        False,
                    ),
                    ("https://doi.org/10.1016/S0377-2217(03)00020-1", False),
                ],
            ),
        ],
    )
    return _workbook_bytes(workbook)


__all__ = [
    "SAW_EXCEL_EXPORT_FILENAME",
    "SAW_EXCEL_EXPORT_REVISION",
    "TOPSIS_EXCEL_EXPORT_REVISION",
    "TOPSIS_EXCEL_EXPORT_FILENAME",
    "VIKOR_EXCEL_EXPORT_FILENAME",
    "VIKOR_EXCEL_EXPORT_REVISION",
    "build_saw_excel_workbook",
    "build_topsis_excel_workbook",
    "build_vikor_excel_workbook",
    "_build_saw_formula_sheet",
    "_build_topsis_formula_sheet",
    "_build_vikor_formula_sheet",
]
