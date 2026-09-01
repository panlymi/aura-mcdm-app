"""Formula-rich multi-method comparison and agreement analysis Excel workbook export."""

from __future__ import annotations

from io import BytesIO
from typing import Any, Mapping, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .agreement import (
    calculate_agreement_table,
    calculate_pairwise_kendall_matrix,
    calculate_pairwise_spearman_matrix,
    get_default_jaccard_cutoffs,
)
from .analysis import compare_methods
from .aras_excel import _build_formula_sheet as _build_aras_sheet
from .arie_excel import _build_formula_sheet as _build_arie_sheet
from .aura_excel import _build_formula_sheet as _build_aura_sheet
from .classical_excel import (
    _build_saw_formula_sheet,
    _build_topsis_formula_sheet,
    _build_vikor_formula_sheet,
)
from .criteria import validate_method_capabilities
from .moora_excel import _build_formula_sheet as _build_moora_sheet
from .syai_excel import (
    _ALTERNATIVE_FILL,
    _BODY_FONT,
    _CALC_NUMBER_FORMAT,
    _FORMULA_FILL,
    _GRID_BORDER,
    _HEADER_FONT,
    _INPUT_FILL,
    _RAW_NUMBER_FORMAT,
    _SECTION_FILL,
    _SUMMARY_FILL,
    _WEIGHT_FORMAT,
    _WINNER_FILL,
    _WHITE_FILL,
    _autofit_column_widths,
    _build_formula_sheet as _build_syai_sheet,
    _set_section_title,
    _set_text,
    _set_title,
    _style_grid,
)
from .validation import validate_crisp_matrix, validate_weights
from .waspas_excel import _build_formula_sheet as _build_waspas_sheet


COMPARISON_EXCEL_EXPORT_REVISION = "v1"
COMPARISON_EXCEL_EXPORT_FILENAME = (
    f"mcdm_method_comparison_{COMPARISON_EXCEL_EXPORT_REVISION}.xlsx"
)


def _build_comparison_summary_sheet(
    workbook: Workbook,
    frame: pd.DataFrame,
    active_methods: Sequence[str],
    excluded_methods: Mapping[str, str],
    benchmark_method: str,
    method_positions: Mapping[str, dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet("Comparison & Agreement", 0)
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 85

    alternatives = [str(index) for index in frame.index]
    alternatives_count = len(alternatives)
    methods_count = len(active_methods)
    last_col = max(7, methods_count + 1)

    _set_title(
        sheet,
        1,
        last_col,
        "MCDM Cross-Method Comparison & Agreement Analysis",
    )
    sheet.merge_cells(
        start_row=2, start_column=1, end_row=2, end_column=last_col
    )
    sheet.cell(
        2,
        1,
        "Comprehensive side-by-side comparison, agreement metrics (Spearman ρ, Kendall τ-b, MARD, Top-k Jaccard), "
        "pairwise correlation matrices, and live links to each method's full formula calculation tab.",
    )
    sheet.cell(2, 1).font = Font(name="Aptos", size=9, italic=True, color="666666")
    sheet.cell(2, 1).alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[2].height = 28

    # Section 1: Overview & Benchmark
    _set_section_title(sheet, 4, 4, "Comparison Parameters")
    params = [
        ("Alternatives", alternatives_count, "Number of evaluated alternatives"),
        ("Criteria", len(frame.columns), "Number of criteria in decision matrix"),
        ("Compared Methods", methods_count, "Number of successfully evaluated MCDM methods"),
        ("Benchmark Method", benchmark_method, "Reference baseline for agreement metrics"),
    ]
    for row_idx, (label, val, desc) in enumerate(params, start=5):
        sheet.cell(row_idx, 1, label)
        sheet.cell(row_idx, 2, val)
        sheet.cell(row_idx, 3, desc)
    _style_grid(sheet, 5, 8, 1, 4)
    sheet.cell(5, 2).fill = _FORMULA_FILL
    sheet.cell(6, 2).fill = _FORMULA_FILL
    sheet.cell(7, 2).fill = _FORMULA_FILL
    sheet.cell(8, 2).fill = _INPUT_FILL
    sheet.cell(8, 2).font = Font(name="Aptos", size=10, bold=True, color="002060")

    # Section 2: Side-by-Side Method Rankings Table
    rank_title_row = 10
    rank_header_row = 11
    rank_data_start = 12
    rank_data_end = rank_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        rank_title_row,
        methods_count + 1,
        "1. Method Ranking Comparison (Live Formula-Linked)",
    )
    sheet.cell(rank_header_row, 1, "Alternative")
    sheet.cell(rank_header_row, 1).fill = _SECTION_FILL
    sheet.cell(rank_header_row, 1).font = _HEADER_FONT

    for col_idx, method in enumerate(active_methods, start=2):
        sheet.cell(rank_header_row, col_idx, method)
        sheet.cell(rank_header_row, col_idx).fill = _SECTION_FILL
        sheet.cell(rank_header_row, col_idx).font = _HEADER_FONT
    sheet.row_dimensions[rank_header_row].height = 24

    for row_offset, alt in enumerate(alternatives):
        row = rank_data_start + row_offset
        _set_text(sheet, row, 1, alt)
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL

        for col_idx, method in enumerate(active_methods, start=2):
            pos = method_positions[method]
            sheet_name = method
            rank_col = pos["rank_column_letter"]
            method_data_start = pos["results_data_start"]
            method_row = method_data_start + row_offset

            formula = f"='{sheet_name}'!{rank_col}{method_row}"
            sheet.cell(row, col_idx, formula)
            sheet.cell(row, col_idx).fill = _WHITE_FILL
            sheet.cell(row, col_idx).number_format = "0"

    _style_grid(
        sheet,
        rank_data_start,
        rank_data_end,
        1,
        methods_count + 1,
        number_format="0",
    )
    for r in range(rank_data_start, rank_data_end + 1):
        for c in range(2, methods_count + 2):
            sheet.cell(r, c).number_format = "0"

    # Conditional formatting for #1 ranks in the comparison table
    sheet.conditional_formatting.add(
        f"B{rank_data_start}:{get_column_letter(methods_count + 1)}{rank_data_end}",
        FormulaRule(formula=["B" + str(rank_data_start) + "=1"], fill=_WINNER_FILL),
    )

    # Section 3: Agreement Table (Table 3)
    k1, k2 = get_default_jaccard_cutoffs(alternatives_count)
    agree_title_row = rank_data_end + 3
    agree_header_row = agree_title_row + 1
    agree_data_start = agree_header_row + 1
    agree_data_end = agree_data_start + methods_count - 1

    _set_section_title(
        sheet,
        agree_title_row,
        6,
        f"2. Agreement of MCDM Methods with {benchmark_method} (Table 3)",
    )
    agree_headers = [
        "Method",
        "Spearman ρ",
        "Kendall τ-b",
        "MARD",
        f"Top-{k1} Jaccard",
        f"Top-{k2} Jaccard",
    ]
    for col_idx, h in enumerate(agree_headers, start=1):
        sheet.cell(agree_header_row, col_idx, h)
        sheet.cell(agree_header_row, col_idx).fill = _SECTION_FILL
        sheet.cell(agree_header_row, col_idx).font = _HEADER_FONT
    sheet.row_dimensions[agree_header_row].height = 24

    # Build canonical rankings DataFrame from formula sheet inputs to compute exact agreement metrics
    rankings_dict = {}
    for m in active_methods:
        # Fetch canonical ranks from method positions
        rankings_dict[m] = method_positions[m]["canonical_ranks"]
    rankings_df = pd.DataFrame(rankings_dict, index=frame.index)
    agreement_df = calculate_agreement_table(rankings_df, benchmark_method=benchmark_method, k1=k1, k2=k2)

    for row_offset, method in enumerate(active_methods):
        row = agree_data_start + row_offset
        _set_text(sheet, row, 1, method)
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL

        if method in agreement_df.index:
            row_data = agreement_df.loc[method]
            sheet.cell(row, 2, float(row_data["Spearman ρ"]))
            sheet.cell(row, 3, float(row_data["Kendall τ-b"]))
            sheet.cell(row, 4, float(row_data["MARD"]))
            sheet.cell(row, 5, float(row_data[f"Top-{k1} Jaccard"]))
            sheet.cell(row, 6, float(row_data[f"Top-{k2} Jaccard"]))

            sheet.cell(row, 2).number_format = "0.0000"
            sheet.cell(row, 3).number_format = "0.0000"
            sheet.cell(row, 4).number_format = "0.00"
            sheet.cell(row, 5).number_format = "0.000"
            sheet.cell(row, 6).number_format = "0.000"

            for col_idx in range(2, 7):
                sheet.cell(row, col_idx).fill = _WHITE_FILL
                if method == benchmark_method:
                    sheet.cell(row, col_idx).fill = _SUMMARY_FILL

    _style_grid(
        sheet,
        agree_data_start,
        agree_data_end,
        1,
        6,
        number_format="0.0000",
    )
    for r in range(agree_data_start, agree_data_end + 1):
        sheet.cell(r, 4).number_format = "0.00"
        sheet.cell(r, 5).number_format = "0.000"
        sheet.cell(r, 6).number_format = "0.000"

    # Note below Table 3
    note_row = agree_data_end + 1
    sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
    sheet.cell(
        note_row,
        1,
        f"Note. MARD = mean absolute rank difference. Benchmark = {benchmark_method}. "
        f"Top-{k1} and Top-{k2} Jaccard coefficients measure subset overlap among top-ranked alternatives.",
    )
    sheet.cell(note_row, 1).font = Font(name="Aptos", size=9, italic=True, color="555555")

    # Section 4: Pairwise Spearman Rank Correlation Matrix
    spearman_title_row = note_row + 3
    spearman_header_row = spearman_title_row + 1
    spearman_data_start = spearman_header_row + 1
    spearman_data_end = spearman_data_start + methods_count - 1

    _set_section_title(
        sheet,
        spearman_title_row,
        methods_count + 1,
        "3. Pairwise Spearman Rank Correlation Matrix (ρ)",
    )
    sheet.cell(spearman_header_row, 1, "Method")
    sheet.cell(spearman_header_row, 1).fill = _SECTION_FILL
    sheet.cell(spearman_header_row, 1).font = _HEADER_FONT
    for col_idx, method in enumerate(active_methods, start=2):
        sheet.cell(spearman_header_row, col_idx, method)
        sheet.cell(spearman_header_row, col_idx).fill = _SECTION_FILL
        sheet.cell(spearman_header_row, col_idx).font = _HEADER_FONT

    spearman_matrix = calculate_pairwise_spearman_matrix(rankings_df)
    for row_offset, m1 in enumerate(active_methods):
        row = spearman_data_start + row_offset
        _set_text(sheet, row, 1, m1)
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for col_idx, m2 in enumerate(active_methods, start=2):
            val = float(spearman_matrix.loc[m1, m2]) if m1 in spearman_matrix.index and m2 in spearman_matrix.columns else 1.0
            sheet.cell(row, col_idx, val)
            sheet.cell(row, col_idx).number_format = "0.0000"
            sheet.cell(row, col_idx).fill = _WHITE_FILL

    _style_grid(
        sheet,
        spearman_data_start,
        spearman_data_end,
        1,
        methods_count + 1,
        number_format="0.0000",
    )
    sheet.conditional_formatting.add(
        f"B{spearman_data_start}:{get_column_letter(methods_count + 1)}{spearman_data_end}",
        ColorScaleRule(
            start_type="num",
            start_value=-1.0,
            start_color="F8696B",
            mid_type="num",
            mid_value=0.0,
            mid_color="FFEB84",
            end_type="num",
            end_value=1.0,
            end_color="63BE7B",
        ),
    )

    # Section 5: Pairwise Kendall Tau-b Correlation Matrix
    kendall_title_row = spearman_data_end + 3
    kendall_header_row = kendall_title_row + 1
    kendall_data_start = kendall_header_row + 1
    kendall_data_end = kendall_data_start + methods_count - 1

    _set_section_title(
        sheet,
        kendall_title_row,
        methods_count + 1,
        "4. Pairwise Kendall Tau-b Rank Correlation Matrix (τ-b)",
    )
    sheet.cell(kendall_header_row, 1, "Method")
    sheet.cell(kendall_header_row, 1).fill = _SECTION_FILL
    sheet.cell(kendall_header_row, 1).font = _HEADER_FONT
    for col_idx, method in enumerate(active_methods, start=2):
        sheet.cell(kendall_header_row, col_idx, method)
        sheet.cell(kendall_header_row, col_idx).fill = _SECTION_FILL
        sheet.cell(kendall_header_row, col_idx).font = _HEADER_FONT

    kendall_matrix = calculate_pairwise_kendall_matrix(rankings_df)
    for row_offset, m1 in enumerate(active_methods):
        row = kendall_data_start + row_offset
        _set_text(sheet, row, 1, m1)
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for col_idx, m2 in enumerate(active_methods, start=2):
            val = float(kendall_matrix.loc[m1, m2]) if m1 in kendall_matrix.index and m2 in kendall_matrix.columns else 1.0
            sheet.cell(row, col_idx, val)
            sheet.cell(row, col_idx).number_format = "0.0000"
            sheet.cell(row, col_idx).fill = _WHITE_FILL

    _style_grid(
        sheet,
        kendall_data_start,
        kendall_data_end,
        1,
        methods_count + 1,
        number_format="0.0000",
    )
    sheet.conditional_formatting.add(
        f"B{kendall_data_start}:{get_column_letter(methods_count + 1)}{kendall_data_end}",
        ColorScaleRule(
            start_type="num",
            start_value=-1.0,
            start_color="F8696B",
            mid_type="num",
            mid_value=0.0,
            mid_color="FFEB84",
            end_type="num",
            end_value=1.0,
            end_color="63BE7B",
        ),
    )

    # Section 6: Top Performing Alternatives (#1 Wins)
    wins_title_row = kendall_data_end + 3
    wins_header_row = wins_title_row + 1
    wins_data_start = wins_header_row + 1
    wins_data_end = wins_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        wins_title_row,
        3,
        "5. Alternative Consensus (#1 Ranks Count)",
    )
    sheet.cell(wins_header_row, 1, "Alternative")
    sheet.cell(wins_header_row, 2, "Times Ranked #1")
    sheet.cell(wins_header_row, 3, "Consensus %")
    sheet.cell(wins_header_row, 1).fill = _SECTION_FILL
    sheet.cell(wins_header_row, 2).fill = _SECTION_FILL
    sheet.cell(wins_header_row, 3).fill = _SECTION_FILL
    sheet.cell(wins_header_row, 1).font = _HEADER_FONT
    sheet.cell(wins_header_row, 2).font = _HEADER_FONT
    sheet.cell(wins_header_row, 3).font = _HEADER_FONT

    last_method_letter = get_column_letter(methods_count + 1)
    for row_offset, alt in enumerate(alternatives):
        row = wins_data_start + row_offset
        comp_rank_row = rank_data_start + row_offset
        _set_text(sheet, row, 1, alt)
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        sheet.cell(
            row,
            2,
            f'=COUNTIF(B{comp_rank_row}:{last_method_letter}{comp_rank_row},1)',
        )
        sheet.cell(
            row,
            3,
            f'=B{row}/{methods_count}',
        )
        sheet.cell(row, 2).fill = _WHITE_FILL
        sheet.cell(row, 3).fill = _WHITE_FILL
        sheet.cell(row, 2).number_format = "0"
        sheet.cell(row, 3).number_format = "0.0%"

    _style_grid(
        sheet,
        wins_data_start,
        wins_data_end,
        1,
        3,
        number_format="0",
    )
    for r in range(wins_data_start, wins_data_end + 1):
        sheet.cell(r, 2).number_format = "0"
        sheet.cell(r, 3).number_format = "0.0%"

    # Excluded methods note if any
    if excluded_methods:
        ex_row = wins_data_end + 2
        _set_section_title(sheet, ex_row, 4, "Excluded Methods (Criterion Type Incompatibility)")
        for off, (ex_m, reason) in enumerate(excluded_methods.items(), start=1):
            sheet.cell(ex_row + off, 1, f"• {ex_m}: {reason}")
            sheet.cell(ex_row + off, 1).font = Font(name="Aptos", size=9, italic=True, color="777777")

    sheet.freeze_panes = "B12"
    sheet.print_area = f"A1:{get_column_letter(last_col)}{wins_data_end + 5}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    _autofit_column_widths(sheet)


def build_comparison_excel_workbook(
    data: pd.DataFrame,
    weights: Mapping[str, Any],
    directions: Mapping[str, Any],
    methods: Sequence[str],
    benchmark_method: str | None = None,
    parameters: Mapping[str, Any] | None = None,
) -> bytes:
    """Return a multi-sheet comparison workbook as XLSX bytes.
    
    Contains a primary 'Comparison & Agreement' summary tab, followed by
    full calculation sheets for every active selected method.
    """
    frame = validate_crisp_matrix(data)
    columns = [str(column) for column in frame.columns]
    normalized_weights = validate_weights(weights, columns, normalize=True)
    params = dict(parameters or {})

    # Run method comparison
    comparison_df, excluded_methods = compare_methods(
        methods,
        frame,
        normalized_weights,
        directions,
        parameters=params,
    )
    active_methods = list(comparison_df.columns)
    if not active_methods:
        raise ValueError("None of the selected methods could be calculated with the current configuration.")

    if benchmark_method is None or benchmark_method not in active_methods:
        benchmark_method = active_methods[0]

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.properties.version = COMPARISON_EXCEL_EXPORT_REVISION
    workbook.properties.creator = "AURA MCDM App"
    workbook.properties.title = "MCDM Multi-Method Comparison & Agreement Analysis"
    workbook.properties.subject = "Auditable Cross-Method MCDM Ranking Comparison"

    method_positions: dict[str, dict[str, Any]] = {}

    # Build individual method sheets
    for method in active_methods:
        m_upper = method.upper()
        preferences = validate_method_capabilities(m_upper, columns, directions)

        if m_upper == "AURA":
            alpha = float(params.get("alpha", 0.5))
            p = int(params.get("p", 1))
            pos = _build_aura_sheet(workbook, frame, normalized_weights, preferences, alpha=alpha, p=p)
            pos["rank_column_letter"] = "I"
        elif m_upper == "SYAI":
            beta = float(params.get("beta", 0.5))
            pos = _build_syai_sheet(workbook, frame, normalized_weights, preferences, beta=beta)
            pos["rank_column_letter"] = "E"
        elif m_upper == "ARIE":
            gamma = float(params.get("gamma", 0.5))
            kappa = float(params.get("kappa", 0.5))
            pos = _build_arie_sheet(workbook, frame, normalized_weights, preferences, gamma=gamma, kappa=kappa)
            pos["rank_column_letter"] = "E"
        elif m_upper == "WASPAS":
            lambda_val = float(params.get("lambda", 0.5))
            pos = _build_waspas_sheet(workbook, frame, normalized_weights, preferences, lambda_value=lambda_val)
            pos["rank_column_letter"] = "E"
        elif m_upper == "MOORA":
            pos = _build_moora_sheet(workbook, frame, normalized_weights, preferences)
            pos["rank_column_letter"] = "E"
        elif m_upper == "SAW":
            saw_norm = params.get("saw_normalization", "ratio_to_max")
            pos = _build_saw_formula_sheet(
                workbook, frame, normalized_weights, preferences, normalization=saw_norm
            )
            pos["rank_column_letter"] = "C"
        elif m_upper == "TOPSIS":
            pos = _build_topsis_formula_sheet(workbook, frame, normalized_weights, preferences)
            pos["rank_column_letter"] = "E"
        elif m_upper == "VIKOR":
            v_param = float(params.get("v", 0.5))
            pos = _build_vikor_formula_sheet(workbook, frame, normalized_weights, preferences, v_param=v_param)
            pos["rank_column_letter"] = "G"
        elif m_upper == "ARAS":
            pos = _build_aras_sheet(workbook, frame, normalized_weights, preferences)
            pos["rank_column_letter"] = "D"
        else:
            continue

        pos["canonical_ranks"] = comparison_df[method]
        method_positions[method] = pos

    # Build primary Comparison & Agreement sheet at index 0
    _build_comparison_summary_sheet(
        workbook,
        frame,
        active_methods,
        excluded_methods,
        benchmark_method,
        method_positions,
    )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


__all__ = [
    "COMPARISON_EXCEL_EXPORT_FILENAME",
    "COMPARISON_EXCEL_EXPORT_REVISION",
    "build_comparison_excel_workbook",
]
