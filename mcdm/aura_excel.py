"""Formula-rich Excel export for complete, auditable AURA calculations."""

from __future__ import annotations

from io import BytesIO
import re
from typing import Any, Mapping

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from aura_calculator import calculate_aura

from .criteria import CriterionPreference, CriterionType, validate_method_capabilities
from .validation import validate_crisp_matrix, validate_weights


_TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
_SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
_WEIGHT_FILL = PatternFill("solid", fgColor="F8CBAD")  # Orange, Accent 6, Lighter 60%
_BENEFIT_FILL = PatternFill("solid", fgColor="00B0F0")  # Light Blue (Standard Colors)
_ALTERNATIVE_FILL = PatternFill("solid", fgColor="FFF200")  # Yellow
_COST_FILL = PatternFill("solid", fgColor="FF3B30")
_TARGET_FILL = PatternFill("solid", fgColor="F4B183")
_SUMMARY_FILL = PatternFill("solid", fgColor="E2F0D9")
_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
_WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
_THIN_BLACK = Side(style="thin", color="000000")
_MEDIUM_NAVY = Side(style="medium", color="1F4E78")
_GRID_BORDER = Border(
    left=_THIN_BLACK,
    right=_THIN_BLACK,
    top=_THIN_BLACK,
    bottom=_THIN_BLACK,
)
_SECTION_BORDER = Border(bottom=_MEDIUM_NAVY)
_BODY_FONT = Font(name="Courier New", size=10, color="000000")
_HEADER_FONT = Font(name="Courier New", size=10, bold=True, color="000000")
_TITLE_FONT = Font(name="Courier New", size=15, bold=True, color="FFFFFF")
_ILLEGAL_EXCEL_TEXT = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

_DECIMAL_FORMAT = "General"
_WEIGHT_FORMAT = "General"

# Increment this whenever the generated workbook changes.  The Streamlit app
# includes it in the cache key so a deployment cannot serve older workbook
# bytes merely because the calculation inputs are unchanged.
AURA_EXCEL_EXPORT_REVISION = "v8"
AURA_EXCEL_EXPORT_FILENAME = (
    f"aura_complete_formula_calculation_{AURA_EXCEL_EXPORT_REVISION}.xlsx"
)


def _safe_excel_text(value: Any) -> str:
    """Return user text that Excel will store as text, never as a formula."""

    return _ILLEGAL_EXCEL_TEXT.sub(" ", str(value))[:32767]


def _set_text(sheet, row: int, column: int, value: Any) -> None:
    cell = sheet.cell(row, column)
    cell.value = _safe_excel_text(value)
    cell.data_type = "s"


def _preference_label(preference: CriterionPreference) -> str:
    if preference.kind is CriterionType.BENEFIT:
        return "Benefit (maximize)"
    if preference.kind is CriterionType.COST:
        return "Cost (minimize)"
    return "Target"


def _preference_fill(preference: CriterionPreference) -> PatternFill:
    if preference.kind is CriterionType.BENEFIT:
        return _BENEFIT_FILL
    if preference.kind is CriterionType.COST:
        return _COST_FILL
    return _TARGET_FILL


def _set_title(sheet, row: int, last_column: int, text: str) -> None:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_column)
    cell = sheet.cell(row=row, column=1, value=text)
    cell.fill = _TITLE_FILL
    cell.font = _TITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _SECTION_BORDER
    sheet.row_dimensions[row].height = 26


def _set_section_title(sheet, row: int, last_column: int, text: str) -> None:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_column)
    cell = sheet.cell(row=row, column=1, value=text)
    cell.fill = _SECTION_FILL
    cell.font = Font(name="Courier New", size=11, bold=True, color="1F1F1F")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _SECTION_BORDER
    sheet.row_dimensions[row].height = 22


def _style_grid(
    sheet,
    min_row: int,
    max_row: int,
    min_column: int,
    max_column: int,
    *,
    number_format: str | None = None,
) -> None:
    for row in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_column,
        max_col=max_column,
    ):
        for cell in row:
            cell.border = _GRID_BORDER
            cell.font = _BODY_FONT
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            if number_format and cell.column > min_column:
                cell.number_format = number_format


def _style_matrix_header(
    sheet,
    header_row: int,
    columns: list[str],
    preferences: Mapping[str, CriterionPreference],
) -> None:
    alternative_cell = sheet.cell(header_row, 1)
    alternative_cell.fill = _ALTERNATIVE_FILL
    alternative_cell.font = _HEADER_FONT
    for offset, criterion in enumerate(columns, start=2):
        cell = sheet.cell(header_row, offset)
        cell.fill = _preference_fill(preferences[criterion])
        cell.font = _HEADER_FONT
    sheet.row_dimensions[header_row].height = 31


def _write_weights_row(
    sheet, row: int, columns: list[str], weights: Mapping[str, float]
) -> None:
    sheet.cell(row, 1, "weightage")
    for offset, criterion in enumerate(columns, start=2):
        sheet.cell(row, offset, float(weights[criterion]))
    for cell in sheet[row][: len(columns) + 1]:
        cell.fill = _WEIGHT_FILL
        cell.font = _HEADER_FONT
        cell.border = _GRID_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.column > 1:
            cell.number_format = _WEIGHT_FORMAT


def _write_indexed_dataframe(
    sheet,
    start_row: int,
    title: str,
    frame: pd.DataFrame,
    *,
    header_label: str = "Alternative",
    matrix_preferences: Mapping[str, CriterionPreference] | None = None,
) -> int:
    columns = [str(column) for column in frame.columns]
    last_column = len(columns) + 1
    _set_section_title(sheet, start_row, last_column, title)
    header_row = start_row + 1
    sheet.cell(header_row, 1, header_label)
    for column_index, column in enumerate(columns, start=2):
        _set_text(sheet, header_row, column_index, column)
    _style_grid(sheet, header_row, header_row + len(frame), 1, last_column)
    if matrix_preferences is None:
        for cell in sheet[header_row][:last_column]:
            cell.fill = _SECTION_FILL
            cell.font = _HEADER_FONT
    else:
        _style_matrix_header(sheet, header_row, columns, matrix_preferences)

    for row_offset, (index, values) in enumerate(frame.iterrows(), start=1):
        excel_row = header_row + row_offset
        _set_text(sheet, excel_row, 1, index)
        sheet.cell(excel_row, 1).fill = _ALTERNATIVE_FILL
        for column_index, value in enumerate(values, start=2):
            sheet.cell(excel_row, column_index, float(value))
            sheet.cell(excel_row, column_index).number_format = _DECIMAL_FORMAT
    return header_row + len(frame) + 2


def _autofit_column_widths(
    sheet,
    *,
    min_width: int = 15,
    padding: int = 4,
    max_width: int = 65,
) -> None:
    """Automatically widen all columns horizontally to fit text and headers without wrapping or clipping."""
    merged_ranges = list(sheet.merged_cells.ranges)

    def is_wide_merged_cell(row_idx: int, col_idx: int) -> bool:
        for rng in merged_ranges:
            if rng.min_row <= row_idx <= rng.max_row and rng.min_col <= col_idx <= rng.max_col:
                if rng.max_col > rng.min_col:
                    return True
        return False

    for col in sheet.columns:
        col_idx = col[0].column
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            if is_wide_merged_cell(cell.row, col_idx):
                continue
            val_str = str(cell.value)
            if val_str.startswith("="):
                val_len = 15
            else:
                lines = val_str.split("\n")
                val_len = max(len(line) for line in lines)
            if val_len > max_len:
                max_len = val_len
        calculated_width = max(min_width, min(max_width, max_len + padding))
        existing_width = sheet.column_dimensions[letter].width or 0
        sheet.column_dimensions[letter].width = max(existing_width, calculated_width)


def _build_formula_sheet(
    workbook: Workbook,
    frame: pd.DataFrame,
    weights: Mapping[str, float],
    preferences: Mapping[str, CriterionPreference],
    *,
    alpha: float,
    p: int,
) -> None:
    sheet = workbook.create_sheet("AURA")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 85

    columns = [str(column) for column in frame.columns]
    alternative_names = [str(index) for index in frame.index]
    last_matrix_column = len(columns) + 1
    last_matrix_letter = get_column_letter(last_matrix_column)
    model_last_column = max(9, last_matrix_column)
    alternatives_count = len(alternative_names)
    criteria_count = len(columns)

    _set_title(
        sheet,
        1,
        model_last_column,
        "AURA — Complete Formula-Driven Calculation Workbook",
    )
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=model_last_column)
    sheet.cell(
        2,
        1,
        "Orange rows contain normalized weights; light blue/red/orange headers represent "
        "benefit/cost/target criteria. Select any derived cell to inspect its Excel formula.",
    )
    sheet.cell(2, 1).font = Font(name="Courier New", size=9, italic=True, color="666666")
    sheet.cell(2, 1).alignment = Alignment(wrap_text=True)

    _set_section_title(sheet, 4, 4, "Model Parameters")
    parameter_rows = [
        (5, "Alpha (α)", float(alpha), "Balance parameter used in the utility score"),
        (6, "Distance metric (p)", int(p), "1 = Manhattan; 2 = Euclidean"),
        (7, "Alternatives", alternatives_count, "Number of decision alternatives"),
        (8, "Criteria", criteria_count, "Number of decision criteria"),
    ]
    for row, label, value, description in parameter_rows:
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
        sheet.cell(row, 3, description)
    sheet.cell(5, 2).number_format = _DECIMAL_FORMAT
    sheet.cell(5, 2).fill = _INPUT_FILL
    sheet.cell(6, 2).fill = _INPUT_FILL
    _style_grid(sheet, 5, 8, 1, 4)
    sheet.cell(5, 2).comment = Comment(
        "Changing alpha in Excel recalculates every utility score and rank.", "User"
    )
    sheet.cell(6, 2).comment = Comment(
        "Use 1 for Manhattan distance or 2 for Euclidean distance.", "User"
    )

    settings_title_row = 10
    settings_header_row = 11
    settings_data_start = 12
    settings_data_end = settings_data_start + criteria_count - 1

    raw_title_row = settings_data_end + 3
    raw_weights_row = raw_title_row + 1
    raw_header_row = raw_title_row + 2
    raw_data_start = raw_title_row + 3
    raw_data_end = raw_data_start + alternatives_count - 1
    raw_min_row = raw_data_end + 1
    raw_max_row = raw_data_end + 2
    raw_reference_row = raw_data_end + 3

    _set_section_title(sheet, settings_title_row, 4, "Criterion Settings")
    settings_headers = ["Criterion", "Normalized Weight", "Preference", "Reference r_j"]
    for column_index, value in enumerate(settings_headers, start=1):
        sheet.cell(settings_header_row, column_index, value)
    _style_grid(sheet, settings_header_row, settings_data_end, 1, 4)
    for cell in sheet[settings_header_row][:4]:
        cell.fill = _SECTION_FILL
        cell.font = _HEADER_FONT

    for criterion_index, criterion in enumerate(columns):
        row = settings_data_start + criterion_index
        matrix_column = get_column_letter(criterion_index + 2)
        preference = preferences[criterion]
        _set_text(sheet, row, 1, criterion)
        sheet.cell(row, 2, float(weights[criterion]))
        sheet.cell(row, 2).number_format = _WEIGHT_FORMAT
        sheet.cell(row, 3, _preference_label(preference))
        sheet.cell(row, 3).fill = _preference_fill(preference)
        if preference.kind is CriterionType.TARGET:
            sheet.cell(row, 4, float(preference.target_value))
        elif preference.kind is CriterionType.BENEFIT:
            sheet.cell(row, 4, f"={matrix_column}${raw_max_row}")
        else:
            sheet.cell(row, 4, f"={matrix_column}${raw_min_row}")
        sheet.cell(row, 4).number_format = _DECIMAL_FORMAT

    _set_section_title(sheet, raw_title_row, last_matrix_column, "Step 0 — Original Decision Matrix")
    _write_weights_row(sheet, raw_weights_row, columns, weights)
    sheet.cell(raw_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, raw_header_row, column_index, criterion)
    _style_grid(
        sheet,
        raw_header_row,
        raw_reference_row,
        1,
        last_matrix_column,
        number_format=_DECIMAL_FORMAT,
    )
    _style_matrix_header(sheet, raw_header_row, columns, preferences)
    for row_offset, (alternative, values) in enumerate(frame.iterrows()):
        row = raw_data_start + row_offset
        _set_text(sheet, row, 1, alternative)
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index, value in enumerate(values, start=2):
            sheet.cell(row, column_index, float(value))
            sheet.cell(row, column_index).fill = _WHITE_FILL
    for summary_row, label in [
        (raw_min_row, "Minimum"),
        (raw_max_row, "Maximum"),
        (raw_reference_row, "Reference r_j"),
    ]:
        sheet.cell(summary_row, 1, label)
        for column_index in range(1, last_matrix_column + 1):
            sheet.cell(summary_row, column_index).fill = _SUMMARY_FILL
            sheet.cell(summary_row, column_index).font = _HEADER_FONT
    for column_index in range(2, last_matrix_column + 1):
        letter = get_column_letter(column_index)
        sheet.cell(
            raw_min_row,
            column_index,
            f"=MIN({letter}${raw_data_start}:{letter}${raw_data_end})",
        )
        sheet.cell(
            raw_max_row,
            column_index,
            f"=MAX({letter}${raw_data_start}:{letter}${raw_data_end})",
        )
        settings_row = settings_data_start + column_index - 2
        sheet.cell(raw_reference_row, column_index, f"=$D${settings_row}")

    normalized_title_row = raw_reference_row + 3
    normalized_weights_row = normalized_title_row + 1
    normalized_header_row = normalized_title_row + 2
    normalized_data_start = normalized_title_row + 3
    normalized_data_end = normalized_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        normalized_title_row,
        last_matrix_column,
        "Step 1 — Normalized Decision Matrix (r_ij)",
    )
    _write_weights_row(sheet, normalized_weights_row, columns, weights)
    sheet.cell(normalized_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, normalized_header_row, column_index, criterion)
    _style_grid(
        sheet,
        normalized_header_row,
        normalized_data_end,
        1,
        last_matrix_column,
        number_format=_DECIMAL_FORMAT,
    )
    _style_matrix_header(sheet, normalized_header_row, columns, preferences)
    for row_offset, alternative in enumerate(alternative_names):
        row = normalized_data_start + row_offset
        raw_row = raw_data_start + row_offset
        sheet.cell(row, 1, f"=A{raw_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index in range(2, last_matrix_column + 1):
            letter = get_column_letter(column_index)
            formula = (
                f"=IF(MAX({letter}${raw_data_start}:{letter}${raw_data_end})="
                f"MIN({letter}${raw_data_start}:{letter}${raw_data_end}),1,"
                f"1-(ABS({letter}{raw_row}-{letter}${raw_reference_row})/"
                f"MAX(MAX({letter}${raw_data_start}:{letter}${raw_data_end})-"
                f"MIN({letter}${raw_data_start}:{letter}${raw_data_end}),1E-9)))"
            )
            sheet.cell(row, column_index, formula)
    sheet.cell(normalized_data_start, 2).comment = Comment(
        "A live Excel implementation of r_ij = 1 - |x_ij-r_j| / "
        "max(max(x_j)-min(x_j), 1E-9). Constant columns return 1.",
        "User",
    )

    weighted_title_row = normalized_data_end + 3
    weighted_weights_row = weighted_title_row + 1
    weighted_header_row = weighted_title_row + 2
    weighted_data_start = weighted_title_row + 3
    weighted_data_end = weighted_data_start + alternatives_count - 1
    pis_row = weighted_data_end + 1
    nis_row = weighted_data_end + 2
    average_row = weighted_data_end + 3

    _set_section_title(
        sheet,
        weighted_title_row,
        last_matrix_column,
        "Step 2 & 3 — Weighted Matrix and Ideal Solutions",
    )
    _write_weights_row(sheet, weighted_weights_row, columns, weights)
    sheet.cell(weighted_header_row, 1, "Alternative")
    for column_index, criterion in enumerate(columns, start=2):
        _set_text(sheet, weighted_header_row, column_index, criterion)
    _style_grid(
        sheet,
        weighted_header_row,
        average_row,
        1,
        last_matrix_column,
        number_format=_DECIMAL_FORMAT,
    )
    _style_matrix_header(sheet, weighted_header_row, columns, preferences)
    for row_offset, alternative in enumerate(alternative_names):
        row = weighted_data_start + row_offset
        normalized_row = normalized_data_start + row_offset
        sheet.cell(row, 1, f"=A{normalized_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        for column_index in range(2, last_matrix_column + 1):
            letter = get_column_letter(column_index)
            sheet.cell(
                row,
                column_index,
                f"={letter}{normalized_row}*{letter}${weighted_weights_row}",
            )
    for solution_row, label, function_name in [
        (pis_row, "PIS", "MAX"),
        (nis_row, "NIS", "MIN"),
        (average_row, "AS (Average)", "AVERAGE"),
    ]:
        sheet.cell(solution_row, 1, label)
        for column_index in range(1, last_matrix_column + 1):
            sheet.cell(solution_row, column_index).fill = _SUMMARY_FILL
            sheet.cell(solution_row, column_index).font = _HEADER_FONT
        for column_index in range(2, last_matrix_column + 1):
            letter = get_column_letter(column_index)
            sheet.cell(
                solution_row,
                column_index,
                f"={function_name}({letter}${weighted_data_start}:"
                f"{letter}${weighted_data_end})",
            )

    distances_title_row = average_row + 3
    sigma_row = distances_title_row + 1
    distances_header_row = distances_title_row + 3
    distances_data_start = distances_title_row + 4
    distances_data_end = distances_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        distances_title_row,
        9,
        "Step 4 & 5 — Distances, Correction, Utility Score, and Rank",
    )
    sigma_labels = [(1, "Correction Factors (σ)"), (2, "Sigma+"), (4, "Sigma-"), (6, "Sigma avg")]
    for column_index, label in sigma_labels:
        sheet.cell(sigma_row, column_index, label)
        sheet.cell(sigma_row, column_index).fill = _SECTION_FILL
        sheet.cell(sigma_row, column_index).font = _HEADER_FONT
    sheet.cell(
        sigma_row,
        3,
        f"=MAX(B${distances_data_start}:B${distances_data_end})-"
        f"MIN(B${distances_data_start}:B${distances_data_end})",
    )
    sheet.cell(
        sigma_row,
        5,
        f"=MAX(C${distances_data_start}:C${distances_data_end})-"
        f"MIN(C${distances_data_start}:C${distances_data_end})",
    )
    sheet.cell(
        sigma_row,
        7,
        f"=MAX(D${distances_data_start}:D${distances_data_end})-"
        f"MIN(D${distances_data_start}:D${distances_data_end})",
    )
    _style_grid(sheet, sigma_row, sigma_row, 1, 7, number_format=_DECIMAL_FORMAT)

    distance_headers = [
        "Alternative",
        "d+ raw",
        "d- raw",
        "d_avg raw",
        "D+ corrected",
        "D- corrected",
        "D_avg corrected",
        "Utility Score",
        "Rank",
    ]
    for column_index, header in enumerate(distance_headers, start=1):
        sheet.cell(distances_header_row, column_index, header)
        sheet.cell(distances_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(distances_header_row, column_index).font = _HEADER_FONT
    _style_grid(
        sheet,
        distances_header_row,
        distances_data_end,
        1,
        9,
        number_format=_DECIMAL_FORMAT,
    )
    sheet.row_dimensions[distances_header_row].height = 31

    weighted_first_letter = "B"
    weighted_last_letter = last_matrix_letter
    for row_offset, alternative in enumerate(alternative_names):
        row = distances_data_start + row_offset
        weighted_row = weighted_data_start + row_offset
        sheet.cell(row, 1, f"=A{weighted_row}")
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        weighted_range = (
            f"{weighted_first_letter}{weighted_row}:{weighted_last_letter}{weighted_row}"
        )
        for distance_column, solution_row in [(2, pis_row), (3, nis_row), (4, average_row)]:
            reference_range = (
                f"${weighted_first_letter}${solution_row}:"
                f"${weighted_last_letter}${solution_row}"
            )
            formula = (
                f"=IF($B$6=1,SUMPRODUCT(ABS({weighted_range}-{reference_range})),"
                f"SQRT(SUMPRODUCT(({weighted_range}-{reference_range})*"
                f"({weighted_range}-{reference_range}))))"
            )
            sheet.cell(row, distance_column, formula)
        sheet.cell(row, 5, f"=B{row}+$C${sigma_row}*B{row}^2")
        sheet.cell(row, 6, f"=C{row}+$E${sigma_row}*C{row}^2")
        sheet.cell(row, 7, f"=D{row}+$G${sigma_row}*D{row}^2")
        sheet.cell(
            row,
            8,
            f"=($B$5*(E{row}-F{row})+(1-$B$5)*G{row})/2",
        )
        sheet.cell(
            row,
            9,
            f"=RANK(H{row},$H${distances_data_start}:$H${distances_data_end},1)+"
            f"COUNTIF($H${distances_data_start}:H{row},H{row})-1",
        )
        sheet.cell(row, 9).number_format = "0"
    sheet.cell(distances_data_start, 8).comment = Comment(
        "Lower AURA utility scores are preferred. Alpha is read from B5.", "User"
    )

    ranking_title_row = distances_data_end + 3
    ranking_header_row = ranking_title_row + 1
    ranking_data_start = ranking_title_row + 2
    ranking_data_end = ranking_data_start + alternatives_count - 1

    _set_section_title(
        sheet,
        ranking_title_row,
        6,
        "Final Ranking — Sorted by Rank",
    )
    ranking_headers = [
        "Alternative",
        "D+ (PIS)",
        "D- (NIS)",
        "D_avg (AS)",
        "Utility Score",
        "Rank",
    ]
    for column_index, header in enumerate(ranking_headers, start=1):
        sheet.cell(ranking_header_row, column_index, header)
        sheet.cell(ranking_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(ranking_header_row, column_index).font = _HEADER_FONT
    _style_grid(
        sheet,
        ranking_header_row,
        ranking_data_end,
        1,
        6,
        number_format=_DECIMAL_FORMAT,
    )
    sheet.row_dimensions[ranking_header_row].height = 31

    for rank_idx in range(1, alternatives_count + 1):
        row = ranking_data_start + rank_idx - 1
        match_formula = (
            f"MATCH({rank_idx},$I${distances_data_start}:$I${distances_data_end},0)"
        )
        sheet.cell(
            row,
            1,
            f"=INDEX($A${distances_data_start}:$A${distances_data_end},{match_formula})",
        )
        sheet.cell(row, 1).fill = _ALTERNATIVE_FILL
        sheet.cell(
            row,
            2,
            f"=INDEX($E${distances_data_start}:$E${distances_data_end},{match_formula})",
        )
        sheet.cell(
            row,
            3,
            f"=INDEX($F${distances_data_start}:$F${distances_data_end},{match_formula})",
        )
        sheet.cell(
            row,
            4,
            f"=INDEX($G${distances_data_start}:$G${distances_data_end},{match_formula})",
        )
        sheet.cell(
            row,
            5,
            f"=INDEX($H${distances_data_start}:$H${distances_data_end},{match_formula})",
        )
        sheet.cell(row, 6, rank_idx)
        sheet.cell(row, 6).font = _HEADER_FONT
        sheet.cell(row, 6).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row, 6).number_format = "0"

    # Keep the alternative labels visible without pinning the whole input
    # section above the matrix, which otherwise looks like a split screen.
    sheet.freeze_panes = "B1"
    sheet.print_area = f"A1:{get_column_letter(model_last_column)}{ranking_data_end}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.oddFooter.center.text = "AURA complete formula model"
    sheet.oddFooter.right.text = "Page &P of &N"

    sheet.column_dimensions["A"].width = max(
        22, min(34, max(len(value) for value in alternative_names) + 3)
    )
    for column_index, criterion in enumerate(columns, start=2):
        letter = get_column_letter(column_index)
        raw_value_width = max(
            len(f"{float(value):.9f}") for value in frame[criterion]
        )
        sheet.column_dimensions[letter].width = max(
            15,
            min(30, max(len(criterion) + 3, raw_value_width + 2)),
        )
    for column_index in range(last_matrix_column + 1, 10):
        sheet.column_dimensions[get_column_letter(column_index)].width = 17
    for row in range(1, ranking_data_end + 1):
        if sheet.row_dimensions[row].height is None:
            sheet.row_dimensions[row].height = 18

    _autofit_column_widths(sheet)


def _build_verified_values_sheet(
    workbook: Workbook,
    frame: pd.DataFrame,
    weights: Mapping[str, float],
    preferences: Mapping[str, CriterionPreference],
    results: pd.DataFrame,
    steps: Mapping[str, Any],
    *,
    alpha: float,
    p: int,
) -> None:
    sheet = workbook.create_sheet("Verified Values")
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 90
    max_columns = max(len(frame.columns) + 1, 9)
    _set_title(
        sheet,
        1,
        max_columns,
        "AURA — Verified Numerical Values",
    )
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_columns)
    sheet.cell(
        2,
        1,
        "Static values calculated by the canonical Python implementation at export time. "
        "Use these values to audit the live Excel formulas on the AURA sheet.",
    )
    sheet.cell(2, 1).font = Font(name="Courier New", size=9, italic=True, color="666666")
    sheet.cell(2, 1).alignment = Alignment(wrap_text=True)

    _set_section_title(sheet, 4, 4, "Parameters")
    parameter_values = [
        ["Alpha (α)", float(alpha), "Distance p", int(p)],
        ["Alternatives", len(frame), "Criteria", len(frame.columns)],
    ]
    for row_offset, values in enumerate(parameter_values, start=5):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row_offset, column_index, value)
    _style_grid(sheet, 5, 6, 1, 4)

    settings_title_row = 8
    settings_header_row = 9
    settings_data_start = 10
    _set_section_title(sheet, settings_title_row, 4, "Criterion Settings")
    settings_headers = ["Criterion", "Weight", "Preference", "Reference"]
    for column_index, header in enumerate(settings_headers, start=1):
        sheet.cell(settings_header_row, column_index, header)
        sheet.cell(settings_header_row, column_index).fill = _SECTION_FILL
        sheet.cell(settings_header_row, column_index).font = _HEADER_FONT
    for row_offset, criterion in enumerate(frame.columns, start=settings_data_start):
        name = str(criterion)
        preference = preferences[name]
        _set_text(sheet, row_offset, 1, name)
        sheet.cell(row_offset, 2, float(weights[name]))
        _set_text(sheet, row_offset, 3, _preference_label(preference))
        sheet.cell(row_offset, 3).fill = _preference_fill(preference)
        reference = (
            float(preference.target_value)
            if preference.kind is CriterionType.TARGET
            else float(frame[criterion].max())
            if preference.kind is CriterionType.BENEFIT
            else float(frame[criterion].min())
        )
        sheet.cell(row_offset, 4, reference)
    settings_data_end = settings_data_start + len(frame.columns) - 1
    _style_grid(
        sheet,
        settings_header_row,
        settings_data_end,
        1,
        4,
        number_format=_DECIMAL_FORMAT,
    )
    sheet.column_dimensions["A"].width = max(
        22,
        min(
            34,
            max(
                max(len(str(value)) for value in frame.index),
                max(len(str(value)) for value in frame.columns),
            )
            + 3,
        ),
    )

    next_row = settings_data_end + 3
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 0 — Original Decision Matrix",
        frame,
        matrix_preferences=preferences,
    )
    normalized = steps["Step 1: Normalized Decision Matrix"].reindex(frame.index)
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 1 — Normalized Decision Matrix",
        normalized,
        matrix_preferences=preferences,
    )
    weighted = steps["Step 2: Weighted Normalized Matrix"].reindex(frame.index)
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 2 — Weighted Normalized Matrix",
        weighted,
        matrix_preferences=preferences,
    )

    ideal_values = pd.DataFrame(
        [
            steps["Step 3: Ideal Solutions"]["PIS (Positive Ideal Solution)"],
            steps["Step 3: Ideal Solutions"]["NIS (Negative Ideal Solution)"],
            steps["Step 3: Ideal Solutions"]["AS (Average Solution)"],
        ],
        index=["PIS", "NIS", "AS (Average)"],
    )
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 3 — Ideal Solutions",
        ideal_values,
        header_label="Solution",
        matrix_preferences=preferences,
    )

    raw_distances = steps["Step 4a: Raw Distances"].reindex(frame.index)
    corrected = steps["Step 4b: Corrected Distances"].reindex(frame.index)
    canonical_results = results.reindex(frame.index)
    distance_values = pd.concat(
        [
            raw_distances,
            corrected.rename(
                columns={
                    "D_plus": "D+ corrected",
                    "D_minus": "D- corrected",
                    "D_avg": "D_avg corrected",
                }
            ),
            canonical_results[["Utility Score", "Rank"]],
        ],
        axis=1,
    )
    next_row = _write_indexed_dataframe(
        sheet,
        next_row,
        "Step 4 & 5 — Distances, Corrected Values, Utility, and Rank",
        distance_values,
    )
    rank_column = distance_values.columns.get_loc("Rank") + 2
    data_start = next_row - len(distance_values) - 1
    for row in range(data_start, data_start + len(distance_values)):
        sheet.cell(row, rank_column).number_format = "0"

    _set_section_title(sheet, next_row, 4, "Correction Factors")
    correction_factors = steps["Step 4b: Correction Factors"]
    sheet.cell(next_row + 1, 1, "Sigma+")
    sheet.cell(next_row + 1, 2, float(correction_factors["Sigma+"]))
    sheet.cell(next_row + 1, 3, "Sigma-")
    sheet.cell(next_row + 1, 4, float(correction_factors["Sigma-"]))
    sheet.cell(next_row + 2, 1, "Sigma avg")
    sheet.cell(next_row + 2, 2, float(correction_factors["Sigma_avg"]))
    _style_grid(sheet, next_row + 1, next_row + 2, 1, 4, number_format=_DECIMAL_FORMAT)

    final_start = next_row + 4
    _write_indexed_dataframe(
        sheet,
        final_start,
        "Final Ranking — Sorted by Rank",
        results[["D+ (PIS)", "D- (NIS)", "D_avg (AS)", "Utility Score", "Rank"]],
    )
    final_rank_col = 6
    final_data_start = final_start + 2
    for row in range(final_data_start, final_data_start + len(results)):
        sheet.cell(row, final_rank_col).number_format = "0"

    for column_index in range(2, max_columns + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 18
    _autofit_column_widths(sheet)


def _build_formula_guide_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Formula Guide")
    sheet.sheet_view.showGridLines = False
    _set_title(sheet, 1, 5, "AURA Formula Guide and Audit Trail")
    sheet.merge_cells("A2:E2")
    sheet["A2"] = (
        "The AURA sheet contains live Excel formulas. The Verified Values sheet contains "
        "the canonical numerical snapshot generated at download time."
    )
    sheet["A2"].alignment = Alignment(wrap_text=True)
    sheet["A2"].font = Font(name="Courier New", size=9, italic=True, color="666666")

    headers = ["Step", "Symbol", "Mathematical formula", "Meaning", "Excel implementation"]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(4, column_index, header)
        sheet.cell(4, column_index).fill = _SECTION_FILL
        sheet.cell(4, column_index).font = _HEADER_FONT

    rows = [
        (
            "1",
            "r_ij",
            "1 - |x_ij - r_j| / max(max(x_j)-min(x_j), 1E-9)",
            "Normalize each criterion around its benefit, cost, or target reference.",
            "IF + ABS + MAX + MIN",
        ),
        (
            "2",
            "v_ij",
            "r_ij × w_j",
            "Apply the normalized criterion weight.",
            "Normalized cell × weight cell",
        ),
        (
            "3",
            "PIS_j",
            "max_i(v_ij)",
            "Positive ideal solution for each criterion.",
            "MAX(weighted criterion range)",
        ),
        (
            "3",
            "NIS_j",
            "min_i(v_ij)",
            "Negative ideal solution for each criterion.",
            "MIN(weighted criterion range)",
        ),
        (
            "3",
            "AS_j",
            "mean_i(v_ij)",
            "Average solution for each criterion.",
            "AVERAGE(weighted criterion range)",
        ),
        (
            "4",
            "d_i",
            "Σ|v_ij-c_j| when p=1; sqrt(Σ(v_ij-c_j)^2) when p=2",
            "Raw distance to PIS, NIS, or AS.",
            "IF + SUMPRODUCT + ABS / SQRT",
        ),
        (
            "4",
            "σ",
            "max_i(d_i) - min_i(d_i)",
            "Spread-based AURA correction factor.",
            "MAX(distance range) - MIN(distance range)",
        ),
        (
            "4",
            "D_i",
            "d_i + σ d_i²",
            "Corrected distance.",
            "Raw distance + sigma × raw distance^2",
        ),
        (
            "5",
            "U_i",
            "[α(D+_i-D-_i) + (1-α)Davg_i] / 2",
            "AURA utility score; lower values are preferred.",
            "References the alpha parameter cell",
        ),
        (
            "5",
            "Rank_i",
            "ascending competition rank of U_i",
            "Rank 1 is the smallest utility score; ties retain the same rank.",
            "RANK(utility, utility range, 1)",
        ),
    ]
    for row_index, values in enumerate(rows, start=5):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row_index, column_index, value)
    _style_grid(sheet, 4, 4 + len(rows), 1, 5)
    for row in range(5, 5 + len(rows)):
        sheet.row_dimensions[row].height = 39

    source_row = 6 + len(rows)
    _set_section_title(sheet, source_row, 5, "Method Reference")
    sheet.cell(source_row + 1, 1, "AURA paper DOI")
    sheet.cell(source_row + 1, 2, "https://doi.org/10.1016/j.softx.2025.102395")
    sheet.merge_cells(
        start_row=source_row + 1,
        start_column=2,
        end_row=source_row + 1,
        end_column=5,
    )
    sheet.cell(source_row + 1, 2).hyperlink = (
        "https://doi.org/10.1016/j.softx.2025.102395"
    )
    sheet.cell(source_row + 1, 2).style = "Hyperlink"

    widths = {"A": 10, "B": 15, "C": 48, "D": 54, "E": 40}
    for letter, width in widths.items():
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A5"


def build_aura_excel_workbook(
    matrix: pd.DataFrame,
    weights: Mapping[str, Any],
    directions: Mapping[str, Any],
    *,
    alpha: float = 0.5,
    p: int = 2,
) -> bytes:
    """Return a complete AURA workbook with live formulas and verified values."""

    frame = validate_crisp_matrix(matrix)
    frame = frame.copy()
    frame.columns = [str(column) for column in frame.columns]
    normalized_weights = validate_weights(weights, frame.columns, normalize=True)
    preferences = validate_method_capabilities("AURA", frame.columns, directions)
    results, steps = calculate_aura(
        frame,
        normalized_weights,
        directions,
        float(alpha),
        int(p),
        return_steps=True,
    )

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "AURA MCDM Application"
    workbook.properties.title = "Complete AURA Formula-Driven Calculation"
    workbook.properties.subject = "Auditable AURA decision model"
    workbook.properties.version = AURA_EXCEL_EXPORT_REVISION
    workbook.properties.description = (
        "Live Excel formulas plus canonical numerical values for every AURA "
        f"calculation step. Export revision {AURA_EXCEL_EXPORT_REVISION}."
    )
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    _build_formula_sheet(
        workbook,
        frame,
        normalized_weights,
        preferences,
        alpha=float(alpha),
        p=int(p),
    )
    _build_verified_values_sheet(
        workbook,
        frame,
        normalized_weights,
        preferences,
        results,
        steps,
        alpha=float(alpha),
        p=int(p),
    )
    _build_formula_guide_sheet(workbook)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


__all__ = [
    "AURA_EXCEL_EXPORT_FILENAME",
    "AURA_EXCEL_EXPORT_REVISION",
    "build_aura_excel_workbook",
]
