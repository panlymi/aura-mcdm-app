import openpyxl
import sys

def main():
    try:
        wb = openpyxl.load_workbook(r'C:\Users\fadzl\Desktop\MCDM\Fuzzy ARAS_Solved.xlsx', data_only=False)
        ws = wb.active

        wb_val = openpyxl.load_workbook(r'C:\Users\fadzl\Desktop\MCDM\Fuzzy ARAS_Solved.xlsx', data_only=True)
        ws_val = wb_val.active

        with open(r'C:\Users\fadzl\Desktop\MCDM\out.txt', 'w', encoding='utf-8') as f:
            for row in range(1, ws.max_row + 1):
                row_data = []
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell_val = ws_val.cell(row=row, column=col)
                    
                    if cell.value is not None:
                        if str(cell.value).startswith('='):
                            val = f"Formula: {cell.value} (Val: {cell_val.value})"
                        else:
                            val = str(cell.value)
                        row_data.append(val)
                    else:
                        row_data.append("")
                
                # Print if row is not completely empty
                if any(row_data):
                    f.write(f"Row {row}: {row_data}\n")
    except Exception as e:
        with open(r'C:\Users\fadzl\Desktop\MCDM\out.txt', 'w', encoding='utf-8') as f:
            f.write(f"Error: {e}\n")

if __name__ == "__main__":
    main()
